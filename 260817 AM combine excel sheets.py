"""
260817 AM combine excel sheets.py
---------------------------------
Combines every data sheet from every Excel/CSV file in a folder into one
de-duplicated dataset, written out as one or more .xlsx parts.

HOW TO RUN
    Open in IDLE and press F5. No command line arguments are needed - you will
    be asked to pick the input folder and then the output folder.
    (Optional: python "260817 AM combine excel sheets.py" <input folder> <output folder>)

RULES IMPLEMENTED
    1. Everything is unhidden first - hidden sheets, hidden rows, hidden columns.
       Hidden content IS included in the combine.
    2. A sheet is SKIPPED if it contains a pivot table.
    3. A sheet is SKIPPED if it contains a chart / graph (including chart sheets).
    4. A sheet is SKIPPED if the header is not in row 1 (see HEADER RULE below).
    5. Headers are matched after trimming, upper/lower-casing and stripping
       special characters, so these land in the SAME column:
           Per Num / Person Number / Per Number / Person Num / Per Nbr   -> Person Number
           Pat Name / Patient Name / Pt Name / Pt.Name                   -> Patient Name
           Encounter Number / Enc Num / Encounter / Enc Nbr              -> Encounter Number
           Date of Service / Enc Dt / Enc Date / Encounter Dt / Servc Dt -> Date of Service
       Every other column is kept too, matched on the same normalised name.
    6. Duplicates are removed per FILE (across all of that file's sheets), then
       once more across the final combined result.
    7. The result is split into parts if it exceeds the Excel row limit.

HEADER RULE (my criteria - a sheet must pass ALL of these or it is skipped)
    * The first row that contains anything must be row 1.
    * Row 1 must have at least 2 filled cells.
    * At least 2 of those cells must be text containing a letter, and at least
      60% of the filled cells must be text (not numbers/dates).
    * No header cell longer than 120 characters.
    * There must be at least one data row under the header.

FILE TYPES
    .xlsx .xlsm .xltx .xltm  - full support (pivot + chart detection)
    .xls                     - data only; Excel 97-2003 does not expose pivot /
                               chart info to Python, so those sheets cannot be
                               auto-detected (they are logged as a warning)
    .xlsb                    - data only, needs the 'pyxlsb' package
    .csv .txt                - single sheet, same header rule

REQUIREMENTS
    pip install pandas openpyxl xlrd
    (optional: pip install pyxlsb   - only needed for .xlsb files)

DATA HANDLING NOTE
    This script only reads and writes files on this machine. If the source data
    contains PHI or other confidential information, keep the input and output
    folders inside the approved Global Insider location - not the desktop, not a
    personal folder, and never paste rows into a chat or email.
"""

from __future__ import annotations

import csv
import datetime as dt
import os
import re
import sys
import traceback
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

# ---------------------------------------------------------------------------
# CONFIGURATION - edit these
# ---------------------------------------------------------------------------

AUTHOR_INITIALS = "AM"          # used in the output file names

RECURSE_SUBFOLDERS = True       # also read files in sub-folders of the input folder

# Rows per output part. Excel's hard limit is 1,048,576 including the header.
MAX_ROWS_PER_PART = 1_000_000

ADD_SOURCE_COLUMNS = True       # add 'Source File' and 'Source Sheet' columns

DEDUPE_ON_KEY_COLUMNS = True    # True  = duplicate means the 4 key columns match
                                # False = duplicate means the whole row matches

DEDUPE_ACROSS_FILES = True      # also remove duplicates across the combined result

# Save unhidden copies of the source workbooks to an "_unhidden" sub-folder of
# the output folder. OFF by default: unhidden data is combined either way, and
# turning this on creates a second copy of the source data on disk.
SAVE_UNHIDDEN_COPIES = False

# Header rule thresholds
MIN_HEADER_CELLS = 2
MIN_TEXT_HEADERS = 2
MIN_TEXT_RATIO = 0.6
MAX_HEADER_LEN = 120

# Header groups. Add your own groups / aliases here - the left side is the
# output column name, the right side is the list of accepted spellings.
CANONICAL_GROUPS = {
    "Person Number": [
        "per num", "per number", "per nbr", "per no", "per #",
        "person num", "person number", "person nbr", "person no",
        "pers num", "pers no", "pers nbr", "personnum", "pernum", "pernbr",
    ],
    "Patient Name": [
        "pat name", "patient name", "pt name", "pt.name", "ptname",
        "pat nm", "patient nm", "pt nm", "patientname", "patname",
    ],
    "Encounter Number": [
        "encounter number", "encounter num", "encounter nbr", "encounter no",
        "enc num", "enc number", "enc nbr", "enc no", "enc #",
        "encounter", "encntr num", "encounter id", "enc id", "encounternumber",
    ],
    "Date of Service": [
        "date of service", "dt of service", "service date", "servc dt",
        "serv dt", "svc dt", "svc date", "enc dt", "enc date",
        "encounter dt", "encounter date", "dos", "dos date", "date of svc",
    ],
}

KEY_COLUMNS = ["Person Number", "Patient Name", "Encounter Number", "Date of Service"]

DATA_EXTENSIONS = {
    ".xlsx", ".xlsm", ".xltx", ".xltm",   # openpyxl
    ".xls",                               # xlrd
    ".xlsb",                              # pyxlsb
    ".csv", ".txt",                       # plain text
}

OPENPYXL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TEXT_EXTENSIONS = {".csv", ".txt"}

# ---------------------------------------------------------------------------
# Imports that need a friendly error message
# ---------------------------------------------------------------------------

try:
    import pandas as pd
except ImportError:                                          # pragma: no cover
    sys.exit("pandas is not installed.  Run:  pip install pandas openpyxl xlrd")

try:
    import openpyxl
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.chartsheet import Chartsheet
except ImportError:                                          # pragma: no cover
    sys.exit("openpyxl is not installed.  Run:  pip install openpyxl")


# ---------------------------------------------------------------------------
# Header normalisation
# ---------------------------------------------------------------------------

_SPECIALS_RE = re.compile(r"[^0-9a-z]+")
_WS_RE = re.compile(r"\s+")

# alias lookup with all spaces removed, so "pt.name" == "pt name" == "ptname"
_ALIAS_LOOKUP: dict[str, str] = {}
for _canon, _aliases in CANONICAL_GROUPS.items():
    _ALIAS_LOOKUP[_canon.lower().replace(" ", "")] = _canon
    for _a in _aliases:
        _ALIAS_LOOKUP[_SPECIALS_RE.sub("", _a.lower())] = _canon

_PERSON_WORDS = {"per", "pers", "person", "persons", "prsn"}
_PATIENT_WORDS = {"pat", "patient", "patients", "pt", "pts"}
_ENCOUNTER_WORDS = {"enc", "encntr", "encounter", "encounters", "encntrs"}
_NUMBER_WORDS = {"num", "no", "nbr", "number", "nmbr", "numb", "id", "nos"}
_NAME_WORDS = {"name", "nm", "names", "nme"}
_DATE_WORDS = {"date", "dt", "dte", "dates"}
_SERVICE_WORDS = {"service", "servc", "svc", "serv", "srvc", "enc", "encounter", "encntr", "dos"}


def normalise_header(raw) -> str:
    """Trim, lower-case and strip special characters: 'Pt.Name ' -> 'pt name'."""
    if raw is None:
        return ""
    text = str(raw).replace(" ", " ").strip().lower()
    text = _SPECIALS_RE.sub(" ", text)
    return _WS_RE.sub(" ", text).strip()


def canonical_header(raw) -> str | None:
    """Return the canonical column name for a header, or None if it isn't one of them."""
    norm = normalise_header(raw)
    if not norm:
        return None

    exact = _ALIAS_LOOKUP.get(norm.replace(" ", ""))
    if exact:
        return exact

    tokens = set(norm.split())

    # Date of service first - "enc dt" must not be read as an encounter number.
    if "dos" in tokens and len(tokens) <= 2:
        return "Date of Service"
    if tokens & _DATE_WORDS and tokens & _SERVICE_WORDS:
        return "Date of Service"

    if tokens & _PERSON_WORDS and tokens & _NUMBER_WORDS:
        return "Person Number"
    if tokens & _PATIENT_WORDS and tokens & _NAME_WORDS:
        return "Patient Name"
    if tokens & _ENCOUNTER_WORDS and tokens & _NUMBER_WORDS:
        return "Encounter Number"

    return None


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, float):
        return value != value          # NaN
    return False


def clean_cell(value):
    """Trim strings, turn empties into None, leave everything else alone."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace(" ", " ").strip()
        return text or None
    if isinstance(value, float) and value != value:
        return None
    return value


def normalise_value(value) -> str:
    """Normalise a cell for duplicate comparison."""
    if is_blank(value):
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and float(value).is_integer():
        return str(int(value))
    text = _WS_RE.sub(" ", str(value).strip()).upper()
    # "12345.0" and "12345" are the same person number
    if re.fullmatch(r"-?\d+\.0+", text):
        text = text.split(".")[0]
    return text


def normalise_date_value(value) -> str:
    """Normalise a date cell so 01/03/2026 and 2026-03-01 compare equal."""
    if is_blank(value):
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=False)
    if parsed is not pd.NaT and not pd.isna(parsed):
        return parsed.strftime("%Y-%m-%d")
    return normalise_value(value)


def trim_grid(grid: list[list]) -> list[list]:
    """Drop trailing blank rows and trailing blank columns."""
    while grid and all(is_blank(c) for c in grid[-1]):
        grid.pop()
    if not grid:
        return grid
    width = 0
    for row in grid:
        for idx in range(len(row) - 1, -1, -1):
            if not is_blank(row[idx]):
                width = max(width, idx + 1)
                break
    return [list(row[:width]) + [None] * (width - len(row[:width])) for row in grid]


# ---------------------------------------------------------------------------
# Pivot table detection straight from the .xlsx package (belt and braces)
# ---------------------------------------------------------------------------

_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_PKG_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"


def _resolve_part(base_dir: str, target: str) -> str:
    """Resolve a relationship target to a path inside the package."""
    target = (target or "").replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    return os.path.normpath(f"{base_dir}/{target}").replace("\\", "/")


def _rels_path(part: str) -> str:
    return f"{os.path.dirname(part)}/_rels/{os.path.basename(part)}.rels"


def analyse_package(path: Path) -> dict[str, dict] | None:
    """
    Read the workbook package (a .xlsx is a zip) and report, per sheet:
        hidden      - the sheet is hidden or very hidden
        chartsheet  - the whole sheet is a chart
        pivot       - a pivot table part is attached to the sheet
        chart       - the sheet's drawing contains a chart object

    Reading the package directly means detection does not depend on openpyxl
    being able to parse pivot/chart definitions, and it lets the values be read
    in low-memory mode afterwards. Returns None if the file is not a readable
    package, in which case the caller falls back to a full openpyxl load.
    """
    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
            if "xl/workbook.xml" not in names:
                return None

            rel_targets: dict[str, str] = {}
            if "xl/_rels/workbook.xml.rels" in names:
                root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                for rel in root.findall(f"{_PKG_REL_NS}Relationship"):
                    rel_targets[rel.get("Id")] = _resolve_part("xl", rel.get("Target"))

            info: dict[str, dict] = {}
            wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
            for sheet in wb_root.iter(f"{_MAIN_NS}sheet"):
                sheet_name = sheet.get("name")
                if not sheet_name:
                    continue
                part = rel_targets.get(sheet.get(f"{_REL_NS}id"), "")
                entry = {
                    "part": part,
                    "hidden": (sheet.get("state") or "visible") != "visible",
                    "chartsheet": "/chartsheets/" in f"/{part}",
                    "pivot": False,
                    "chart": False,
                }
                info[sheet_name] = entry

                if not part or entry["chartsheet"]:
                    continue
                rels_part = _rels_path(part)
                if rels_part not in names:
                    continue
                for rel in ET.fromstring(zf.read(rels_part)).findall(f"{_PKG_REL_NS}Relationship"):
                    target = (rel.get("Target") or "").replace("\\", "/")
                    if "pivotTable" in target:
                        entry["pivot"] = True
                    elif "drawing" in target:
                        drawing_rels = _rels_path(_resolve_part(os.path.dirname(part), target))
                        if drawing_rels not in names:
                            continue
                        for drel in ET.fromstring(zf.read(drawing_rels)).findall(
                                f"{_PKG_REL_NS}Relationship"):
                            dtarget = (drel.get("Target") or "").replace("\\", "/")
                            if "charts/chart" in dtarget:
                                entry["chart"] = True
                                break
            return info
    except Exception:
        # A malformed package should never stop the combine - fall back instead.
        return None


def sheets_with_pivots_from_package(path: Path) -> set[str]:
    """Names of sheets that have a pivot table attached (package-level check)."""
    info = analyse_package(path)
    return {name for name, entry in (info or {}).items() if entry["pivot"]}


def sheets_with_charts_from_package(path: Path) -> set[str]:
    """Names of sheets that contain a chart, including chart-only sheets."""
    info = analyse_package(path)
    return {name for name, entry in (info or {}).items() if entry["chart"] or entry["chartsheet"]}


# ---------------------------------------------------------------------------
# Readers - each yields (sheet_name, grid, skip_reason)
# ---------------------------------------------------------------------------

class SheetBlock:
    __slots__ = ("name", "grid", "skip_reason")

    def __init__(self, name, grid=None, skip_reason=None):
        self.name = name
        self.grid = grid
        self.skip_reason = skip_reason


def read_openpyxl_file(path: Path, log) -> list[SheetBlock]:
    """Read .xlsx / .xlsm. Hidden content is included, pivots and charts skipped."""
    info = analyse_package(path)
    if info is None:
        return read_openpyxl_file_full(path, log)

    hidden = [name for name, entry in info.items() if entry["hidden"]]
    if hidden:
        log(f"    unhid {len(hidden)} hidden sheet(s): {', '.join(hidden)}")

    if SAVE_UNHIDDEN_COPIES:
        save_unhidden_copy(path, log)

    blocks: list[SheetBlock] = []
    # data_only=True gives the values Excel cached for formula cells (VLOOKUP etc.);
    # read_only=True keeps memory down on very large workbooks.
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True, keep_links=False)
    except Exception as exc:
        log(f"    low-memory read failed ({exc}); retrying with a full load")
        return read_openpyxl_file_full(path, log)

    try:
        for name, entry in info.items():
            if entry["chartsheet"]:
                blocks.append(SheetBlock(name, skip_reason="chart sheet (graph only, no data)"))
                continue
            if entry["pivot"]:
                blocks.append(SheetBlock(name, skip_reason="contains a pivot table"))
                continue
            if entry["chart"]:
                blocks.append(SheetBlock(name, skip_reason="contains a chart/graph object"))
                continue
            try:
                ws = wb[name]
            except KeyError:
                continue
            # some files record the wrong used-range; this forces a real scan
            try:
                ws.reset_dimensions = True
            except AttributeError:
                pass
            grid = trim_grid([list(row) for row in ws.iter_rows(values_only=True)])
            blocks.append(SheetBlock(name, grid=grid))
    finally:
        wb.close()

    return blocks


def read_openpyxl_file_full(path: Path, log) -> list[SheetBlock]:
    """Fallback reader: full openpyxl load, using its own pivot/chart objects."""
    wb = openpyxl.load_workbook(path, data_only=True, keep_links=False)
    blocks: list[SheetBlock] = []
    hidden_sheets = hidden_rows = hidden_cols = 0

    for sheet in wb._sheets:                       # includes chart sheets, in tab order
        if getattr(sheet, "sheet_state", "visible") != "visible":
            hidden_sheets += 1
            sheet.sheet_state = "visible"

        if isinstance(sheet, Chartsheet):
            blocks.append(SheetBlock(sheet.title, skip_reason="chart sheet (graph only, no data)"))
            continue

        for dim in sheet.row_dimensions.values():
            if dim.hidden:
                hidden_rows += 1
                dim.hidden = False
        for dim in sheet.column_dimensions.values():
            if dim.hidden:
                hidden_cols += 1
                dim.hidden = False

        if getattr(sheet, "_pivots", None):
            blocks.append(SheetBlock(sheet.title, skip_reason="contains a pivot table"))
            continue
        if getattr(sheet, "_charts", None):
            n = len(sheet._charts)
            blocks.append(SheetBlock(sheet.title, skip_reason=f"contains {n} chart/graph object(s)"))
            continue

        grid = trim_grid([list(row) for row in sheet.iter_rows(values_only=True)])
        blocks.append(SheetBlock(sheet.title, grid=grid))

    if hidden_sheets or hidden_rows or hidden_cols:
        log(f"    unhid {hidden_sheets} sheet(s), {hidden_rows} row(s), {hidden_cols} column(s)")

    if SAVE_UNHIDDEN_COPIES:
        try:
            out_dir = Path(log.unhidden_dir)
            out_dir.mkdir(parents=True, exist_ok=True)
            wb.save(out_dir / path.name)
            log(f"    saved unhidden copy -> {out_dir / path.name}")
        except Exception as exc:
            log(f"    could not save unhidden copy: {exc}")

    wb.close()
    return blocks


def save_unhidden_copy(path: Path, log) -> None:
    """Write a copy of the workbook with every sheet, row and column unhidden."""
    try:
        out_dir = Path(log.unhidden_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.load_workbook(path, keep_links=False)     # keep formulas intact
        for sheet in wb._sheets:
            sheet.sheet_state = "visible"
            for dim in getattr(sheet, "row_dimensions", {}).values():
                dim.hidden = False
            for dim in getattr(sheet, "column_dimensions", {}).values():
                dim.hidden = False
        wb.save(out_dir / path.name)
        wb.close()
        log(f"    saved unhidden copy -> {out_dir / path.name}")
    except Exception as exc:
        log(f"    could not save unhidden copy: {exc}")


def read_xls_file(path: Path, log) -> list[SheetBlock]:
    """Read legacy .xls. Hidden sheets/rows/columns are read regardless."""
    try:
        import xlrd
    except ImportError:
        raise RuntimeError("reading .xls needs xlrd - run:  pip install xlrd")

    book = xlrd.open_workbook(path)
    blocks: list[SheetBlock] = []
    hidden_sheets = 0

    for sheet in book.sheets():
        if getattr(sheet, "visibility", 0) != 0:
            hidden_sheets += 1                     # read anyway = effectively unhidden

        grid = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
                    value = None
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = None
                row.append(value)
            grid.append(row)
        blocks.append(SheetBlock(sheet.name, grid=trim_grid(grid)))

    if hidden_sheets:
        log(f"    {hidden_sheets} hidden sheet(s) included")
    log("    NOTE: .xls format - pivot tables and charts cannot be auto-detected, "
        "check these sheets manually")
    return blocks


def read_xlsb_file(path: Path, log) -> list[SheetBlock]:
    try:
        import pyxlsb  # noqa: F401
    except ImportError:
        raise RuntimeError("reading .xlsb needs pyxlsb - run:  pip install pyxlsb")

    blocks: list[SheetBlock] = []
    with pd.ExcelFile(path, engine="pyxlsb") as xl:
        for name in xl.sheet_names:
            frame = xl.parse(name, header=None, dtype=object)
            grid = trim_grid(frame.values.tolist())
            blocks.append(SheetBlock(name, grid=grid))
    log("    NOTE: .xlsb format - pivot tables and charts cannot be auto-detected, "
        "check these sheets manually")
    return blocks


def read_text_file(path: Path, log) -> list[SheetBlock]:
    """Read .csv / .txt with encoding and delimiter detection."""
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise RuntimeError("could not decode this text file")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    grid = [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]
    return [SheetBlock(path.stem, grid=trim_grid(grid))]


def read_any_file(path: Path, log) -> list[SheetBlock]:
    ext = path.suffix.lower()
    if ext in OPENPYXL_EXTENSIONS:
        return read_openpyxl_file(path, log)
    if ext == ".xls":
        return read_xls_file(path, log)
    if ext == ".xlsb":
        return read_xlsb_file(path, log)
    if ext in TEXT_EXTENSIONS:
        return read_text_file(path, log)
    raise RuntimeError(f"unsupported file type '{ext}'")


# ---------------------------------------------------------------------------
# Header rule + sheet -> DataFrame
# ---------------------------------------------------------------------------

def check_header_row(grid: list[list]) -> str | None:
    """Return None if row 1 is a valid header row, otherwise the reason it isn't."""
    if not grid:
        return "sheet is empty"

    first_filled = next((i for i, row in enumerate(grid)
                         if any(not is_blank(c) for c in row)), None)
    if first_filled is None:
        return "sheet is empty"
    if first_filled != 0:
        return f"header is not in row 1 (first filled row is row {first_filled + 1})"

    header = grid[0]
    filled = [c for c in header if not is_blank(c)]
    if len(filled) < MIN_HEADER_CELLS:
        return f"row 1 has only {len(filled)} filled cell(s) - not a header row"

    texty = [c for c in filled if isinstance(c, str) and re.search(r"[A-Za-z]", c)]
    if len(texty) < MIN_TEXT_HEADERS:
        return "row 1 does not look like a header row (not enough text labels)"
    if len(texty) / len(filled) < MIN_TEXT_RATIO:
        return "row 1 looks like data, not headers (mostly numbers/dates)"
    if any(len(str(c)) > MAX_HEADER_LEN for c in filled):
        return "row 1 contains a very long value - looks like a title or data, not headers"
    if len(grid) < 2 or not any(any(not is_blank(c) for c in row) for row in grid[1:]):
        return "header row found but there are no data rows under it"
    return None


def sheet_to_frame(block: SheetBlock, source_file: str, name_registry: dict[str, str]):
    """Turn a validated sheet grid into a DataFrame with mapped column names."""
    grid = block.grid
    header = grid[0]

    # map each source column index to an output column name
    targets: dict[str, list[int]] = {}
    order: list[str] = []
    for idx, raw in enumerate(header):
        if is_blank(raw):
            continue
        canon = canonical_header(raw)
        if canon:
            name = canon
        else:
            key = normalise_header(raw)
            display = _WS_RE.sub(" ", str(raw).strip())
            name = name_registry.setdefault(key, display)
        if name not in targets:
            targets[name] = []
            order.append(name)
        targets[name].append(idx)

    if not targets:
        return None

    header_signature = tuple(normalise_value(c) for c in header)

    columns: dict[str, list] = {name: [] for name in order}
    if ADD_SOURCE_COLUMNS:
        src_file_col, src_sheet_col = [], []

    kept = 0
    for row in grid[1:]:
        if all(is_blank(c) for c in row):
            continue
        # a repeated header row inside the data is not data
        if tuple(normalise_value(c) for c in row[:len(header_signature)]) == header_signature:
            continue

        for name in order:
            value = None
            for idx in targets[name]:            # several spellings -> first non-empty wins
                if idx < len(row):
                    candidate = clean_cell(row[idx])
                    if candidate is not None:
                        value = candidate
                        break
            columns[name].append(value)
        if ADD_SOURCE_COLUMNS:
            src_file_col.append(source_file)
            src_sheet_col.append(block.name)
        kept += 1

    if kept == 0:
        return None

    if ADD_SOURCE_COLUMNS:
        columns["Source File"] = src_file_col
        columns["Source Sheet"] = src_sheet_col

    return pd.DataFrame(columns, dtype=object)


# ---------------------------------------------------------------------------
# Duplicate removal
# ---------------------------------------------------------------------------

def build_dedupe_keys(frame: "pd.DataFrame") -> "pd.Series":
    """
    Build the comparison key for each row. Rows where every key column is blank
    get a unique key so they are never collapsed into each other.
    """
    if not DEDUPE_ON_KEY_COLUMNS or not any(c in frame.columns for c in KEY_COLUMNS):
        cols = [c for c in frame.columns if c not in ("Source File", "Source Sheet")]
        return frame[cols].apply(lambda r: "\x1f".join(normalise_value(v) for v in r), axis=1)

    parts = []
    for col in KEY_COLUMNS:
        if col not in frame.columns:
            parts.append(pd.Series([""] * len(frame), index=frame.index))
        elif col == "Date of Service":
            parts.append(frame[col].map(normalise_date_value))
        else:
            parts.append(frame[col].map(normalise_value))

    keys = parts[0].astype(str)
    for part in parts[1:]:
        keys = keys + "\x1f" + part.astype(str)

    blank = keys.map(lambda k: k.replace("\x1f", "") == "")
    if blank.any():
        keys = keys.where(~blank, pd.Series([f"__blank__{i}" for i in range(len(keys))],
                                            index=keys.index))
    return keys


def drop_duplicates(frame: "pd.DataFrame", label: str, log) -> "pd.DataFrame":
    if frame is None or frame.empty:
        return frame
    keys = build_dedupe_keys(frame)
    mask = ~keys.duplicated(keep="first")
    removed = int((~mask).sum())
    if removed:
        log(f"    removed {removed:,} duplicate row(s) {label}")
    return frame.loc[mask]


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def excel_safe(value):
    if value is None:
        return None
    if isinstance(value, float) and value != value:
        return None
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if len(value) > 32767:
            value = value[:32764] + "..."
        return value
    if isinstance(value, (int, float, bool, dt.datetime, dt.date, dt.time)):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return str(value)


def write_parts(frame: "pd.DataFrame", out_dir: Path, stamp: str, log) -> list[Path]:
    if len(frame.columns) > 16384:
        raise RuntimeError(f"{len(frame.columns)} columns exceeds the Excel limit of 16,384")

    headers = list(frame.columns)
    total = len(frame)
    parts = max(1, -(-total // MAX_ROWS_PER_PART))     # ceiling division
    written: list[Path] = []

    for part in range(parts):
        chunk = frame.iloc[part * MAX_ROWS_PER_PART:(part + 1) * MAX_ROWS_PER_PART]
        suffix = f" part {part + 1} of {parts}" if parts > 1 else ""
        out_path = out_dir / f"{stamp} {AUTHOR_INITIALS} combined data{suffix}.xlsx"

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Combined")
        ws.append(headers)
        for row in chunk.itertuples(index=False, name=None):
            ws.append([excel_safe(v) for v in row])
        wb.save(out_path)
        wb.close()

        written.append(out_path)
        log(f"  wrote {len(chunk):,} rows -> {out_path.name}")

    return written


def write_report(rows: list[dict], out_dir: Path, stamp: str, log) -> Path:
    path = out_dir / f"{stamp} {AUTHOR_INITIALS} combine report.xlsx"
    frame = pd.DataFrame(rows, columns=["File", "Sheet", "Action", "Detail", "Rows"])
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Report")
    ws.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        ws.append([excel_safe(v) for v in row])
    wb.save(path)
    wb.close()
    log(f"  wrote report -> {path.name}")
    return path


# ---------------------------------------------------------------------------
# Main combine
# ---------------------------------------------------------------------------

def find_input_files(folder: Path) -> list[Path]:
    pattern = "**/*" if RECURSE_SUBFOLDERS else "*"
    files = []
    for path in sorted(folder.glob(pattern)):
        if not path.is_file():
            continue
        if path.name.startswith("~$") or path.name.startswith("."):
            continue
        if path.suffix.lower() not in DATA_EXTENSIONS:
            continue
        # never re-read our own output
        if re.search(r" (combined data|combine report)", path.stem):
            continue
        files.append(path)
    return files


def combine(input_dir: Path, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.date.today().strftime("%y%m%d")

    def log(message):
        print(message, flush=True)
    log.unhidden_dir = output_dir / "_unhidden"

    files = find_input_files(input_dir)
    if not files:
        print(f"No Excel or CSV files found in: {input_dir}")
        return

    print(f"Input folder : {input_dir}")
    print(f"Output folder: {output_dir}")
    print(f"Found {len(files)} file(s)\n")

    name_registry: dict[str, str] = {}
    report: list[dict] = []
    all_frames: list["pd.DataFrame"] = []

    for file_no, path in enumerate(files, 1):
        rel = str(path.relative_to(input_dir))
        print(f"[{file_no}/{len(files)}] {rel}")

        try:
            blocks = read_any_file(path, log)
        except Exception as exc:
            print(f"    ERROR: {exc}")
            report.append({"File": rel, "Sheet": "", "Action": "FILE SKIPPED",
                           "Detail": str(exc), "Rows": 0})
            continue

        file_frames: list["pd.DataFrame"] = []

        for block in blocks:
            if block.skip_reason:
                print(f"    skip [{block.name}] - {block.skip_reason}")
                report.append({"File": rel, "Sheet": block.name, "Action": "SHEET SKIPPED",
                               "Detail": block.skip_reason, "Rows": 0})
                continue

            reason = check_header_row(block.grid)
            if reason:
                print(f"    skip [{block.name}] - {reason}")
                report.append({"File": rel, "Sheet": block.name, "Action": "SHEET SKIPPED",
                               "Detail": reason, "Rows": 0})
                continue

            frame = sheet_to_frame(block, rel, name_registry)
            if frame is None or frame.empty:
                print(f"    skip [{block.name}] - no usable data rows")
                report.append({"File": rel, "Sheet": block.name, "Action": "SHEET SKIPPED",
                               "Detail": "no usable data rows", "Rows": 0})
                continue

            mapped = [c for c in KEY_COLUMNS if c in frame.columns]
            print(f"    use  [{block.name}] {len(frame):,} rows, {len(frame.columns)} cols, "
                  f"key columns found: {', '.join(mapped) if mapped else 'none'}")
            report.append({"File": rel, "Sheet": block.name, "Action": "COMBINED",
                           "Detail": f"key columns: {', '.join(mapped) or 'none'}",
                           "Rows": len(frame)})
            file_frames.append(frame)

        if not file_frames:
            continue

        file_frame = pd.concat(file_frames, ignore_index=True, sort=False)
        before = len(file_frame)
        file_frame = drop_duplicates(file_frame, "within this file", log)
        report.append({"File": rel, "Sheet": "(whole file)", "Action": "DE-DUPLICATED",
                       "Detail": f"{before:,} rows in, {len(file_frame):,} rows out",
                       "Rows": len(file_frame)})
        all_frames.append(file_frame)

    if not all_frames:
        print("\nNothing to combine - every sheet was skipped.")
        write_report(report, output_dir, stamp, log)
        return

    print("\nCombining files...")
    combined = pd.concat(all_frames, ignore_index=True, sort=False)

    # column order: the 4 key columns first, then everything else, source last
    key_first = [c for c in KEY_COLUMNS if c in combined.columns]
    tail = [c for c in ("Source File", "Source Sheet") if c in combined.columns]
    middle = [c for c in combined.columns if c not in key_first and c not in tail]
    combined = combined[key_first + middle + tail]

    before = len(combined)
    if DEDUPE_ACROSS_FILES:
        combined = drop_duplicates(combined, "across all files", log)
    combined = combined.reset_index(drop=True)

    report.append({"File": "(all files)", "Sheet": "", "Action": "FINAL",
                   "Detail": f"{before:,} rows in, {len(combined):,} rows out, "
                             f"{len(combined.columns)} columns",
                   "Rows": len(combined)})

    print(f"Final: {len(combined):,} rows x {len(combined.columns)} columns\n")
    write_parts(combined, output_dir, stamp, log)
    write_report(report, output_dir, stamp, log)

    print("\nDone.")
    print("Reminder: if this data is confidential, keep the output in the approved "
          "Global Insider folder.")


# ---------------------------------------------------------------------------
# Folder pickers / entry point
# ---------------------------------------------------------------------------

def pick_folder(title: str) -> Path | None:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        return None
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    chosen = filedialog.askdirectory(title=title)
    root.destroy()
    return Path(chosen) if chosen else None


def main(argv: list[str]) -> int:
    if len(argv) >= 2:
        input_dir = Path(argv[0]).expanduser().resolve()
        output_dir = Path(argv[1]).expanduser().resolve()
    else:
        print("Pick the folder that contains the Excel/CSV files to combine...")
        input_dir = pick_folder("Select the folder containing the files to combine")
        if not input_dir:
            print("Cancelled.")
            return 1
        print("Pick where the combined output should be saved...")
        output_dir = pick_folder("Select the folder to save the combined output in")
        if not output_dir:
            output_dir = input_dir / "Combined output"
            print(f"No output folder picked - using {output_dir}")

    if not input_dir.is_dir():
        print(f"Not a folder: {input_dir}")
        return 1

    try:
        combine(input_dir, output_dir)
    except Exception:
        traceback.print_exc()
        input("\nSomething went wrong. Press Enter to close...")
        return 1

    if len(argv) < 2:
        input("\nPress Enter to close...")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
