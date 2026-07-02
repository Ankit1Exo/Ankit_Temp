"""
extract_wage_form.py
---------------------
Extracts the following fields from NY Quarterly Employee/Payee Wage
Reporting PDFs (Part C section):

    a  Social Security Number
    b  Last Name
    c  First Name
    d  MI  (Middle Initial)
    g  Gross Federal Wages or Distribution

HOW COLUMN SPLITTING WORKS
---------------------------
The NY-45 form uses "|" pipe characters as visual column separators.
pdfplumber sometimes reads these as the letter "I" or merges them with the
adjacent word (e.g. "| SCOTT" becomes "ISCOTT").

This script uses a TWO-LAYER approach to handle both cases:

  Layer 1 — Dynamic pipe detection:
    On each data row, find all narrow single-character words that are "|",
    "I", or "l" at column-boundary positions. Use their x-coordinates as
    the actual column split points. This is more reliable than fixed % bands
    because it adapts to each PDF's exact coordinate space.

  Layer 2 — Merged-word splitting:
    After extracting each field value, if it still starts with "I" followed
    by uppercase letters (e.g. "ISCOTT", "IC"), strip the leading "I" pipe
    character that pdfplumber merged into the word.

Requirements:
    pip install pdfplumber pdf2image pytesseract pillow pandas openpyxl

Run:
    python extract_wage_form.py
"""

import os
import re
import sys
import csv

try:
    import pdfplumber
    import pandas as pd
except ImportError:
    sys.exit("Run: pip install pdfplumber pandas openpyxl")


# ── File / Folder Picker ──────────────────────────────────────────────────────

def pick_source() -> list:
    """
    GUI dialog: choose a folder (all PDFs inside) or individual PDF files.
    Falls back to console input in headless environments.
    """
    try:
        import tkinter as tk
        from tkinter import filedialog, simpledialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        mode = simpledialog.askstring(
            "Select Source",
            "Type  folder  → process all PDFs in a folder\n"
            "Type  files   → pick individual PDFs",
            parent=root,
        )
        mode = (mode or "").strip().lower()

        if "folder" in mode:
            folder = filedialog.askdirectory(title="Select folder containing PDFs")
            root.destroy()
            if not folder:
                sys.exit("No folder selected.")
            paths = sorted([
                os.path.join(folder, f) for f in os.listdir(folder)
                if f.lower().endswith(".pdf") and not f.startswith("~")
            ])
            if not paths:
                sys.exit(f"No PDFs found in: {folder}")
        else:
            paths = list(filedialog.askopenfilenames(
                title="Select PDF files (Ctrl/Cmd+click for multiple)",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            ))
            root.destroy()
            if not paths:
                sys.exit("No files selected.")

        return paths

    except Exception:
        print("\nNo display found. Please type your selection:")
        print("  1 - Enter a folder path (all PDFs inside will be processed)")
        print("  2 - Enter individual file paths one by one")
        choice = input("Choice [1/2]: ").strip()
        if choice == "1":
            folder = input("Folder path: ").strip()
            paths = sorted([
                os.path.join(folder, f) for f in os.listdir(folder)
                if f.lower().endswith(".pdf")
            ])
            if not paths:
                sys.exit(f"No PDFs found in: {folder}")
            return paths
        else:
            paths = []
            print("Enter PDF paths one per line (blank line to finish):")
            while True:
                p = input("  Path: ").strip()
                if not p:
                    break
                if os.path.exists(p):
                    paths.append(p)
                else:
                    print(f"  ⚠  Not found, skipping: {p}")
            if not paths:
                sys.exit("No valid files entered.")
            return paths


# ── Row Grouping ──────────────────────────────────────────────────────────────

def get_words(page) -> list:
    """
    Extract words with coordinates. Use tight x_tolerance so that
    pipe characters are NOT merged into adjacent words where possible.
    """
    return page.extract_words(
        x_tolerance=2,       # tight: keeps "|" separate from adjacent text
        y_tolerance=3,
        keep_blank_chars=False,
        use_text_flow=False,
    ) or []


def group_into_rows(words: list, gap: float = 5.0) -> list:
    """
    Cluster words into horizontal rows by top coordinate.
    Returns list of (avg_top, [words]) sorted top-to-bottom.
    """
    if not words:
        return []
    sorted_words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    rows, current_y, current_row = [], sorted_words[0]["top"], [sorted_words[0]]
    for w in sorted_words[1:]:
        if abs(w["top"] - current_y) <= gap:
            current_row.append(w)
        else:
            rows.append((current_y, current_row))
            current_y, current_row = w["top"], [w]
    rows.append((current_y, current_row))
    return sorted(rows, key=lambda r: r[0])


# ── Pipe / Delimiter Detection ────────────────────────────────────────────────

# Characters that can represent a column-separator pipe in PDF text
PIPE_CHARS = frozenset({"|", "¦", "‖"})

# A word is treated as a pipe delimiter when:
#   - Its text is one of PIPE_CHARS, OR
#   - Its text is a single "I" or "l" AND it is very narrow (width < 5pts)
def is_pipe_word(w: dict) -> bool:
    t    = w["text"].strip()
    wdth = w["x1"] - w["x0"]
    if t in PIPE_CHARS:
        return True
    if t in ("I", "l", "1") and wdth < 5.0:
        return True
    return False


def find_pipe_xs(row_words: list) -> list:
    """
    Return sorted list of x-centres of pipe delimiter words in a row.
    These become the column split boundaries.
    """
    xs = [(w["x0"] + w["x1"]) / 2 for w in row_words if is_pipe_word(w)]
    return sorted(xs)


def split_pipe_prefix(text: str) -> str:
    """
    Layer 2 fix: when pdfplumber merges a pipe into the start of a word
    (e.g. "ISCOTT" → pipe + "SCOTT", or "IC" → pipe + "C"),
    strip the leading pipe character.

    Rule: strip leading "I", "|", or "l" only when:
      - The character is a single prefix (not the whole word)
      - The remainder starts with an uppercase letter (it's a name/initial)
    """
    if len(text) > 1 and text[0] in ("I", "|", "l", "¦"):
        remainder = text[1:]
        if remainder and remainder[0].isupper():
            return remainder
    return text


# ── Field Extraction from a Row ───────────────────────────────────────────────

def words_before(row_words: list, x_boundary: float) -> str:
    """All non-pipe words whose centre is LEFT of x_boundary."""
    result = [
        w for w in row_words
        if not is_pipe_word(w)
        and (w["x0"] + w["x1"]) / 2 < x_boundary
    ]
    return " ".join(w["text"] for w in sorted(result, key=lambda w: w["x0"])).strip()


def words_between(row_words: list, x_lo: float, x_hi: float) -> str:
    """All non-pipe words whose centre is strictly between x_lo and x_hi."""
    result = [
        w for w in row_words
        if not is_pipe_word(w)
        and x_lo < (w["x0"] + w["x1"]) / 2 < x_hi
    ]
    return " ".join(w["text"] for w in sorted(result, key=lambda w: w["x0"])).strip()


def extract_name_fields(row_words: list) -> tuple:
    """
    Given a name data row, return (last_name, first_name, mi).

    Uses dynamic pipe detection:
      pipes[0] → boundary between Last Name and First Name
      pipes[1] → boundary between First Name and MI
      pipes[2] → boundary between MI and Wage Type  (optional)

    Falls back to x-percentage bands if no pipes are found.
    """
    pipes = find_pipe_xs(row_words)
    pw    = max((w["x1"] for w in row_words), default=612.0)

    if len(pipes) >= 2:
        # Happy path: found at least 2 pipe separators
        last_raw  = words_before(row_words, pipes[0])
        first_raw = words_between(row_words, pipes[0], pipes[1])
        mi_raw    = words_between(row_words, pipes[1], pipes[2]) \
                    if len(pipes) >= 3 else ""

    elif len(pipes) == 1:
        # Only one pipe found — split last/first on it, no MI
        last_raw  = words_before(row_words, pipes[0])
        first_raw = " ".join(
            w["text"] for w in sorted(row_words, key=lambda w: w["x0"])
            if not is_pipe_word(w) and (w["x0"]+w["x1"])/2 > pipes[0]
        ).strip()
        mi_raw    = ""

    else:
        # No pipes found — fall back to x-percentage bands
        # (last resort for unusual PDFs)
        last_raw  = " ".join(
            w["text"] for w in sorted(row_words, key=lambda w: w["x0"])
            if not is_pipe_word(w) and (w["x0"]+w["x1"])/2 < pw * 0.42
        ).strip()
        first_raw = " ".join(
            w["text"] for w in sorted(row_words, key=lambda w: w["x0"])
            if not is_pipe_word(w)
            and pw*0.43 < (w["x0"]+w["x1"])/2 < pw * 0.65
        ).strip()
        mi_raw    = ""

    # Apply Layer 2: strip any merged pipe prefix from each field
    last_name  = split_pipe_prefix(last_raw)
    first_name = split_pipe_prefix(first_raw)
    mi         = split_pipe_prefix(mi_raw)

    # Extra guard: if MI is more than 2 chars, it's probably noise
    if len(mi) > 2:
        mi = mi[0] if mi[0].isupper() else ""

    return last_name, first_name, mi


def extract_ssn(row_words: list, page_width: float) -> str:
    """
    SSN sits in the leftmost column (0–20% of page width).
    It may be masked/redacted in the PDF.
    """
    ssn_words = [
        w for w in row_words
        if not is_pipe_word(w)
        and (w["x0"] + w["x1"]) / 2 < page_width * 0.20
    ]
    return " ".join(w["text"] for w in sorted(ssn_words, key=lambda w: w["x0"])).strip()


def extract_gross_wages(row_words: list, page_width: float) -> str:
    """
    Gross federal wages (column g) sits roughly at 32–55% of page width
    on the wages data row (the row below the 'g Gross federal...' label).
    Use pipe detection here too.
    """
    pipes = find_pipe_xs(row_words)

    if len(pipes) >= 2:
        # Column g is between pipe[0] (after 'f UI' column) and pipe[1]
        raw = words_between(row_words, pipes[0], pipes[1])
    else:
        # Fallback to x-band
        raw = " ".join(
            w["text"] for w in sorted(row_words, key=lambda w: w["x0"])
            if not is_pipe_word(w)
            and page_width * 0.28 < (w["x0"]+w["x1"])/2 < page_width * 0.56
        ).strip()

    return split_pipe_prefix(raw)


# ── Label Row Detection ───────────────────────────────────────────────────────

def row_text(row_words: list) -> str:
    return " ".join(w["text"].lower() for w in row_words)


def is_ssn_label_row(row_words: list) -> bool:
    t = row_text(row_words)
    return "social" in t and "security" in t


def is_gross_label_row(row_words: list) -> bool:
    t = row_text(row_words)
    return "gross" in t and "federal" in t


# ── Per-Page Extraction ───────────────────────────────────────────────────────

def extract_records_from_page(page) -> list:
    """
    Extract all employee records from one PDF page.

    Form row structure per employee:
      ROW A (label):  "a Social Security number | b Last name | c First name | d MI | e Wage"
      ROW B (data):    <SSN>  |  <LAST>  |  <FIRST>  |  <MI>  |  <wage type>
      ROW C (label):  "f Total UI... | g Gross federal wages... | h NYS... | ..."
      ROW D (data):    <UI amt>  |  <GROSS WAGES>  |  <NYS>  |  ...
    """
    pw     = float(page.width)
    words  = get_words(page)
    rows   = group_into_rows(words, gap=5.0)
    records = []

    i = 0
    while i < len(rows):
        _, row_words = rows[i]

        if is_ssn_label_row(row_words):
            if i + 1 >= len(rows):
                i += 1
                continue

            # ROW B: name/SSN data
            _, data_row = rows[i + 1]

            ssn                          = extract_ssn(data_row, pw)
            last_name, first_name, mi    = extract_name_fields(data_row)

            # ROW D: gross wages (scan up to 6 rows ahead for gross label)
            gross_wages = ""
            for j in range(i + 2, min(i + 8, len(rows))):
                _, candidate = rows[j]
                if is_gross_label_row(candidate):
                    if j + 1 < len(rows):
                        _, wage_row = rows[j + 1]
                        gross_wages = extract_gross_wages(wage_row, pw)
                    break

            if last_name or first_name:
                records.append({
                    "ssn":         clean_ssn(ssn),
                    "last_name":   last_name,
                    "first_name":  first_name,
                    "mi":          mi,
                    "gross_wages": clean_number(gross_wages),
                })

            i += 2
            continue

        i += 1

    return records


# ── Value Cleaning ────────────────────────────────────────────────────────────

def clean_ssn(raw: str) -> str:
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 9:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return raw


def clean_number(raw: str) -> str:
    return re.sub(r"[^\d.,]", "", raw).strip()


# ── OCR Fallback ──────────────────────────────────────────────────────────────

def ocr_fallback(pdf_path: str) -> list:
    try:
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import ImageEnhance, ImageFilter
    except ImportError:
        print("  ⚠  OCR requires: pip install pdf2image pytesseract pillow")
        return []

    print("  → No text layer found. Running OCR (may take a moment)...")
    pages = convert_from_path(pdf_path, dpi=300)
    all_records = []

    for img in pages:
        img = img.resize((img.width * 2, img.height * 2))
        img = ImageEnhance.Contrast(img).enhance(2.0)
        img = img.filter(ImageFilter.SHARPEN)

        # Use pdfplumber-style word extraction via pytesseract bounding boxes
        import pytesseract
        data = pytesseract.image_to_data(
            img, config="--psm 6",
            output_type=pytesseract.Output.DICT
        )
        # Convert to pdfplumber word format
        ocr_words = []
        h = img.height
        for k, text in enumerate(data["text"]):
            if not text.strip() or data["conf"][k] < 20:
                continue
            ocr_words.append({
                "text":   text,
                "x0":     float(data["left"][k]),
                "x1":     float(data["left"][k] + data["width"][k]),
                "top":    float(data["top"][k]),
                "bottom": float(data["top"][k] + data["height"][k]),
            })

        # Run the same row-grouping + field extraction
        ocr_rows = group_into_rows(ocr_words, gap=8.0)
        pw = img.width

        j = 0
        while j < len(ocr_rows):
            _, row_words = ocr_rows[j]
            if is_ssn_label_row(row_words):
                if j + 1 < len(ocr_rows):
                    _, data_row = ocr_rows[j + 1]
                    ssn = extract_ssn(data_row, pw)
                    last_name, first_name, mi = extract_name_fields(data_row)
                    gross_wages = ""
                    for k2 in range(j + 2, min(j + 8, len(ocr_rows))):
                        _, cand = ocr_rows[k2]
                        if is_gross_label_row(cand):
                            if k2 + 1 < len(ocr_rows):
                                _, wage_row = ocr_rows[k2 + 1]
                                gross_wages = extract_gross_wages(wage_row, pw)
                            break
                    if last_name or first_name:
                        all_records.append({
                            "ssn":         clean_ssn(ssn),
                            "last_name":   last_name,
                            "first_name":  first_name,
                            "mi":          mi,
                            "gross_wages": clean_number(gross_wages),
                        })
                j += 2
                continue
            j += 1

    return all_records


# ── Per-File Driver ───────────────────────────────────────────────────────────

def process_pdf(pdf_path: str) -> list:
    filename = os.path.basename(pdf_path)
    records  = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            has_text = any((p.extract_text() or "").strip() for p in pdf.pages)

            if not has_text:
                raw = ocr_fallback(pdf_path)
                for r in raw:
                    r["file_name"] = filename
                    r.setdefault("page", "")
                return raw

            for page_num, page in enumerate(pdf.pages, start=1):
                page_records = extract_records_from_page(page)
                for r in page_records:
                    r["file_name"] = filename
                    r["page"]      = page_num
                records.extend(page_records)

    except Exception as e:
        print(f"  ⚠  Error reading {filename}: {e}")

    return records


# ── Output ────────────────────────────────────────────────────────────────────

COLUMNS = ["file_name", "page", "ssn", "last_name", "first_name", "mi", "gross_wages"]

LABELS = {
    "file_name":   "File Name",
    "page":        "Page",
    "ssn":         "a Social Security Number",
    "last_name":   "b Last Name",
    "first_name":  "c First Name",
    "mi":          "d MI",
    "gross_wages": "g Gross Federal Wages",
}


def save_csv(records: list, out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writerow(LABELS)
        writer.writerows(records)
    print(f"  CSV   → {out_path}")


def save_excel(records: list, out_path: str) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(records, columns=COLUMNS)
    df.rename(columns=LABELS, inplace=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Wage Data")
        ws = writer.sheets["Wage Data"]

        HEADER_FILL = PatternFill("solid", start_color="1F4E79")
        HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        DATA_FONT   = Font(name="Arial", size=10)
        ALT_FILL    = PatternFill("solid", start_color="DEEAF1")

        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.font = DATA_FONT
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL

        for col in ws.columns:
            max_w = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = min(max_w + 4, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"  Excel → {out_path}")


def print_summary(files: list, all_records: list) -> None:
    print(f"\n{'='*60}")
    print(f"  Extraction Summary")
    print(f"{'='*60}")
    print(f"  Files processed : {len(files)}")
    print(f"  Total records   : {len(all_records)}")
    print(f"\n  Records per file:")
    for f in files:
        n = sum(1 for r in all_records if r["file_name"] == os.path.basename(f))
        print(f"    • {os.path.basename(f)}: {n} employee(s)")
    no_ssn = sum(1 for r in all_records if not r.get("ssn"))
    no_mi  = sum(1 for r in all_records if not r.get("mi"))
    if no_ssn:
        print(f"\n  ⚠  {no_ssn} record(s) with no SSN (redacted in source)")
    if no_mi:
        print(f"  ℹ  {no_mi} record(s) with no MI (employee has no middle initial)")
    print(f"{'='*60}\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  NY Quarterly Wage Form — Field Extractor")
    print("=" * 60)

    pdf_paths = pick_source()
    print(f"\nFiles selected: {len(pdf_paths)}")
    for p in pdf_paths:
        print(f"  • {os.path.basename(p)}")

    all_records = []
    for path in pdf_paths:
        print(f"\nProcessing: {os.path.basename(path)}")
        records = process_pdf(path)
        print(f"  → {len(records)} record(s) found")
        all_records.extend(records)

    if not all_records:
        print("\nNo records extracted. Check PDF content.")
        sys.exit(1)

    print_summary(pdf_paths, all_records)

    out_dir  = os.path.dirname(os.path.abspath(pdf_paths[0]))
    csv_path = os.path.join(out_dir, "wage_extraction.csv")
    xls_path = os.path.join(out_dir, "wage_extraction.xlsx")

    # Avoid overwriting existing files
    for suffix in ["csv", "xls"]:
        path_var = csv_path if suffix == "csv" else xls_path
        if os.path.exists(path_var):
            root, ext = os.path.splitext(path_var)
            n = 1
            while os.path.exists(path_var):
                path_var = f"{root}_{n}{ext}"
                n += 1
            if suffix == "csv":
                csv_path = path_var
            else:
                xls_path = path_var

    print("Saving outputs...")
    save_csv(all_records, csv_path)
    save_excel(all_records, xls_path)
    print("\nDone.")


if __name__ == "__main__":
    main()
