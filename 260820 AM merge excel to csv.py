"""
260820 AM merge excel to csv.py
-------------------------------
Merges many Excel files into ONE CSV, row for row, exactly as they are.
Built for very large volumes (target shape: ~50 lakh / 5 million rows).

HOW TO RUN
    Open in IDLE and press F5. No command line arguments are needed - you will
    be asked to pick the input folder and then the output folder.
    (Optional: python "260820 AM merge excel to csv.py" <input folder> <output folder>)
    (Optional: python "260820 AM merge excel to csv.py" --selftest)

RULES IMPLEMENTED
    1. Everything is appended AS IT IS. No de-duplication, no header matching,
       no column re-ordering, no trimming, no cleaning.
    2. A completely blank row is appended as a completely blank row. Blank rows
       inside the data and blank rows at the end are both kept.
    3. Every value is written as TEXT. Numbers keep their digits (leading zeros
       in 007 stay, a 16 digit account number stays a 16 digit number), and every
       field is quoted so no reader treats a column as a number.
    4. Every sheet of every workbook is appended, in file name order, then in
       sheet order (set SHEETS = "first" to take only the first sheet).
    5. Hidden sheets, hidden rows and hidden columns are appended too - hidden is
       not a reason to drop data.
    6. A progress bar shows rows done / rows total, elapsed time and ETA.
    7. A run report CSV lists how many rows came out of every file and sheet.

MEMORY / SPEED
    The workbooks are streamed (openpyxl read_only) and written straight to the
    CSV one row at a time, so 5 million rows use the same memory as 500 rows.
    Measured at 65 columns on this machine:
        file declares its size    2,350 rows/sec  -> about 35 min for 50 lakh rows
        file does not declare it  1,220 rows/sec  -> about 70 min for 50 lakh rows
    Those test rows were sparse; densely filled rows run slower. Every row comes
    out the same width, so the CSV loads into pandas or a database with no ragged
    row errors. Note that 50 lakh rows is about 1.5 GB and is far past Excel's
    1,048,576 row limit - the result is for pandas, Power Query or a database.

FILES THAT DO NOT DECLARE THEIR SIZE
    A workbook saved by Excel records its used range, so the row and column counts
    are read instantly. A workbook written by an export tool often does not. For
    those the script counts the rows first and reads the file twice, because
    without the true column count a row ending in empty cells comes back short and
    lands in the CSV ragged. Set FAST_SCAN_UNDECLARED = True to read once instead,
    at the cost of an approximate progress bar.

FILE TYPES
    .xlsx .xlsm .xltx .xltm  - streamed with openpyxl
    .xls                     - read with xlrd (Excel 97-2003)
    .xlsb                    - needs the 'pyxlsb' package, else the file is
                               logged as skipped

REQUIREMENTS
    pip install openpyxl xlrd
    (optional: pip install pyxlsb   - only needed for .xlsb files)

DATA HANDLING NOTE
    This script only reads and writes files on this machine. If the source data
    contains PHI, SSNs or other confidential information, keep the input and
    output folders inside the approved Global Insider location - not the desktop,
    not a personal folder - and never paste rows into a chat or an email. The
    merged CSV is data, not code: do not commit it to the repository.
"""

from __future__ import annotations

import csv
import datetime as dt
import numbers
import sys
import time
import traceback
from pathlib import Path

# ---------------------------------------------------------------------------
# CONFIGURATION - edit these
# ---------------------------------------------------------------------------

AUTHOR_INITIALS = "AM"          # used in the output file names

OUTPUT_DESCRIPTION = "merged excel data"       # description part of the file name

RECURSE_SUBFOLDERS = True       # also read files in sub-folders of the input folder

SHEETS = "all"                  # "all"   = append every sheet of every workbook
                                # "first" = append only the first sheet of each

HEADER_ROWS = 0                 # 0 = append literally everything, so every file
                                #     keeps its own header row (this is the
                                #     "as it is" behaviour that was asked for)
                                # 1 = keep the header row of the FIRST sheet only
                                #     and skip the top row of every sheet after it

# The columns every file is expected to have, in order. The check compares row 1
# of each sheet against this list, ignoring case and extra spaces.
#   * Set to [] to turn the check off completely.
#   * Check the spelling against your real files - these were transcribed from a
#     screenshot, so one wrong character here means every file reports a mismatch.
EXPECTED_HEADERS = [
    "Doc ID",
    "First Name",
    "Middle Name",
    "Last Name",
    "Suffix",
    "Entity Type_Patient",
    "Entity Type_Employee",
    "Entity Type_Other",
    "Is Minor",
    "Is Deceased",
    "Street Address",
    "City",
    "State",
    "Zip Code",
    "International Street Address",
    "International City",
    "International State/Province",
    "Postal Code",
    "Country",
    "Email Address",
    "Phone Number",
    "Date of Birth",
    "Social Security Number (SSN)",
    "Individual Taxpayer Identification Number (ITIN)",
    "Driver's License Number",
    "Tax ID PIN",
    "Other Government ID Number",
    "Passport Number",
    "State ID Number",
    "Student ID Number (CO, WA, DC)",
    "Employee ID and Password (ND,SD)",
    "Fin Acct Number ONLY",
    "Fin Acct Number WITH Access",
    "Fin Acct Number Last Four Digits",
    "Payment Card Number ONLY",
    "Payment Card WITH Access",
    "Payment Card Number Last Four Digits",
    "Username/Email Address and Password for Non-Fin Acc",
    "Username/Email Address and Password for Fin Acc",
    "Diagnosis/Diagnosis Code",
    "Health Insurance Individual Policy Number",
    "Medical History/Treatment/Condition/Prescription",
    "Hospital Unit/Physician Name",
    "Date of Service",
    "Medical Record Number (MRN)",
    "Patient ID",
    "Patient Account Number (PAN)",
    "Other Unique Health Identifier",
    "Medicare/Medicaid Number",
    "Patient Personal Representative",
    "Referral",
    "Biometric Identifiers",
    "Digital Cryptographic Signatures (AZ, NC, ND, WA)",
    "Certificate/License Number",
    "DNA Profile",
    "Full Face Photo",
    "IP Address",
    "Medical Device Identifier",
    "URL",
    "Vehicle ID Number",
    "Birth OR Marriage Certificate (WY)",
    "General Tax Info",
    "Parent's Birth Name (ND, NC)",
    "Work-Related Evaluations (PR)",
    "Review Comments",
]

ON_HEADER_MISMATCH = "warn"     # what to do when a sheet's row 1 is not an exact
                                # match for EXPECTED_HEADERS:
                                # "warn"    = say so and append the rows anyway,
                                #             untouched (the "as it is" choice)
                                # "realign" = move each column into the expected
                                #             order, so a file with its columns
                                #             shuffled still lines up. Missing
                                #             columns are written blank.
                                # "skip"    = do not append that sheet at all
                                # "stop"    = stop the whole run

FAST_SCAN_UNDECLARED = False    # Some files (usually written by an export tool
                                # rather than by Excel) do not declare how big
                                # they are. For those, the default counts the rows
                                # first, so the row count and the column count are
                                # both right - that file gets read twice.
                                # True = do not count, guess the size from the file
                                # size and take the width from the first 500 rows.
                                # One read instead of two, but the progress bar is
                                # approximate and an unusually wide row further
                                # down the sheet would be missed.

WIDTH_PROBE_ROWS = 500          # rows read to measure the width when FAST_SCAN is on

KEEP_TRAILING_BLANK_ROWS = True  # True  = blank rows at the bottom of a sheet are
                                 #         appended as blank rows
                                 # False = drop them (Excel often leaves hundreds
                                 #         of empty rows behind in a used range)

QUOTE_EVERY_FIELD = True        # wrap every field in " " so a reader sees text

FORCE_TEXT_FOR_EXCEL = False    # True writes ="0071" instead of 0071, which stops
                                # Excel re-typing a column when you double click
                                # the CSV. It makes the file harder to use in other
                                # tools, so it is OFF by default.

ENCODING = "utf-8-sig"          # utf-8 with BOM, so Excel opens accents correctly

LINE_ENDING = "\r\n"            # Windows line endings

# How dates and times are written. Once a value is text, Excel has no idea what a
# date "looks like", so these formats decide it.
DATE_FMT = "%Y-%m-%d"
DATETIME_FMT = "%Y-%m-%d %H:%M:%S"
TIME_FMT = "%H:%M:%S"

WRITE_RUN_REPORT = True         # write a second CSV listing rows per file / sheet

OPENPYXL_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
XLS_EXTS = {".xls"}
XLSB_EXTS = {".xlsb"}
INPUT_EXTS = OPENPYXL_EXTS | XLS_EXTS | XLSB_EXTS


# ---------------------------------------------------------------------------
# Imports that need a friendly error message
# ---------------------------------------------------------------------------

try:
    import openpyxl
except ImportError:                                          # pragma: no cover
    sys.exit("openpyxl is not installed.  Run:  pip install openpyxl xlrd")


# ---------------------------------------------------------------------------
# Cell -> text
# ---------------------------------------------------------------------------

def cell_to_text(value) -> str:
    """Turn one cell into the text that goes in the CSV, changing nothing else."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return str(value)
        # whole numbers must not turn into 1.0 or 1.23e+15
        if value.is_integer() and abs(value) < 1e18:
            return f"{value:.0f}"
        return repr(value)
    if isinstance(value, dt.datetime):
        if value.time() == dt.time(0, 0):
            return value.strftime(DATE_FMT)
        return value.strftime(DATETIME_FMT)
    if isinstance(value, dt.date):
        return value.strftime(DATE_FMT)
    if isinstance(value, dt.time):
        return value.strftime(TIME_FMT)
    if isinstance(value, dt.timedelta):
        return str(value)
    if isinstance(value, numbers.Number):
        return str(value)
    return str(value)


def for_csv(text: str) -> str:
    if FORCE_TEXT_FOR_EXCEL and text != "":
        return '="' + text.replace('"', '""') + '"'
    return text


def normalise_header(text) -> str:
    """Compare header names without tripping over case or double spaces."""
    return " ".join(str(text).split()).casefold()


def describe_header_mismatch(header_row: list[str]) -> tuple[str, list[int | None] | None]:
    """Compare one sheet's row 1 against EXPECTED_HEADERS.

    Returns (description, mapping). The description is "" when the header is an
    exact match. The mapping says, for each expected column, which column of this
    file holds it - None meaning the file does not have that column at all.
    """
    wanted = [normalise_header(name) for name in EXPECTED_HEADERS]
    found = [normalise_header(value) for value in header_row]
    while found and found[-1] == "":         # ignore padding on the right
        found.pop()

    if found == wanted:
        return "", None

    missing = [EXPECTED_HEADERS[i] for i, name in enumerate(wanted) if name not in found]
    extra = [header_row[i] for i, name in enumerate(found) if name not in wanted]
    parts = []
    if len(found) != len(wanted):
        parts.append(f"{len(found)} columns, expected {len(wanted)}")
    if missing:
        parts.append(f"missing: {', '.join(missing[:5])}"
                     + (f" (+{len(missing) - 5} more)" if len(missing) > 5 else ""))
    if extra:
        parts.append(f"not expected: {', '.join(extra[:5])}"
                     + (f" (+{len(extra) - 5} more)" if len(extra) > 5 else ""))
    if not parts:
        # same names, different order - this is the dangerous one
        moved = [EXPECTED_HEADERS[i] for i, name in enumerate(wanted)
                 if i < len(found) and found[i] != name]
        parts.append(f"same columns in a different order, from: {', '.join(moved[:5])}"
                     + (f" (+{len(moved) - 5} more)" if len(moved) > 5 else ""))

    mapping = [found.index(name) if name in found else None for name in wanted]
    return "; ".join(parts), mapping


def realign(row: list[str], mapping: list[int | None]) -> list[str]:
    return [row[index] if index is not None and index < len(row) else ""
            for index in mapping]


def row_is_blank(row) -> bool:
    for value in row:
        if value != "":
            return False
    return True


# ---------------------------------------------------------------------------
# Progress bar (works in IDLE, which cannot redraw a line)
# ---------------------------------------------------------------------------

def _inline_supported() -> bool:
    if "idlelib" in sys.modules:        # IDLE prints \r as a new line
        return False
    try:
        return bool(sys.stdout.isatty())
    except Exception:
        return False


def _hms(seconds: float) -> str:
    seconds = int(max(seconds, 0))
    return f"{seconds // 3600:02d}:{seconds % 3600 // 60:02d}:{seconds % 60:02d}"


class ProgressBar:
    """Row based progress bar. Prints one line per update when run from IDLE."""

    BAR_WIDTH = 34

    def __init__(self, total_rows: int, exact: bool = True, label: str = "Merging"):
        self.total = max(int(total_rows), 1)
        self.exact = exact
        self.label = label
        self.done = 0
        self.start = time.time()
        self.inline = _inline_supported()
        self.min_gap = 0.2 if self.inline else 2.0
        self.last_draw = 0.0
        self.last_pct = -1.0
        self.draw(force=True)

    def advance(self, rows: int = 1) -> None:
        self.done += rows
        now = time.time()
        if now - self.last_draw < self.min_gap:
            return
        if not self.inline and 100.0 * self.done / self.total - self.last_pct < 1.0:
            return
        self.draw()

    def draw(self, force: bool = False) -> None:
        now = time.time()
        pct = 100.0 * self.done / self.total
        if not force:
            pct = min(pct, 99.9)        # only show 100% when it really is finished
        filled = int(self.BAR_WIDTH * min(pct, 100.0) / 100.0)
        bar = "#" * filled + "-" * (self.BAR_WIDTH - filled)
        elapsed = now - self.start
        eta = _hms(elapsed * (100.0 - pct) / pct) if pct > 0 else "--:--:--"
        about = "" if self.exact else "~"
        line = (f"{self.label}  [{bar}] {pct:5.1f}%   "
                f"{self.done:,} / {about}{self.total:,} rows   "
                f"elapsed {_hms(elapsed)}   eta {eta}")
        if self.inline:
            sys.stdout.write("\r" + line + "   ")
        else:
            sys.stdout.write(line + "\n")
        sys.stdout.flush()
        self.last_draw = now
        self.last_pct = pct

    def close(self) -> None:
        self.total = max(self.done, 1)
        self.exact = True
        self.draw(force=True)
        if self.inline:
            sys.stdout.write("\n")
        sys.stdout.flush()


# ---------------------------------------------------------------------------
# Scanning - how many rows and columns are we about to read
# ---------------------------------------------------------------------------

class SheetPlan:
    """One sheet that will be appended, plus its size as reported by the file."""

    def __init__(self, name, n_rows, n_cols, exact=True):
        self.name = name
        self.n_rows = n_rows
        self.n_cols = n_cols
        self.exact = exact          # False = size is a guess, so the bar is approximate


def count_rows(row_iter, limit: int = 0) -> tuple[int, int]:
    """Count rows and find the widest row. Stops early when limit is set.

    Needed for files that do not declare their used range. Reading a sheet
    without telling openpyxl the bounds keeps every row, including blank ones,
    but it trims the empty cells off the end of a row - so the width has to be
    measured here or short rows end up ragged in the CSV.
    """
    n_rows = 0
    n_cols = 0
    for row in row_iter:
        n_rows += 1
        if len(row) > n_cols:
            n_cols = len(row)
        if limit and n_rows >= limit:
            break
    return n_rows, n_cols


def scan_openpyxl(path: Path, log) -> list[SheetPlan]:
    plans: list[SheetPlan] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        names = wb.sheetnames if SHEETS == "all" else wb.sheetnames[:1]
        for name in names:
            ws = wb[name]
            n_rows, n_cols, exact = ws.max_row, ws.max_column, True
            if not n_rows or not n_cols:
                if FAST_SCAN_UNDECLARED:
                    log(f"    {name}: size is not declared in this file - guessing "
                        f"the row count, width from the first {WIDTH_PROBE_ROWS} rows")
                    _, n_cols = count_rows(ws.iter_rows(values_only=True),
                                           limit=WIDTH_PROBE_ROWS)
                    n_rows = max(int(path.stat().st_size / 120), 1)
                    exact = False
                else:
                    log(f"    {name}: size is not declared in this file - counting "
                        f"the rows first, so this file is read twice")
                    n_rows, n_cols = count_rows(ws.iter_rows(values_only=True))
            plans.append(SheetPlan(name, n_rows, n_cols, exact))
    finally:
        wb.close()
    return plans


def scan_xls(path: Path, log) -> list[SheetPlan]:
    import xlrd
    plans: list[SheetPlan] = []
    book = xlrd.open_workbook(path, on_demand=True)
    try:
        names = book.sheet_names() if SHEETS == "all" else book.sheet_names()[:1]
        for name in names:
            sh = book.sheet_by_name(name)
            plans.append(SheetPlan(name, sh.nrows, sh.ncols, True))
            book.unload_sheet(name)
    finally:
        book.release_resources()
    return plans


def scan_xlsb(path: Path, log) -> list[SheetPlan]:
    from pyxlsb import open_workbook
    plans: list[SheetPlan] = []
    with open_workbook(str(path)) as book:
        names = book.sheets if SHEETS == "all" else book.sheets[:1]
        for name in names:
            # .xlsb does not report its size either, so it is counted the same way
            with book.get_sheet(name) as sh:
                limit = WIDTH_PROBE_ROWS if FAST_SCAN_UNDECLARED else 0
                n_rows, n_cols = count_rows(sh.rows(), limit=limit)
            if FAST_SCAN_UNDECLARED:
                n_rows = max(int(path.stat().st_size / 90), 1)
            plans.append(SheetPlan(name, n_rows, n_cols, not FAST_SCAN_UNDECLARED))
    return plans


def scan_file(path: Path, log) -> list[SheetPlan]:
    ext = path.suffix.lower()
    if ext in OPENPYXL_EXTS:
        return scan_openpyxl(path, log)
    if ext in XLS_EXTS:
        return scan_xls(path, log)
    if ext in XLSB_EXTS:
        return scan_xlsb(path, log)
    return []


# ---------------------------------------------------------------------------
# Streaming - the rows themselves
# ---------------------------------------------------------------------------

def _openpyxl_rows(ws, plan: SheetPlan, width: int):
    if plan.exact and plan.n_rows and plan.n_cols:
        # Giving explicit bounds makes openpyxl fill in the rows and columns that
        # are missing from the XML. That is what keeps a blank row blank instead
        # of letting it disappear.
        rows = ws.iter_rows(min_row=1, max_row=plan.n_rows,
                            min_col=1, max_col=max(plan.n_cols, width),
                            values_only=True)
    else:
        rows = ws.iter_rows(values_only=True)
    for row in rows:
        yield [cell_to_text(value) for value in row]


def stream_openpyxl(path: Path, plans: list[SheetPlan], width: int):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        for plan in plans:
            yield plan, _openpyxl_rows(wb[plan.name], plan, width)
    finally:
        wb.close()


def _xls_rows(book, sh, n_cols: int):
    import xlrd
    for r in range(sh.nrows):
        out = []
        for c in range(n_cols):
            if c >= sh.ncols:
                out.append("")
                continue
            ctype = sh.cell_type(r, c)
            value = sh.cell_value(r, c)
            if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                out.append("")
            elif ctype == xlrd.XL_CELL_DATE:
                out.append(cell_to_text(xlrd.xldate_as_datetime(value, book.datemode)))
            elif ctype == xlrd.XL_CELL_BOOLEAN:
                out.append("TRUE" if value else "FALSE")
            elif ctype == xlrd.XL_CELL_ERROR:
                out.append("#ERROR")
            else:
                out.append(cell_to_text(value))
        yield out


def stream_xls(path: Path, plans: list[SheetPlan], width: int):
    import xlrd
    book = xlrd.open_workbook(path, on_demand=True)
    try:
        for plan in plans:
            sh = book.sheet_by_name(plan.name)
            yield plan, _xls_rows(book, sh, max(plan.n_cols, width))
            book.unload_sheet(plan.name)
    finally:
        book.release_resources()


def stream_xlsb(path: Path, plans: list[SheetPlan], width: int):
    from pyxlsb import open_workbook
    with open_workbook(str(path)) as book:
        for plan in plans:
            with book.get_sheet(plan.name) as sh:
                yield plan, ([cell_to_text(cell.v) for cell in row] for row in sh.rows())


def stream_file(path: Path, plans: list[SheetPlan], width: int):
    ext = path.suffix.lower()
    if ext in OPENPYXL_EXTS:
        return stream_openpyxl(path, plans, width)
    if ext in XLS_EXTS:
        return stream_xls(path, plans, width)
    return stream_xlsb(path, plans, width)


# ---------------------------------------------------------------------------
# The merge
# ---------------------------------------------------------------------------

def find_input_files(folder: Path) -> list[Path]:
    pattern = "**/*" if RECURSE_SUBFOLDERS else "*"
    found = []
    for path in folder.glob(pattern):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):          # Excel lock file
            continue
        if path.suffix.lower() in INPUT_EXTS:
            found.append(path)
    return sorted(found, key=lambda p: str(p).lower())


def scan_all(files: list[Path], input_dir: Path, report_rows: list[dict]):
    """Pass 1 - ask every file how big it is, so the progress bar means something."""
    plan_by_file: dict[Path, list[SheetPlan]] = {}
    total_rows = 0
    width = 0
    exact = True

    for index, path in enumerate(files, 1):
        rel = str(path.relative_to(input_dir))
        print(f"  [{index}/{len(files)}] {rel}")
        try:
            plans = scan_file(path, print)
        except ImportError as exc:
            print(f"    SKIPPED - {exc}")
            report_rows.append({"File": rel, "Sheet": "", "Rows": 0,
                                "Note": f"skipped, {exc}"})
            continue
        except Exception as exc:
            print(f"    SKIPPED - could not be opened: {exc}")
            report_rows.append({"File": rel, "Sheet": "", "Rows": 0,
                                "Note": f"skipped, could not be opened: {exc}"})
            continue
        if not plans:
            continue
        plan_by_file[path] = plans
        for plan in plans:
            total_rows += plan.n_rows
            width = max(width, plan.n_cols)
            exact = exact and plan.exact
        print(f"    {len(plans)} sheet(s), {sum(p.n_rows for p in plans):,} rows, "
              f"{max(p.n_cols for p in plans)} columns")

    return plan_by_file, total_rows, max(width, 1), exact


def merge(input_dir: Path, output_dir: Path) -> Path | None:
    stamp = dt.date.today().strftime("%y%m%d")
    output_dir.mkdir(parents=True, exist_ok=True)
    out_csv = output_dir / f"{stamp} {AUTHOR_INITIALS} {OUTPUT_DESCRIPTION}.csv"
    report_rows: list[dict] = []

    files = find_input_files(input_dir)
    if not files:
        print(f"No Excel files found in {input_dir}")
        return None

    print(f"Found {len(files)} file(s) in {input_dir}")
    print("Reading the size of every sheet first, so the progress bar is real...\n")
    plan_by_file, total_rows, width, exact = scan_all(files, input_dir, report_rows)

    if not plan_by_file:
        print("\nNothing could be read.")
        return None

    if EXPECTED_HEADERS:
        # never write fewer columns than the expected header list, so a realigned
        # sheet and an untouched one still produce the same shaped CSV
        width = max(width, len(EXPECTED_HEADERS))

    print(f"\nTotal to merge: {'' if exact else 'about '}{total_rows:,} rows "
          f"x {width} columns")
    print(f"Writing: {out_csv}\n")

    quoting = csv.QUOTE_ALL if QUOTE_EVERY_FIELD else csv.QUOTE_MINIMAL
    bar = ProgressBar(total_rows, exact=exact)
    written = 0
    first_sheet = True

    with open(out_csv, "w", newline="", encoding=ENCODING, buffering=1 << 20) as handle:
        writer = csv.writer(handle, quoting=quoting, lineterminator=LINE_ENDING)

        for path in files:
            plans = plan_by_file.get(path)
            if not plans:
                continue
            rel = str(path.relative_to(input_dir))
            try:
                for plan, rows in stream_file(path, plans, width):
                    sheet_written = 0
                    to_skip = 0 if first_sheet else HEADER_ROWS
                    pending_blanks: list[list[str]] = []
                    row_number = 0
                    mapping: list[int | None] | None = None
                    drop_sheet = False

                    for row in rows:
                        bar.advance()
                        row_number += 1

                        if row_number == 1 and EXPECTED_HEADERS:
                            problem, found_mapping = describe_header_mismatch(row)
                            if problem:
                                label = f"{rel} [{plan.name}]"
                                print(f"\n  HEADER MISMATCH  {label}\n"
                                      f"      {problem}")
                                report_rows.append(
                                    {"File": rel, "Sheet": plan.name, "Rows": 0,
                                     "Note": f"header mismatch - {problem}"})
                                if ON_HEADER_MISMATCH == "stop":
                                    raise RuntimeError(
                                        f"header mismatch in {label}: {problem}")
                                if ON_HEADER_MISMATCH == "skip":
                                    print("      this sheet was NOT appended")
                                    drop_sheet = True
                                elif ON_HEADER_MISMATCH == "realign":
                                    print("      columns moved into the expected order")
                                    mapping = found_mapping
                                else:
                                    print("      appended anyway, exactly as it is")

                        if drop_sheet:
                            continue
                        if mapping is not None:
                            row = realign(row, mapping)
                        if to_skip > 0:
                            to_skip -= 1
                            continue
                        if len(row) < width:
                            row = row + [""] * (width - len(row))
                        if not KEEP_TRAILING_BLANK_ROWS and row_is_blank(row):
                            # hold it back - if data follows, it was not trailing
                            pending_blanks.append(row)
                            continue
                        if pending_blanks:
                            for blank in pending_blanks:
                                writer.writerow([for_csv(v) for v in blank])
                            sheet_written += len(pending_blanks)
                            pending_blanks.clear()
                        writer.writerow([for_csv(v) for v in row])
                        sheet_written += 1

                    written += sheet_written
                    first_sheet = False
                    report_rows.append({"File": rel, "Sheet": plan.name,
                                        "Rows": sheet_written, "Note": ""})
            except Exception as exc:
                print(f"\n  PROBLEM in {rel}: {exc}")
                print("  The rows read before this point are already in the CSV.")
                report_rows.append({"File": rel, "Sheet": "", "Rows": 0,
                                    "Note": f"failed part way: {exc}"})

    bar.close()
    print(f"\nDone. {written:,} rows written to:\n    {out_csv}")
    print(f"    {out_csv.stat().st_size / (1024 * 1024):,.1f} MB")

    if WRITE_RUN_REPORT:
        report_path = output_dir / (f"{stamp} {AUTHOR_INITIALS} {OUTPUT_DESCRIPTION} "
                                   f"run report.csv")
        with open(report_path, "w", newline="", encoding=ENCODING) as handle:
            report = csv.DictWriter(handle, fieldnames=["File", "Sheet", "Rows", "Note"],
                                    lineterminator=LINE_ENDING)
            report.writeheader()
            report.writerows(report_rows)
            report.writerow({"File": "(all files)", "Sheet": "", "Rows": written,
                             "Note": "total rows written"})
        print(f"    run report: {report_path}")

    print("\nReminder: if this data is confidential, keep the output in the approved "
          "Global Insider folder, and do not commit the CSV to the repository.")
    return out_csv


# ---------------------------------------------------------------------------
# Self test - builds throwaway workbooks with a decoy and checks the CSV
# ---------------------------------------------------------------------------

def _strip_row_element(source: Path, target: Path, row_number: int) -> None:
    """Copy a workbook, deleting one <row> element so that row becomes a true gap.

    This fakes what some export tools produce: a blank row that has no XML at all.
    """
    import re
    import zipfile

    member = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(source) as zin:
        with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == member:
                    xml = data.decode("utf-8")
                    xml = re.sub(rf'<row r="{row_number}"[^>]*/>', "", xml)
                    xml = re.sub(rf'<row r="{row_number}".*?</row>', "", xml, flags=re.S)
                    data = xml.encode("utf-8")
                zout.writestr(item, data)


def selftest() -> int:
    """Prove the rules on fabricated data. No real / confidential data is used."""
    global EXPECTED_HEADERS, ON_HEADER_MISMATCH
    import shutil
    import tempfile

    work = Path(tempfile.mkdtemp(prefix="merge_selftest_"))
    src = work / "in"
    out = work / "out"
    src.mkdir()
    print(f"Self test working folder: {work}\n")

    # the test files have their own small header, not the real 65 column one
    EXPECTED_HEADERS = ["Emp ID", "Name", "Amount", "Joined"]
    ON_HEADER_MISMATCH = "warn"

    # File 1: normal data, a blank row in the middle, two blank rows at the end
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Emp ID", "Name", "Amount", "Joined"])
    ws.append(["007", "Aditi Rao", 1500.5, dt.date(2024, 4, 1)])
    ws.append([None, None, None, None])                       # blank row in the middle
    ws.append(["1234567890123456", "Ravi Kumar", 200, dt.datetime(2024, 4, 2, 9, 30)])
    ws.append([None, None, None, None])                       # trailing blanks
    ws.append([None, None, None, None])
    # decoy: a second sheet that must also be appended, with a different width
    ws2 = wb.create_sheet("Extra")
    ws2.append(["Emp ID", "Name"])
    ws2.append(["008", "Sunita Devi"])
    wb.save(src / "file a.xlsx")

    # File 2: starts with a blank row, has a hidden row, unicode and a comma
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([None, None, None, None])                       # leading blank row
    ws.append(["Emp ID", "Name", "Amount", "Joined"])          # its own header, kept
    ws.append(["009", "Rao, Meena", 0, dt.date(2024, 5, 9)])
    ws.append(["010", "Jose Muñoz", True, None])
    ws.row_dimensions[4].hidden = True                         # hidden stays in
    wb.save(src / "file b.xlsx")

    # File 3: the shape an export tool produces - the used range is not declared,
    # rows end in empty cells, and one blank row has no XML row element at all.
    # This is the case that silently produced ragged rows before it was fixed.
    raw = work / "raw c.xlsx"
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append(["Emp ID", "Name", "Amount", "Joined"])
    ws.append(["011", "Blank Trail", None, None])              # trailing empty cells
    ws.append([None, None, None, None])                        # becomes a true gap
    ws.append(["012", None, None, None])
    wb.save(raw)
    _strip_row_element(raw, src / "file c.xlsx", row_number=3)

    # decoy: an Excel lock file that must be ignored
    (src / "~$file b.xlsx").write_text("lock", encoding="utf-8")

    out_csv = merge(src, out)
    if out_csv is None:
        print("FAILED: nothing was written")
        return 1

    with open(out_csv, newline="", encoding=ENCODING) as handle:
        rows = list(csv.reader(handle))

    expected = [
        ["Emp ID", "Name", "Amount", "Joined"],
        ["007", "Aditi Rao", "1500.5", "2024-04-01"],
        ["", "", "", ""],
        ["1234567890123456", "Ravi Kumar", "200", "2024-04-02 09:30:00"],
        ["", "", "", ""],
        ["", "", "", ""],
        ["Emp ID", "Name", "", ""],
        ["008", "Sunita Devi", "", ""],
        ["", "", "", ""],
        ["Emp ID", "Name", "Amount", "Joined"],
        ["009", "Rao, Meena", "0", "2024-05-09"],
        ["010", "Jose Muñoz", "TRUE", ""],
        ["Emp ID", "Name", "Amount", "Joined"],
        ["011", "Blank Trail", "", ""],
        ["", "", "", ""],
        ["012", "", "", ""],
    ]

    failures = []
    if len(rows) != len(expected):
        failures.append(f"row count is {len(rows)}, expected {len(expected)}")
    for index, want in enumerate(expected):
        got = rows[index] if index < len(rows) else None
        if got != want:
            failures.append(f"row {index + 1}: got {got}, expected {want}")

    ragged = {len(row) for row in rows} - {len(expected[0])}
    if ragged:
        failures.append(f"not every row has {len(expected[0])} fields, also saw {ragged}")

    raw_csv = open(out_csv, newline="", encoding=ENCODING).read()
    if '"007"' not in raw_csv:
        failures.append("leading zeros were not written as quoted text")

    print()
    for line in failures:
        print(f"  FAIL  {line}")
    if failures:
        print(f"\nSELF TEST FAILED ({len(failures)} problem(s)). Files kept at {work}")
        return 1

    print(f"  PASS  {len(rows)} rows, all {len(expected[0])} fields wide, blank rows "
          f"kept, every sheet appended, leading zeros kept")

    # ---- second run: a file whose columns are in a different order ----------
    print("\nChecking the shuffled column case (ON_HEADER_MISMATCH = realign)...\n")
    src2 = work / "in2"
    out2 = work / "out2"
    src2.mkdir()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Emp ID", "Name", "Amount", "Joined"])
    ws.append(["020", "In Order", "50", "2024-06-01"])
    wb.save(src2 / "file a.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Joined", "Emp ID", "Amount", "Name"])      # same columns, shuffled
    ws.append(["2024-06-02", "021", "60", "Shuffled"])
    wb.save(src2 / "file b.xlsx")

    ON_HEADER_MISMATCH = "realign"
    out_csv2 = merge(src2, out2)
    with open(out_csv2, newline="", encoding=ENCODING) as handle:
        rows2 = list(csv.reader(handle))

    expected2 = [
        ["Emp ID", "Name", "Amount", "Joined"],
        ["020", "In Order", "50", "2024-06-01"],
        ["Emp ID", "Name", "Amount", "Joined"],
        ["021", "Shuffled", "60", "2024-06-02"],
    ]
    for index, want in enumerate(expected2):
        got = rows2[index] if index < len(rows2) else None
        if got != want:
            failures.append(f"realign row {index + 1}: got {got}, expected {want}")

    print()
    for line in failures:
        print(f"  FAIL  {line}")
    if failures:
        print(f"\nSELF TEST FAILED ({len(failures)} problem(s)). Files kept at {work}")
        return 1

    print("  PASS  shuffled columns were put back into the expected order")
    print("\nSELF TEST PASSED")
    shutil.rmtree(work, ignore_errors=True)
    return 0


# ---------------------------------------------------------------------------
# Folder pickers / entry point
# ---------------------------------------------------------------------------

def pick_folder(title: str) -> Path | None:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        return None
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    chosen = filedialog.askdirectory(title=title)
    root.destroy()
    return Path(chosen) if chosen else None


def main(argv: list[str]) -> int:
    if argv and argv[0] in ("--selftest", "-t"):
        return selftest()

    if len(argv) >= 2:
        input_dir = Path(argv[0]).expanduser().resolve()
        output_dir = Path(argv[1]).expanduser().resolve()
    else:
        print("Pick the folder that contains the Excel files to merge...")
        input_dir = pick_folder("Select the folder containing the Excel files")
        if not input_dir:
            print("Cancelled.")
            return 1
        print("Pick where the merged CSV should be saved...")
        output_dir = pick_folder("Select the folder to save the merged CSV in")
        if not output_dir:
            output_dir = input_dir / "Merged output"
            print(f"No output folder picked - using {output_dir}")

    if not input_dir.is_dir():
        print(f"Not a folder: {input_dir}")
        return 1

    try:
        merge(input_dir, output_dir)
    except Exception:
        traceback.print_exc()
        input("\nSomething went wrong. Press Enter to close...")
        return 1

    if len(argv) < 2:
        input("\nPress Enter to close...")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
