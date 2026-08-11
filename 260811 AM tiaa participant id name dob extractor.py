r"""
TIAA participant reports  --  identity extractor
================================================
Reads searchable TIAA-CREF participant PDFs and writes Part ID (SSN),
Name, Status, Date of Birth, Date of Death and Address to one combined
Excel workbook.

LAYOUTS, DETECTED PER PAGE
    A  JOINED TABLE     e.g. "Small Cash Out Monitoring Report"
       One participant per printed row, the name printed as one field:

         PART ID       NAME                DIV/LOC  STATUS      DATE OF   DATE OF ...
                                                                BIRTH     HIRE
         123-45-6789   ACKERMAN, KENTON C  0001     Terminated  01/15/63  03/01/90

    B  LABEL           e.g. "Participant Contribution Investment Allocation Report"
       One participant per block of "LABEL: value" pairs:

         NAME: ABBEY, BORIS S   DATE OF BIRTH: 04/12/1973  DATE OF TERMINATION:
         PART ID: 987654321     DATE OF HIRE: 06/01/2015   DATE OF DEATH:
         STATUS: Active         DATE OF ENTRY: 07/01/2015  DATE OF REHIRE:
         ADDRESS: 136 BETHESDA RD                          DIV/LOC:
                  STATESVILLE, NC 28677

    C  SPLIT TABLE     e.g. "Terminated in prior year with a vested balance"
       Same as A, but the identity columns are captioned differently: the
       Part ID column is headed SSN, and the name arrives already split:

         ITEM          MIDDLE                        DATE OF   DATE OF
         COUNT  SSN    FIRST NAME  INITIAL LAST NAME STATUS    TERMINATION ...
         1      ...    CHRISTINE   R       CHAN      Terminated 02/28/2014

    Nothing about the parsing is layout-specific beyond the captions --
    A and C are the same code reading a different header. Detection is per
    PAGE, not per file, so a single PDF may hold all three. A page that is
    none of them (a detail or continuation page with no participant
    identity on it) simply produces no rows.

WHY BLOCKS ARE GROUPED BY POSITION, NOT BY CAPTION
    These reports OMIT a caption entirely when its value is empty. A
    participant with no name on file prints no "NAME:" at all -- not
    "NAME:" followed by blank. Same for PART ID and the rest.

    So a participant block cannot be delimited by watching for a caption
    to repeat. The caption that would have marked the boundary is exactly
    the one that is missing, and the next participant's name gets read
    backwards onto the previous participant's Part ID:

         printed                        read as
         PART ID: ...4444               ...4444  CLIFTON, BERA S   <- wrong
         DATE OF BIRTH: 03/04/1971               person
         NAME: CLIFTON, BERA S          (row lost entirely)
         DATE OF BIRTH: 05/06/1972

    Vertical position is never omitted. Blocks are therefore grouped by
    vertical gap and each is parsed in isolation, so a missing caption
    leaves a blank field and cannot leak a value across a boundary. A
    repeated caption inside one block still splits it, which covers two
    blocks printed with no gap between them.

    selftest() asserts this on a synthetic page built from the case above.
    Raising BLOCK_GAP_HEIGHTS until the grouping collapses reproduces the
    original fault, which is how that test was checked to have teeth.

RECORDS THAT SPAN A PAGE BREAK
    A participant can start at the foot of one page and finish at the head
    of the next -- the name printed on page 4, the Part ID on page 5.
    Read page by page those are two half-participants.

    So the last record on a page is carried forward, and if it is
    incomplete it is offered the first record of the next page. The merge
    only happens when the two share NO field with two different values.
    On any conflict both are written separately and flagged, because
    fusing two different people into one row is worse than leaving a split
    one for a human to join.

    The same carry joins a name that wrapped across a page break in the
    table layouts.

NAMES THAT WRAP ONTO A SECOND LINE
    A long name in the table layouts prints across two lines, the second
    holding nothing but the rest of the name. That line has no Part ID, no
    dates and no status, and every word on it falls inside a name column
    -- which is how it is recognised. It is appended to the row above with
    a space and the row is flagged, rather than being dropped or, worse,
    written out as a nameless participant of its own.

WHY THE DATE COLUMN NEEDS GEOMETRY
    The table layouts print up to five date columns -- BIRTH, DEATH, HIRE,
    REHIRE, TERMINATION -- and any of them may be blank on any given row.
    Reading "the first date on the line" is therefore wrong the moment a
    birth date is missing: the next date slides left into its place and is
    silently written out as a date of birth. Nothing about the resulting
    workbook looks broken.

    So the column header band is read first, clustered into columns by x
    position, and every date token on a data row is assigned to whichever
    column caption it sits under. Only the token under BIRTH is kept as a
    date of birth, only the one under DEATH as a date of death.

    Header captions stack across up to three printed lines ("DATE OF"
    above "TERMINATION" above "PROCESSED"), so the band is grown upward
    from the caption row while the line above still looks like captions.

    Continuation pages that repeat no header reuse the last column map
    seen in the same file. If no header has been seen at all, the parser
    falls back to pattern order (SSN, then name up to the first date or
    status word, then the leftmost date) and every row it produces is
    flagged in the Notes column, because that fallback is exactly the
    guess described above.

A BLANK PART ID KEEPS THE PARTICIPANT
    A row or block with no Part ID is still written out, with the cell
    empty, the reason in Notes and a count in Reconciliation. Dropping it
    would be a silent loss. In the table layouts such a row is only
    accepted if it still looks like a participant -- a name plus a date or
    a status -- otherwise report titles and page footers would become rows.

LINE RECONSTRUCTION
    Lines are rebuilt from extract_words() coordinates and clustered by y.
    page.extract_text(layout=True) is NOT used anywhere: it paints
    characters into a fixed grid sized from page.width, and on a wide
    landscape report characters landing in the same cell are dropped.
    That failure is page-geometry dependent, so it shows up on some files
    and not others -- the worst kind of extraction bug to inherit.

ENGINES
    pdfplumber is the default. PyMuPDF reads the same positioned words
    far faster and is available as --engine mupdf, but speed here comes
    from processing files in parallel, which cannot change WHAT is read,
    rather than from swapping the engine, which can. Run --verify over
    your own files first if you intend to switch.

COMPLETENESS
    The Reconciliation sheet counts, per file, how many Part IDs were
    sitting on the pages against how many rows were written. A file is
    Complete only when

        rows written == Part IDs on page + rows with no Part ID printed
        no row is missing a name
        no row had dates on it but none under the BIRTH column
        no name fragment was left unattached

    A blank date of birth does NOT flag the file. Two of these reports
    carry no DATE OF BIRTH column at all, and a flag that fires on every
    file is a flag people learn to ignore. A date that failed to land
    under the BIRTH column does flag it, because that is a possible
    misread rather than an empty field. Both counts are on the sheet
    either way.

    Check that sheet before signing off a run. A workbook holding 40 of 47
    participants looks completely healthy otherwise -- there is nothing in
    it that says "incomplete".

HANDLING OF PERSONAL DATA
    Every row is name + SSN + date of birth + home address, which is a
    complete identity record rather than a re-identification key. Write it
    to a controlled location, not a desktop or a personal drive, give it a
    retention and disposal date, and limit access to those who need it.
    Nothing here should be pasted into chat tools or e-mail.

    The console, the QA dump and the log NEVER print an SSN or an address.
    --mask writes the workbook itself with masked Part IDs
    (***-**-6789) for copies that need to circulate more widely.

USAGE
    python "260811 AM tiaa participant id name dob extractor.py"
        Tkinter folder picker, progress bar, combined workbook.

    python "260811 AM tiaa participant id name dob extractor.py" \
        --src <folder> --out <file.xlsx> [--mask] [--engine pdfplumber|mupdf]
        Headless run, no window.

    ... --selftest                Build a synthetic PDF (fabricated data,
                                  no real PII) and assert the parse.
    ... --qa <file.pdf>           Masked dump of how each line was split.
    ... --lines <file.pdf> [--page N]
                                  Masked geometry dump: every rebuilt line,
                                  each word's x span and the column it was
                                  read as. Letters print as A and digits as
                                  9, so the output can be shared to diagnose
                                  a misread without moving any PII.
    ... --verify <folder>         Read every PDF with both engines and
                                  report disagreements. Counts only.

ADDING A REPORT VARIANT
    A new caption spelling needs one line, and nothing else:
        a table column   -> COLUMN_KEYWORDS
        a block caption  -> LABELS
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import traceback
import unicodedata
from concurrent.futures import ProcessPoolExecutor
from pathlib import Path

try:
    import pdfplumber
except ImportError:  # pragma: no cover - environment dependent
    pdfplumber = None

try:
    import pymupdf as fitz
except ImportError:  # pragma: no cover - environment dependent
    try:
        import fitz
    except ImportError:
        fitz = None

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# configuration
# ---------------------------------------------------------------------------

OUTPUT_XLSX_NAME = "260811 AM tiaa participant identity.xlsx"

DEFAULT_ENGINE = "pdfplumber"

OUTPUT_COLUMNS = [
    "File Name",
    "Page",
    "Layout",
    "Part ID (SSN)",
    "Part ID As Printed",
    "Name As Printed",
    "Last Name",
    "First Name",
    "Middle Name",
    "Suffix",
    "Status",
    "Date Of Birth",
    "Date Of Birth As Printed",
    "Date Of Death",
    "Date Of Death As Printed",
    "Address",
    "Notes",
]

RECON_COLUMNS = [
    "File Name",
    "Pages",
    "Layouts Seen",
    "Part IDs On Page",
    "Rows Written",
    "Rows Missing Name",
    "Rows Missing DOB",
    "Rows With Suspicious DOB",
    "Rows Without Part ID",
    "Rows With Non-SSN Part ID",
    "Unattached Name Fragments",
    "Complete",
    "Error",
]

# Written into Notes when a row carried dates but none of them sat under the
# DATE OF BIRTH column. Distinct from a row that printed no dates at all: the
# first is a possible misread, the second is just an empty field.
SUSPICIOUS_DOB_NOTE = "no date sat under the DATE OF BIRTH column"
NO_PART_ID_NOTE = "no Part ID printed"
NO_NAME_NOTE = "no name printed in this block"

# Notes that describe a field as missing, and so stop being true the moment
# two halves of a page-split record are joined.
_STALE_ON_MERGE = (NO_PART_ID_NOTE, NO_NAME_NOTE)

# horizontal gap, in points, that separates one header caption from the next.
# Captions inside one column ("DATE OF HIRE") sit a few points apart; adjacent
# columns on these reports are 20pt or more apart.
HEADER_COLUMN_GAP = 14.0

# Slack allowed at a column's left edge. Data is left aligned under its
# caption, so a column owns everything from just left of its own caption up to
# just left of the next one.
COLUMN_PAD = 2.0

# A caption row is grown upward while the line above still looks like part of
# the header. A band line may not sit further than this many glyph heights
# above, and no single caption on it may straddle this many columns -- which
# is what stops a section title spanning the table from being swallowed.
CAPTION_MAX_GAP_HEIGHTS = 2.5
CAPTION_MAX_COLUMNS_STRADDLED = 3

MIN_FILES_FOR_PARALLEL = 4


# ---------------------------------------------------------------------------
# text normalisation
# ---------------------------------------------------------------------------

# Every dash-like codepoint a PDF may carry, folded to ASCII "-" before any
# matching happens. A U+2010 in "123‐45‐6789" otherwise hides an SSN
# from every pattern below.
_DASHES = "‐‑‒–—―−﹘﹣－­"
_SPACES = "       "
_FOLD_MAP = {ord(c): "-" for c in _DASHES}
_FOLD_MAP.update({ord(c): " " for c in _SPACES})


def fold(text: str) -> str:
    """Normalise unicode punctuation and whitespace to ASCII equivalents."""
    if not text:
        return ""
    return unicodedata.normalize("NFKC", str(text)).translate(_FOLD_MAP)


def squeeze(text: str) -> str:
    """Collapse runs of whitespace to single spaces."""
    return re.sub(r"\s+", " ", fold(text)).strip()


SSN_TOKEN_RE = re.compile(
    r"^(?:"
    r"\d{3}-\d{2}-\d{4}"          # 123-45-6789
    r"|[Xx*]{3}-[Xx*]{2}-\d{4}"   # XXX-XX-6789
    r"|\d{9}"                     # 123456789
    r"|[Xx*]{5}\d{4}"             # XXXXX6789
    r")$"
)

SSN_ANY_RE = re.compile(
    r"(?<![\d-])(?:"
    r"\d{3}-\d{2}-\d{4}"
    r"|[Xx*]{3}-[Xx*]{2}-\d{4}"
    r"|\d{9}"
    r"|[Xx*]{5}\d{4}"
    r")(?![\d-])"
)

DATE_TOKEN_RE = re.compile(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2}|\d{4})$")
DATE_ANY_RE = re.compile(r"(?<!\d)(\d{1,2})[/-](\d{1,2})[/-](\d{2}|\d{4})(?!\d)")

STATUS_WORDS = {
    "ACTIVE", "TERMINATED", "TERMINATE", "RETIRED", "DECEASED", "INACTIVE",
    "LEAVE", "DISABLED", "SUSPENDED", "REHIRED", "ELIGIBLE", "INELIGIBLE",
}

NAME_SUFFIXES = {
    "JR", "SR", "II", "III", "IV", "V", "VI", "MD", "PHD", "DDS", "DVM",
    "ESQ", "CPA", "RN",
}

# Plausibility window for a date of birth. Anything outside it is still
# written out, but the row carries a note -- a wrong DOB that looks
# reasonable is far more dangerous than one that is obviously junk.
DOB_MIN_YEAR = 1890
DOB_MAX_YEAR = 2030


def norm_ssn(raw: str) -> str:
    """Return an SSN as 123-45-6789 (or XXX-XX-6789) regardless of input form."""
    compact = re.sub(r"[^0-9Xx*]", "", fold(raw))
    if len(compact) == 9:
        compact = compact.upper().replace("*", "X")
        return f"{compact[0:3]}-{compact[3:5]}-{compact[5:9]}"
    return squeeze(raw).upper()


def mask_ssn(value: str) -> str:
    """***-**-6789 -- safe for console, logs and circulated copies."""
    digits = re.sub(r"\D", "", fold(value))
    if len(digits) >= 4:
        return f"***-**-{digits[-4:]}"
    return "***-**-****" if value else ""


def norm_date(raw: str):
    """Return (normalised MM/DD/YYYY, note). Empty string if unparseable."""
    m = DATE_ANY_RE.search(fold(raw))
    if not m:
        return "", ""
    month, day, year = int(m.group(1)), int(m.group(2)), m.group(3)
    if len(year) == 2:
        # A two-digit year on a date of birth is a 20th century year far more
        # often than not; 30 is the usual pivot on these reports.
        year = 1900 + int(year) if int(year) > 30 else 2000 + int(year)
    else:
        year = int(year)
    if not (1 <= month <= 12) or not (1 <= day <= 31):
        return "", "date does not parse"
    note = ""
    if not (DOB_MIN_YEAR <= year <= DOB_MAX_YEAR):
        note = f"year {year} outside plausible range"
    return f"{month:02d}/{day:02d}/{year}", note


def split_name(printed: str):
    """Split "LAST, FIRST MIDDLE" into parts. Returns (last, first, middle,
    suffix, note)."""
    text = squeeze(printed).strip(" ,")
    if not text:
        # No note: whichever parser produced the blank has already said so,
        # and in more specific terms than this function could.
        return "", "", "", "", ""

    note = ""
    if "," in text:
        last_part, _, rest = text.partition(",")
    else:
        # No comma. The report format is Last, First -- so this is either a
        # single-word name or something the column split got wrong. Assume
        # "FIRST MIDDLE LAST" and flag it, rather than guessing silently.
        tokens = text.split()
        if len(tokens) == 1:
            return tokens[0], "", "", "", "no comma in name; treated as surname only"
        last_part, rest = tokens[-1], " ".join(tokens[:-1])
        note = "no comma in name; word order assumed First Middle Last"

    last_tokens = last_part.split()
    rest_tokens = rest.split()

    suffix = ""
    for bucket in (rest_tokens, last_tokens):
        while bucket and bucket[-1].strip(".").upper() in NAME_SUFFIXES:
            suffix = bucket.pop().strip(".").upper()

    last = " ".join(last_tokens)
    first = rest_tokens[0] if rest_tokens else ""
    middle = " ".join(rest_tokens[1:]) if len(rest_tokens) > 1 else ""
    if not last:
        note = (note + "; " if note else "") + "surname missing"
    return last, first, middle, suffix, note


_SSN_BODY = (
    r"\d{3}-\d{2}-\d{4}|\d{9}|[Xx*]{3}-[Xx*]{2}-\d{4}|[Xx*]{5}\d{4}"
)
GLUED_SSN_HEAD_RE = re.compile(r"^(" + _SSN_BODY + r")([A-Za-z].*)$")
GLUED_SSN_TAIL_RE = re.compile(r"^(.*[A-Za-z])(" + _SSN_BODY + r")$")


def split_glued(word):
    """Split a word that holds an SSN glued to text from the next column.

    When two columns are printed close enough together, the extractor
    returns "789-01-2345CHRISTINE" as a single word. The SSN still
    normalises correctly, so the Part ID looks right and the first name
    simply disappears -- a silent partial loss, and the workbook gives no
    sign of it.

    The x span is apportioned by character count, which is exact for the
    monospaced fonts these reports use and close enough elsewhere to put
    each piece in the right column.
    """
    text = word["text"]
    m = GLUED_SSN_HEAD_RE.match(text) or GLUED_SSN_TAIL_RE.match(text)
    if not m:
        return [word]
    span = word["x1"] - word["x0"]
    pieces, cursor = [], word["x0"]
    for part in (m.group(1), m.group(2)):
        width = span * len(part) / len(text)
        pieces.append({**word, "text": part, "x0": cursor, "x1": cursor + width,
                       "glued": True})
        cursor += width
    return pieces


def looks_like_part_id(text: str) -> bool:
    """True for anything that could be a printed Part ID.

    Deliberately looser than the SSN patterns: a row whose Part ID is not
    SSN-shaped is still a participant, and dropping it would be a silent
    loss. Loose enough to catch those, tight enough to reject page footers
    and running totals.
    """
    compact = re.sub(r"[^0-9A-Za-z*-]", "", fold(text))
    if len(compact) < 4:
        return False
    return len(re.findall(r"[0-9Xx*]", compact)) >= 4


# ---------------------------------------------------------------------------
# word extraction and line rebuilding
# ---------------------------------------------------------------------------

def words_pdfplumber(page):
    out = []
    for w in page.extract_words(use_text_flow=False, keep_blank_chars=False):
        out.append({
            "text": fold(w["text"]),
            "x0": float(w["x0"]),
            "x1": float(w["x1"]),
            "top": float(w["top"]),
            "bottom": float(w["bottom"]),
        })
    return out


def words_pymupdf(page):
    out = []
    for x0, y0, x1, y1, text, *_ in page.get_text("words"):
        out.append({
            "text": fold(text),
            "x0": float(x0),
            "x1": float(x1),
            "top": float(y0),
            "bottom": float(y1),
        })
    return out


def cluster_rows(words, y_tol=None):
    """Group words into printed lines by vertical centre.

    Tolerance defaults to half the median glyph height, so it adapts to the
    point size instead of assuming one.
    """
    if not words:
        return []
    tol = y_tol if y_tol is not None else max(1.5, median_height(words) * 0.5)

    ordered = sorted(words, key=lambda w: ((w["top"] + w["bottom"]) / 2.0, w["x0"]))
    rows, current, centre = [], [], None
    for w in ordered:
        c = (w["top"] + w["bottom"]) / 2.0
        if current and abs(c - centre) > tol:
            rows.append(sorted(current, key=lambda x: x["x0"]))
            current = []
        current.append(w)
        centre = sum((x["top"] + x["bottom"]) / 2.0 for x in current) / len(current)
    if current:
        rows.append(sorted(current, key=lambda x: x["x0"]))
    return rows


def median_height(words) -> float:
    heights = sorted(w["bottom"] - w["top"] for w in words)
    return heights[len(heights) // 2] if heights else 8.0


def row_text(row) -> str:
    return " ".join(w["text"] for w in row)


def row_centre(row) -> float:
    return sum((w["top"] + w["bottom"]) / 2.0 for w in row) / len(row)


def centre_x(word) -> float:
    return (word["x0"] + word["x1"]) / 2.0


def read_pages(path: Path, engine: str):
    """Yield (page_number, words) for every page. One place per engine."""
    if engine == "mupdf":
        if fitz is None:
            raise RuntimeError("PyMuPDF is not installed")
        with fitz.open(str(path)) as doc:
            for i, page in enumerate(doc, start=1):
                yield i, words_pymupdf(page)
    else:
        if pdfplumber is None:
            raise RuntimeError("pdfplumber is not installed")
        with pdfplumber.open(str(path)) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                yield i, words_pdfplumber(page)


def available_engines():
    found = []
    if pdfplumber is not None:
        found.append("pdfplumber")
    if fitz is not None:
        found.append("mupdf")
    return found


# ---------------------------------------------------------------------------
# the column header
# ---------------------------------------------------------------------------

# Checked in order, so the more specific caption wins. REHIRE must be tested
# before HIRE and the three split-name captions before plain NAME, or
# "DATE OF REHIRE" reads as the hire column and "FIRST NAME" as the name one.
COLUMN_KEYWORDS = [
    ("item_count", ("ITEM COUNT", "ITEM")),
    ("part_id", ("PART ID", "PARTID", "PARTICIPANT ID", "SSN",
                 "SOCIAL SECURITY")),
    ("first_name", ("FIRST NAME", "FIRST")),
    ("middle_name", ("MIDDLE INITIAL", "MIDDLE NAME", "MIDDLE")),
    ("last_name", ("LAST NAME", "LAST")),
    ("name", ("NAME",)),
    ("div_loc", ("DIV/LOC", "DIVLOC", "DIV / LOC", "DIV")),
    ("status", ("STATUS",)),
    ("dob", ("BIRTH",)),
    ("death", ("DEATH",)),
    ("rehire", ("REHIRE",)),
    ("hire", ("HIRE",)),
    ("term", ("TERMINATION", "TERM DATE")),
    ("balance", ("BALANCE",)),
]

DATE_COLUMN_KEYS = ("dob", "death", "hire", "rehire", "term")

# Columns that hold a piece of the participant's name, and the record field
# each fills.
NAME_COLUMN_FIELD = {
    "name": "name",
    "first_name": "first",
    "middle_name": "middle",
    "last_name": "last",
}

# What makes a row the header row: an identity caption and a name caption on
# the same printed line.
_ID_CAPTIONS = ("SSN", "PARTID")
_NAME_CAPTIONS = ("NAME", "FIRST", "LAST")


def cluster_columns(band, gap=HEADER_COLUMN_GAP):
    """Merge header words into columns by horizontal overlap / proximity."""
    clusters = []
    for w in sorted(band, key=lambda x: x["x0"]):
        if clusters and w["x0"] <= clusters[-1]["x1"] + gap:
            c = clusters[-1]
            c["x0"] = min(c["x0"], w["x0"])
            c["x1"] = max(c["x1"], w["x1"])
            c["words"].append(w)
        else:
            clusters.append({"x0": w["x0"], "x1": w["x1"], "words": [w]})
    for c in clusters:
        # Caption words are ordered left-to-right then top-to-bottom so
        # "DATE OF" + "BIRTH" reads as one caption.
        ordered = sorted(c["words"], key=lambda w: (w["top"], w["x0"]))
        c["text"] = " ".join(w["text"] for w in ordered).upper()
        c["centre"] = (c["x0"] + c["x1"]) / 2.0
    return clusters


def _is_anchor_row(row) -> bool:
    upper = [w["text"].upper() for w in row]
    flat = "".join(upper)
    has_id = ("PART" in upper and "ID" in upper) or any(c in flat for c in _ID_CAPTIONS)
    has_name = any(c in upper for c in _NAME_CAPTIONS)
    return has_id and has_name


def _is_caption_row(row, anchor_clusters) -> bool:
    """True when a line above the header still looks like part of the header.

    The discriminating test is straddling. A caption belongs to exactly one
    column, so it overlaps one or two of the columns below it. A section
    title printed across the table -- "SECTION 1: TERMINATED IN PRIOR YEAR
    WITH A VESTED BALANCE" -- overlaps most of them, and merging it into the
    band would fuse every column into one.
    """
    if len(row) < 2:
        return False
    for w in row:
        text = w["text"]
        if DATE_TOKEN_RE.match(text) or SSN_TOKEN_RE.match(text) or "$" in text:
            return False
    for cluster in cluster_columns(row):
        straddled = sum(
            1 for a in anchor_clusters
            if cluster["x0"] < a["x1"] and a["x0"] < cluster["x1"]
        )
        if straddled >= CAPTION_MAX_COLUMNS_STRADDLED:
            return False
    return True


def find_header_band(rows):
    """Return (header words, index of the header row) or (None, -1).

    The caption block spans up to three printed lines on these reports --
    "DATE OF" above "TERMINATION" above "PROCESSED" -- so the band is grown
    upward from the caption row for as long as the line above still reads
    as captions.
    """
    for i, row in enumerate(rows):
        if not _is_anchor_row(row):
            continue
        band = list(row)
        anchor_clusters = cluster_columns(row)
        height = median_height(row)
        top = row_centre(row)
        j = i - 1
        while j >= 0:
            candidate = rows[j]
            gap = top - row_centre(candidate)
            if gap > height * CAPTION_MAX_GAP_HEIGHTS:
                break
            if not _is_caption_row(candidate, anchor_clusters):
                break
            band += candidate
            top = row_centre(candidate)
            j -= 1
        return band, i
    return None, -1


def build_column_map(rows):
    """Return (column map, header row index). Map is key -> {x0,x1,centre}."""
    band, index = find_header_band(rows)
    if band is None:
        return None, -1

    colmap = {}
    for cluster in cluster_columns(band):
        text = cluster["text"]
        flat = text.replace(" ", "")
        for key, keywords in COLUMN_KEYWORDS:
            if key in colmap:
                continue
            if any(kw in text or kw.replace(" ", "") in flat for kw in keywords):
                colmap[key] = {
                    "x0": cluster["x0"],
                    "x1": cluster["x1"],
                    "centre": cluster["centre"],
                }
                break
    has_name = any(k in colmap for k in NAME_COLUMN_FIELD)
    if "part_id" not in colmap or not has_name:
        return None, -1
    return colmap, index


def make_column_of(colmap):
    """Return a function mapping a word to the column it was printed in.

    A column owns everything from just left of its own caption up to just
    left of the next caption. Data on these reports is left aligned under
    its caption and runs much wider than the caption itself -- "NAME" is
    four characters over a field holding "ACKERMAN, KENTON CLYDE" -- so a
    midpoint boundary would truncate names.
    """
    ordered = sorted(colmap.items(), key=lambda kv: kv[1]["x0"])
    bounds = []
    for i, (key, col) in enumerate(ordered):
        left = col["x0"] - COLUMN_PAD
        right = (ordered[i + 1][1]["x0"] - COLUMN_PAD
                 if i + 1 < len(ordered) else float("inf"))
        bounds.append((key, left, right))

    def column_of(word):
        x = centre_x(word)
        for key, left, right in bounds:
            if left <= x < right:
                return key
        return None

    return column_of


def nearest_date_column(colmap, x):
    """Which date column a token at centre x sits under.

    Dates are narrow and their captions are wide and stacked, so these are
    matched on centre distance rather than by the interval rule above.
    """
    anchors = [(key, colmap[key]["centre"]) for key in DATE_COLUMN_KEYS if key in colmap]
    if not anchors:
        return None
    return min(anchors, key=lambda kv: abs(kv[1] - x))[0]


# ---------------------------------------------------------------------------
# records
# ---------------------------------------------------------------------------

RECORD_FIELDS = (
    "part_id", "name", "first", "middle", "last", "status", "dob", "death",
    "address",
)


def new_record(page_num, layout, **fields):
    record = {"page": page_num, "layout": layout, "notes": []}
    record.update(fields)
    return record


def record_is_complete(record) -> bool:
    return bool(record.get("part_id")) and bool(
        record.get("name") or record.get("last") or record.get("first")
    )


def merge_across_pages(carry, record) -> bool:
    """Join a record split over a page break. Returns True if merged.

    Refuses on ANY field where both sides hold different values. Two
    half-participants left side by side for a human to join is a nuisance;
    two different people fused into one row is a data integrity failure
    that nothing downstream would catch.
    """
    if carry is None or record_is_complete(carry):
        return False
    for key in RECORD_FIELDS:
        a, b = carry.get(key, ""), record.get(key, "")
        if a and b and a != b:
            return False
    for key in RECORD_FIELDS:
        if record.get(key) and not carry.get(key):
            carry[key] = record[key]

    # Both halves were incomplete, so both carry "no X printed" notes that the
    # merge may have just answered. Drop them and re-state whichever is still
    # true of the joined record -- a row that names a missing field it now has
    # sends the reader hunting for a problem that no longer exists.
    notes = [n for n in carry["notes"] + record.get("notes", [])
             if not n.startswith(_STALE_ON_MERGE)]
    if not carry.get("part_id"):
        notes.append(NO_PART_ID_NOTE + " in this block")
    if not (carry.get("name") or carry.get("last") or carry.get("first")):
        notes.append(NO_NAME_NOTE)
    notes.append(f"joined with a block continued onto page {record['page']}")
    carry["notes"] = list(dict.fromkeys(notes))
    return True


def make_row(file_name, record):
    """Turn one gathered record into the output row."""
    notes = list(record.get("notes", []))
    printed_name = squeeze(record.get("name", ""))
    last = squeeze(record.get("last", ""))
    first = squeeze(record.get("first", ""))
    middle = squeeze(record.get("middle", ""))
    suffix = ""

    if last or first or middle:
        # The report printed the parts in their own columns. Use them as
        # given rather than joining and re-splitting, which could only lose
        # information the source already had right.
        rest = " ".join(p for p in (first, middle) if p)
        if not printed_name:
            printed_name = f"{last}, {rest}".strip(", ") if last else rest
            notes.append("name composed from separate first/middle/last columns")
    else:
        last, first, middle, suffix, name_note = split_name(printed_name)
        if name_note:
            notes.append(name_note)

    raw_dob = record.get("dob", "")
    dob, dob_note = norm_date(raw_dob)
    if dob_note:
        notes.append("date of birth: " + dob_note)
    elif raw_dob and not dob:
        notes.append("date of birth does not parse")
    elif not raw_dob and SUSPICIOUS_DOB_NOTE not in notes:
        notes.append("no date of birth printed")

    # A blank date of death is the normal case, so it is never noted. Only a
    # value that will not parse is worth reporting.
    raw_death = record.get("death", "")
    death, death_note = norm_date(raw_death)
    if raw_death and not death:
        notes.append("date of death: " + (death_note or "does not parse"))

    raw_id = record.get("part_id", "")
    return {
        "File Name": file_name,
        "Page": record["page"],
        "Layout": record["layout"],
        "Part ID (SSN)": norm_ssn(raw_id),
        "Part ID As Printed": squeeze(raw_id),
        "Name As Printed": squeeze(printed_name),
        "Last Name": last,
        "First Name": first,
        "Middle Name": middle,
        "Suffix": suffix,
        "Status": squeeze(record.get("status", "")),
        "Date Of Birth": dob,
        "Date Of Birth As Printed": squeeze(raw_dob),
        "Date Of Death": death,
        "Date Of Death As Printed": squeeze(raw_death),
        "Address": squeeze(record.get("address", "")),
        "Notes": "; ".join(n for n in notes if n),
    }


# ---------------------------------------------------------------------------
# the table layouts
# ---------------------------------------------------------------------------

def parse_table_page(rows, colmap, header_index, page_num, layout, carry=None):
    """Extract one record per participant from a column-table page."""
    records, candidates, fragments = [], 0, 0
    column_of = make_column_of(colmap)
    inherited = header_index < 0
    last_record = carry if (carry and carry["layout"].startswith("table")) else None

    for printed_row in rows[header_index + 1:]:
        row = [piece for w in printed_row for piece in split_glued(w)]
        buckets, date_words, glued = {}, [], False
        for w in row:
            if DATE_TOKEN_RE.match(w["text"]):
                date_words.append(w)
                continue
            key = column_of(w)
            if key:
                buckets.setdefault(key, []).append(w)
                glued = glued or w.get("glued", False)

        id_word = next(
            (w for w in buckets.get("part_id", []) if looks_like_part_id(w["text"])),
            None,
        )
        has_status_word = any(
            w["text"].strip(".,").upper() in STATUS_WORDS for w in row
        )

        def column_text(key, stop_at_other_fields=True):
            words = sorted(buckets.get(key, []), key=lambda w: w["x0"])
            out = []
            for w in words:
                text = w["text"]
                if stop_at_other_fields and (
                    SSN_TOKEN_RE.match(text)
                    or text.strip(".,").upper() in STATUS_WORDS
                ):
                    break
                out.append(text)
            return squeeze(" ".join(out))

        # --- a name that wrapped onto a second printed line ------------------
        name_only = (
            id_word is None
            and not date_words
            and not has_status_word
            and buckets
            and all(key in NAME_COLUMN_FIELD for key in buckets)
        )
        if name_only:
            if last_record is None:
                fragments += 1
                continue
            for key, field in NAME_COLUMN_FIELD.items():
                if key in buckets:
                    tail = column_text(key)
                    if tail:
                        last_record[field] = squeeze(
                            f"{last_record.get(field, '')} {tail}"
                        )
            note = "name continued onto a second printed line"
            if note not in last_record["notes"]:
                last_record["notes"].append(note)
            continue

        fields = {field: column_text(key)
                  for key, field in NAME_COLUMN_FIELD.items() if key in buckets}
        fields["status"] = column_text("status", stop_at_other_fields=False)

        for w in date_words:
            column = nearest_date_column(colmap, centre_x(w))
            if column in ("dob", "death") and not fields.get(column):
                fields[column] = w["text"]

        has_name = any(fields.get(f) for f in ("name", "first", "middle", "last"))
        notes = []
        if inherited:
            notes.append("column layout inherited from an earlier page")
        if glued:
            notes.append("Part ID was printed with no space before the next "
                         "column; the two were separated on their x positions")

        if id_word is None:
            # With no Part ID to anchor on, only accept the row if it still
            # looks like a participant. Otherwise report titles, period
            # headings and page footers would all become rows.
            if not (has_name and (date_words or has_status_word)):
                continue
            notes.append(NO_PART_ID_NOTE + " on this row")
        else:
            candidates += 1
            fields["part_id"] = id_word["text"]
            if not SSN_TOKEN_RE.match(id_word["text"]):
                notes.append("Part ID is not SSN-shaped")

        if not has_name:
            notes.append("no name found in the name column")
        if not fields.get("dob") and date_words and "dob" in colmap:
            notes.append(SUSPICIOUS_DOB_NOTE)

        record = new_record(page_num, layout, **fields)
        record["notes"] = notes
        records.append(record)
        last_record = record

    return records, candidates, fragments


def parse_pattern_rows(rows, page_num):
    """Last resort: read a table row by pattern order alone.

    Every row produced here is flagged, because with no header the birth
    date is only "the leftmost date", and that is wrong on any row whose
    birth date is blank.
    """
    records, candidates = [], 0
    for row in rows:
        texts = [w["text"] for w in row]
        idx = next((i for i, t in enumerate(texts) if SSN_TOKEN_RE.match(t)), None)
        if idx is None:
            continue
        candidates += 1
        name_tokens, dob_raw, status = [], "", ""
        for t in texts[idx + 1:]:
            if DATE_TOKEN_RE.match(t):
                dob_raw = t
                break
            if t.strip(".,").upper() in STATUS_WORDS:
                status = t
                continue
            if SSN_TOKEN_RE.match(t) or any(ch.isdigit() for ch in t):
                continue
            name_tokens.append(t)
        record = new_record(
            page_num, "table (no header)",
            part_id=texts[idx], name=squeeze(" ".join(name_tokens)),
            dob=dob_raw, status=status,
        )
        record["notes"] = [
            "no column header found; date of birth taken as the leftmost date"
        ]
        records.append(record)
    return records, candidates


# ---------------------------------------------------------------------------
# the label layout
# ---------------------------------------------------------------------------

# Every caption these reports print. Matched longest-first so "DATE OF BIRTH"
# is never read as "DATE OF" plus stray text. Add new captions here; the
# parser needs no other change.
LABELS = {
    "AFTER-TAX DEFERRAL % OR $": "after_tax",
    "PRE-TAX DEFERRAL % OR $": "pre_tax",
    "SOCIAL SECURITY NUMBER": "part_id",
    "DATE OF TERMINATION": "term",
    "PARTICIPANT NAME": "name",
    "MAILING ADDRESS": "address",
    "DATE OF REHIRE": "rehire",
    "HOME ADDRESS": "address",
    "DATE OF BIRTH": "dob",
    "DATE OF DEATH": "death",
    "DATE OF ENTRY": "entry",
    "PARTICIPANT ID": "part_id",
    "DATE OF HIRE": "hire",
    "EMPLOYEE ID": "employee_id",
    "PLAN NUMBER": "plan",
    "DIV / LOC": "div_loc",
    "ADDRESS": "address",
    "PART ID": "part_id",
    "DIV/LOC": "div_loc",
    "STATUS": "status",
    "NAME": "name",
    "SSN": "part_id",
}

_LABEL_ALTERNATION = "|".join(
    re.escape(label) for label in sorted(LABELS, key=len, reverse=True)
)
LABEL_RE = re.compile(r"(?<![A-Z])(" + _LABEL_ALTERNATION + r")\s*:", re.IGNORECASE)

LABEL_LAYOUT_RE = re.compile(r"(?:PART\s*ID|PARTICIPANT\s*ID|SSN)\s*:", re.IGNORECASE)

# Labels that belong to a participant's identity block. Anything else on the
# page (deferral percentages, fund allocations) is read but not used to decide
# where one participant ends and the next begins.
PARTICIPANT_LABEL_KEYS = {
    "part_id", "name", "dob", "death", "status", "hire", "rehire", "term",
    "entry", "div_loc", "employee_id", "plan", "address",
}

# The only field on these reports that wraps onto a line of its own.
CONTINUABLE_KEY = "address"

# Words that turn "NAME:" into somebody else's caption. The allocation detail
# below a participant block prints fund and plan names with the same word, and
# reading one of those as a participant name would invent a record.
NAME_PREFIX_BLOCK = {
    "FUND", "PLAN", "INVESTMENT", "OPTION", "SOURCE", "CONTRACT", "EMPLOYER",
    "COMPANY", "ACCOUNT", "PRODUCT", "VENDOR", "SUBACCOUNT", "PORTFOLIO",
    "FIRST", "LAST", "MIDDLE",
}

# How many glyph heights of blank space end a participant block. Lines inside
# one block are printed about 1.5 heights apart; the gap to the next block, or
# down to the allocation detail, is far larger.
BLOCK_GAP_HEIGHTS = 2.5

# A block may end with wrapped text carrying no caption of its own -- the
# second line of an address. At most this many such lines are taken in, and
# only while they sit at normal line spacing.
BLOCK_TRAILING_LINES = 2
BLOCK_TRAILING_GAP_HEIGHTS = 1.6


def label_pairs(line: str):
    """Split "A: x   B: y   C:" into [(key, value, match_start), ...]."""
    matches = list(LABEL_RE.finditer(line))
    pairs = []
    for i, m in enumerate(matches):
        end = matches[i + 1].start() if i + 1 < len(matches) else len(line)
        key = LABELS[m.group(1).upper()]
        pairs.append((key, line[m.end():end].strip(" \t:"), m.start()))
    return pairs


def name_is_participant(line: str, position: int) -> bool:
    """False when "NAME:" is the tail of a compound caption like "FUND NAME:"."""
    before = line[:position].rstrip()
    if not before:
        return True
    return before.split()[-1].strip(":").upper() not in NAME_PREFIX_BLOCK


def label_blocks(rows):
    """Group label-bearing lines into participant blocks by vertical gap.

    This is the whole reason the parser does not lose track of who owns which
    value. These reports omit a caption entirely when its value is empty, so a
    block cannot be delimited by watching for a caption to repeat -- the
    caption that would have marked the boundary is exactly the one that is
    missing. See the module docstring.

    Lines with no caption at all are pulled in when they fall inside a block,
    or immediately below it, because that is how the second line of an address
    is printed.
    """
    lines = []
    for row in rows:
        line = squeeze(row_text(row))
        if not line:
            continue
        lines.append({
            "y": row_centre(row),
            "height": median_height(row),
            "line": line,
            "pairs": label_pairs(line),
        })
    if not lines:
        return []

    labelled = [i for i, e in enumerate(lines) if e["pairs"]]
    if not labelled:
        return []
    height = sorted(e["height"] for e in lines)[len(lines) // 2] or 8.0

    groups, current = [], [labelled[0]]
    for index in labelled[1:]:
        if lines[index]["y"] - lines[current[-1]]["y"] > height * BLOCK_GAP_HEIGHTS:
            groups.append(current)
            current = []
        current.append(index)
    groups.append(current)

    blocks = []
    for group in groups:
        start, end = group[0], group[-1]
        # Interior unlabelled lines come in with the slice. Trailing ones are
        # taken only at normal line spacing, so the header of the table below
        # a block is never swallowed.
        taken = 0
        while (end + 1 < len(lines) and taken < BLOCK_TRAILING_LINES
               and not lines[end + 1]["pairs"]
               and lines[end + 1]["y"] - lines[end]["y"]
               <= height * BLOCK_TRAILING_GAP_HEIGHTS):
            end += 1
            taken += 1
        blocks.append(lines[start:end + 1])
    return blocks


def records_from_block(block):
    """One block usually holds one participant. A repeated caption inside it
    means two blocks were printed with no gap between them, so split there.

    Returns [(fields, orphan_texts)].
    """
    out, record, orphans = [], {}, []

    def flush():
        if record:
            out.append((dict(record), list(orphans)))

    for entry in block:
        pairs = entry["pairs"]
        head = entry["line"][:pairs[0][2]].strip() if pairs else entry["line"].strip()
        if head:
            # Text with no caption in front of it: the wrapped remainder of
            # the field above. Only the address is known to wrap, so anything
            # else is reported rather than guessed at.
            if record.get(CONTINUABLE_KEY):
                record[CONTINUABLE_KEY] = squeeze(
                    f"{record[CONTINUABLE_KEY]}, {head}"
                )
            else:
                orphans.append(head)

        for key, value, position in pairs:
            if key not in PARTICIPANT_LABEL_KEYS:
                continue
            if key == "name" and not name_is_participant(entry["line"], position):
                continue
            if value and record.get(key) and value != record[key]:
                flush()
                record, orphans = {}, []
            # Blank values are still stored: a caption that was printed with
            # nothing after it is evidence this is a participant block.
            if value or key not in record:
                record[key] = value
    flush()
    return out


def parse_label_page(rows, page_num, carry=None):
    """Extract participant records from a "LABEL: value" page."""
    records, candidates = [], 0
    first_on_page = True
    for block in label_blocks(rows):
        for fields, orphans in records_from_block(block):
            part_id = fields.get("part_id", "")
            name = fields.get("name", "")
            # A lone caption is not a participant. Two or more identity
            # captions in one block is.
            if not part_id and not (name and len(fields) >= 2):
                continue
            if part_id:
                candidates += 1

            notes = []
            if not part_id:
                notes.append(NO_PART_ID_NOTE + " in this block")
            if not name:
                notes.append(NO_NAME_NOTE)
            if orphans:
                notes.append(
                    f"{len(orphans)} unlabelled line(s) in this block were ignored"
                )

            record = new_record(
                page_num, "label",
                **{k: v for k, v in fields.items() if k in RECORD_FIELDS},
            )
            record["notes"] = notes

            if first_on_page and carry is not None and carry["layout"] == "label":
                first_on_page = False
                if merge_across_pages(carry, record):
                    continue
            first_on_page = False
            records.append(record)
    return records, candidates


# ---------------------------------------------------------------------------
# per-file processing
# ---------------------------------------------------------------------------

def detect_layout(rows):
    """"label", "table" or None, decided from the page's own content."""
    for row in rows:
        text = row_text(row)
        if LABEL_LAYOUT_RE.search(text) or re.match(r"^\s*NAME\s*:", text, re.I):
            return "label"
    if find_header_band(rows)[0] is not None:
        return "table"
    return None


def process_pdf(path, engine: str = DEFAULT_ENGINE):
    """Read one PDF. Returns (rows, diagnostics)."""
    path = Path(path)
    all_records = []
    diag = {
        "File Name": path.name,
        "Pages": 0,
        "Layouts Seen": "",
        "Part IDs On Page": 0,
        "Rows Written": 0,
        "Rows Missing Name": 0,
        "Rows Missing DOB": 0,
        "Rows With Suspicious DOB": 0,
        "Rows Without Part ID": 0,
        "Rows With Non-SSN Part ID": 0,
        "Unattached Name Fragments": 0,
        "Complete": "",
        "Error": "",
    }
    layouts = []
    last_colmap = None
    carry = None

    try:
        for page_num, words in read_pages(path, engine):
            diag["Pages"] += 1
            rows = cluster_rows(words)
            if not rows:
                continue

            layout = detect_layout(rows)
            fragments = 0
            if layout == "label":
                page_records, candidates = parse_label_page(rows, page_num, carry)
            elif layout == "table":
                colmap, header_index = build_column_map(rows)
                if colmap:
                    last_colmap = colmap
                    page_records, candidates, fragments = parse_table_page(
                        rows, colmap, header_index, page_num, "table", carry
                    )
                else:
                    page_records, candidates = parse_pattern_rows(rows, page_num)
            elif last_colmap is not None and any(
                SSN_TOKEN_RE.match(w["text"]) for w in words
            ):
                # A continuation page: participant rows, header printed only
                # on the first page of the run.
                layout = "table (inherited)"
                page_records, candidates, fragments = parse_table_page(
                    rows, last_colmap, -1, page_num, layout, carry
                )
            else:
                continue

            if page_records or candidates:
                layouts.append(layout)
            all_records.extend(page_records)
            diag["Part IDs On Page"] += candidates
            diag["Unattached Name Fragments"] += fragments
            # Carried forward so a participant split over the page break, or a
            # name that wrapped across it, can still be joined up.
            if page_records:
                carry = page_records[-1]
    except Exception as exc:  # pragma: no cover - depends on the file
        diag["Error"] = f"{type(exc).__name__}: {exc}"

    rows_out = [make_row(path.name, r) for r in all_records]

    diag["Rows Written"] = len(rows_out)
    diag["Rows Missing Name"] = sum(1 for r in rows_out if not r["Name As Printed"])
    diag["Rows Missing DOB"] = sum(1 for r in rows_out if not r["Date Of Birth"])
    diag["Rows With Suspicious DOB"] = sum(
        1 for r in rows_out if SUSPICIOUS_DOB_NOTE in r["Notes"]
    )
    diag["Rows Without Part ID"] = sum(1 for r in rows_out if not r["Part ID (SSN)"])
    diag["Rows With Non-SSN Part ID"] = sum(
        1 for r in rows_out
        if r["Part ID (SSN)"]
        and not re.match(r"^[\dX]{3}-[\dX]{2}-\d{4}$", r["Part ID (SSN)"])
    )
    diag["Layouts Seen"] = ", ".join(sorted(set(layouts))) or "none"
    # Every Part ID on the page became a row, plus the rows that legitimately
    # had no Part ID printed. A blank date of birth is normal -- two of these
    # reports have no DOB column at all -- and does not flag the file; a date
    # that failed to land under the BIRTH column does.
    diag["Complete"] = "YES" if (
        not diag["Error"]
        and diag["Rows Written"] == diag["Part IDs On Page"] + diag["Rows Without Part ID"]
        and diag["Rows Missing Name"] == 0
        and diag["Rows With Suspicious DOB"] == 0
        and diag["Unattached Name Fragments"] == 0
    ) else "CHECK"
    return rows_out, diag


def _worker(payload):
    """Top-level so Windows spawn can pickle it."""
    path, engine = payload
    return process_pdf(path, engine)


def process_folder(pdfs, engine=DEFAULT_ENGINE, workers=None, progress=None):
    """Read every PDF. Returns (rows, diagnostics)."""
    pdfs = list(pdfs)
    rows, diags = [], []

    if workers is None:
        workers = max(1, (os.cpu_count() or 2) - 1)
    # Windows has no fork, so each worker is a fresh interpreter costing about
    # a second to start. On a handful of files that start-up cost exceeds the
    # work itself.
    if len(pdfs) < MIN_FILES_FOR_PARALLEL or workers == 1:
        for i, path in enumerate(pdfs, start=1):
            r, d = process_pdf(path, engine)
            rows.extend(r)
            diags.append(d)
            if progress:
                progress(i, len(pdfs), Path(path).name)
        return rows, diags

    with ProcessPoolExecutor(max_workers=workers) as pool:
        payloads = [(str(p), engine) for p in pdfs]
        for i, (r, d) in enumerate(pool.map(_worker, payloads), start=1):
            rows.extend(r)
            diags.append(d)
            if progress:
                progress(i, len(pdfs), d["File Name"])
    return rows, diags


# ---------------------------------------------------------------------------
# workbook
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FLAG_FILL = PatternFill("solid", fgColor="FCE4E4")


def _write_sheet(ws, columns, records, text_columns=()):
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    text_indexes = {columns.index(c) + 1 for c in text_columns if c in columns}
    for record in records:
        ws.append([record.get(c, "") for c in columns])
        for idx in text_indexes:
            # Excel eats leading zeros and reformats 123-45-6789 unless the
            # cell is explicitly text.
            ws.cell(row=ws.max_row, column=idx).number_format = "@"

    widths = [len(c) for c in columns]
    for record in records:
        for i, c in enumerate(columns):
            widths[i] = max(widths[i], min(60, len(str(record.get(c, "")))))
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width + 2
    ws.freeze_panes = "A2"


def write_workbook(rows, diagnostics, dest: Path, masked=False) -> Path:
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)

    if masked:
        rows = [dict(r) for r in rows]
        for r in rows:
            r["Part ID (SSN)"] = mask_ssn(r["Part ID (SSN)"])
            r["Part ID As Printed"] = mask_ssn(r["Part ID As Printed"])

    wb = Workbook()
    _write_sheet(
        wb.active, OUTPUT_COLUMNS, rows,
        text_columns=("Part ID (SSN)", "Part ID As Printed",
                      "Date Of Birth", "Date Of Birth As Printed",
                      "Date Of Death", "Date Of Death As Printed"),
    )
    wb.active.title = "Data"

    recon = wb.create_sheet("Reconciliation")
    _write_sheet(recon, RECON_COLUMNS, diagnostics)
    complete_col = RECON_COLUMNS.index("Complete") + 1
    for row_idx in range(2, recon.max_row + 1):
        if recon.cell(row=row_idx, column=complete_col).value != "YES":
            for col_idx in range(1, len(RECON_COLUMNS) + 1):
                recon.cell(row=row_idx, column=col_idx).fill = FLAG_FILL

    wb.save(dest)
    return dest


# ---------------------------------------------------------------------------
# QA and verification -- masked output only
# ---------------------------------------------------------------------------

def _shape(text: str) -> str:
    """Reduce a word to its shape: letters to A, digits to 9.

    "CLARK,LISA" becomes "AAAAA,AAAA" and "177-72-5725" becomes
    "999-99-9999". Enough to diagnose a column or line-grouping fault,
    nothing that identifies anybody.
    """
    return re.sub(r"[A-Za-z]", "A", re.sub(r"\d", "9", fold(text)))


def dump_lines(path, engine=DEFAULT_ENGINE, page_no=None, limit=80):
    """Masked geometry dump: every rebuilt line, its y, and each word's x
    span and column. Diagnoses a misread without moving any PII."""
    path = Path(path)
    print(f"{path.name}  --  shapes only, no participant data")
    colmap = None
    for page_num, words in read_pages(path, engine):
        if page_no and page_num != page_no:
            continue
        rows = cluster_rows(words)
        found, header_index = build_column_map(rows)
        if found:
            colmap = found
        column_of = make_column_of(colmap) if colmap else (lambda w: "?")
        print(f"\n--- page {page_num}   layout={detect_layout(rows)}   "
              f"header row={header_index}   lines={len(rows)}")
        if colmap:
            ordered = sorted(colmap.items(), key=lambda kv: kv[1]["x0"])
            print("    columns  " + "  ".join(
                f"{k}[{c['x0']:.0f}-{c['x1']:.0f}]" for k, c in ordered))
        for row in rows[:limit]:
            body = " ".join(
                f"[{w['x0']:.0f}-{w['x1']:.0f} {column_of(w) or '-'}]{_shape(w['text'])}"
                for w in row
            )
            print(f"  y={row_centre(row):7.1f}  {body}")
        if len(rows) > limit:
            print(f"  ... {len(rows) - limit} more line(s) not shown")


def qa_file(path, engine=DEFAULT_ENGINE, limit=25):
    """Masked dump of how each participant line was split.

    Prints no SSN and no address -- the address is the one field that
    cannot be usefully masked, so it is simply not shown.
    """
    rows, diag = process_pdf(path, engine)
    print(f"\n{Path(path).name}  --  {diag['Pages']} page(s), "
          f"layout(s): {diag['Layouts Seen']}")
    print(f"  Part IDs on page {diag['Part IDs On Page']}, "
          f"rows written {diag['Rows Written']}, complete: {diag['Complete']}")
    if diag["Error"]:
        print(f"  ERROR  {diag['Error']}")
    for r in rows[:limit]:
        print(f"    p{r['Page']:<4} {mask_ssn(r['Part ID (SSN)']):<12} "
              f"{r['Last Name']:<14} {r['First Name']:<12} "
              f"{r['Middle Name']:<8} {r['Suffix']:<4} {r['Status']:<11} "
              f"{r['Date Of Birth']:<11} {r['Date Of Death']:<11} "
              f"{'addr' if r['Address'] else '':<5}")
        if r["Notes"]:
            print(f"          note: {r['Notes']}")
    if len(rows) > limit:
        print(f"    ... {len(rows) - limit} more row(s) not shown")


def verify(folder, limit=None):
    """Read every PDF with both engines and report disagreements.

    Counts only -- no participant data is printed.
    """
    engines = available_engines()
    if len(engines) < 2:
        print("Both pdfplumber and PyMuPDF must be installed to verify.")
        return 1
    pdfs = sorted(Path(folder).rglob("*.pdf"))
    if limit:
        pdfs = pdfs[:limit]
    if not pdfs:
        print(f"No PDFs under {folder}")
        return 1

    def key(rs):
        return sorted(
            (r["Part ID (SSN)"], r["Name As Printed"], r["Status"],
             r["Date Of Birth"], r["Date Of Death"], r["Address"])
            for r in rs
        )

    mismatches = 0
    for path in pdfs:
        a_rows, _ = process_pdf(path, "pdfplumber")
        b_rows, _ = process_pdf(path, "mupdf")
        same = key(a_rows) == key(b_rows)
        mismatches += 0 if same else 1
        print(f"  {'identical' if same else 'DIFFERS':<10} {path.name}  "
              f"pdfplumber={len(a_rows)}  mupdf={len(b_rows)}")
    print(f"\n{len(pdfs)} file(s), {mismatches} disagreement(s).")
    if mismatches == 0:
        print("Both engines agree on every file. --engine mupdf is safe here.")
    else:
        print("Engines disagree. Keep the default engine.")
    return 0


# ---------------------------------------------------------------------------
# selftest -- synthetic PDF, fabricated data, no real PII
# ---------------------------------------------------------------------------

def _build_test_pdf(dest: Path) -> Path:
    """Write a 9-page synthetic PDF. Every value below is made up."""
    if fitz is None:
        raise RuntimeError("PyMuPDF is required to build the selftest PDF")
    doc = fitz.open()

    def put(page, x, y, text):
        page.insert_text((x, y), text, fontsize=8, fontname="cour")

    # --- page 1: joined-name table, header present ---------------------------
    p1 = doc.new_page(width=792, height=612)
    put(p1, 40, 60, "Small Cash Out Monitoring Report")
    put(p1, 355, 100, "DATE OF")
    put(p1, 585, 100, "DATE OF")
    put(p1, 660, 100, "VESTED TOTAL")
    for x, caption in [
        (40, "PART ID"), (110, "NAME"), (250, "DIV/LOC"), (305, "STATUS"),
        (360, "BIRTH"), (410, "DATE OF HIRE"), (490, "DATE OF REHIRE"),
        (580, "TERMINATION"), (655, "ACCOUNT BALANCE"),
    ]:
        put(p1, x, 110, caption)

    rows = [
        # part id, name, div, status, dob, hire, rehire, term, balance
        ("123-45-6789", "ACKERMAN, KENTON CLYDE", "0001", "Terminated",
         "01/15/1963", "03/01/1990", "", "06/30/2015", "$1,234.18"),
        # no birth date printed -- the hire date must NOT be read as the DOB
        ("234-56-7890", "CHILDERS,MORGAN N", "0002", "Terminated",
         "", "07/12/1988", "", "01/31/2015", "$88.15"),
        # running nine digits, and a suffix on the surname
        ("345678901", "SMITH JR, JOHN A", "0003", "Active",
         "12/25/1977", "05/02/2001", "05/02/2009", "", "$4,010.06"),
    ]
    y = 130
    for pid, name, div, status, dob, hire, rehire, term, bal in rows:
        put(p1, 40, y, pid)
        put(p1, 110, y, name)
        put(p1, 250, y, div)
        put(p1, 305, y, status)
        if dob:
            put(p1, 358, y, dob)
        if hire:
            put(p1, 415, y, hire)
        if rehire:
            put(p1, 495, y, rehire)
        if term:
            put(p1, 582, y, term)
        put(p1, 660, y, bal)
        y += 14
    put(p1, 40, y + 20, "Page 1 of 9")

    # --- page 2: labelled block, with a wrapped address ----------------------
    p2 = doc.new_page(width=792, height=612)
    put(p2, 40, 60, "Participant Contribution Investment Allocation Report")
    put(p2, 40, 100, "NAME: ABBEY, BORIS S")
    put(p2, 300, 100, "DATE OF BIRTH: 04/12/1973")
    put(p2, 520, 100, "DATE OF TERMINATION:")
    put(p2, 40, 112, "PART ID: 987654321")
    put(p2, 300, 112, "DATE OF HIRE: 06/01/2015")
    put(p2, 520, 112, "DATE OF DEATH: 03/22/2024")
    put(p2, 40, 124, "STATUS: Active")
    put(p2, 300, 124, "DATE OF ENTRY: 07/01/2015")
    put(p2, 520, 124, "DATE OF REHIRE:")
    put(p2, 40, 136, "ADDRESS: 41 SYNTHETIC RD")
    put(p2, 300, 136, "DIV/LOC:")
    # the second address line carries no caption of its own
    put(p2, 40, 148, "FAKETOWN, NC 200000000")
    # a fund name below the block, which must not start a second record
    put(p2, 40, 200, "FUND NAME: TIAA TRADITIONAL")

    # --- page 3: table continuation, no header printed -----------------------
    p3 = doc.new_page(width=792, height=612)
    put(p3, 40, 60, "Small Cash Out Monitoring Report (continued)")
    put(p3, 40, 130, "456-78-9012")
    put(p3, 110, 130, "O'LEARY, MARY JANE")
    put(p3, 250, 130, "0004")
    put(p3, 305, 130, "Retired")
    put(p3, 358, 130, "07/04/1955")
    put(p3, 415, 130, "09/09/1979")
    put(p3, 582, 130, "12/31/2014")

    # --- page 4: table carrying BOTH a birth and a death column --------------
    # The two captions sit side by side, so this is where a date-of-death
    # would be misread as a date of birth if the columns were not anchored.
    p4 = doc.new_page(width=792, height=612)
    put(p4, 40, 60, "Small Cash Out Monitoring Report -- with date of death")
    put(p4, 305, 100, "DATE OF")
    put(p4, 380, 100, "DATE OF")
    put(p4, 455, 100, "DATE OF")
    for x, caption in [
        (40, "PART ID"), (110, "NAME"), (250, "STATUS"),
        (310, "BIRTH"), (385, "DEATH"), (450, "TERMINATION"),
    ]:
        put(p4, x, 110, caption)
    put(p4, 40, 130, "567-89-0123")
    put(p4, 110, 130, "DENNING, JAMES DARYL")
    put(p4, 250, 130, "Deceased")
    put(p4, 303, 130, "02/09/1962")
    put(p4, 378, 130, "11/30/2019")
    put(p4, 452, 130, "11/30/2019")
    put(p4, 40, 144, "678-90-1234")
    put(p4, 110, 144, "DAVIS, CATHRYN R")
    put(p4, 250, 144, "Active")
    put(p4, 303, 144, "08/21/1960")
    # a participant whose PART ID cell is empty: must still be written out,
    # flagged, rather than dropped
    put(p4, 110, 158, "BAYLOCK, CINDY KING")
    put(p4, 250, 158, "Active")
    put(p4, 303, 158, "04/05/1959")

    # --- page 5: label layout with captions OMITTED, not blank ---------------
    # This is the case that mis-assigned names. The report prints no caption
    # at all when a value is empty, so block 2 has no "NAME:" line and block 3
    # has no "PART ID:" line. Block 3's name must NOT land on block 2's
    # Part ID.
    p5 = doc.new_page(width=792, height=612)
    put(p5, 40, 60, "Participant Contribution Investment Allocation Report")
    put(p5, 40, 100, "NAME: CLARK, CHRISTOPHER R")
    put(p5, 300, 100, "DATE OF BIRTH: 01/02/1970")
    put(p5, 40, 112, "PART ID: 111223333")
    put(p5, 300, 112, "DATE OF HIRE: 01/01/2000")
    # no NAME caption anywhere in this block
    put(p5, 40, 170, "PART ID: 222334444")
    put(p5, 300, 170, "DATE OF BIRTH: 03/04/1971")
    put(p5, 40, 182, "STATUS: Active")
    put(p5, 300, 182, "DATE OF HIRE: 02/02/2001")
    # no PART ID caption anywhere in this block
    put(p5, 40, 240, "NAME: CLIFTON, BERA S")
    put(p5, 300, 240, "DATE OF BIRTH: 05/06/1972")
    put(p5, 40, 252, "STATUS: Active")
    put(p5, 300, 252, "DATE OF HIRE: 03/03/2002")

    # --- page 6: split-name table, SSN caption, three-line header ------------
    # Also has no DATE OF BIRTH column at all, and a section title printed
    # across the full table width which must NOT be read as a caption row.
    p6 = doc.new_page(width=792, height=612)
    put(p6, 40, 60, "Termination Activity Summary")
    put(p6, 40, 80, "SECTION 1: TERMINATED IN PRIOR YEAR WITH A VESTED BALANCE")
    put(p6, 530, 90, "DATE OF")
    for x, caption in [
        (40, "ITEM"), (230, "MIDDLE"), (440, "DATE OF"), (530, "TERMINATION"),
        (620, "BEGINNING"),
    ]:
        put(p6, x, 100, caption)
    for x, caption in [
        (40, "COUNT"), (100, "SSN"), (150, "FIRST NAME"), (230, "INITIAL"),
        (300, "LAST NAME"), (380, "STATUS"), (440, "TERMINATION"),
        (530, "PROCESSED"), (620, "BALANCE"),
    ]:
        put(p6, x, 110, caption)
    put(p6, 45, 130, "1")
    put(p6, 88, 130, "789-01-2345")
    put(p6, 150, 130, "CHRISTINE")
    put(p6, 240, 130, "R")
    put(p6, 300, 130, "CHAN")
    put(p6, 380, 130, "Terminated")
    put(p6, 440, 130, "02/28/2014")
    put(p6, 530, 130, "03/07/2014")
    put(p6, 620, 130, "$275.09")
    # the SSN printed hard against the first name, so the extractor returns
    # them as one word. The first name must not disappear.
    put(p6, 45, 144, "2")
    put(p6, 95, 144, "246-80-1357ANNA")
    put(p6, 240, 144, "S")
    put(p6, 300, 144, "LOPEZ")
    put(p6, 380, 144, "Terminated")
    put(p6, 440, 144, "04/04/2014")

    # --- page 7: joined-name table where the name wraps onto a second line ---
    p7 = doc.new_page(width=792, height=612)
    put(p7, 40, 60, "Activity for the Reporting Period: 01/01/2015 to 12/31/2015")
    for x, caption in [(40, "ITEM"), (330, "DIV/"), (540, "DISTRIBUTION")]:
        put(p7, x, 100, caption)
    for x, caption in [
        (40, "COUNT"), (100, "PART ID"), (170, "NAME"), (330, "LOC"),
        (380, "STATUS"), (450, "TERM DATE"), (540, "AMOUNT"),
    ]:
        put(p7, x, 110, caption)
    put(p7, 45, 130, "1")
    put(p7, 100, 130, "012-34-5678")
    put(p7, 170, 130, "SMITH, BRECK HOWARD")
    put(p7, 330, 130, "0001")
    put(p7, 380, 130, "Terminated")
    put(p7, 450, 130, "05/31/2009")
    put(p7, 540, 130, "($10.05)")
    # the rest of the name, on its own line, with nothing else on it
    put(p7, 170, 144, "BRECKENRIDGE")

    # --- pages 8 and 9: one participant split across the page break ----------
    p8 = doc.new_page(width=792, height=612)
    put(p8, 40, 60, "Participant Contribution Investment Allocation Report")
    put(p8, 40, 520, "NAME: WATSON, GEORGE T")
    put(p8, 300, 520, "DATE OF BIRTH: 09/09/1949")
    put(p8, 40, 532, "STATUS: Terminated")
    put(p8, 300, 532, "ADDRESS: 12 MAPLE ST")

    p9 = doc.new_page(width=792, height=612)
    put(p9, 40, 60, "PART ID: 890123456")
    put(p9, 300, 60, "DATE OF HIRE: 01/01/1980")
    put(p9, 40, 72, "DIV/LOC:")
    put(p9, 300, 72, "DATE OF TERMINATION: 05/05/2005")
    # a complete participant further down: must NOT be merged into anything
    put(p9, 40, 200, "NAME: YOUNG, ALICE M")
    put(p9, 300, 200, "DATE OF BIRTH: 03/03/1966")
    put(p9, 40, 212, "PART ID: 901234567")
    put(p9, 300, 212, "STATUS: Active")

    doc.save(str(dest))
    doc.close()
    return dest


def selftest() -> int:
    import tempfile

    if fitz is None:
        print("SKIP  PyMuPDF is not installed, cannot build the test PDF.")
        return 0

    failures = []

    def check(label, got, want):
        if got != want:
            failures.append(f"{label}: got {got!r}, want {want!r}")

    # pure functions first
    check("norm_ssn dashed", norm_ssn("123-45-6789"), "123-45-6789")
    check("norm_ssn running", norm_ssn("123456789"), "123-45-6789")
    check("norm_ssn unicode dash", norm_ssn("123‐45‐6789"), "123-45-6789")
    check("norm_ssn masked", norm_ssn("XXX-XX-6789"), "XXX-XX-6789")
    check("norm_date 4-digit", norm_date("01/15/1963")[0], "01/15/1963")
    check("norm_date 2-digit", norm_date("1/5/63")[0], "01/05/1963")
    check("mask", mask_ssn("123-45-6789"), "***-**-6789")
    check("split comma", split_name("ACKERMAN, KENTON CLYDE")[:4],
          ("ACKERMAN", "KENTON", "CLYDE", ""))
    check("split no space", split_name("CHILDERS,MORGAN N")[:4],
          ("CHILDERS", "MORGAN", "N", ""))
    check("split suffix", split_name("SMITH JR, JOHN A")[:4],
          ("SMITH", "JOHN", "A", "JR"))

    with tempfile.TemporaryDirectory() as tmp:
        pdf = _build_test_pdf(Path(tmp) / "synthetic.pdf")
        for engine in available_engines():
            rows, diag = process_pdf(pdf, engine)
            tag = f"[{engine}]"
            by_id = {r["Part ID (SSN)"]: r for r in rows if r["Part ID (SSN)"]}
            by_name = {r["Name As Printed"]: r for r in rows}

            check(f"{tag} row count", len(rows), 16)
            check(f"{tag} error", diag["Error"], "")
            check(f"{tag} no stray fragments", diag["Unattached Name Fragments"], 0)

            r = by_id.get("123-45-6789", {})
            check(f"{tag} table name", r.get("Name As Printed"),
                  "ACKERMAN, KENTON CLYDE")
            check(f"{tag} table dob", r.get("Date Of Birth"), "01/15/1963")
            check(f"{tag} table status", r.get("Status"), "Terminated")

            # the row with no birth date must NOT pick up the hire date
            r = by_id.get("234-56-7890", {})
            check(f"{tag} blank dob", r.get("Date Of Birth"), "")
            check(f"{tag} blank dob name", r.get("Name As Printed"), "CHILDERS,MORGAN N")

            r = by_id.get("345-67-8901", {})
            check(f"{tag} running ssn dob", r.get("Date Of Birth"), "12/25/1977")
            check(f"{tag} running ssn suffix", r.get("Suffix"), "JR")

            # --- label layout, address wrapped onto an uncaptioned line -----
            r = by_id.get("987-65-4321", {})
            check(f"{tag} label layout", r.get("Layout"), "label")
            check(f"{tag} label name", r.get("Name As Printed"), "ABBEY, BORIS S")
            check(f"{tag} label dob", r.get("Date Of Birth"), "04/12/1973")
            check(f"{tag} label death", r.get("Date Of Death"), "03/22/2024")
            check(f"{tag} label status", r.get("Status"), "Active")
            check(f"{tag} label address", r.get("Address"),
                  "41 SYNTHETIC RD, FAKETOWN, NC 200000000")

            r = by_id.get("456-78-9012", {})
            check(f"{tag} inherited dob", r.get("Date Of Birth"), "07/04/1955")
            check(f"{tag} inherited layout", r.get("Layout"), "table (inherited)")

            # birth and death columns side by side must not be confused
            r = by_id.get("567-89-0123", {})
            check(f"{tag} table dob beside death", r.get("Date Of Birth"), "02/09/1962")
            check(f"{tag} table death", r.get("Date Of Death"), "11/30/2019")

            r = by_id.get("678-90-1234", {})
            check(f"{tag} dob with no death", r.get("Date Of Birth"), "08/21/1960")
            check(f"{tag} blank death", r.get("Date Of Death"), "")

            # a table row whose PART ID cell was empty: kept, and flagged
            r = by_name.get("BAYLOCK, CINDY KING", {})
            check(f"{tag} blank part id kept", r.get("Date Of Birth"), "04/05/1959")
            check(f"{tag} blank part id blank", r.get("Part ID (SSN)"), "")
            check(f"{tag} blank part id flagged",
                  NO_PART_ID_NOTE in r.get("Notes", ""), True)

            # --- the omitted-caption case that mis-assigned names -----------
            r = by_id.get("111-22-3333", {})
            check(f"{tag} block 1 name", r.get("Name As Printed"),
                  "CLARK, CHRISTOPHER R")
            check(f"{tag} block 1 dob", r.get("Date Of Birth"), "01/02/1970")

            # block 2 printed no NAME caption at all. Its name must be blank,
            # NOT block 3's name.
            r = by_id.get("222-33-4444", {})
            check(f"{tag} block 2 name stays blank", r.get("Name As Printed"), "")
            check(f"{tag} block 2 dob", r.get("Date Of Birth"), "03/04/1971")

            # block 3 printed no PART ID caption. Its name must survive with a
            # blank Part ID, not be swallowed by block 2.
            r = by_name.get("CLIFTON, BERA S", {})
            check(f"{tag} block 3 kept", r.get("Date Of Birth"), "05/06/1972")
            check(f"{tag} block 3 part id blank", r.get("Part ID (SSN)"), "")

            # --- split-name table, SSN caption, no DOB column ---------------
            r = by_id.get("789-01-2345", {})
            check(f"{tag} split first", r.get("First Name"), "CHRISTINE")
            check(f"{tag} split middle", r.get("Middle Name"), "R")
            check(f"{tag} split last", r.get("Last Name"), "CHAN")
            check(f"{tag} split composed", r.get("Name As Printed"),
                  "CHAN, CHRISTINE R")
            check(f"{tag} split status", r.get("Status"), "Terminated")
            check(f"{tag} split no dob column", r.get("Date Of Birth"), "")
            check(f"{tag} split dob not flagged suspicious",
                  SUSPICIOUS_DOB_NOTE in r.get("Notes", ""), False)

            # SSN glued to the first name: both must survive, and be flagged
            r = by_id.get("246-80-1357", {})
            check(f"{tag} glued first", r.get("First Name"), "ANNA")
            check(f"{tag} glued last", r.get("Last Name"), "LOPEZ")
            check(f"{tag} glued middle", r.get("Middle Name"), "S")
            check(f"{tag} glued flagged",
                  "no space before the next column" in r.get("Notes", ""), True)

            # --- a name wrapped onto a second printed line ------------------
            r = by_id.get("012-34-5678", {})
            check(f"{tag} wrapped name joined", r.get("Name As Printed"),
                  "SMITH, BRECK HOWARD BRECKENRIDGE")
            check(f"{tag} wrapped name last", r.get("Last Name"), "SMITH")
            check(f"{tag} wrapped name flagged",
                  "continued onto a second printed line" in r.get("Notes", ""), True)
            check(f"{tag} wrapped name status", r.get("Status"), "Terminated")

            # --- a participant split across the page break ------------------
            r = by_id.get("890-12-3456", {})
            check(f"{tag} page split name", r.get("Name As Printed"),
                  "WATSON, GEORGE T")
            check(f"{tag} page split dob", r.get("Date Of Birth"), "09/09/1949")
            check(f"{tag} page split address", r.get("Address"), "12 MAPLE ST")
            check(f"{tag} page split status", r.get("Status"), "Terminated")
            check(f"{tag} page split page number", r.get("Page"), 8)
            check(f"{tag} page split flagged",
                  "continued onto page 9" in r.get("Notes", ""), True)
            check(f"{tag} page split part id resolved",
                  NO_PART_ID_NOTE in r.get("Notes", ""), False)
            # both halves were flagged incomplete; the joined row must not
            # still claim a field it now has
            check(f"{tag} page split name note cleared",
                  NO_NAME_NOTE in r.get("Notes", ""), False)

            # the complete participant below it must stay separate
            r = by_id.get("901-23-4567", {})
            check(f"{tag} unmerged neighbour", r.get("Name As Printed"),
                  "YOUNG, ALICE M")
            check(f"{tag} unmerged neighbour dob", r.get("Date Of Birth"),
                  "03/03/1966")

    if failures:
        print("SELFTEST FAILED")
        for f in failures:
            print("  " + f)
        return 1
    print(f"SELFTEST PASSED  (engines: {', '.join(available_engines())})")
    return 0


# ---------------------------------------------------------------------------
# headless run
# ---------------------------------------------------------------------------

def find_pdfs(folder):
    return sorted(Path(folder).rglob("*.pdf"))


def recon_summary(d) -> str:
    """One line per flagged file. No participant data, counts only."""
    return (
        f"{d['File Name']}: ids={d['Part IDs On Page']} rows={d['Rows Written']} "
        f"no-part-id={d['Rows Without Part ID']} "
        f"no-name={d['Rows Missing Name']} "
        f"no-dob={d['Rows Missing DOB']} "
        f"dob-suspect={d['Rows With Suspicious DOB']} "
        f"fragments={d['Unattached Name Fragments']} {d['Error']}".rstrip()
    )


PII_REMINDER = (
    "This workbook holds names, SSNs, dates of birth and home addresses. "
    "Save it to the appropriate Global Insider folder with restricted access "
    "-- not to a desktop or personal drive -- and give it a retention date."
)


def run_headless(src, out, engine=DEFAULT_ENGINE, workers=None, masked=False) -> int:
    pdfs = find_pdfs(src)
    if not pdfs:
        print(f"No PDFs found under {src}")
        return 1

    def progress(done, total, name):
        print(f"  [{done}/{total}] {name}")

    print(f"Reading {len(pdfs)} PDF(s) with {engine} ...")
    rows, diags = process_folder(pdfs, engine, workers, progress)
    dest = write_workbook(rows, diags, Path(out), masked=masked)

    flagged = [d for d in diags if d["Complete"] != "YES"]
    print(f"\n{len(rows)} row(s) written to {dest}")
    if masked:
        print("Part IDs are MASKED in this workbook.")
    if flagged:
        print(f"{len(flagged)} file(s) need checking on the Reconciliation sheet:")
        for d in flagged:
            print("  " + recon_summary(d))
    else:
        print("Reconciliation: every file complete.")
    print("\n" + PII_REMINDER)
    return 0


# ---------------------------------------------------------------------------
# Tkinter front end
# ---------------------------------------------------------------------------

class App:
    def __init__(self, root):
        import tkinter as tk
        from tkinter import scrolledtext, ttk

        self.tk = tk
        self.root = root
        root.title("TIAA participant identity extractor")
        root.geometry("900x580")

        frame = ttk.Frame(root, padding=12)
        frame.pack(fill="both", expand=True)

        self.src = tk.StringVar()
        self.out = tk.StringVar()
        self.engine = tk.StringVar(value=DEFAULT_ENGINE)
        self.masked = tk.BooleanVar(value=False)

        ttk.Label(frame, text="Folder of PDFs").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame, textvariable=self.src, width=78).grid(row=0, column=1, padx=6)
        ttk.Button(frame, text="Browse", command=self.pick_src).grid(row=0, column=2)

        ttk.Label(frame, text="Output workbook").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Entry(frame, textvariable=self.out, width=78).grid(row=1, column=1, padx=6)
        ttk.Button(frame, text="Save as", command=self.pick_out).grid(row=1, column=2)

        options = ttk.Frame(frame)
        options.grid(row=2, column=1, sticky="w", pady=4)
        ttk.Label(options, text="Engine").pack(side="left")
        ttk.Combobox(
            options, textvariable=self.engine, width=12, state="readonly",
            values=available_engines() or [DEFAULT_ENGINE],
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            options, text="Mask Part IDs in the workbook (***-**-6789)",
            variable=self.masked,
        ).pack(side="left", padx=16)

        self.run_button = ttk.Button(frame, text="Extract", command=self.run)
        self.run_button.grid(row=3, column=1, sticky="w", pady=8)

        self.progress = ttk.Progressbar(frame, length=640, mode="determinate")
        self.progress.grid(row=4, column=0, columnspan=3, sticky="we", pady=4)

        self.status = ttk.Label(frame, text="Ready.")
        self.status.grid(row=5, column=0, columnspan=3, sticky="w")

        self.log_box = scrolledtext.ScrolledText(frame, height=20, width=108)
        self.log_box.grid(row=6, column=0, columnspan=3, sticky="nsew", pady=8)
        frame.rowconfigure(6, weight=1)
        frame.columnconfigure(1, weight=1)

        self.log(PII_REMINDER)
        self.log("No SSN and no address is ever written to this log.")

    def log(self, message):
        self.log_box.insert("end", message + "\n")
        self.log_box.see("end")
        self.root.update_idletasks()

    def pick_src(self):
        from tkinter import filedialog
        path = filedialog.askdirectory(title="Select the folder holding the PDFs")
        if path:
            self.src.set(path)
            if not self.out.get():
                self.out.set(str(Path(path) / OUTPUT_XLSX_NAME))

    def pick_out(self):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            title="Save the workbook as", defaultextension=".xlsx",
            initialfile=OUTPUT_XLSX_NAME,
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if path:
            self.out.set(path)

    def run(self):
        from tkinter import messagebox
        src, out = self.src.get().strip(), self.out.get().strip()
        if not src or not Path(src).is_dir():
            messagebox.showerror("Folder", "Pick a folder of PDFs first.")
            return
        if not out:
            out = str(Path(src) / OUTPUT_XLSX_NAME)
            self.out.set(out)

        pdfs = find_pdfs(src)
        if not pdfs:
            messagebox.showerror("Folder", f"No PDFs found under {src}")
            return

        self.run_button.state(["disabled"])
        self.progress["maximum"] = len(pdfs)
        self.progress["value"] = 0
        self.log(f"\nReading {len(pdfs)} PDF(s) with {self.engine.get()} ...")

        def progress(done, total, name):
            self.progress["value"] = done
            self.status.config(text=f"[{done}/{total}] {name}")
            self.root.update_idletasks()

        try:
            rows, diags = process_folder(pdfs, self.engine.get(), None, progress)
            dest = write_workbook(rows, diags, Path(out), masked=self.masked.get())
        except Exception as exc:
            self.log(f"FAILED  {type(exc).__name__}: {exc}")
            self.log(traceback.format_exc())
            messagebox.showerror("Extraction failed", str(exc))
            self.run_button.state(["!disabled"])
            return

        flagged = [d for d in diags if d["Complete"] != "YES"]
        self.log(f"{len(rows)} row(s) written to {dest}")
        if self.masked.get():
            self.log("Part IDs are MASKED in this workbook.")
        if flagged:
            self.log(f"{len(flagged)} file(s) need checking on the "
                     f"Reconciliation sheet:")
            for d in flagged:
                self.log("  " + recon_summary(d))
        else:
            self.log("Reconciliation: every file complete.")
        self.status.config(text="Done.")
        self.run_button.state(["!disabled"])
        messagebox.showinfo("Done", f"{len(rows)} row(s) written.\n\n{dest}")


def launch_gui() -> int:
    try:
        import tkinter as tk
    except ImportError:
        print("Tkinter is unavailable. Use --src and --out for a headless run.")
        return 1
    root = tk.Tk()
    App(root)
    root.mainloop()
    return 0


# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract Part ID, Name, Status, Date of Birth, Date of "
                    "Death and Address from TIAA participant PDFs.",
    )
    parser.add_argument("--src", help="folder of PDFs (headless run)")
    parser.add_argument("--out", help="output .xlsx path")
    parser.add_argument("--engine", choices=["pdfplumber", "mupdf"],
                        default=DEFAULT_ENGINE)
    parser.add_argument("--workers", type=int, default=None)
    parser.add_argument("--mask", action="store_true",
                        help="write masked Part IDs (***-**-6789)")
    parser.add_argument("--selftest", action="store_true")
    parser.add_argument("--qa", metavar="PDF", help="masked dump of one file")
    parser.add_argument("--lines", metavar="PDF",
                        help="masked geometry dump: rebuilt lines, x spans "
                             "and column assignment (shapes only, no PII)")
    parser.add_argument("--page", type=int, default=None,
                        help="restrict --lines to one page")
    parser.add_argument("--verify", metavar="FOLDER",
                        help="compare both engines over a folder")
    args = parser.parse_args()

    if not available_engines():
        print("Install pdfplumber or PyMuPDF first:  pip install pdfplumber")
        return 1

    if args.selftest:
        return selftest()
    if args.qa:
        qa_file(args.qa, args.engine)
        return 0
    if args.lines:
        dump_lines(args.lines, args.engine, args.page)
        return 0
    if args.verify:
        return verify(args.verify)
    if args.src:
        out = args.out or str(Path(args.src) / OUTPUT_XLSX_NAME)
        return run_headless(args.src, out, args.engine, args.workers, args.mask)
    return launch_gui()


if __name__ == "__main__":
    sys.exit(main())
