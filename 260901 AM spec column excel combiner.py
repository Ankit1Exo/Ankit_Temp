"""
260901 AM spec column excel combiner.py

Combines many Excel files in a folder into ONE table, driven by a column SPEC
that you type in when the tool starts.

THE RULE THIS TOOL FOLLOWS
    You paste the list of columns a file is SUPPOSED to have. Then, sheet by
    sheet:

    * every expected column present  -> the sheet is APPENDED
    * one or more expected missing   -> the sheet is SKIPPED, and the exact
                                        missing columns are written into the
                                        summary workbook
    * extra columns not in your list -> kept, and placed at the RIGHT-HAND END
                                        of the combined table
    * a file with more than one matching sheet -> all of them are appended and
                                        the file is listed in the summary

HEADERS
    The header is ROW 1 of every sheet. Nothing above it, nothing scanned for.
    Header names match after trimming and ignoring case, so "Client ID",
    "client id" and "  CLIENT   ID " all count as the same column. The output
    uses YOUR spelling from the spec box.

    Duplicate headings inside one sheet are kept apart as "Name", "Name (2)".
    A data column whose heading cell is blank is named after its Excel column
    letter, e.g. "Column T".

DATA IS COPIED AS IT IS
    Every value is written as TEXT, exactly as it is stored in the file.
    Leading zeros survive (00123 stays 00123), nothing is converted to a
    number, a date or a code, nothing is trimmed, nothing is de-duplicated.
    Dates already stored as real Excel dates are written as YYYY-MM-DD because
    that is the only lossless text form for them.

    The ONLY rows removed are rows that are completely blank across every
    column being written.

OUTPUT (written into a "combined output" folder inside the source folder)
    <base> data.csv          every row, no size limit - use this one for the
                             normalisation step
    <base> data.xlsx         the same rows for eyeballing. If the data does not
                             fit inside Excel's 1,048,576-row limit it is split
                             into <base> data part 01.xlsx, part 02, ... The
                             run does NOT stop or fail when that happens.
    <base> summary.xlsx      Summary / Sheet Details / Extra Columns

    Column 1-3 of the combined table are Source File, Source Sheet, Source Row
    so any row can be traced back to where it came from.

SUPPORTED INPUT
    .xlsx and .xlsm in the chosen folder only (no subfolders). Files locked by
    Excel (~$...) are ignored. A file that cannot be read is reported in the
    summary; the run carries on.

USAGE
    Open in IDLE and press F5, or double-click. No command line needed.
    A folder can also be passed as the first argument.

NOTE ON FORMULAS
    Cells are read as their cached VALUES. A workbook that has never been
    opened and saved by Excel may have no cached values, in which case formula
    cells come through blank.

NOTE ON SENSITIVE DATA
    These extracts usually hold employee or client identifiers. Run this
    locally and save the output only in the approved Global Insider folder.
    Do not e-mail or upload the combined file, and keep the output folder out
    of git.
"""

import csv
import json
import os
import re
import sys
import threading
import traceback
from datetime import date, datetime, time, timedelta
from decimal import Decimal

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

# Extensions read by this tool.
READABLE = (".xlsx", ".xlsm")

# Excel's hard limit is 1,048,576 rows including the header. Parts are cut a
# little below it so there is room to add a note row later if you ever need to.
PART_ROWS = 1_000_000

# Tracking columns written in front of your own columns.
TRACKING_COLUMNS = ("Source File", "Source Sheet", "Source Row")

# House document formatting for the summary workbook.
BODY_FONT = Font(name="Calibri", size=11)
HEAD_FONT = Font(name="Calibri", size=11, bold=True)
TITLE_FONT = Font(name="Calibri", size=14, bold=True)

# Where the last-used spec is remembered between runs.
SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".excel_spec_combiner.json")

# Status values used in the summary.
APPENDED = "Appended"
SKIPPED_MISSING = "Skipped - missing columns"
SKIPPED_EMPTY = "Skipped - no header row"
FILE_ERROR = "Error - file could not be read"


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------

def is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def match_key(text):
    """Key used to line a heading up against the spec: non-breaking spaces
    normalised, outer and repeated spaces collapsed, case ignored."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\xa0", " ").split()).casefold()


def tidy(text):
    """The heading as it will be shown: trimmed, inner runs of spaces collapsed."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\xa0", " ").split())


def cell_text(value):
    """Render one cell as text WITHOUT changing what it says.

    Numbers keep every digit they were stored with, whole floats lose the
    pointless .0, real dates become YYYY-MM-DD. Strings are passed through
    untouched - no trimming - so a value stored as "00123" stays "00123"."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return str(value)
        if value.is_integer() and abs(value) < 1e15:
            return str(int(value))
        return repr(value)
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second or value.microsecond:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return str(value)
    return str(value)


def parse_spec(text):
    """Turn the pasted spec into an ordered list of column names. Accepts one
    name per line, or comma / tab / semicolon separated, or any mixture.
    Case-insensitive duplicates are dropped, keeping the first spelling."""
    pieces = []
    for line in str(text).splitlines():
        for piece in re.split(r"[,;\t]", line):
            name = tidy(piece)
            if name:
                pieces.append(name)
    ordered = []
    seen = set()
    for name in pieces:
        key = match_key(name)
        if key not in seen:
            seen.add(key)
            ordered.append(name)
    return ordered


# ---------------------------------------------------------------------------
# Planning - decide, per sheet, whether it qualifies (pass 1: headers only)
# ---------------------------------------------------------------------------

class SheetPlan:
    """One sheet of one file, and what is going to happen to it."""

    def __init__(self, path, sheet):
        self.path = path
        self.file = os.path.basename(path)
        self.sheet = sheet
        self.status = ""
        self.detail = ""
        self.hidden = False
        self.mapping = {}       # spec key -> column index in the sheet
        self.extras = []        # (column index, display name, key)
        self.missing = []       # spec names not found
        self.duplicates = []    # headings seen more than once in this sheet
        self.rows_out = 0
        self.rows_blank = 0
        self.filled = set()     # output columns that ever held a value
        self.empty_columns = [] # its own columns that came through blank

    @property
    def ok(self):
        return self.status == APPENDED


def plan_sheet(plan, header_row, spec_keys, spec_names):
    """Work out the column mapping for one sheet from its row 1."""
    cells = list(header_row or ())
    last = -1
    for index, value in enumerate(cells):
        if not is_blank(value):
            last = index
    if last < 0:
        plan.status = SKIPPED_EMPTY
        plan.detail = "row 1 is empty"
        return plan

    seen = {}
    for index in range(last + 1):
        raw = cells[index]
        if is_blank(raw):
            # A gap inside the table often still holds data underneath it, so
            # keep it and name it after its Excel column.
            display = "Column " + get_column_letter(index + 1)
        else:
            display = tidy(raw)
        key = match_key(display)
        count = seen.get(key, 0) + 1
        seen[key] = count
        if count > 1:
            plan.duplicates.append(display)
            display = "%s (%d)" % (display, count)
            key = match_key(display)
        if key in spec_keys and key not in plan.mapping:
            plan.mapping[key] = index
        else:
            plan.extras.append((index, display, key))

    plan.missing = [spec_names[key] for key in spec_keys if key not in plan.mapping]
    if plan.missing:
        plan.status = SKIPPED_MISSING
        plan.detail = "missing: " + ", ".join(plan.missing)
    else:
        plan.status = APPENDED
    return plan


def scan_file(path, spec_keys, spec_names, log):
    """Pass 1 over one workbook: read row 1 of every sheet, decide its fate."""
    plans = []
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            plan = SheetPlan(path, worksheet.title)
            plan.hidden = getattr(worksheet, "sheet_state", "visible") != "visible"
            header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            plans.append(plan_sheet(plan, header, spec_keys, spec_names))
    except Exception as error:
        plan = SheetPlan(path, "")
        plan.status = FILE_ERROR
        plan.detail = "%s: %s" % (type(error).__name__, error)
        log("  ! %s could not be read - %s" % (os.path.basename(path), error))
        plans = [plan]
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
    return plans


def collect_files(folder):
    """Every readable workbook sitting directly in the folder, sorted by name.
    Excel's own ~$ lock files are ignored."""
    found = []
    for name in sorted(os.listdir(folder), key=str.lower):
        if name.startswith("~$"):
            continue
        path = os.path.join(folder, name)
        if os.path.isfile(path) and name.lower().endswith(READABLE):
            found.append(path)
    return found


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

class PartWriter:
    """Writes the combined table to .xlsx, rolling over to a new part file
    whenever Excel's row limit is reached. Every value goes in as text."""

    def __init__(self, folder, base, headers):
        self.folder = folder
        self.base = base
        self.headers = headers
        self.paths = []
        self.workbook = None
        self.sheet = None
        self.rows_in_part = 0

    def _start(self):
        self.workbook = Workbook(write_only=True)
        self.sheet = self.workbook.create_sheet("Combined")
        self.sheet.freeze_panes = "A2"
        cells = []
        for text in self.headers:
            cell = WriteOnlyCell(self.sheet, value=text)
            cell.font = HEAD_FONT
            cells.append(cell)
        self.sheet.append(cells)
        self.rows_in_part = 0

    def _finish(self):
        if self.workbook is None:
            return
        path = os.path.join(
            self.folder, "%s data part %02d.xlsx" % (self.base, len(self.paths) + 1))
        self.workbook.save(path)
        self.workbook.close()
        self.paths.append(path)
        self.workbook = None
        self.sheet = None

    def append(self, values):
        if self.workbook is None or self.rows_in_part >= PART_ROWS:
            self._finish()
            self._start()
        row = []
        for text in values:
            if text == "":
                row.append(None)
            elif text.startswith("="):
                # Stop openpyxl storing a value that merely looks like a
                # formula as a formula.
                cell = WriteOnlyCell(self.sheet, value=text)
                cell.data_type = "s"
                row.append(cell)
            else:
                row.append(text)
        self.sheet.append(row)
        self.rows_in_part += 1

    def close(self):
        """Finish the last part. A single part is renamed to drop 'part 01'."""
        if self.workbook is None and not self.paths:
            self._start()          # nothing was written - still leave a header
        self._finish()
        if len(self.paths) == 1:
            plain = os.path.join(self.folder, "%s data.xlsx" % self.base)
            try:
                os.replace(self.paths[0], plain)
                self.paths[0] = plain
            except OSError:
                pass
        return self.paths


def unique_base(folder, base):
    """Avoid overwriting an earlier run in the same folder."""
    candidate = base
    counter = 2
    while True:
        clash = any(
            os.path.exists(os.path.join(folder, "%s %s" % (candidate, tail)))
            for tail in ("data.csv", "data.xlsx", "data part 01.xlsx", "summary.xlsx")
        )
        if not clash:
            return candidate
        candidate = "%s (%d)" % (base, counter)
        counter += 1


# ---------------------------------------------------------------------------
# The combine itself (pass 2: stream the data rows out)
# ---------------------------------------------------------------------------

def combine(folder, initials, spec_names, log, progress):
    started = datetime.now()
    spec_keys = [match_key(name) for name in spec_names]
    spec_lookup = dict(zip(spec_keys, spec_names))

    files = collect_files(folder)
    log("%d file(s) found in %s" % (len(files), folder))
    if not files:
        raise RuntimeError("No .xlsx or .xlsm files in that folder.")

    # ---- pass 1 -----------------------------------------------------------
    log("")
    log("Checking headers ...")
    plans = []
    for index, path in enumerate(files, 1):
        progress(index, len(files) * 2)
        log("  %s" % os.path.basename(path))
        for plan in scan_file(path, spec_keys, spec_lookup, log):
            plans.append(plan)
            if plan.status == APPENDED:
                log("      [%s] all %d spec columns present%s"
                    % (plan.sheet, len(spec_names),
                       ", %d extra" % len(plan.extras) if plan.extras else ""))
            elif plan.status != FILE_ERROR:
                log("      [%s] SKIPPED - %s" % (plan.sheet, plan.detail))

    # Extra columns, in the order they were first met, become the tail of the
    # output table. The first spelling seen wins.
    extra_order = []
    extra_names = {}
    for plan in plans:
        if plan.ok:
            for _, display, key in plan.extras:
                if key not in extra_names:
                    extra_names[key] = display
                    extra_order.append(key)

    headers = list(TRACKING_COLUMNS) + list(spec_names) + [extra_names[k] for k in extra_order]
    extra_position = {key: index for index, key in enumerate(extra_order)}
    good = [plan for plan in plans if plan.ok]
    log("")
    log("%d sheet(s) qualify, %d skipped. Output table is %d column(s) wide."
        % (len(good), len(plans) - len(good), len(headers)))

    # ---- output files -----------------------------------------------------
    out_folder = os.path.join(folder, "combined output")
    os.makedirs(out_folder, exist_ok=True)
    base = unique_base(out_folder, "%s %s combined" % (started.strftime("%y%m%d"), initials))
    csv_path = os.path.join(out_folder, "%s data.csv" % base)

    total_rows = 0
    total_blank = 0
    parts = PartWriter(out_folder, base, headers)

    # ---- pass 2 -----------------------------------------------------------
    log("")
    log("Combining ...")
    by_path = {}
    for plan in good:
        by_path.setdefault(plan.path, []).append(plan)

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)

        for index, path in enumerate(sorted(by_path, key=str.lower), 1):
            progress(len(files) + index, len(files) * 2)
            wanted = {plan.sheet: plan for plan in by_path[path]}
            workbook = None
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                for worksheet in workbook.worksheets:
                    plan = wanted.get(worksheet.title)
                    if plan is None:
                        continue
                    log("  %s [%s]" % (plan.file, plan.sheet))
                    targets = [plan.mapping[key] for key in spec_keys]
                    tail = len(TRACKING_COLUMNS) + len(spec_names)
                    extra_targets = [(source, tail + extra_position[key])
                                     for source, _, key in plan.extras]
                    width = len(headers)
                    row_number = 1
                    for cells in worksheet.iter_rows(min_row=2, values_only=True):
                        row_number += 1
                        size = len(cells)
                        out = [""] * width
                        empty = True
                        for offset, source in enumerate(targets):
                            column = len(TRACKING_COLUMNS) + offset
                            text = cell_text(cells[source]) if source < size else ""
                            out[column] = text
                            if text != "":
                                empty = False
                                plan.filled.add(column)
                        for source, column in extra_targets:
                            text = cell_text(cells[source]) if source < size else ""
                            out[column] = text
                            if text != "":
                                empty = False
                                plan.filled.add(column)
                        if empty:
                            plan.rows_blank += 1
                            total_blank += 1
                            continue
                        out[0] = plan.file
                        out[1] = plan.sheet
                        out[2] = str(row_number)
                        writer.writerow(out)
                        parts.append(out)
                        plan.rows_out += 1
                        total_rows += 1
                    if plan.rows_out:
                        own = set(range(len(TRACKING_COLUMNS),
                                        len(TRACKING_COLUMNS) + len(spec_names)))
                        own.update(column for _, column in extra_targets)
                        plan.empty_columns = [headers[i]
                                              for i in sorted(own - plan.filled)]
                        if plan.empty_columns:
                            log("      note: %s came through empty - check for "
                                "uncached formulas"
                                % ", ".join(plan.empty_columns))
                    log("      %d row(s) appended%s"
                        % (plan.rows_out,
                           ", %d blank row(s) dropped" % plan.rows_blank
                           if plan.rows_blank else ""))
            except Exception as error:
                # Never abandon the run because of one bad file.
                broken = SheetPlan(path, "")
                broken.status = FILE_ERROR
                broken.detail = "%s: %s" % (type(error).__name__, error)
                plans.append(broken)
                log("  ! %s failed part way through - %s"
                    % (os.path.basename(path), error))
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass

    part_paths = parts.close()
    summary_path = write_summary(
        out_folder, base, folder, started, spec_names, plans, headers,
        extra_order, extra_names, files, total_rows, total_blank,
        csv_path, part_paths)

    log("")
    log("Done. %d data row(s) written, %d blank row(s) dropped."
        % (total_rows, total_blank))
    if len(part_paths) > 1:
        log("The data did not fit in one workbook, so it was split into %d parts."
            % len(part_paths))
    log("Output folder: %s" % out_folder)
    progress(1, 1)
    return out_folder, csv_path, part_paths, summary_path, total_rows


# ---------------------------------------------------------------------------
# Summary workbook
# ---------------------------------------------------------------------------

def style_row(sheet, row_index, bold=False, title=False):
    for cell in sheet[row_index]:
        cell.font = TITLE_FONT if title else (HEAD_FONT if bold else BODY_FONT)
        cell.alignment = Alignment(vertical="top")


def add_table(sheet, headings, rows, start_row):
    row_index = start_row
    for column, text in enumerate(headings, 1):
        sheet.cell(row=row_index, column=column, value=text.upper())
    style_row(sheet, row_index, bold=True)
    row_index += 1
    for row in rows:
        for column, value in enumerate(row, 1):
            sheet.cell(row=row_index, column=column, value=value)
        style_row(sheet, row_index)
        row_index += 1
    return row_index


def add_block(sheet, row_index, title):
    row_index += 1
    sheet.cell(row=row_index, column=1, value=title)
    style_row(sheet, row_index, bold=True)
    return row_index + 1


def autosize(sheet, limit=70):
    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = min(len(str(cell.value)), limit)
            if length > widths.get(cell.column, 0):
                widths[cell.column] = length
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width + 2


def write_summary(out_folder, base, source_folder, started, spec_names, plans,
                  headers, extra_order, extra_names, files, total_rows,
                  total_blank, csv_path, part_paths):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    appended = [p for p in plans if p.ok]
    skipped = [p for p in plans if p.status in (SKIPPED_MISSING, SKIPPED_EMPTY)]
    errors = [p for p in plans if p.status == FILE_ERROR]

    # Files whose data came from more than one sheet.
    multi = {}
    for plan in appended:
        multi.setdefault(plan.file, []).append(plan.sheet)
    multi = {name: sheets for name, sheets in multi.items() if len(sheets) > 1}

    contributed = {plan.file for plan in appended}
    no_data = [os.path.basename(p) for p in files
               if os.path.basename(p) not in contributed]

    summary["A1"] = "EXCEL COMBINE SUMMARY"
    style_row(summary, 1, title=True)

    facts = [
        ("Run started", started.strftime("%Y-%m-%d %H:%M:%S")),
        ("Source folder", source_folder),
        ("Output folder", out_folder),
        ("", ""),
        ("Files found", len(files)),
        ("Files that contributed data", len(contributed)),
        ("Files that contributed nothing", len(no_data)),
        ("Files that could not be read", len(errors)),
        ("", ""),
        ("Sheets seen", len([p for p in plans if p.status != FILE_ERROR])),
        ("Sheets appended", len(appended)),
        ("Sheets skipped", len(skipped)),
        ("Files with more than one matching sheet", len(multi)),
        ("", ""),
        ("Data rows written", total_rows),
        ("Blank rows dropped", total_blank),
        ("Spec columns", len(spec_names)),
        ("Extra columns added at the end", len(extra_order)),
        ("Output table width", len(headers)),
        ("", ""),
        ("Combined CSV", os.path.basename(csv_path)),
        ("Combined workbook", ", ".join(os.path.basename(p) for p in part_paths)),
        ("Split into parts", "Yes - data exceeded Excel's row limit"
                             if len(part_paths) > 1 else "No"),
    ]
    row_index = 3
    for label, value in facts:
        summary.cell(row=row_index, column=1, value=label or None)
        summary.cell(row=row_index, column=2, value=value if value != "" else None)
        style_row(summary, row_index)
        if label:
            summary.cell(row=row_index, column=1).font = HEAD_FONT
        row_index += 1

    row_index = add_block(summary, row_index, "COLUMN SPEC YOU SUPPLIED")
    for position, name in enumerate(spec_names, 1):
        summary.cell(row=row_index, column=1, value=position)
        summary.cell(row=row_index, column=2, value=name)
        style_row(summary, row_index)
        row_index += 1

    if multi:
        row_index = add_block(summary, row_index,
                              "FILES WITH MORE THAN ONE MATCHING SHEET")
        row_index = add_table(
            summary, ("File", "Matching sheets", "Sheet names"),
            [(name, len(sheets), ", ".join(sheets))
             for name, sheets in sorted(multi.items())],
            row_index)

    if no_data:
        row_index = add_block(summary, row_index, "FILES THAT CONTRIBUTED NO DATA")
        for name in no_data:
            summary.cell(row=row_index, column=1, value=name)
            style_row(summary, row_index)
            row_index += 1

    autosize(summary)

    # ---- sheet by sheet ---------------------------------------------------
    detail = workbook.create_sheet("Sheet Details")
    detail["A1"] = "SHEET BY SHEET"
    style_row(detail, 1, title=True)
    rows = []
    for plan in plans:
        rows.append((
            plan.file,
            plan.sheet,
            plan.status,
            "Yes" if plan.hidden else "",
            plan.rows_out,
            plan.rows_blank,
            ", ".join(plan.missing),
            ", ".join(display for _, display, _ in plan.extras),
            ", ".join(sorted(set(plan.duplicates))),
            ", ".join(plan.empty_columns),
            plan.detail,
        ))
    add_table(detail, ("File", "Sheet", "Status", "Hidden sheet", "Rows appended",
                       "Blank rows dropped", "Missing spec columns",
                       "Extra columns in this sheet", "Duplicate headings",
                       "Columns that came through empty", "Detail"), rows, 3)
    detail.freeze_panes = "A4"
    autosize(detail)

    # ---- extra columns ----------------------------------------------------
    extras = workbook.create_sheet("Extra Columns")
    extras["A1"] = "EXTRA COLUMNS NOT IN YOUR SPEC"
    style_row(extras, 1, title=True)
    rows = []
    for position, key in enumerate(extra_order, 1):
        sources = [p for p in plans if p.ok and any(k == key for _, _, k in p.extras)]
        rows.append((
            position,
            extra_names[key],
            get_column_letter(len(TRACKING_COLUMNS) + len(spec_names) + position),
            len(sources),
            ", ".join(sorted({p.file for p in sources})),
        ))
    add_table(extras, ("Order", "Column name", "Output column",
                       "Sheets it came from", "Files"), rows, 3)
    extras.freeze_panes = "A4"
    autosize(extras)

    path = os.path.join(out_folder, "%s summary.xlsx" % base)
    workbook.save(path)
    return path


# ---------------------------------------------------------------------------
# Remembered settings
# ---------------------------------------------------------------------------

def load_settings():
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return {}


def save_settings(data):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as handle:
            json.dump(data, handle, indent=2)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Window
# ---------------------------------------------------------------------------

class CombinerApp:

    def __init__(self, root, start_folder=""):
        self.root = root
        root.title("Spec column Excel combiner")
        root.geometry("980x720")

        settings = load_settings()
        self.folder = tk.StringVar(value=start_folder or settings.get("folder", ""))
        self.initials = tk.StringVar(value=settings.get("initials", ""))
        self.status = tk.StringVar(value="Pick a folder, paste your column list, then Run.")
        self.busy = False
        self.last_output = ""

        pad = {"padx": 8, "pady": 4}
        top = ttk.Frame(root)
        top.pack(fill="x", **pad)

        ttk.Label(top, text="Folder of Excel files").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.folder, width=90).grid(row=0, column=1, sticky="we", padx=6)
        ttk.Button(top, text="Browse ...", command=self.pick_folder).grid(row=0, column=2)

        ttk.Label(top, text="Your 2-letter author code").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(top, textvariable=self.initials, width=8).grid(row=1, column=1, sticky="w", padx=6, pady=(6, 0))
        top.columnconfigure(1, weight=1)

        ttk.Label(
            root,
            text=("Columns every file SHOULD have - one per line (commas also accepted). "
                  "A sheet missing any of these is skipped and listed in the summary; "
                  "columns not on this list are added at the end."),
            wraplength=940, justify="left",
        ).pack(fill="x", padx=8, pady=(8, 2))

        self.spec = scrolledtext.ScrolledText(root, height=9, font=("Consolas", 10))
        self.spec.pack(fill="x", padx=8)
        self.spec.insert("1.0", settings.get("spec", ""))

        buttons = ttk.Frame(root)
        buttons.pack(fill="x", **pad)
        self.check_button = ttk.Button(buttons, text="Check headers only", command=self.on_check)
        self.check_button.pack(side="left")
        self.run_button = ttk.Button(buttons, text="Run and combine", command=self.on_run)
        self.run_button.pack(side="left", padx=8)
        ttk.Button(buttons, text="Open output folder", command=self.open_output).pack(side="left")
        ttk.Button(buttons, text="Clear log",
                   command=lambda: self.log_box.delete("1.0", "end")).pack(side="right")

        self.progress = ttk.Progressbar(root, mode="determinate", maximum=100)
        self.progress.pack(fill="x", padx=8)

        self.log_box = scrolledtext.ScrolledText(root, height=20, font=("Consolas", 9))
        self.log_box.pack(fill="both", expand=True, padx=8, pady=6)

        ttk.Label(root, textvariable=self.status, anchor="w").pack(fill="x", padx=8, pady=(0, 6))

    # -- small helpers ------------------------------------------------------

    def log(self, text=""):
        self.root.after(0, self._log, text)

    def _log(self, text):
        self.log_box.insert("end", text + "\n")
        self.log_box.see("end")

    def say(self, text):
        self.root.after(0, self.status.set, text)

    def step(self, done, total):
        value = 0 if not total else min(100, int(100 * done / total))
        self.root.after(0, self.progress.configure, {"value": value})

    def pick_folder(self):
        chosen = filedialog.askdirectory(
            title="Folder holding the Excel files",
            initialdir=self.folder.get() or os.path.expanduser("~"))
        if chosen:
            self.folder.set(os.path.normpath(chosen))

    def open_output(self):
        target = self.last_output or self.folder.get()
        if target and os.path.isdir(target):
            try:
                os.startfile(target)
            except Exception as error:
                messagebox.showerror("Cannot open folder", str(error))
        else:
            messagebox.showinfo("Nothing yet", "Run the combine first.")

    def inputs(self):
        folder = self.folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Folder", "Pick the folder holding the Excel files.")
            return None
        spec = parse_spec(self.spec.get("1.0", "end"))
        if not spec:
            messagebox.showerror("Columns", "Paste the list of expected column names first.")
            return None
        initials = self.initials.get().strip().upper()
        if len(initials) != 2 or not initials.isalpha():
            messagebox.showerror("Author code",
                                 "Please provide your 2-letter author code (your initials).")
            return None
        save_settings({"folder": folder, "initials": initials,
                       "spec": self.spec.get("1.0", "end").strip()})
        return folder, initials, spec

    def lock(self, running):
        self.busy = running
        state = "disabled" if running else "normal"
        self.run_button.configure(state=state)
        self.check_button.configure(state=state)

    # -- actions ------------------------------------------------------------

    def on_check(self):
        got = self.inputs()
        if got and not self.busy:
            folder, _, spec = got
            self.lock(True)
            threading.Thread(target=self._check, args=(folder, spec), daemon=True).start()

    def _check(self, folder, spec):
        try:
            self.say("Checking headers ...")
            spec_keys = [match_key(name) for name in spec]
            lookup = dict(zip(spec_keys, spec))
            files = collect_files(folder)
            self.log("=" * 78)
            self.log("HEADER CHECK - nothing is written")
            self.log("%d file(s) found" % len(files))
            appended = skipped = 0
            for index, path in enumerate(files, 1):
                self.step(index, len(files))
                self.log("")
                self.log(os.path.basename(path))
                for plan in scan_file(path, spec_keys, lookup, self.log):
                    if plan.status == FILE_ERROR:
                        continue
                    self.log("  %s [%s] %s" % ("OK  " if plan.ok else "SKIP",
                                               plan.sheet,
                                               "" if plan.ok else plan.detail))
                    if plan.ok and plan.extras:
                        self.log("         extra: %s"
                                 % ", ".join(d for _, d, _ in plan.extras))
                    if plan.ok:
                        appended += 1
                    else:
                        skipped += 1
            self.log("")
            self.log("%d sheet(s) would be appended, %d skipped." % (appended, skipped))
            self.say("Header check finished - %d sheet(s) would be appended." % appended)
        except Exception as error:
            self.log("")
            self.log(traceback.format_exc())
            self.say("Header check failed - see the log.")
            messagebox.showerror("Header check", str(error))
        finally:
            self.root.after(0, self.lock, False)

    def on_run(self):
        got = self.inputs()
        if got and not self.busy:
            folder, initials, spec = got
            self.lock(True)
            threading.Thread(target=self._run, args=(folder, initials, spec),
                             daemon=True).start()

    def _run(self, folder, initials, spec):
        try:
            self.say("Combining ...")
            self.log("=" * 78)
            self.log("COMBINE - %s" % datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            out_folder, csv_path, parts, summary_path, rows = combine(
                folder, initials, spec, self.log, self.step)
            self.last_output = out_folder
            self.say("Finished - %d row(s). Output in %s" % (rows, out_folder))
            messagebox.showinfo(
                "Finished",
                "%d data row(s) combined.\n\n%s\n%s\n%s\n\nSave the output in the "
                "approved Global Insider folder."
                % (rows, os.path.basename(csv_path),
                   "\n".join(os.path.basename(p) for p in parts),
                   os.path.basename(summary_path)))
        except Exception as error:
            self.log("")
            self.log(traceback.format_exc())
            self.say("Failed - see the log.")
            messagebox.showerror("Combine failed", str(error))
        finally:
            self.root.after(0, self.lock, False)


def main():
    start = ""
    if len(sys.argv) > 1 and os.path.isdir(sys.argv[1]):
        start = os.path.normpath(sys.argv[1])
    root = tk.Tk()
    CombinerApp(root, start)
    root.mainloop()


if __name__ == "__main__":
    main()
