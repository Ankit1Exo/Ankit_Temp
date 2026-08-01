"""
260731 te am excel consolidator.py

Compiles ALL sheets (including hidden / very-hidden sheets, hidden rows, and
hidden columns) from one or more Excel files (.xlsx, .xlsm, .xls) into a
SINGLE output sheet.

  - Column 1: Source File   (the original file name)
  - Column 2: Source Sheet  (the original sheet name)
  - Columns match across sheets/files by a fuzzy header comparison
    (case-insensitive, punctuation/extra-whitespace ignored). Matching
    columns are stacked under one output column; unmatched columns are
    appended to the right-hand end of the table.
  - Duplicate headers WITHIN the same sheet are kept as separate columns
    (e.g. "Name", "Name (2)").
  - Completely blank sheets and completely blank rows are skipped.
  - Cell values are copied through unchanged (no type coercion), so text
    stored with leading zeros, numbers, and dates keep their original type.
  - Merged cells: only the top-left cell of a merge holds the value (this is
    how Excel stores it); the other cells in the merge are left blank,
    exactly as they are in the source file.

USAGE
    pip install openpyxl xlrd   # xlrd only needed for legacy .xls files
    python "260731 te am excel consolidator.py" file1.xlsx file2.xls "folder with files" -o compiled.xlsx

NOTES
  - Hidden sheets/rows/columns are included automatically: this script reads
    every sheet in the workbook and every row/column in each sheet's used
    range, regardless of hidden state - nothing is filtered out based on
    visibility.
  - Header row is assumed to be row 1 of every sheet (per your confirmation).
  - For legacy .xls files, hidden-sheet/row/column state is not always
    exposed by the xlrd library, but all DATA (visible or hidden) is still
    read and included since xlrd reads the full used range regardless.
"""

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


# ---------------------------------------------------------------------------
# Path resolution (files and/or folders)
# ---------------------------------------------------------------------------
def resolve_excel_paths(inputs):
    resolved = []
    for item in inputs:
        p = Path(item)
        if p.is_dir():
            found = sorted(
                f for f in p.iterdir()
                if f.suffix.lower() in SUPPORTED_EXTENSIONS and not f.name.startswith("~$")
            )
            if not found:
                print(f"  (no .xlsx/.xlsm/.xls files found in folder: {p})")
            resolved.extend(found)
        elif p.is_file():
            if p.suffix.lower() in SUPPORTED_EXTENSIONS:
                resolved.append(p)
            else:
                print(f"  (skipping - unsupported extension: {p})")
        else:
            print(f"  (skipping - not found: {p})")
    return resolved


# ---------------------------------------------------------------------------
# Header normalization / dedup
# ---------------------------------------------------------------------------
def normalize_header(h):
    """Case-insensitive, punctuation/extra-whitespace-insensitive key for matching."""
    s = "" if h is None else str(h)
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = " ".join(s.split())
    return s


def dedup_headers(raw_headers):
    """Given a sheet's raw header row, make duplicates within that SAME sheet unique
    by appending ' (2)', ' (3)', etc. Blank headers become 'Column N'."""
    seen = {}
    result = []
    for i, h in enumerate(raw_headers, start=1):
        text = "" if h is None else str(h).strip()
        if not text:
            text = f"Column {i}"
        if text in seen:
            seen[text] += 1
            text = f"{text} ({seen[text]})"
        else:
            seen[text] = 1
        result.append(text)
    return result


# ---------------------------------------------------------------------------
# Sheet readers - yield (sheet_name, header_row, data_rows)
# data_rows is a list of lists, values copied through unchanged.
# ---------------------------------------------------------------------------
def read_xlsx_like(path):
    is_xlsm = path.suffix.lower() == ".xlsm"
    wb = load_workbook(path, data_only=True, keep_vba=is_xlsm)
    for ws in wb.worksheets:  # includes hidden and veryHidden sheets
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            continue

        raw_header = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in raw_header):
            continue  # nothing at all in this sheet

        data_rows = []
        for r in range(2, max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue  # blank row - skip
            data_rows.append(row_vals)

        if not data_rows:
            continue  # header only, no data - treat as blank sheet

        yield ws.title, raw_header, data_rows
    wb.close()


def read_xls(path):
    try:
        import xlrd
    except ImportError:
        print("xlrd not installed - run: pip install xlrd", file=sys.stderr)
        raise

    book = xlrd.open_workbook(str(path))
    for sheet in book.sheets():
        if sheet.nrows == 0 or sheet.ncols == 0:
            continue

        def cell_val(r, c):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                import datetime
                return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
            if cell.ctype == xlrd.XL_CELL_BOOLEAN:
                return bool(cell.value)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                return None
            return cell.value

        raw_header = [cell_val(0, c) for c in range(sheet.ncols)]
        if all(v is None or str(v).strip() == "" for v in raw_header):
            continue

        data_rows = []
        for r in range(1, sheet.nrows):
            row_vals = [cell_val(r, c) for c in range(sheet.ncols)]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
            data_rows.append(row_vals)

        if not data_rows:
            continue

        yield sheet.name, raw_header, data_rows


def read_workbook(path):
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        yield from read_xlsx_like(path)
    elif ext == ".xls":
        yield from read_xls(path)
    else:
        raise ValueError(f"Unsupported extension: {ext}")


# ---------------------------------------------------------------------------
# Consolidation
# ---------------------------------------------------------------------------
def consolidate(paths):
    column_order = []           # display names, in first-seen order
    normalized_to_display = {}  # normalized key -> display name used in output
    rows = []                   # list of dicts: {"Source File":.., "Source Sheet":.., <col>: val, ...}

    for path in paths:
        print(f"Reading {path} ...")
        try:
            sheets = list(read_workbook(path))
        except Exception as e:
            print(f"  ERROR reading {path}: {e}", file=sys.stderr)
            continue

        if not sheets:
            print("  (no non-blank sheets found)")

        for sheet_name, raw_header, data_rows in sheets:
            headers = dedup_headers(raw_header)

            # map each column index -> display name used in the consolidated output
            col_display_names = []
            for h in headers:
                key = normalize_header(h)
                if key not in normalized_to_display:
                    normalized_to_display[key] = h
                    column_order.append(h)
                col_display_names.append(normalized_to_display[key])

            for row_vals in data_rows:
                record = {"Source File": path.name, "Source Sheet": sheet_name}
                for col_name, val in zip(col_display_names, row_vals):
                    # if a duplicate maps to a column already set for this row
                    # (shouldn't normally happen), last value wins
                    record[col_name] = val
                rows.append(record)

            print(f"  Sheet '{sheet_name}': {len(data_rows)} row(s), {len(headers)} column(s)")

    return column_order, rows


def write_output(column_order, rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    full_headers = ["Source File", "Source Sheet"] + column_order
    ws.append(full_headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for record in rows:
        ws.append([record.get(h, None) for h in full_headers])

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(
        description="Consolidate all sheets from multiple Excel files into one sheet."
    )
    parser.add_argument("inputs", nargs="+", help="Excel file(s) and/or folder(s) to compile")
    parser.add_argument("-o", "--output", default="compiled.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    paths = resolve_excel_paths(args.inputs)
    if not paths:
        print("No Excel files to process. Exiting.")
        return

    column_order, rows = consolidate(paths)
    write_output(column_order, rows, args.output)
    print(f"\nDone. {len(rows)} row(s), {len(column_order)} data column(s) -> {args.output}")


if __name__ == "__main__":
    main()
