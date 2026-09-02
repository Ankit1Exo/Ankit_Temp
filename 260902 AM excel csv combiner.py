"""
260902 AM excel csv combiner.py

Combines every Excel and CSV file in a folder into ONE table, keeping the data
exactly as it looks on screen.

WHAT IT READS
    .xlsx  .xlsm  .xltx  .xltm     (openpyxl)
    .xls                           (xlrd, if installed)
    .xlsb                          (pyxlsb, if installed - see the caveat below)
    .csv   .tsv                    (encoding and delimiter sniffed)

    Every VISIBLE sheet of a workbook is read. Pivot table sheets are found by
    looking at the workbook's own relationships and are always left out, as are
    chart sheets and hidden or very-hidden sheets. Excel's ~$ lock files are
    ignored, and so is the tool's own "combined output" folder.

TWO WAYS TO FIND THE HEADER ROW
    1. FIRST ROW - the first row of the sheet that has anything in it is the
       header. Nothing is searched for.

    2. BY COLUMN NAME - you type one or more column names. The first row within
       the top 20 that contains ALL of those names is the header row. Rows above
       it are thrown away. A sheet with no such row in its top 20 is skipped and
       named in the summary, so you can see exactly what was left out.

WHAT COUNTS AS A COLUMN
    Once the header row is known, every column of that sheet that holds data
    becomes a column of the output - not just the ones you typed. A column whose
    heading cell is blank but which has data underneath is carried through as
    "Unknown Column <letter>", e.g. "Unknown Column T". Two columns with the
    same heading in one sheet are kept apart as "Name" and "Name (2)".

HOW COLUMNS LINE UP BETWEEN FILES
    By heading name, ignoring case and any difference in spacing, so "Client ID",
    "client id" and "  CLIENT   ID " all land in the same column. A heading that
    no earlier file had is added at the RIGHT-HAND END of the table. Nothing is
    ever dropped; a file with no such column simply leaves those cells blank.

DATA IS COPIED AS IT IS
    Every value is written as TEXT, showing what Excel shows. The cell's own
    number format is applied, so
        a date formatted dd/mm/yyyy   stays  01/03/2026
        an ID formatted 000000        stays  007123
        an SSN formatted 000-00-0000  stays  123-45-6789
        1.50 formatted 0.00           stays  1.50
        45%                           stays  45%
    Text is passed through untouched - not trimmed, not converted, not
    de-duplicated. "00123" typed as text stays "00123".

    The ONLY rows removed are rows that are completely blank.

OUTPUT (written to a "combined output" folder inside the folder you picked)
    If the data fits inside one worksheet:
        <base> data.xlsx
    If it does NOT fit in Excel's 1,048,576-row limit, you choose beforehand:
        one CSV        -> <base> data.csv, every row in a single file
        Excel parts    -> <base> data part 01.xlsx, part 02, ...
                          A part is never cut through the middle of a source
                          file: each file lands whole in one part. The only
                          exception is a single file that is itself bigger than
                          a whole worksheet, which has to be split; it is
                          flagged loudly in the log and in the summary.
    Always:
        <base> summary.xlsx     Summary / Sheet Details / Extra Columns / Skipped

    Columns 1 and 2 of the combined table are Source File and Source Sheet, so
    any row can be traced back to where it came from.

USAGE
    Open in IDLE and press F5, or double-click. No command line needed.
    A folder can also be passed as the first argument.

CAVEAT ON .xlsb
    The .xlsb reader gives values but no number formats, so a date in a .xlsb
    file comes through as its underlying serial number. If you need .xlsb dates
    to read correctly, save the file as .xlsx first. Every .xlsb file read is
    flagged in the summary.

NOTE ON FORMULAS
    Cells are read as their cached VALUES. A workbook that has never been opened
    and saved by Excel may hold no cached values, in which case formula cells
    come through blank. Any column that comes through entirely empty is called
    out in the log and in the summary.

NOTE ON SENSITIVE DATA
    These extracts usually hold employee or client identifiers such as names and
    SSNs. Run this locally, and save the output only in the approved Global
    Insider folder. Do not e-mail or upload the combined file, and keep the
    output folder out of git.
"""

import csv
import json
import os
import re
import sys
import threading
import traceback
import zipfile
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from xml.etree import ElementTree

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Optional readers. The tool runs without them and says so in the summary.
try:
    import xlrd
except Exception:
    xlrd = None
try:
    import pyxlsb
except Exception:
    pyxlsb = None
try:
    from charset_normalizer import from_path as sniff_bytes
except Exception:
    sniff_bytes = None


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

# File types this tool understands, grouped by the reader that handles them.
OPENPYXL_EXTS = (".xlsx", ".xlsm", ".xltx", ".xltm")
XLRD_EXTS = (".xls",)
XLSB_EXTS = (".xlsb",)
TEXT_EXTS = (".csv", ".tsv")
READABLE = OPENPYXL_EXTS + XLRD_EXTS + XLSB_EXTS + TEXT_EXTS

# Excel's own ceilings. A worksheet holds 1,048,576 rows INCLUDING the header.
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384

# Rows of data per part file. Kept a little under the ceiling so there is room
# for the header row and for a note row later if one is ever needed.
# The self-test lowers this to exercise the splitting code on small files.
PART_ROWS = 1_000_000

# How far down a sheet the header row is hunted for in "by column name" mode.
HEADER_SCAN_ROWS = 20

# Tracking columns written in front of your own columns.
TRACKING_COLUMNS = ("Source File", "Source Sheet")

# The folder this tool writes into, and therefore never reads from.
OUTPUT_FOLDER_NAME = "combined output"

# House document formatting for the summary workbook.
BODY_FONT = Font(name="Calibri", size=11)
HEAD_FONT = Font(name="Calibri", size=11, bold=True)
TITLE_FONT = Font(name="Calibri", size=14, bold=True)

# Where the last-used choices are remembered between runs.
SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".excel_csv_combiner.json")

# Header modes.
HEADER_FIRST_ROW = "first"
HEADER_BY_NAME = "byname"

# What to do when the combined table will not fit in one worksheet.
OVERFLOW_CSV = "csv"
OVERFLOW_PARTS = "parts"

# Status values used in the log and in the summary.
APPENDED = "Appended"
SKIPPED_NO_HEADER = "Skipped - header row not found"
SKIPPED_EMPTY = "Skipped - sheet is empty"
SKIPPED_PIVOT = "Skipped - pivot table sheet"
SKIPPED_HIDDEN = "Skipped - hidden sheet"
SKIPPED_NO_READER = "Skipped - reader not installed"
FILE_ERROR = "Error - file could not be read"


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------

def is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def match_key(text):
    """Key used to line one heading up against another: non-breaking spaces
    normalised, outer and repeated spaces collapsed, case ignored."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\xa0", " ").split()).casefold()


def tidy(text):
    """The heading as it will be shown: trimmed, inner runs of spaces collapsed."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\xa0", " ").split())


def parse_spec(text):
    """Turn the typed column list into an ordered list of names. Accepts one
    name per line, or comma / tab / semicolon separated, or any mixture.
    Case-insensitive duplicates are dropped, keeping the first spelling."""
    pieces = []
    for line in str(text).splitlines():
        for piece in re.split(r"[,;\t]", line):
            name = tidy(piece)
            if name:
                pieces.append(name)
    ordered = []
    seen = set()
    for name in pieces:
        key = match_key(name)
        if key not in seen:
            seen.add(key)
            ordered.append(name)
    return ordered


# ---------------------------------------------------------------------------
# Number format engine - turns a stored value into the text Excel shows
# ---------------------------------------------------------------------------
#
# Excel renders a cell by applying its number format code to the stored value.
# To copy data "as displayed" that has to be reproduced. A format code is up to
# four sections separated by semicolons - positive; negative; zero; text - and
# each section is a mix of digit placeholders (0 # ?), literals and date
# letters. The subset handled here covers everything these extracts throw at it:
# zero-padded IDs, punctuated IDs such as SSNs, thousands separators, fixed
# decimals, percentages, currency, scientific notation, and every ordinary date
# and time code.

# Parsed formats are reused across millions of cells, so they are cached.
_FORMAT_CACHE = {}

# Excel's day zero. Serial 60 is Excel's non-existent 29 Feb 1900, which is why
# serials above it are one day further along than a naive count suggests.
_EXCEL_EPOCH = datetime(1899, 12, 30)


def _split_sections(code):
    """Split a format code on the semicolons that separate its sections,
    ignoring semicolons inside quotes or square brackets."""
    sections = []
    current = []
    depth = 0
    in_quote = False
    index = 0
    while index < len(code):
        char = code[index]
        if in_quote:
            current.append(char)
            if char == '"':
                in_quote = False
        elif char == '"':
            current.append(char)
            in_quote = True
        elif char == "\\":
            current.append(char)
            if index + 1 < len(code):
                index += 1
                current.append(code[index])
        elif char == "[":
            depth += 1
            current.append(char)
        elif char == "]":
            depth = max(0, depth - 1)
            current.append(char)
        elif char == ";" and depth == 0:
            sections.append("".join(current))
            current = []
        else:
            current.append(char)
        index += 1
    sections.append("".join(current))
    return sections


def _tokenize(section):
    """Break one format section into (kind, text) tokens.

    kinds: lit     a literal character to print as-is
           ph      a run of digit placeholders, 0 # or ?
           dot     the decimal point
           comma   a comma, either a thousands separator or a /1000 scale
           pct     a per-cent sign, which also scales the value by 100
           sci     an E+ or E- scientific exponent marker
           date    a run of the same date letter, lower-cased
           elapsed a bracketed elapsed-time code such as [h]
           ampm    a 12-hour clock marker
           at      the @ text placeholder
    """
    tokens = []
    index = 0
    size = len(section)
    while index < size:
        char = section[index]
        if char == '"':
            end = section.find('"', index + 1)
            if end < 0:
                end = size
            tokens.append(("lit", section[index + 1:end]))
            index = end + 1
        elif char == "\\":
            if index + 1 < size:
                tokens.append(("lit", section[index + 1]))
            index += 2
        elif char == "_":
            # _x means "a space as wide as x".
            tokens.append(("lit", " "))
            index += 2
        elif char == "*":
            # *x means "repeat x to fill the cell" - nothing to copy.
            index += 2
        elif char == "[":
            end = section.find("]", index + 1)
            if end < 0:
                end = size
            body = section[index + 1:end]
            if body and set(body.lower()) <= set("hms"):
                tokens.append(("elapsed", body.lower()))
            # Anything else in brackets is a colour, a condition or a locale.
            index = end + 1
        elif char == "@":
            tokens.append(("at", "@"))
            index += 1
        elif char in "0#?":
            end = index
            while end < size and section[end] in "0#?":
                end += 1
            tokens.append(("ph", section[index:end]))
            index = end
        elif char == ".":
            tokens.append(("dot", "."))
            index += 1
        elif char == ",":
            tokens.append(("comma", ","))
            index += 1
        elif char == "%":
            tokens.append(("pct", "%"))
            index += 1
        elif char in "yYmMdDhHsS":
            letter = char.lower()
            end = index
            while end < size and section[end].lower() == letter:
                end += 1
            tokens.append(("date", section[index:end].lower()))
            index = end
        elif section[index:index + 5].upper() == "AM/PM":
            tokens.append(("ampm", "AM/PM"))
            index += 5
        elif section[index:index + 3].upper() == "A/P":
            tokens.append(("ampm", "A/P"))
            index += 3
        elif char in "eE" and index + 1 < size and section[index + 1] in "+-":
            tokens.append(("sci", section[index:index + 2]))
            index += 2
        else:
            tokens.append(("lit", char))
            index += 1
    return tokens


def _resolve_months(tokens):
    """Decide, for each run of m's, whether it means months or minutes.

    Excel's rule: an m is minutes when it sits next to an hour or a second,
    otherwise it is a month. Returns a token list in which minute runs have
    been re-kinded as 'minute'."""
    places = [i for i, (kind, _) in enumerate(tokens)
              if kind in ("date", "elapsed")]
    out = list(tokens)
    for spot, index in enumerate(places):
        kind, text = tokens[index]
        if kind != "date" or not text.startswith("m"):
            continue
        before = tokens[places[spot - 1]][1] if spot > 0 else ""
        after = tokens[places[spot + 1]][1] if spot + 1 < len(places) else ""
        if before.startswith("h") or after.startswith("s"):
            out[index] = ("minute", text)
    return out


class NumberFormat:
    """One parsed number format code, ready to render values."""

    def __init__(self, code):
        self.code = code or "General"
        self.sections = [_resolve_months(_tokenize(part))
                         for part in _split_sections(self.code)]
        self.general = self.code.strip().lower() == "general"
        self.date_sections = [
            any(kind in ("date", "minute", "elapsed", "ampm") for kind, _ in section)
            for section in self.sections
        ]

    def pick(self, value):
        """Excel chooses the section by sign: positive; negative; zero; text.
        Returns (index, section, use_absolute)."""
        count = len(self.sections)
        if isinstance(value, str):
            if count >= 4:
                return 3, self.sections[3], False
            return -1, None, False
        if not isinstance(value, (int, float, Decimal)):
            # A date, time or duration has no sign to choose a section by.
            return 0, self.sections[0], False
        negative = value < 0
        if negative and count >= 2 and self.sections[1]:
            # A negative section renders the magnitude; the code supplies its
            # own minus sign or brackets.
            return 1, self.sections[1], True
        if not negative and value == 0 and count >= 3 and self.sections[2]:
            return 2, self.sections[2], False
        return 0, self.sections[0], False


def parse_format(code):
    parsed = _FORMAT_CACHE.get(code)
    if parsed is None:
        try:
            parsed = NumberFormat(code)
        except Exception:
            parsed = NumberFormat("General")
        _FORMAT_CACHE[code] = parsed
    return parsed


def general_text(value):
    """How a value reads with no format of its own: every digit it was stored
    with, no thousands separator, whole floats without a pointless .0."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return str(value)
        if value.is_integer() and abs(value) < 1e15:
            return str(int(value))
        return repr(value)
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second or value.microsecond:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return str(value)
    return str(value)


def _serial_to_datetime(serial):
    """Excel serial number -> datetime, allowing for the 1900 leap year bug."""
    days = float(serial)
    if days < 60:
        # Serials below the phantom 29 Feb 1900 are one day later.
        return _EXCEL_EPOCH + timedelta(days=days + 1)
    return _EXCEL_EPOCH + timedelta(days=days)


MONTH_NAMES = ("January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December")
DAY_NAMES = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
             "Saturday", "Sunday")


def _render_date(value, tokens):
    """Render a date, time or duration through its format tokens."""
    elapsed = None
    if isinstance(value, timedelta):
        elapsed = value
        moment = _EXCEL_EPOCH + value
    elif isinstance(value, datetime):
        moment = value
    elif isinstance(value, date):
        moment = datetime(value.year, value.month, value.day)
    elif isinstance(value, time):
        moment = datetime(1899, 12, 30, value.hour, value.minute, value.second,
                          value.microsecond)
        elapsed = timedelta(hours=value.hour, minutes=value.minute,
                            seconds=value.second)
    else:
        moment = _serial_to_datetime(value)
        elapsed = timedelta(days=float(value))

    twelve_hour = any(kind == "ampm" for kind, _ in tokens)
    out = []
    index = 0
    while index < len(tokens):
        kind, text = tokens[index]
        if kind == "date":
            letter = text[0]
            width = len(text)
            if letter == "y":
                out.append("%04d" % moment.year if width > 2
                           else "%02d" % (moment.year % 100))
            elif letter == "m":
                if width >= 5:
                    out.append(MONTH_NAMES[moment.month - 1][0])
                elif width == 4:
                    out.append(MONTH_NAMES[moment.month - 1])
                elif width == 3:
                    out.append(MONTH_NAMES[moment.month - 1][:3])
                elif width == 2:
                    out.append("%02d" % moment.month)
                else:
                    out.append(str(moment.month))
            elif letter == "d":
                if width >= 4:
                    out.append(DAY_NAMES[moment.weekday()])
                elif width == 3:
                    out.append(DAY_NAMES[moment.weekday()][:3])
                elif width == 2:
                    out.append("%02d" % moment.day)
                else:
                    out.append(str(moment.day))
            elif letter == "h":
                hour = moment.hour
                if twelve_hour:
                    hour = hour % 12 or 12
                out.append("%02d" % hour if width >= 2 else str(hour))
            elif letter == "s":
                # Seconds followed by ".00" carry a fraction.
                seconds = "%02d" % moment.second if width >= 2 else str(moment.second)
                fraction = ""
                if (index + 2 < len(tokens) and tokens[index + 1][0] == "dot"
                        and tokens[index + 2][0] == "ph"):
                    places = len(tokens[index + 2][1])
                    fraction = ("%.*f" % (places, moment.microsecond / 1e6))[1:]
                    index += 2
                out.append(seconds + fraction)
        elif kind == "minute":
            out.append("%02d" % moment.minute if len(text) >= 2 else str(moment.minute))
        elif kind == "elapsed":
            span = elapsed if elapsed is not None else (moment - _EXCEL_EPOCH)
            total = span.total_seconds()
            if text[0] == "h":
                out.append("%0*d" % (len(text), int(total // 3600)))
            elif text[0] == "m":
                out.append("%0*d" % (len(text), int(total // 60)))
            else:
                out.append("%0*d" % (len(text), int(total)))
        elif kind == "ampm":
            morning = moment.hour < 12
            if text == "A/P":
                out.append("A" if morning else "P")
            else:
                out.append("AM" if morning else "PM")
        elif kind == "lit":
            out.append(text)
        elif kind == "dot":
            out.append(".")
        elif kind == "comma":
            out.append(",")
        index += 1
    return "".join(out)


def _fill_digits(pattern, digits):
    """Lay a string of digits into the placeholders of an integer pattern,
    right to left, so that punctuated formats such as 000-00-0000 put their
    separators in the right places. Digits left over once the placeholders run
    out are pushed out to the left, which is how Excel widens 0000 to hold a
    six-digit number. A 0 with no digit left pads with a zero, a ? pads with a
    space, and a # pads with nothing at all."""
    out = []
    spot = len(digits) - 1
    for kind, text in reversed(pattern):
        if kind == "ph":
            for placeholder in reversed(text):
                while spot >= 0 and digits[spot] == ",":
                    out.append(",")
                    spot -= 1
                if spot >= 0:
                    out.append(digits[spot])
                    spot -= 1
                elif placeholder == "0":
                    out.append("0")
                elif placeholder == "?":
                    out.append(" ")
        elif kind == "comma":
            continue            # grouping is already baked into `digits`
        else:
            out.append(text)
    while spot >= 0:
        out.append(digits[spot])
        spot -= 1
    return "".join(reversed(out))


def _render_number(value, tokens):
    """Render a number through its format tokens. Returns None if the section
    holds no digit placeholders at all, in which case the caller falls back to
    the general rendering."""
    places = [i for i, (kind, _) in enumerate(tokens) if kind == "ph"]
    if not places:
        return None

    number = float(value)

    # A per-cent sign scales by 100; commas sitting after the last placeholder
    # scale down by 1000 each.
    for kind, _ in tokens:
        if kind == "pct":
            number *= 100
    last = places[-1]
    trailing_commas = 0
    for kind, _ in tokens[last + 1:]:
        if kind == "comma":
            trailing_commas += 1
        else:
            break
    if trailing_commas:
        number /= 1000 ** trailing_commas
        tokens = tokens[:last + 1] + tokens[last + 1 + trailing_commas:]
        places = [i for i, (kind, _) in enumerate(tokens) if kind == "ph"]
        last = places[-1]

    dot = next((i for i, (kind, _) in enumerate(tokens) if kind == "dot"), None)
    if dot is not None and dot < places[0]:
        dot = None               # a dot before any placeholder is a literal

    # The exponent marker ends the number proper: in 0.00E+00 the decimals are
    # the two before the E, and the 00 after it sizes the exponent.
    sci = next((i for i, (kind, _) in enumerate(tokens) if kind == "sci"), None)

    integer_pattern = tokens[:dot] if dot is not None else tokens[:last + 1]
    if dot is None:
        decimal_pattern = []
    elif sci is not None:
        decimal_pattern = tokens[dot + 1:sci]
    else:
        decimal_pattern = tokens[dot + 1:]

    integer_places = [text for kind, text in integer_pattern if kind == "ph"]
    decimal_places = "".join(text for kind, text in decimal_pattern if kind == "ph")
    forced_left = sum(text.count("0") for text in integer_places)
    decimals = len(decimal_places)
    forced_right = decimal_places.count("0")

    # Thousands grouping is asked for by a comma between two placeholders.
    grouped = any(kind == "comma" and places[0] < index < last
                  for index, (kind, _) in enumerate(tokens))

    if sci is not None:
        body = "%.*E" % (decimals, number)
        mantissa, _, exponent = body.partition("E")
        width = len(next((text for kind, text in tokens[sci + 1:] if kind == "ph"),
                         "00"))
        power = int(exponent)
        marker = tokens[sci][1]
        sign = "-" if power < 0 else ("+" if marker.endswith("+") else "")
        prefix = "".join(text for kind, text in tokens[:places[0]] if kind == "lit")
        return "%s%sE%s%0*d" % (prefix, mantissa, sign, width, abs(power))

    # Excel rounds a half away from zero, where Python rounds it to the
    # nearest even number, so 2.5 under "0" has to come out as 3 not 2.
    rounded = str(Decimal(str(abs(number))).quantize(
        Decimal(1).scaleb(-decimals), rounding=ROUND_HALF_UP))
    if decimals:
        whole, _, fraction = rounded.partition(".")
    else:
        whole, fraction = rounded, ""

    if int(whole) == 0 and forced_left == 0:
        # No 0 placeholder on the left means Excel shows no digit at all for a
        # zero integer part: "#.##" shows .5, and an accounting zero shows " - ".
        whole = ""
    elif grouped:
        whole = "{:,}".format(int(whole))

    out = [_fill_digits(integer_pattern, whole)]
    if decimals:
        # Trailing zeros only survive where the format forced them with a 0.
        keep = fraction
        while len(keep) > forced_right and keep.endswith("0"):
            keep = keep[:-1]
        if keep:
            out.append(".")
            out.append(keep)
        elif forced_right:
            out.append(".")
            out.append("0" * forced_right)
    # Everything after the last placeholder that is not part of the number.
    for kind, text in tokens[last + 1:]:
        if kind == "lit":
            out.append(text)
        elif kind == "pct":
            out.append("%")
        elif kind == "dot" and not decimals:
            out.append(".")
    body = "".join(out)
    if number < 0:
        body = "-" + body
    return body


def display_text(value, number_format):
    """The heart of "as displayed": one stored value plus its number format
    code, rendered to the text Excel puts on screen."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, str):
        # Text is never reformatted. Whatever was typed is what is copied.
        return value

    fmt = parse_format(number_format)
    if fmt.general:
        return general_text(value)

    index, section, absolute = fmt.pick(value)
    if section is None:
        return general_text(value)

    try:
        if fmt.date_sections[index]:
            return _render_date(value, section)
        if isinstance(value, (datetime, date, time, timedelta)):
            # A date value under a numeric format is shown by Excel as its
            # serial number; keeping the readable form is the safer copy.
            return general_text(value)
        rendered = _render_number(abs(value) if absolute else value, section)
        if rendered is None:
            return general_text(value)
        return rendered
    except Exception:
        # No cell is ever allowed to break a run.
        return general_text(value)


# ---------------------------------------------------------------------------
# Pivot table detection
# ---------------------------------------------------------------------------

_NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_NS_REL_DOC = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_NS_REL_PKG = "{http://schemas.openxmlformats.org/package/2006/relationships}"


def pivot_sheet_names(path):
    """Names of the sheets in an xlsx-family workbook that hold a pivot table.

    Found from the package itself rather than from openpyxl, because a pivot
    sheet is only identifiable by the relationship its own sheet part declares,
    and openpyxl does not surface that in read-only mode. A workbook that
    cannot be inspected returns an empty set - better to combine a pivot sheet
    by mistake than to abandon the run."""
    names = set()
    try:
        with zipfile.ZipFile(path) as package:
            listing = set(package.namelist())
            if "xl/workbook.xml" not in listing:
                return names

            # rId -> part name, from the workbook's own relationships.
            targets = {}
            if "xl/_rels/workbook.xml.rels" in listing:
                tree = ElementTree.fromstring(package.read("xl/_rels/workbook.xml.rels"))
                for node in tree:
                    rel_id = node.get("Id")
                    target = node.get("Target") or ""
                    if not rel_id:
                        continue
                    target = target.replace("\\", "/")
                    if target.startswith("/"):
                        part = target.lstrip("/")
                    elif target.startswith("../"):
                        part = target[3:]
                    else:
                        part = "xl/" + target
                    targets[rel_id] = part

            book = ElementTree.fromstring(package.read("xl/workbook.xml"))
            for node in book.iter(_NS_MAIN + "sheet"):
                title = node.get("name")
                rel_id = node.get(_NS_REL_DOC + "id")
                part = targets.get(rel_id)
                if not title or not part:
                    continue
                folder, _, filename = part.rpartition("/")
                rels = "%s/_rels/%s.rels" % (folder, filename)
                if rels not in listing:
                    continue
                body = package.read(rels).decode("utf-8", "replace")
                if "pivotTable" in body:
                    names.add(title)
    except Exception:
        return set()
    return names


# ---------------------------------------------------------------------------
# Readers - one small class per family, all with the same three methods
# ---------------------------------------------------------------------------
#
# sheets()          the visible, non-pivot sheets worth reading, in book order
# values(sheet)     raw values, row by row - used by pass 1, which only needs
#                   to find the header row and count the data
# display(sheet)    the same rows rendered to display text - used by pass 2
#
# Both generators yield lists, and neither pads short rows: a row shorter than
# the header is simply missing its tail, exactly as Excel stores it.

class SheetInfo:
    def __init__(self, name, status="", detail=""):
        self.name = name
        self.status = status        # "" means readable
        self.detail = detail


def _shut(book):
    try:
        book.close()
    except Exception:
        pass


class OpenpyxlReader:
    """.xlsx .xlsm .xltx .xltm - the only family that carries number formats
    and pivot table relationships, so the only one that is fully faithful."""

    def __init__(self, path):
        self.path = path
        self.notes = []

    def sheets(self):
        found = []
        pivots = pivot_sheet_names(self.path)
        book = load_workbook(self.path, read_only=True, data_only=True)
        try:
            for sheet in book.worksheets:
                state = getattr(sheet, "sheet_state", "visible")
                if sheet.title in pivots:
                    found.append(SheetInfo(sheet.title, SKIPPED_PIVOT,
                                           "holds a pivot table"))
                elif state != "visible":
                    found.append(SheetInfo(sheet.title, SKIPPED_HIDDEN, state))
                else:
                    found.append(SheetInfo(sheet.title))
            # Chart sheets never hold cell data and openpyxl keeps them apart.
            for sheet in getattr(book, "chartsheets", []):
                found.append(SheetInfo(sheet.title, SKIPPED_PIVOT, "chart sheet"))
        finally:
            _shut(book)
        return found

    def values(self, name):
        book = load_workbook(self.path, read_only=True, data_only=True)
        try:
            for row in book[name].iter_rows(values_only=True):
                yield list(row)
        finally:
            _shut(book)

    def display(self, name):
        book = load_workbook(self.path, read_only=True, data_only=True)
        try:
            for row in book[name].iter_rows():
                yield [display_text(cell.value, cell.number_format) for cell in row]
        finally:
            _shut(book)


class XlrdReader:
    """.xls through xlrd. Number formats are available, so values render the
    same way they do for .xlsx. Pivot sheets cannot be identified in the old
    binary format, which is noted in the summary."""

    def __init__(self, path):
        self.path = path
        self.notes = ["%s is a .xls file: pivot sheets cannot be detected in "
                      "the old format, so check the sheet list."
                      % os.path.basename(path)]

    def _open(self):
        try:
            return xlrd.open_workbook(self.path, formatting_info=True), True
        except Exception:
            # Some .xls files refuse formatting_info; values still come through.
            return xlrd.open_workbook(self.path), False

    def sheets(self):
        book, _ = self._open()
        found = []
        try:
            for sheet in book.sheets():
                if getattr(sheet, "visibility", 0):
                    found.append(SheetInfo(sheet.name, SKIPPED_HIDDEN, "hidden"))
                else:
                    found.append(SheetInfo(sheet.name))
        finally:
            _shut(book)
        return found

    def _rows(self, name, rendered):
        book, styled = self._open()
        try:
            sheet = book.sheet_by_name(name)
            formats = {}
            if styled:
                for index, record in enumerate(book.xf_list):
                    entry = book.format_map.get(record.format_key)
                    formats[index] = getattr(entry, "format_str", "General")
            for index in range(sheet.nrows):
                row = []
                for column in range(sheet.ncols):
                    kind = sheet.cell_type(index, column)
                    value = sheet.cell_value(index, column)
                    if kind in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        value = None
                    elif kind == xlrd.XL_CELL_BOOLEAN:
                        value = bool(value)
                    elif kind == xlrd.XL_CELL_ERROR:
                        value = xlrd.error_text_from_code.get(value, "#ERROR")
                    elif kind == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                        except Exception:
                            pass
                    if not rendered:
                        row.append(value)
                    else:
                        code = "General"
                        if styled:
                            code = formats.get(sheet.cell_xf_index(index, column),
                                               "General")
                        row.append(display_text(value, code))
                yield row
        finally:
            _shut(book)

    def values(self, name):
        return self._rows(name, False)

    def display(self, name):
        return self._rows(name, True)


class XlsbReader:
    """.xlsb through pyxlsb. Values only - the format table is not exposed, so
    real dates arrive as serial numbers. Flagged in the summary."""

    def __init__(self, path):
        self.path = path
        self.notes = ["%s is a .xlsb file: number formats are not readable, so "
                      "dates come through as serial numbers and pivot sheets "
                      "cannot be detected. Re-save it as .xlsx for a faithful "
                      "copy." % os.path.basename(path)]

    def sheets(self):
        with pyxlsb.open_workbook(self.path) as book:
            return [SheetInfo(name) for name in book.sheets]

    def _rows(self, name, rendered):
        with pyxlsb.open_workbook(self.path) as book:
            with book.get_sheet(name) as sheet:
                for cells in sheet.rows():
                    row = []
                    for cell in cells:
                        value = cell.v
                        if isinstance(value, str) and value == "":
                            value = None
                        row.append(general_text(value) if rendered else value)
                    yield row

    def values(self, name):
        return self._rows(name, False)

    def display(self, name):
        return self._rows(name, True)


class TextReader:
    """.csv and .tsv. Every field is already text, so this is the one reader
    that cannot lose anything - nothing is parsed, converted or trimmed."""

    SHEET = "CSV"

    def __init__(self, path):
        self.path = path
        self.notes = []
        self.encoding = None
        self.delimiter = None

    def _settle(self):
        """Work out the encoding and the delimiter once, and say so in the log."""
        if self.encoding:
            return
        with open(self.path, "rb") as handle:
            head = handle.read(64 * 1024)
        if head.startswith(b"\xef\xbb\xbf"):
            self.encoding = "utf-8-sig"
        elif head.startswith((b"\xff\xfe", b"\xfe\xff")):
            self.encoding = "utf-16"
        else:
            self.encoding = "utf-8"
            try:
                head.decode("utf-8")
            except UnicodeDecodeError:
                guess = None
                if sniff_bytes is not None:
                    try:
                        best = sniff_bytes(self.path).best()
                        guess = best.encoding if best else None
                    except Exception:
                        guess = None
                self.encoding = guess or "cp1252"

        sample = ""
        try:
            with open(self.path, "r", encoding=self.encoding, errors="replace",
                      newline="") as handle:
                sample = handle.read(64 * 1024)
        except Exception:
            sample = ""
        if self.path.lower().endswith(".tsv"):
            self.delimiter = "\t"
        else:
            try:
                self.delimiter = csv.Sniffer().sniff(
                    sample, delimiters=",;\t|").delimiter
            except Exception:
                self.delimiter = ","
        self.notes.append(
            "%s read as %s, delimiter %r."
            % (os.path.basename(self.path), self.encoding, self.delimiter))

    def sheets(self):
        self._settle()
        return [SheetInfo(self.SHEET)]

    def _rows(self):
        self._settle()
        with open(self.path, "r", encoding=self.encoding, errors="replace",
                  newline="") as handle:
            for row in csv.reader(handle, delimiter=self.delimiter):
                yield row

    def values(self, name):
        return self._rows()

    def display(self, name):
        return self._rows()


def reader_for(path):
    """The right reader for a path, or None with the reason it cannot be read."""
    lower = path.lower()
    if lower.endswith(OPENPYXL_EXTS):
        return OpenpyxlReader(path), ""
    if lower.endswith(XLRD_EXTS):
        if xlrd is None:
            return None, "xlrd is not installed - run: pip install xlrd"
        return XlrdReader(path), ""
    if lower.endswith(XLSB_EXTS):
        if pyxlsb is None:
            return None, "pyxlsb is not installed - run: pip install pyxlsb"
        return XlsbReader(path), ""
    if lower.endswith(TEXT_EXTS):
        return TextReader(path), ""
    return None, "unsupported file type"


def _wanted(name):
    return (not name.startswith("~$")
            and not name.startswith(".")
            and name.lower().endswith(READABLE))


def collect_files(folder, subfolders=False):
    """Every readable file in the folder, sorted by name. Excel's own ~$ lock
    files and this tool's own output folder are left out."""
    found = []
    if subfolders:
        for root, directories, names in os.walk(folder):
            directories[:] = sorted(
                d for d in directories if d.lower() != OUTPUT_FOLDER_NAME)
            for name in sorted(names, key=str.lower):
                if _wanted(name):
                    found.append(os.path.join(root, name))
    else:
        for name in sorted(os.listdir(folder), key=str.lower):
            path = os.path.join(folder, name)
            if os.path.isfile(path) and _wanted(name):
                found.append(path)
    return found


# ---------------------------------------------------------------------------
# Header hunting and per-sheet planning (pass 1)
# ---------------------------------------------------------------------------

def find_header_row(rows, mode, spec_keys):
    """Which of the scanned rows is the header. Returns its index in `rows`,
    or None.

    FIRST ROW mode takes the first row with anything in it. BY COLUMN NAME mode
    takes the first row that carries EVERY name asked for - one name means one
    name has to be there, three names means all three do."""
    if mode == HEADER_FIRST_ROW:
        for index, row in enumerate(rows):
            if any(not is_blank(value) for value in row):
                return index
        return None
    wanted = set(spec_keys)
    for index, row in enumerate(rows):
        present = {match_key(value) for value in row if not is_blank(value)}
        if wanted <= present:
            return index
    return None


class SheetPlan:
    """One sheet of one file, and what is going to happen to it."""

    def __init__(self, path, sheet):
        self.path = path
        self.file = os.path.basename(path)
        self.sheet = sheet
        self.status = ""
        self.detail = ""
        self.header_offset = 0      # rows above the header, thrown away
        self.columns = []           # (source index, display name, match key)
        self.duplicates = []        # headings seen more than once here
        self.unknown = []           # data columns that had no heading
        self.data_rows = 0          # non-blank rows counted in pass 1
        self.rows_out = 0           # rows actually written in pass 2
        self.rows_blank = 0
        self.filled = set()         # output columns that ever held a value
        self.empty_columns = []

    @property
    def ok(self):
        return self.status == APPENDED

    @property
    def label(self):
        return "%s [%s]" % (self.file, self.sheet)


def _tally(row, with_data, plan):
    """Count one data row and remember which of its columns held something."""
    empty = True
    for index, value in enumerate(row):
        if not is_blank(value):
            with_data.add(index)
            empty = False
    if empty:
        plan.rows_blank += 1
    else:
        plan.data_rows += 1


def plan_sheet(plan, rows, mode, spec_keys):
    """Read one sheet through once: find the header, work out its columns, and
    count the data rows so the output can be sized before anything is written."""
    scanned = []
    stream = iter(rows)
    for row in stream:
        scanned.append(row)
        if len(scanned) >= HEADER_SCAN_ROWS:
            break

    if not any(any(not is_blank(v) for v in row) for row in scanned):
        plan.status = SKIPPED_EMPTY
        plan.detail = "no data in the first %d row(s)" % max(1, len(scanned))
        return plan

    offset = find_header_row(scanned, mode, spec_keys)
    if offset is None:
        plan.status = SKIPPED_NO_HEADER
        plan.detail = ("no row in the first %d holds all of: %s"
                       % (HEADER_SCAN_ROWS, ", ".join(spec_keys)))
        return plan

    plan.header_offset = offset
    header = scanned[offset]

    # Walk the rest of the sheet to count rows and to learn which columns hold
    # data. Columns with data but no heading become "Unknown Column X", and
    # data sitting to the right of the last heading is picked up here too.
    with_data = set()
    for row in scanned[offset + 1:]:
        _tally(row, with_data, plan)
    for row in stream:
        _tally(row, with_data, plan)

    last_heading = -1
    for index, value in enumerate(header):
        if not is_blank(value):
            last_heading = index
    width = max(last_heading + 1, (max(with_data) + 1) if with_data else 0)

    seen = {}
    for index in range(width):
        raw = header[index] if index < len(header) else None
        if is_blank(raw):
            if index not in with_data:
                continue        # no heading and no data - not a column at all
            display = "Unknown Column " + get_column_letter(index + 1)
            plan.unknown.append(display)
        else:
            display = tidy(raw)
        key = match_key(display)
        count = seen.get(key, 0) + 1
        seen[key] = count
        if count > 1:
            plan.duplicates.append(display)
            display = "%s (%d)" % (display, count)
            key = match_key(display)
        plan.columns.append((index, display, key))

    if not plan.columns:
        plan.status = SKIPPED_EMPTY
        plan.detail = "header row found but the sheet holds no columns"
        return plan

    plan.status = APPENDED
    bits = ["%d column(s)" % len(plan.columns), "%d data row(s)" % plan.data_rows]
    if offset:
        bits.append("header on row %d" % (offset + 1))
    if plan.unknown:
        bits.append("%d unnamed" % len(plan.unknown))
    plan.detail = ", ".join(bits)
    return plan


def scan_file(path, mode, spec_keys, log):
    """Pass 1 over one file: every sheet's fate and size, decided up front."""
    plans = []
    reader, problem = reader_for(path)
    if reader is None:
        plan = SheetPlan(path, "")
        plan.status = SKIPPED_NO_READER
        plan.detail = problem
        log("  ! %s - %s" % (os.path.basename(path), problem))
        return [plan], []
    try:
        for info in reader.sheets():
            plan = SheetPlan(path, info.name)
            if info.status:
                plan.status = info.status
                plan.detail = info.detail
                plans.append(plan)
                continue
            plans.append(plan_sheet(plan, reader.values(info.name), mode, spec_keys))
    except Exception as error:
        plan = SheetPlan(path, "")
        plan.status = FILE_ERROR
        plan.detail = "%s: %s" % (type(error).__name__, error)
        log("  ! %s could not be read - %s" % (os.path.basename(path), error))
        plans.append(plan)
    return plans, list(getattr(reader, "notes", []))


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

class PartWriter:
    """Writes the combined table to .xlsx, starting a new part file whenever
    the row limit would be crossed.

    A part is never cut through the middle of a source file. Before each file
    the writer is told how many rows are coming, and if they will not fit in
    what is left of the current part it rolls over first. A single file bigger
    than a whole worksheet cannot be kept together; it is split and recorded in
    `split_files` so the run can say so."""

    def __init__(self, folder, base, headers, limit=None):
        self.folder = folder
        self.base = base
        self.headers = headers
        self.limit = limit or PART_ROWS
        self.paths = []
        self.workbook = None
        self.sheet = None
        self.rows_in_part = 0
        self.split_files = []
        self.current_file = ""

    def _start(self):
        self.workbook = Workbook(write_only=True)
        self.sheet = self.workbook.create_sheet("Combined")
        self.sheet.freeze_panes = "A2"
        cells = []
        for text in self.headers:
            cell = WriteOnlyCell(self.sheet, value=text)
            cell.font = HEAD_FONT
            cells.append(cell)
        self.sheet.append(cells)
        self.rows_in_part = 0

    def _finish(self):
        if self.workbook is None:
            return
        path = os.path.join(
            self.folder, "%s data part %02d.xlsx" % (self.base, len(self.paths) + 1))
        self.workbook.save(path)
        self.workbook.close()
        self.paths.append(path)
        self.workbook = None
        self.sheet = None

    def begin_file(self, name, rows_expected):
        """Keep this whole file in one part if it can possibly be done."""
        self.current_file = name
        if self.workbook is None:
            return
        if self.rows_in_part and self.rows_in_part + rows_expected > self.limit:
            self._finish()

    def append(self, values):
        if self.workbook is None:
            self._start()
        elif self.rows_in_part >= self.limit:
            # Only reachable for a file too big to fit a worksheet on its own.
            if self.current_file and self.current_file not in self.split_files:
                self.split_files.append(self.current_file)
            self._finish()
            self._start()
        row = []
        for text in values:
            if text == "":
                row.append(None)
            elif text[0] in "=+-@":
                # Stop Excel and openpyxl treating a value that merely looks
                # like a formula as one. The text itself is unchanged.
                cell = WriteOnlyCell(self.sheet, value=text)
                cell.data_type = "s"
                row.append(cell)
            else:
                row.append(text)
        self.sheet.append(row)
        self.rows_in_part += 1

    def close(self):
        """Finish the last part. A lone part drops the 'part 01' from its name."""
        if self.workbook is None and not self.paths:
            self._start()               # nothing qualified - still leave a header
        self._finish()
        if len(self.paths) == 1:
            plain = os.path.join(self.folder, "%s data.xlsx" % self.base)
            try:
                os.replace(self.paths[0], plain)
                self.paths[0] = plain
            except OSError:
                pass
        return self.paths


class CsvOutput:
    """Writes the combined table to a single CSV, which has no row limit."""

    def __init__(self, folder, base, headers):
        self.path = os.path.join(folder, "%s data.csv" % base)
        self.split_files = []
        self.handle = open(self.path, "w", newline="", encoding="utf-8-sig")
        self.writer = csv.writer(self.handle)
        self.writer.writerow(headers)

    def begin_file(self, name, rows_expected):
        pass

    def append(self, values):
        self.writer.writerow(values)

    def close(self):
        self.handle.close()
        return [self.path]


def unique_base(folder, base):
    """Avoid overwriting an earlier run in the same folder."""
    candidate = base
    counter = 2
    while True:
        clash = any(
            os.path.exists(os.path.join(folder, "%s %s" % (candidate, tail)))
            for tail in ("data.csv", "data.xlsx", "data part 01.xlsx", "summary.xlsx")
        )
        if not clash:
            return candidate
        candidate = "%s (%d)" % (base, counter)
        counter += 1


# ---------------------------------------------------------------------------
# The combine itself
# ---------------------------------------------------------------------------

def build_headers(plans, spec_names, mode):
    """The output table's columns.

    Tracking columns first, then - in "by column name" mode - the names that
    were typed in, spelled the way they were typed, because every appended
    sheet is guaranteed to have them. Everything else follows in the order it
    was first met, so a heading no earlier file had lands at the end."""
    order = []
    names = {}
    if mode == HEADER_BY_NAME:
        for name in spec_names:
            key = match_key(name)
            if key not in names:
                names[key] = name
                order.append(key)
    for plan in plans:
        if not plan.ok:
            continue
        for _, display, key in plan.columns:
            if key not in names:
                names[key] = display
                order.append(key)
    headers = list(TRACKING_COLUMNS) + [names[key] for key in order]
    position = {key: len(TRACKING_COLUMNS) + index
                for index, key in enumerate(order)}
    return headers, position


def combine(folder, initials, mode, spec_names, overflow, subfolders, log, progress):
    started = datetime.now()
    spec_keys = [match_key(name) for name in spec_names]

    files = collect_files(folder, subfolders)
    log("%d file(s) found in %s%s"
        % (len(files), folder, " and its subfolders" if subfolders else ""))
    if not files:
        raise RuntimeError(
            "No Excel or CSV files in that folder. Looked for: %s"
            % ", ".join(READABLE))

    # ---- pass 1: headers, columns and sizes -------------------------------
    log("")
    log("Reading headers ...")
    plans = []
    notes = []
    for index, path in enumerate(files, 1):
        progress(index, len(files) * 2)
        log("  %s" % os.path.relpath(path, folder))
        found, extra = scan_file(path, mode, spec_keys, log)
        notes.extend(extra)
        for plan in found:
            plans.append(plan)
            if plan.status == FILE_ERROR or plan.status == SKIPPED_NO_READER:
                continue
            marker = "OK  " if plan.ok else "SKIP"
            log("      %s [%s] %s" % (marker, plan.sheet, plan.detail))

    good = [plan for plan in plans if plan.ok]
    if not good:
        raise RuntimeError(
            "Nothing to combine - no sheet produced a usable header row. "
            "The summary was not written; check the log above.")

    headers, position = build_headers(plans, spec_names, mode)
    total_rows = sum(plan.data_rows for plan in good)

    log("")
    log("%d sheet(s) qualify, %d skipped. The table is %d column(s) wide and "
        "about %d row(s) deep." % (len(good), len(plans) - len(good),
                                   len(headers), total_rows))
    if len(headers) > EXCEL_MAX_COLUMNS:
        raise RuntimeError(
            "The combined table would be %d columns wide, past Excel's limit "
            "of %d. Narrow the input before combining."
            % (len(headers), EXCEL_MAX_COLUMNS))

    # ---- decide where it is going before writing a single row -------------
    out_folder = os.path.join(folder, OUTPUT_FOLDER_NAME)
    os.makedirs(out_folder, exist_ok=True)
    base = unique_base(out_folder, "%s %s combined" % (started.strftime("%y%m%d"),
                                                       initials))
    fits = total_rows + 1 <= EXCEL_MAX_ROWS
    if fits:
        writer = PartWriter(out_folder, base, headers, limit=EXCEL_MAX_ROWS - 1)
        log("It fits in one worksheet, so one workbook is being written.")
    elif overflow == OVERFLOW_CSV:
        writer = CsvOutput(out_folder, base, headers)
        log("It does NOT fit in one worksheet (%d rows against a limit of %d), "
            "so everything is going into one CSV as you chose."
            % (total_rows, EXCEL_MAX_ROWS - 1))
    else:
        writer = PartWriter(out_folder, base, headers)
        log("It does NOT fit in one worksheet (%d rows against a limit of %d), "
            "so it is being split into parts of up to %d rows, each source "
            "file kept whole." % (total_rows, EXCEL_MAX_ROWS - 1, PART_ROWS))

    # ---- pass 2: stream the data out --------------------------------------
    log("")
    log("Combining ...")
    by_path = {}
    for plan in good:
        by_path.setdefault(plan.path, []).append(plan)

    written = 0
    dropped = 0
    order = [path for path in files if path in by_path]
    try:
        for index, path in enumerate(order, 1):
            progress(len(files) + index, len(files) * 2)
            sheets = by_path[path]
            name = os.path.basename(path)
            writer.begin_file(name, sum(plan.data_rows for plan in sheets))
            reader, problem = reader_for(path)
            if reader is None:
                continue
            for plan in sheets:
                try:
                    rows, blanks = _pour(plan, reader, position, len(headers),
                                         headers, writer)
                except Exception as error:
                    plan.status = FILE_ERROR
                    plan.detail = "%s: %s" % (type(error).__name__, error)
                    log("  ! %s failed part way through - %s"
                        % (plan.label, error))
                    continue
                written += rows
                dropped += blanks
                log("  %s -> %d row(s)%s"
                    % (plan.label, rows,
                       ", %d blank row(s) dropped" % blanks if blanks else ""))
                if plan.empty_columns:
                    log("      note: %s came through empty - check for "
                        "uncached formulas"
                        % ", ".join(plan.empty_columns))
    finally:
        paths = writer.close()

    if writer.split_files:
        log("")
        log("WARNING: %s is bigger on its own than one worksheet can hold, so "
            "its rows had to be spread over more than one part."
            % ", ".join(writer.split_files))

    summary_path = write_summary(
        out_folder, base, folder, started, mode, spec_names, overflow,
        subfolders, plans, headers, files, written, dropped, paths,
        writer.split_files, notes)

    log("")
    log("Done. %d data row(s) written, %d blank row(s) dropped."
        % (written, dropped))
    for path in paths:
        log("  %s" % os.path.basename(path))
    log("  %s" % os.path.basename(summary_path))
    log("Output folder: %s" % out_folder)
    log("Save the output in the approved Global Insider folder, not here.")
    progress(1, 1)
    return out_folder, paths, summary_path, written


def _pour(plan, reader, position, width, headers, writer):
    """Stream one sheet's data rows into the writer."""
    targets = [(source, position[key]) for source, _, key in plan.columns]
    skip = plan.header_offset + 1
    rows = 0
    blanks = 0
    for index, cells in enumerate(reader.display(plan.sheet)):
        if index < skip:
            continue
        size = len(cells)
        out = [""] * width
        empty = True
        for source, target in targets:
            if source >= size:
                continue
            text = cells[source]
            if text != "":
                out[target] = text
                empty = False
                plan.filled.add(target)
        if empty:
            blanks += 1
            continue
        out[0] = plan.file
        out[1] = plan.sheet
        writer.append(out)
        rows += 1
    plan.rows_out = rows
    plan.rows_blank = blanks
    if rows:
        own = {target for _, target in targets}
        plan.empty_columns = [headers[i] for i in sorted(own - plan.filled)]
    return rows, blanks


# ---------------------------------------------------------------------------
# Summary workbook
# ---------------------------------------------------------------------------

def style_row(sheet, row_index, bold=False, title=False):
    for cell in sheet[row_index]:
        cell.font = TITLE_FONT if title else (HEAD_FONT if bold else BODY_FONT)
        cell.alignment = Alignment(vertical="top", wrap_text=not (bold or title))


def add_block(sheet, row_index, title):
    """A subheading, in house style: Calibri 11, uppercase, bold."""
    sheet.cell(row=row_index, column=1, value=title.upper())
    style_row(sheet, row_index, bold=True)
    return row_index + 2


def add_table(sheet, headings, rows, start_row):
    for column, text in enumerate(headings, 1):
        sheet.cell(row=start_row, column=column, value=text)
    style_row(sheet, start_row, bold=True)
    row_index = start_row + 1
    for values in rows:
        for column, value in enumerate(values, 1):
            sheet.cell(row=row_index, column=column, value=value)
        style_row(sheet, row_index)
        row_index += 1
    if rows:
        sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1).coordinate
    return row_index + 1


def autosize(sheet, limit=70):
    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            longest = max(len(line) for line in str(cell.value).split("\n"))
            widths[cell.column] = min(limit, max(widths.get(cell.column, 10),
                                                 longest + 2))
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def page_furniture(sheet, description):
    """House header and footer: EXO Edge plus the document description at the
    top, Page X of Y bottom left, Calibri 11 throughout."""
    sheet.oddHeader.left.text = "EXO Edge - %s" % description
    sheet.oddHeader.left.font = "Calibri,Regular"
    sheet.oddHeader.left.size = 11
    sheet.oddFooter.left.text = "Page &P of &N"
    sheet.oddFooter.left.font = "Calibri,Regular"
    sheet.oddFooter.left.size = 11


def write_summary(out_folder, base, source_folder, started, mode, spec_names,
                  overflow, subfolders, plans, headers, files, written, dropped,
                  paths, split_files, notes):
    """The run's own record: what was combined, what was left out and why."""
    description = "Combined Excel and CSV data summary"
    book = Workbook()

    # ---- Summary ----------------------------------------------------------
    sheet = book.active
    sheet.title = "Summary"
    page_furniture(sheet, description)
    sheet["A1"] = "COMBINED EXCEL AND CSV DATA - SUMMARY"
    style_row(sheet, 1, title=True)

    good = [plan for plan in plans if plan.ok]
    skipped = [plan for plan in plans if not plan.ok]
    if mode == HEADER_BY_NAME:
        how = ("Header row found by column name, scanning the first %d rows for: %s"
               % (HEADER_SCAN_ROWS, ", ".join(spec_names)))
    else:
        how = "Header row = the first row of each sheet that holds anything"

    rows = [
        ("Run started", started.strftime("%Y-%m-%d %H:%M:%S")),
        ("Source folder", source_folder),
        ("Subfolders included", "Yes" if subfolders else "No"),
        ("How the header was found", how),
        ("Files examined", len(files)),
        ("Sheets appended", len(good)),
        ("Sheets skipped", len(skipped)),
        ("Data rows written", written),
        ("Blank rows dropped", dropped),
        ("Columns in the output", len(headers)),
        ("Excel row limit", EXCEL_MAX_ROWS),
        ("Overflow choice", "One CSV" if overflow == OVERFLOW_CSV
                            else "Excel parts of up to %d rows" % PART_ROWS),
        ("Output files", "\n".join(os.path.basename(p) for p in paths)),
    ]
    row_index = add_table(sheet, ("Item", "Value"), rows, 3)

    warnings = []
    if split_files:
        warnings.append(
            "These files were too big to keep in one part on their own, so "
            "their rows are spread across parts: %s" % ", ".join(split_files))
    if any(plan.status == FILE_ERROR for plan in plans):
        warnings.append("Some files could not be read - see Skipped.")
    if any(plan.status == SKIPPED_NO_READER for plan in plans):
        warnings.append("Some files need an extra reader installed - see Skipped.")
    for plan in good:
        if plan.empty_columns:
            warnings.append("%s: %s came through empty (uncached formulas?)"
                            % (plan.label, ", ".join(plan.empty_columns)))
    warnings.extend(notes)
    warnings.append("This output may hold employee or client identifiers. Save "
                    "it only in the approved Global Insider folder; do not "
                    "e-mail or upload it, and keep it out of git.")
    row_index = add_block(sheet, row_index, "Notes and warnings")
    row_index = add_table(sheet, ("#", "Note"),
                          [(n + 1, text) for n, text in enumerate(warnings)],
                          row_index)
    sheet.column_dimensions["B"].width = 100
    autosize(sheet, limit=100)

    # ---- Sheet Details ----------------------------------------------------
    sheet = book.create_sheet("Sheet Details")
    page_furniture(sheet, description)
    sheet["A1"] = "EVERY SHEET LOOKED AT"
    style_row(sheet, 1, title=True)
    rows = []
    for plan in plans:
        rows.append((
            plan.file,
            plan.sheet,
            plan.status,
            plan.detail,
            plan.header_offset + 1 if plan.ok else "",
            len(plan.columns) if plan.ok else "",
            plan.rows_out if plan.ok else "",
            plan.rows_blank,
            ", ".join(plan.unknown),
            ", ".join(sorted(set(plan.duplicates))),
            ", ".join(plan.empty_columns),
        ))
    add_table(sheet, ("Source File", "Source Sheet", "Status", "Detail",
                      "Header Row", "Columns", "Rows Appended",
                      "Blank Rows Dropped", "Unnamed Columns",
                      "Duplicate Headings", "Columns That Came Through Empty"),
              rows, 3)
    autosize(sheet)

    # ---- Columns ----------------------------------------------------------
    sheet = book.create_sheet("Columns")
    page_furniture(sheet, description)
    sheet["A1"] = "COLUMNS IN THE COMBINED TABLE"
    style_row(sheet, 1, title=True)
    supplied = {}
    for plan in good:
        for _, _, key in plan.columns:
            supplied[key] = supplied.get(key, 0) + 1
    rows = []
    for spot, name in enumerate(headers, 1):
        if spot <= len(TRACKING_COLUMNS):
            rows.append((spot, name, "Added by this tool", ""))
            continue
        key = match_key(name)
        count = supplied.get(key, 0)
        rows.append((spot, name, count,
                     "In every appended sheet" if count == len(good)
                     else "Blank for the sheets that lack it"))
    add_table(sheet, ("Position", "Column", "Sheets Supplying It", "Note"),
              rows, 3)
    autosize(sheet)

    # ---- Skipped ----------------------------------------------------------
    sheet = book.create_sheet("Skipped")
    page_furniture(sheet, description)
    sheet["A1"] = "WHAT WAS LEFT OUT, AND WHY"
    style_row(sheet, 1, title=True)
    rows = [(plan.file, plan.sheet, plan.status, plan.detail)
            for plan in skipped]
    if not rows:
        rows = [("", "", "Nothing was skipped", "")]
    add_table(sheet, ("Source File", "Source Sheet", "Status", "Detail"),
              rows, 3)
    autosize(sheet, limit=90)

    path = os.path.join(out_folder, "%s summary.xlsx" % base)
    book.save(path)
    book.close()
    return path


# ---------------------------------------------------------------------------
# Remembered settings
# ---------------------------------------------------------------------------

def load_settings():
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return {}


def save_settings(data):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as handle:
            json.dump(data, handle, indent=2)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Window
# ---------------------------------------------------------------------------

class CombinerApp:

    def __init__(self, root, start_folder=""):
        self.root = root
        root.title("Excel and CSV combiner")
        root.geometry("1000x780")

        settings = load_settings()
        self.folder = tk.StringVar(value=start_folder or settings.get("folder", ""))
        self.initials = tk.StringVar(value=settings.get("initials", ""))
        self.subfolders = tk.BooleanVar(value=settings.get("subfolders", False))
        self.mode = tk.StringVar(value=settings.get("mode", HEADER_BY_NAME))
        self.overflow = tk.StringVar(value=settings.get("overflow", OVERFLOW_PARTS))
        self.status = tk.StringVar(
            value="Pick a folder, choose how the header is found, then Run.")
        self.busy = False
        self.last_output = ""

        pad = {"padx": 8, "pady": 4}
        top = ttk.Frame(root)
        top.pack(fill="x", **pad)

        ttk.Label(top, text="Folder of Excel / CSV files").grid(
            row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.folder, width=88).grid(
            row=0, column=1, sticky="we", padx=6)
        ttk.Button(top, text="Browse ...", command=self.pick_folder).grid(
            row=0, column=2)
        ttk.Checkbutton(top, text="Include subfolders",
                        variable=self.subfolders).grid(
            row=1, column=1, sticky="w", padx=6, pady=(6, 0))
        ttk.Label(top, text="Your 2-letter author code").grid(
            row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(top, textvariable=self.initials, width=8).grid(
            row=2, column=1, sticky="w", padx=6, pady=(6, 0))
        top.columnconfigure(1, weight=1)

        # ---- how the header is found --------------------------------------
        box = ttk.LabelFrame(root, text="How is the header row found?")
        box.pack(fill="x", padx=8, pady=(8, 2))
        ttk.Radiobutton(
            box, text="The first row of each sheet is the header",
            variable=self.mode, value=HEADER_FIRST_ROW,
            command=self.sync).pack(anchor="w", padx=8, pady=(4, 0))
        ttk.Radiobutton(
            box,
            text=("Find it by column name - the first row in the top %d that "
                  "holds ALL of these:" % HEADER_SCAN_ROWS),
            variable=self.mode, value=HEADER_BY_NAME,
            command=self.sync).pack(anchor="w", padx=8)
        ttk.Label(
            box,
            text=("One name per line, or comma separated. One name means that "
                  "one name identifies the header; several means every one of "
                  "them has to be in the row. Every other column of that row is "
                  "carried through as well."),
            wraplength=940, justify="left").pack(fill="x", padx=26, pady=(2, 2))
        self.spec = scrolledtext.ScrolledText(box, height=6, font=("Consolas", 10))
        self.spec.pack(fill="x", padx=26, pady=(0, 8))
        self.spec.insert("1.0", settings.get("spec", ""))

        # ---- what to do if it will not fit --------------------------------
        box = ttk.LabelFrame(
            root, text="If the combined table will not fit in one Excel sheet "
                       "(%d rows)" % EXCEL_MAX_ROWS)
        box.pack(fill="x", padx=8, pady=(6, 2))
        ttk.Radiobutton(
            box, text="Split into several Excel files - each source file kept "
                      "whole in one part",
            variable=self.overflow, value=OVERFLOW_PARTS).pack(
            anchor="w", padx=8, pady=(4, 0))
        ttk.Radiobutton(
            box, text="Put everything in one CSV instead (no row limit)",
            variable=self.overflow, value=OVERFLOW_CSV).pack(
            anchor="w", padx=8, pady=(0, 6))

        buttons = ttk.Frame(root)
        buttons.pack(fill="x", **pad)
        self.check_button = ttk.Button(buttons, text="Check headers only",
                                       command=self.on_check)
        self.check_button.pack(side="left")
        self.run_button = ttk.Button(buttons, text="Run and combine",
                                     command=self.on_run)
        self.run_button.pack(side="left", padx=8)
        ttk.Button(buttons, text="Open output folder",
                   command=self.open_output).pack(side="left")
        ttk.Button(buttons, text="Clear log",
                   command=lambda: self.log_box.delete("1.0", "end")).pack(
            side="right")

        self.progress = ttk.Progressbar(root, mode="determinate", maximum=100)
        self.progress.pack(fill="x", padx=8)

        self.log_box = scrolledtext.ScrolledText(root, height=18,
                                                 font=("Consolas", 9))
        self.log_box.pack(fill="both", expand=True, padx=8, pady=6)

        ttk.Label(root, textvariable=self.status, anchor="w").pack(
            fill="x", padx=8, pady=(0, 6))
        self.sync()

    # -- small helpers ------------------------------------------------------

    def sync(self):
        """The column box only matters in "by column name" mode."""
        state = "normal" if self.mode.get() == HEADER_BY_NAME else "disabled"
        self.spec.configure(state=state)

    def log(self, text=""):
        self.root.after(0, self._log, text)

    def _log(self, text):
        self.log_box.insert("end", text + "\n")
        self.log_box.see("end")

    def say(self, text):
        self.root.after(0, self.status.set, text)

    def step(self, done, total):
        value = 0 if not total else min(100, int(100 * done / total))
        self.root.after(0, self.progress.configure, {"value": value})

    def pick_folder(self):
        chosen = filedialog.askdirectory(
            title="Folder holding the Excel and CSV files",
            initialdir=self.folder.get() or os.path.expanduser("~"))
        if chosen:
            self.folder.set(os.path.normpath(chosen))

    def open_output(self):
        target = self.last_output or self.folder.get()
        if target and os.path.isdir(target):
            try:
                os.startfile(target)
            except Exception as error:
                messagebox.showerror("Cannot open folder", str(error))
        else:
            messagebox.showinfo("Nothing yet", "Run the combine first.")

    def inputs(self):
        folder = self.folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Folder",
                                 "Pick the folder holding the Excel and CSV files.")
            return None
        mode = self.mode.get()
        spec = parse_spec(self.spec.get("1.0", "end")) if mode == HEADER_BY_NAME else []
        if mode == HEADER_BY_NAME and not spec:
            messagebox.showerror(
                "Column names",
                "Type at least one column name to identify the header row, or "
                "switch to the first-row option.")
            return None
        initials = self.initials.get().strip().upper()
        if len(initials) != 2 or not initials.isalpha():
            messagebox.showerror(
                "Author code",
                "Please provide your 2-letter author code (your initials).")
            return None
        save_settings({
            "folder": folder,
            "initials": initials,
            "mode": mode,
            "overflow": self.overflow.get(),
            "subfolders": bool(self.subfolders.get()),
            "spec": self.spec.get("1.0", "end").strip(),
        })
        return folder, initials, mode, spec

    def lock(self, running):
        self.busy = running
        state = "disabled" if running else "normal"
        self.run_button.configure(state=state)
        self.check_button.configure(state=state)

    # -- actions ------------------------------------------------------------

    def on_check(self):
        got = self.inputs()
        if got and not self.busy:
            folder, _, mode, spec = got
            self.lock(True)
            threading.Thread(target=self._check, args=(folder, mode, spec),
                             daemon=True).start()

    def _check(self, folder, mode, spec):
        """A dry run: every sheet's fate and size, nothing written."""
        try:
            self.say("Checking headers ...")
            keys = [match_key(name) for name in spec]
            files = collect_files(folder, bool(self.subfolders.get()))
            self.log("=" * 78)
            self.log("HEADER CHECK - nothing is written")
            self.log("%d file(s) found" % len(files))
            appended = skipped = rows = 0
            plans = []
            for index, path in enumerate(files, 1):
                self.step(index, len(files))
                self.log("")
                self.log(os.path.relpath(path, folder))
                found, notes = scan_file(path, mode, keys, self.log)
                for note in notes:
                    self.log("      note: %s" % note)
                for plan in found:
                    plans.append(plan)
                    if plan.status in (FILE_ERROR, SKIPPED_NO_READER):
                        continue
                    self.log("  %s [%s] %s" % ("OK  " if plan.ok else "SKIP",
                                               plan.sheet, plan.detail))
                    if plan.ok:
                        appended += 1
                        rows += plan.data_rows
                        self.log("         columns: %s"
                                 % ", ".join(d for _, d, _ in plan.columns))
                    else:
                        skipped += 1
            headers, _ = build_headers(plans, spec, mode)
            self.log("")
            self.log("%d sheet(s) would be appended, %d skipped." % (appended, skipped))
            self.log("About %d data row(s) and %d column(s) in the result."
                     % (rows, len(headers)))
            if rows + 1 > EXCEL_MAX_ROWS:
                self.log("That is past Excel's %d-row limit, so the overflow "
                         "choice above will be used." % EXCEL_MAX_ROWS)
            else:
                self.log("That fits in one worksheet.")
            self.say("Header check finished - %d sheet(s) would be appended." % appended)
        except Exception as error:
            self.log("")
            self.log(traceback.format_exc())
            self.say("Header check failed - see the log.")
            messagebox.showerror("Header check", str(error))
        finally:
            self.root.after(0, self.lock, False)

    def on_run(self):
        got = self.inputs()
        if got and not self.busy:
            folder, initials, mode, spec = got
            self.lock(True)
            threading.Thread(
                target=self._run,
                args=(folder, initials, mode, spec, self.overflow.get(),
                      bool(self.subfolders.get())),
                daemon=True).start()

    def _run(self, folder, initials, mode, spec, overflow, subfolders):
        try:
            self.say("Combining ...")
            self.log("=" * 78)
            self.log("COMBINE - %s" % datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            out_folder, paths, summary_path, rows = combine(
                folder, initials, mode, spec, overflow, subfolders,
                self.log, self.step)
            self.last_output = out_folder
            self.say("Finished - %d row(s). Output in %s" % (rows, out_folder))
            messagebox.showinfo(
                "Finished",
                "%d data row(s) combined.\n\n%s\n%s\n\nSave the output in the "
                "approved Global Insider folder."
                % (rows,
                   "\n".join(os.path.basename(p) for p in paths),
                   os.path.basename(summary_path)))
        except Exception as error:
            self.log("")
            self.log(traceback.format_exc())
            self.say("Failed - see the log.")
            messagebox.showerror("Combine failed", str(error))
        finally:
            self.root.after(0, self.lock, False)


def main():
    start = ""
    if len(sys.argv) > 1 and os.path.isdir(sys.argv[1]):
        start = os.path.normpath(sys.argv[1])
    root = tk.Tk()
    CombinerApp(root, start)
    root.mainloop()


if __name__ == "__main__":
    main()
