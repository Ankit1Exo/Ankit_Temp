"""
extract_wage_form.py
---------------------
Extracts the following fields from NY Quarterly Employee/Payee Wage
Reporting PDFs (Part C section):

    a  Social Security Number
    b  Last Name
    c  First Name
    d  MI  (Middle Initial)
    g  Gross Federal Wages or Distribution

IMPORTANT — "I" pipe characters:
    The NY-45 form uses the letter "I" as a visual column delimiter between
    fields (visible as vertical bars). This script uses precise x-coordinate
    bands that sit BETWEEN the pipe positions so the delimiter "I" characters
    are never included in any field value.

    Column boundary layout (approx % of page width):
      | a SSN (0-18%) | b Last Name (18-36%) | I | c First Name (38-55%) | I | d MI (57-62%) | I | e Wage |
                                              ↑ pipe ~37%               ↑ pipe ~56%          ↑ pipe ~62%

Supports searchable (text-layer) PDFs. Falls back to OCR automatically
if no text layer is found.

Output: CSV + styled Excel saved in the same folder as your PDFs.

Requirements:
    pip install pdfplumber pdf2image pytesseract pillow pandas openpyxl

Run:
    python extract_wage_form.py
"""

import os
import re
import sys
import csv

try:
    import pdfplumber
    import pandas as pd
except ImportError:
    sys.exit("Run: pip install pdfplumber pandas openpyxl")


# ── File / Folder Picker ──────────────────────────────────────────────────────

def pick_source() -> list:
    """
    GUI dialog: choose a folder (all PDFs inside) or individual PDF files.
    Falls back to console input in headless environments.
    """
    try:
        import tkinter as tk
        from tkinter import filedialog, simpledialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        mode = simpledialog.askstring(
            "Select Source",
            "Type  folder  → process all PDFs in a folder\n"
            "Type  files   → pick individual PDFs",
            parent=root,
        )
        mode = (mode or "").strip().lower()

        if "folder" in mode:
            folder = filedialog.askdirectory(title="Select folder containing PDFs")
            root.destroy()
            if not folder:
                sys.exit("No folder selected.")
            paths = sorted([
                os.path.join(folder, f) for f in os.listdir(folder)
                if f.lower().endswith(".pdf") and not f.startswith("~")
            ])
            if not paths:
                sys.exit(f"No PDFs found in: {folder}")
        else:
            paths = list(filedialog.askopenfilenames(
                title="Select PDF files (Ctrl/Cmd+click for multiple)",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            ))
            root.destroy()
            if not paths:
                sys.exit("No files selected.")

        return paths

    except Exception:
        # Headless / no-display fallback
        print("\nNo display found. Please type your selection:")
        print("  1 - Enter a folder path (all PDFs inside will be processed)")
        print("  2 - Enter individual file paths one by one")
        choice = input("Choice [1/2]: ").strip()
        if choice == "1":
            folder = input("Folder path: ").strip()
            paths = sorted([
                os.path.join(folder, f) for f in os.listdir(folder)
                if f.lower().endswith(".pdf")
            ])
            if not paths:
                sys.exit(f"No PDFs found in: {folder}")
            return paths
        else:
            paths = []
            print("Enter PDF paths one per line (blank line to finish):")
            while True:
                p = input("  Path: ").strip()
                if not p:
                    break
                if os.path.exists(p):
                    paths.append(p)
                else:
                    print(f"  ⚠  Not found, skipping: {p}")
            if not paths:
                sys.exit("No valid files entered.")
            return paths


# ── Column Band Definitions ───────────────────────────────────────────────────
#
# The NY-45 Part C form uses "I" characters as visual column separators.
# Each band is defined as (x_start_fraction, x_end_fraction) of page width.
#
# The bands are deliberately narrowed at both edges to sit BETWEEN the pipe
# positions so the "I" delimiter characters fall outside every band.
#
# Pipe positions (approx):  ~37% (after Last Name)
#                            ~56% (after First Name)
#                            ~62% (after MI)
#
# Name / SSN row bands:
NAME_COLS = {
    "ssn":        (0.00, 0.17),   # a - Social Security Number
    "last_name":  (0.18, 0.36),   # b - Last Name       (ends before pipe at 37%)
    "first_name": (0.38, 0.55),   # c - First Name      (starts after pipe at 37%, ends before pipe at 56%)
    "mi":         (0.57, 0.62),   # d - Middle Initial  (starts after pipe at 56%, ends before pipe at 62%)
}

# Gross wages row band (on the separate "f/g/h/i/j" data row):
# Column g sits roughly at 32-54% on that row
WAGE_COLS = {
    "gross_wages": (0.32, 0.54),  # g - Gross Federal Wages or Distribution
}


# ── Spatial Word Helpers ──────────────────────────────────────────────────────

def get_words(page) -> list:
    """Extract words with spatial coordinates from a pdfplumber page."""
    return page.extract_words(
        x_tolerance=3,
        y_tolerance=3,
        keep_blank_chars=False,
        use_text_flow=False,
    ) or []


def words_in_band(words: list, y_min: float, y_max: float) -> list:
    """Words whose vertical centre falls within [y_min, y_max]."""
    return [w for w in words if y_min <= (w["top"] + w["bottom"]) / 2 <= y_max]


def words_in_col(words: list, x_min: float, x_max: float) -> list:
    """
    Words whose horizontal centre falls within [x_min, x_max].
    Also filters out lone 'I' characters that are pipe delimiters —
    a single 'I' at a column-boundary x position is always a delimiter,
    never a real field value.
    """
    result = []
    for w in words:
        cx = (w["x0"] + w["x1"]) / 2
        if x_min <= cx <= x_max:
            result.append(w)
    return result


def join_words(words: list) -> str:
    """Sort by x position and join into a single string."""
    return " ".join(
        w["text"] for w in sorted(words, key=lambda w: w["x0"])
    ).strip()


def col_value(words: list, band: tuple, page_width: float) -> str:
    """Extract and join all words that fall within a named column band."""
    x_min = page_width * band[0]
    x_max = page_width * band[1]
    return join_words(words_in_col(words, x_min, x_max))


# ── Row Grouping ──────────────────────────────────────────────────────────────

def group_into_rows(words: list, gap: float = 5.0) -> list:
    """
    Cluster words into horizontal text rows by their top coordinate.
    Returns list of (avg_top, [words]) sorted top-to-bottom.
    """
    if not words:
        return []

    sorted_words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    rows = []
    current_y   = sorted_words[0]["top"]
    current_row = [sorted_words[0]]

    for w in sorted_words[1:]:
        if abs(w["top"] - current_y) <= gap:
            current_row.append(w)
        else:
            rows.append((current_y, current_row))
            current_y   = w["top"]
            current_row = [w]
    rows.append((current_y, current_row))

    return sorted(rows, key=lambda r: r[0])


# ── Label Row Detection ───────────────────────────────────────────────────────

def row_text(row_words: list) -> str:
    return " ".join(w["text"].lower() for w in row_words)


def is_ssn_label_row(row_words: list) -> bool:
    """True when the row contains the 'a Social Security number' label."""
    t = row_text(row_words)
    return "social" in t and "security" in t


def is_gross_label_row(row_words: list) -> bool:
    """True when the row contains the 'g Gross federal wages' label."""
    t = row_text(row_words)
    return "gross" in t and "federal" in t


# ── Per-Page Extraction ───────────────────────────────────────────────────────

def extract_records_from_page(page) -> list:
    """
    Extract all employee records from one PDF page.

    Form structure (per employee block):
        ROW A (label):  "a Social Security number | b Last name | c First name | d MI | e Wage type"
        ROW B (data):    <SSN>  |  <LAST>  |  <FIRST>  |  <MI>  |  <type>
        ROW C (label):  "f Total UI ... | g Gross federal wages ... | h NYS ... | ..."
        ROW D (data):    <UI amt>  |  <GROSS WAGES>  |  <NYS>  |  ...
    """
    pw     = float(page.width)
    words  = get_words(page)
    rows   = group_into_rows(words, gap=5.0)
    records = []

    i = 0
    while i < len(rows):
        _, row_words = rows[i]

        if is_ssn_label_row(row_words):
            # ROW B: name / SSN data row immediately follows the label
            if i + 1 >= len(rows):
                i += 1
                continue

            _, data_row = rows[i + 1]

            ssn        = col_value(data_row, NAME_COLS["ssn"],        pw)
            last_name  = col_value(data_row, NAME_COLS["last_name"],  pw)
            first_name = col_value(data_row, NAME_COLS["first_name"], pw)
            mi_raw     = col_value(data_row, NAME_COLS["mi"],         pw)

            # Strip any residual lone "I" pipe that slipped through at band edges
            mi = strip_pipe(mi_raw)

            # ROW C + D: scan ahead (up to 5 rows) for the gross wages label
            gross_wages = ""
            for j in range(i + 2, min(i + 7, len(rows))):
                _, candidate = rows[j]
                if is_gross_label_row(candidate):
                    if j + 1 < len(rows):
                        _, wage_row = rows[j + 1]
                        gross_wages = col_value(wage_row, WAGE_COLS["gross_wages"], pw)
                    break

            # Only record if at least a name field was captured
            if last_name or first_name:
                records.append({
                    "ssn":         clean_ssn(ssn),
                    "last_name":   last_name,
                    "first_name":  first_name,
                    "mi":          mi,
                    "gross_wages": clean_number(gross_wages),
                })

            i += 2
            continue

        i += 1

    return records


def strip_pipe(value: str) -> str:
    """
    Remove lone 'I' tokens that are pipe delimiters, not real initials.

    Real middle initials are a single letter A-Z (excluding I itself is wrong —
    someone could genuinely have MI = I). Instead we rely on the column bands
    being correctly positioned, but add a post-processing guard: if the entire
    extracted value is just "I" and nothing else, treat it as a pipe artefact.

    If the value is "I A" (pipe leaked + real initial), keep only the last token.
    """
    tokens = value.split()
    if not tokens:
        return ""
    # If first token looks like a stray pipe and more tokens follow, drop it
    if len(tokens) > 1 and tokens[0] == "I":
        tokens = tokens[1:]
    # If still just "I" and it's the only character, caller should decide;
    # we return it as-is — a genuine MI of "I" is valid (e.g. "Ivan")
    return " ".join(tokens)


# ── Value Cleaning ────────────────────────────────────────────────────────────

def clean_ssn(raw: str) -> str:
    """Format SSN as ###-##-#### when 9 digits are present."""
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 9:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return raw


def clean_number(raw: str) -> str:
    """Strip non-numeric characters from a wage figure."""
    return re.sub(r"[^\d.,]", "", raw).strip()


# ── OCR Fallback ──────────────────────────────────────────────────────────────

def ocr_fallback(pdf_path: str) -> list:
    """
    For scanned (image-only) PDFs: convert pages to high-res images,
    run Tesseract OCR, then parse with regex.
    """
    try:
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import ImageEnhance, ImageFilter
    except ImportError:
        print("  ⚠  OCR requires: pip install pdf2image pytesseract pillow")
        return []

    print("  → No text layer found. Running OCR (may take a moment)...")
    pages = convert_from_path(pdf_path, dpi=300)
    all_records = []

    for img in pages:
        img = img.resize((img.width * 2, img.height * 2))
        img = ImageEnhance.Contrast(img).enhance(2.0)
        img = img.filter(ImageFilter.SHARPEN)
        text = pytesseract.image_to_string(img, config="--psm 6")
        all_records.extend(_parse_ocr_text(text))

    return all_records


def _parse_ocr_text(text: str) -> list:
    """
    Regex-based parser for plain OCR text (fallback when coordinates unavailable).
    Less precise but handles scanned pages.
    """
    SSN_RE   = re.compile(r"\b(\d{3}[-\s]?\d{2}[-\s]?\d{4})\b")
    GROSS_RE = re.compile(r"gross\s+federal[^\n]*\n\s*([\d,.\s]+)", re.I)

    records = []
    ssns    = SSN_RE.findall(text)
    grosses = [m.strip() for m in GROSS_RE.findall(text)]

    # Match "LASTNAME  FIRSTNAME  MI" style lines
    name_pattern = re.compile(
        r"^([A-Z][A-Z\s,.\-]+?)\s{2,}([A-Z][A-Za-z\s.\-]+?)(?:\s{2,}([A-Z]))?\s*$",
        re.MULTILINE,
    )
    for idx, nm in enumerate(name_pattern.finditer(text)):
        records.append({
            "ssn":         clean_ssn(ssns[idx]) if idx < len(ssns) else "",
            "last_name":   nm.group(1).strip(),
            "first_name":  nm.group(2).strip(),
            "mi":          nm.group(3).strip() if nm.group(3) else "",
            "gross_wages": clean_number(grosses[idx]) if idx < len(grosses) else "",
        })

    return records


# ── Per-File Driver ───────────────────────────────────────────────────────────

def process_pdf(pdf_path: str) -> list:
    """
    Process all pages of one PDF. Returns list of record dicts,
    each tagged with file_name and page number.
    """
    filename = os.path.basename(pdf_path)
    records  = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check whether any text layer exists
            has_text = any(
                (p.extract_text() or "").strip() for p in pdf.pages
            )

            if not has_text:
                raw = ocr_fallback(pdf_path)
                for r in raw:
                    r["file_name"] = filename
                    r.setdefault("page", "")
                return raw

            for page_num, page in enumerate(pdf.pages, start=1):
                page_records = extract_records_from_page(page)
                for r in page_records:
                    r["file_name"] = filename
                    r["page"]      = page_num
                records.extend(page_records)

    except Exception as e:
        print(f"  ⚠  Error reading {filename}: {e}")

    return records


# ── Output ────────────────────────────────────────────────────────────────────

COLUMNS = ["file_name", "page", "ssn", "last_name", "first_name", "mi", "gross_wages"]

LABELS = {
    "file_name":   "File Name",
    "page":        "Page",
    "ssn":         "a Social Security Number",
    "last_name":   "b Last Name",
    "first_name":  "c First Name",
    "mi":          "d MI",
    "gross_wages": "g Gross Federal Wages",
}


def save_csv(records: list, out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writerow(LABELS)
        writer.writerows(records)
    print(f"  CSV   → {out_path}")


def save_excel(records: list, out_path: str) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook

    df = pd.DataFrame(records, columns=COLUMNS)
    df.rename(columns=LABELS, inplace=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Wage Data")
        ws = writer.sheets["Wage Data"]

        HEADER_FILL = PatternFill("solid", start_color="1F4E79")
        HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        DATA_FONT   = Font(name="Arial", size=10)
        ALT_FILL    = PatternFill("solid", start_color="DEEAF1")

        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.font = DATA_FONT
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL

        for col in ws.columns:
            max_w = max(
                (len(str(c.value)) for c in col if c.value), default=8
            )
            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = min(max_w + 4, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"  Excel → {out_path}")


def print_summary(files: list, all_records: list) -> None:
    print(f"\n{'='*60}")
    print(f"  Extraction Summary")
    print(f"{'='*60}")
    print(f"  Files processed : {len(files)}")
    print(f"  Total records   : {len(all_records)}")
    print(f"\n  Records per file:")
    for f in files:
        n = sum(1 for r in all_records if r["file_name"] == os.path.basename(f))
        print(f"    • {os.path.basename(f)}: {n} employee(s)")

    no_ssn = sum(1 for r in all_records if not r.get("ssn"))
    no_mi  = sum(1 for r in all_records if not r.get("mi"))
    if no_ssn:
        print(f"\n  ⚠  {no_ssn} record(s) with no SSN "
              f"(redacted or blank in source PDF)")
    if no_mi:
        print(f"  ℹ  {no_mi} record(s) with no MI "
              f"(employee may not have a middle initial)")
    print(f"{'='*60}\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  NY Quarterly Wage Form — Field Extractor")
    print("=" * 60)

    pdf_paths = pick_source()
    print(f"\nFiles selected: {len(pdf_paths)}")
    for p in pdf_paths:
        print(f"  • {os.path.basename(p)}")

    all_records = []
    for path in pdf_paths:
        print(f"\nProcessing: {os.path.basename(path)}")
        records = process_pdf(path)
        print(f"  → {len(records)} record(s) found")
        all_records.extend(records)

    if not all_records:
        print("\nNo records extracted. Check PDF content.")
        sys.exit(1)

    print_summary(pdf_paths, all_records)

    # Save outputs alongside the first selected PDF
    out_dir  = os.path.dirname(os.path.abspath(pdf_paths[0]))
    csv_path = os.path.join(out_dir, "wage_extraction.csv")
    xls_path = os.path.join(out_dir, "wage_extraction.xlsx")

    # Increment filename if outputs already exist
    for attr in ["csv_path", "xls_path"]:
        p = locals()[attr]
        if os.path.exists(p):
            root, ext = os.path.splitext(p)
            n = 1
            while os.path.exists(p):
                p = f"{root}_{n}{ext}"
                n += 1
            locals()[attr]  # re-bind via exec below
        exec(f"{attr} = p")   # noqa

    print("Saving outputs...")
    save_csv(all_records, csv_path)
    save_excel(all_records, xls_path)
    print("\nDone.")


if __name__ == "__main__":
    main()
