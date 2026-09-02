"""
260902 AM excel csv combiner selftest.py

Proves "260902 AM excel csv combiner.py" on files this script builds itself.

Open in IDLE and press F5. Nothing is asked for and nothing outside a temporary
folder is touched; the folder is deleted at the end.

Every name, SSN and client number below is INVENTED. No real personal data is
used, produced or needed by this test.

What it plants, and therefore what it proves:

    * two junk rows ABOVE the header, so header hunting has to skip them
    * a DECOY sheet with no header at all, which must be skipped and reported
    * a DECOY file holding First Name but no SSN, which must also be skipped,
      proving that several column names all have to be present
    * a real PIVOT TABLE relationship on one sheet, which must never be combined
    * a HIDDEN sheet, which must never be combined
    * a blank heading over a column that has data, which must arrive as
      "Unknown Column D"
    * data in a column to the RIGHT of the last heading, which must arrive as
      "Unknown Column E"
    * the same columns in a different ORDER, in different CASE and with double
      spaces in another file, which must all land in the same output columns
    * columns no other file has, which must be added at the END
    * .xlsx, .csv and .tsv in one run
    * number formats - a zero-padded ID, an SSN mask, a date, a percentage,
      a thousands separator - which must come through as they are DISPLAYED
    * a completely blank row, which must be dropped
    * dates in five different formats, every one of which must be copied
      exactly as the cell shows it, including Excel's locale-linked
      "*3/14/2012" format whose stored code disagrees with the screen
    * an overflow that does not fit one worksheet, written as parts with no
      source file split across two parts, and the same overflow written as one
      CSV instead
"""

import csv
import importlib.util
import os
import shutil
import sys
import tempfile
import traceback
import zipfile
from datetime import datetime

from openpyxl import Workbook, load_workbook

TOOL = "260902 AM excel csv combiner.py"

FAILURES = []
CHECKS = [0]


def load_tool():
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, TOOL)
    if not os.path.isfile(path):
        raise SystemExit("Cannot find %s next to this script." % TOOL)
    spec = importlib.util.spec_from_file_location("combiner_under_test", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def check(label, got, want):
    CHECKS[0] += 1
    if got == want:
        print("  ok    %s" % label)
        return True
    print("  FAIL  %s" % label)
    print("          got  %r" % (got,))
    print("          want %r" % (want,))
    FAILURES.append(label)
    return False


def check_true(label, got):
    return check(label, bool(got), True)


# ---------------------------------------------------------------------------
# Building the test files
# ---------------------------------------------------------------------------

PIVOT_REL = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    '<Relationship Id="rId1" '
    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
    'Target="../pivotTables/pivotTable1.xml"/>'
    '</Relationships>'
)


def inject_pivot(path, sheet_name):
    """Give one sheet of a saved workbook a genuine pivotTable relationship,
    which is the only thing that marks a sheet as a pivot sheet in the file
    format. Rewrites the package in place."""
    from xml.etree import ElementTree
    main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    doc = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

    with zipfile.ZipFile(path) as package:
        entries = [(item, package.read(item.filename))
                   for item in package.infolist()]
        book = ElementTree.fromstring(dict((n, b) for n, b in
                                           ((i.filename, d) for i, d in entries)
                                           )["xl/workbook.xml"])
        rels = ElementTree.fromstring(
            dict((i.filename, d) for i, d in entries)["xl/_rels/workbook.xml.rels"])

    targets = {node.get("Id"): node.get("Target") for node in rels}
    part = None
    for node in book.iter(main + "sheet"):
        if node.get("name") == sheet_name:
            part = targets.get(node.get(doc + "id"))
    if not part:
        raise RuntimeError("sheet %r not found in %s" % (sheet_name, path))
    filename = part.replace("\\", "/").rsplit("/", 1)[-1]
    rel_path = "xl/worksheets/_rels/%s.rels" % filename

    temporary = path + ".tmp"
    with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as out:
        for item, data in entries:
            if item.filename == rel_path:
                continue
            out.writestr(item, data)
        out.writestr(rel_path, PIVOT_REL)
        out.writestr("xl/pivotTables/pivotTable1.xml",
                     '<?xml version="1.0"?><pivotTableDefinition '
                     'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                     'name="TestPivot" dataOnRows="0"/>')
    os.replace(temporary, path)


def put(sheet, row, column, value, number_format=None):
    cell = sheet.cell(row=row, column=column, value=value)
    if number_format:
        cell.number_format = number_format
    return cell


def build_alpha(folder):
    """Junk rows above the header, a blank heading with data under it, a blank
    row, a decoy sheet, a hidden sheet and a pivot sheet."""
    path = os.path.join(folder, "260902 AM test alpha.xlsx")
    book = Workbook()

    sheet = book.active
    sheet.title = "Data"
    put(sheet, 1, 1, "Fabricated payroll extract - test data only")
    # row 2 deliberately left blank
    for column, name in enumerate(
            ["First Name", "SSN", "Client ID", None, "Amount"], 1):
        if name:
            put(sheet, 3, column, name)
    put(sheet, 4, 1, "Ada")
    put(sheet, 4, 2, 123456789, "000-00-0000")
    put(sheet, 4, 3, 7123, "000000")
    put(sheet, 4, 4, "note x")
    put(sheet, 4, 5, 1.5, "0.00")
    put(sheet, 5, 1, "Grace")
    put(sheet, 5, 2, 987654321, "000-00-0000")
    put(sheet, 5, 3, 42, "000000")
    put(sheet, 5, 4, "note y")
    put(sheet, 5, 5, 1234567.891, "#,##0.00")
    # row 6 completely blank - must be dropped
    put(sheet, 7, 1, "Alan")
    put(sheet, 7, 2, 111223333, "000-00-0000")
    put(sheet, 7, 3, 5, "000000")
    put(sheet, 7, 5, 0.45, "0%")

    decoy = book.create_sheet("Notes")
    put(decoy, 1, 1, "Handover notes")
    put(decoy, 2, 1, "no header row anywhere on this sheet")

    pivot = book.create_sheet("PivotTable1")
    put(pivot, 1, 1, "First Name")
    put(pivot, 1, 2, "SSN")
    put(pivot, 2, 1, "Grand Total")
    put(pivot, 2, 2, 3)

    hidden = book.create_sheet("HiddenStuff")
    put(hidden, 1, 1, "First Name")
    put(hidden, 1, 2, "SSN")
    put(hidden, 2, 1, "Should Not Appear")
    put(hidden, 2, 2, 999887777)
    hidden.sheet_state = "hidden"

    book.save(path)
    book.close()
    inject_pivot(path, "PivotTable1")
    return path


def build_bravo(folder):
    """Same columns, different order, different case, a double space, plus a
    column no other file has."""
    path = os.path.join(folder, "260902 AM test bravo.xlsx")
    book = Workbook()
    sheet = book.active
    for column, name in enumerate(["SSN", "first  name", "City", "Client ID"], 1):
        put(sheet, 1, column, name)
    put(sheet, 2, 1, "555-66-7777")          # already text - must not change
    put(sheet, 2, 2, "Linus")
    put(sheet, 2, 3, "Helsinki")
    put(sheet, 2, 4, 9, "000000")
    put(sheet, 3, 1, "444-55-6666")
    put(sheet, 3, 2, "Radia")
    put(sheet, 3, 3, "Stockholm")
    put(sheet, 3, 4, 10, "000000")
    book.save(path)
    book.close()
    return path


def build_charlie(folder):
    """A CSV, including a leading-zero code and a field with a comma in it."""
    path = os.path.join(folder, "260902 AM test charlie.csv")
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["First Name", "SSN", "Client ID", "Amount"])
        writer.writerow(["Katherine", "222-33-4444", "007123", "12.00"])
        writer.writerow(["Margaret", "333-44-5555", "000045", "Smith, J. credit"])
    return path


def build_delta(folder):
    """A tab separated file with a column no other file has."""
    path = os.path.join(folder, "260902 AM test delta.tsv")
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["First Name", "SSN", "Region"])
        writer.writerow(["Barbara", "666-77-8888", "North"])
        writer.writerow(["Annie", "777-88-9999", "South"])
    return path


def build_echo(folder):
    """A decoy: First Name is there but SSN is not, so every column asked for
    is NOT present and the sheet must be skipped."""
    path = os.path.join(folder, "260902 AM test echo.xlsx")
    book = Workbook()
    sheet = book.active
    put(sheet, 1, 1, "First Name")
    put(sheet, 1, 2, "Client ID")
    put(sheet, 2, 1, "Should Not Appear")
    put(sheet, 2, 2, 1)
    book.save(path)
    book.close()
    return path


def build_foxtrot(folder):
    """A real date format, and data in a column to the right of the last
    heading."""
    path = os.path.join(folder, "260902 AM test foxtrot.xlsx")
    book = Workbook()
    sheet = book.active
    for column, name in enumerate(["First Name", "SSN", "Hire Date"], 1):
        put(sheet, 1, column, name)
    put(sheet, 2, 1, "Joan")
    put(sheet, 2, 2, 888990000, "000-00-0000")
    put(sheet, 2, 3, datetime(2026, 3, 1), "dd/mm/yyyy")
    put(sheet, 2, 5, "orphan value")          # column E, past the last heading
    put(sheet, 3, 1, "Shafi")
    put(sheet, 3, 2, 999001111, "000-00-0000")
    put(sheet, 3, 3, datetime(2025, 12, 25), "dd/mm/yyyy")
    put(sheet, 3, 5, "another orphan")
    book.save(path)
    book.close()
    return path


def build_all(folder):
    return [build_alpha(folder), build_bravo(folder), build_charlie(folder),
            build_delta(folder), build_echo(folder), build_foxtrot(folder)]


# ---------------------------------------------------------------------------
# Reading the result back
# ---------------------------------------------------------------------------

def read_xlsx(path):
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = [["" if v is None else v for v in row]
                for row in book.worksheets[0].iter_rows(values_only=True)]
    finally:
        book.close()
    return rows


def read_csv_file(path):
    with open(path, "r", newline="", encoding="utf-8-sig") as handle:
        return [row for row in csv.reader(handle)]


def as_records(rows):
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def find(records, name, key="First Name"):
    for record in records:
        if record.get(key) == name:
            return record
    return {}


# ---------------------------------------------------------------------------
# The scenarios
# ---------------------------------------------------------------------------

def scenario_by_name(mod, folder):
    print("")
    print("-" * 72)
    print("HEADER FOUND BY COLUMN NAME - First Name + SSN")
    print("-" * 72)
    spec = ["First Name", "SSN"]
    out_folder, paths, summary, written = mod.combine(
        folder, "AM", mod.HEADER_BY_NAME, spec, mod.OVERFLOW_PARTS, False,
        lambda *a: None, lambda *a: None)

    check("one workbook written, no parts needed", len(paths), 1)
    check("output is a single .xlsx", os.path.basename(paths[0]).endswith("data.xlsx"), True)

    rows = read_xlsx(paths[0])
    headers = rows[0]
    records = as_records(rows)

    check("column order - typed columns first, new ones at the end",
          headers,
          ["Source File", "Source Sheet", "First Name", "SSN", "Client ID",
           "Unknown Column D", "Amount", "City", "Region", "Hire Date",
           "Unknown Column E"])

    check("data rows written", written, 11)
    check("blank row was dropped", len(records), 11)

    # --- the sheets that must NOT be there -------------------------------
    sheets = {(r["Source File"], r["Source Sheet"]) for r in records}
    check_true("pivot table sheet excluded",
               not any(s == "PivotTable1" for _, s in sheets))
    check_true("hidden sheet excluded",
               not any(s == "HiddenStuff" for _, s in sheets))
    check_true("decoy sheet with no header excluded",
               not any(s == "Notes" for _, s in sheets))
    check_true("decoy file missing SSN excluded",
               not any("echo" in f for f, _ in sheets))
    check_true("nothing from the hidden sheet leaked",
               not find(records, "Should Not Appear"))

    # --- junk rows above the header ---------------------------------------
    check_true("junk row above the header was not treated as data",
               not find(records, "Fabricated payroll extract - test data only"))

    # --- values exactly as displayed --------------------------------------
    ada = find(records, "Ada")
    check("SSN shown through its 000-00-0000 mask", ada.get("SSN"), "123-45-6789")
    check("client ID keeps its leading zeros", ada.get("Client ID"), "007123")
    check("1.5 under 0.00 stays 1.50", ada.get("Amount"), "1.50")
    check("blank heading became Unknown Column D", ada.get("Unknown Column D"), "note x")
    check("source file recorded", ada.get("Source File"),
          "260902 AM test alpha.xlsx")
    check("source sheet recorded", ada.get("Source Sheet"), "Data")

    grace = find(records, "Grace")
    check("thousands separator kept", grace.get("Amount"), "1,234,567.89")
    alan = find(records, "Alan")
    check("percentage kept as a percentage", alan.get("Amount"), "45%")
    check("empty cell stays empty", alan.get("Unknown Column D"), "")

    # --- matching across files --------------------------------------------
    linus = find(records, "Linus")
    check("lower case and double spaced heading matched First Name",
          linus.get("First Name"), "Linus")
    check("column in a different order landed in the right column",
          linus.get("SSN"), "555-66-7777")
    check("a column no other file had went to the end", linus.get("City"),
          "Helsinki")
    check("that file's own Client ID format applied", linus.get("Client ID"),
          "000009")
    check("columns this file lacks are left blank", linus.get("Amount"), "")

    # --- csv and tsv -------------------------------------------------------
    katherine = find(records, "Katherine")
    check("CSV leading zeros survive", katherine.get("Client ID"), "007123")
    margaret = find(records, "Margaret")
    check("CSV field containing a comma survives",
          margaret.get("Amount"), "Smith, J. credit")
    barbara = find(records, "Barbara")
    check("TSV read with the right delimiter", barbara.get("SSN"), "666-77-8888")
    check("TSV's own extra column present", barbara.get("Region"), "North")

    # --- dates and columns past the header --------------------------------
    joan = find(records, "Joan")
    check("date shown through its dd/mm/yyyy format",
          joan.get("Hire Date"), "01/03/2026")
    check("data to the right of the last heading became Unknown Column E",
          joan.get("Unknown Column E"), "orphan value")

    # --- the summary -------------------------------------------------------
    book = load_workbook(summary)
    try:
        titles = book.sheetnames
        skipped = [[c.value for c in row] for row in book["Skipped"].iter_rows()]
    finally:
        book.close()
    check("summary has all four tabs", titles,
          ["Summary", "Sheet Details", "Columns", "Skipped"])
    text = "\n".join(str(v) for row in skipped for v in row if v)
    check_true("summary names the skipped pivot sheet", "PivotTable1" in text)
    check_true("summary names the skipped hidden sheet", "HiddenStuff" in text)
    check_true("summary names the skipped decoy sheet", "Notes" in text)
    check_true("summary names the skipped decoy file", "echo" in text)
    return len(records)


def scenario_first_row(mod, folder):
    print("")
    print("-" * 72)
    print("HEADER IS THE FIRST ROW")
    print("-" * 72)
    _, paths, _, written = mod.combine(
        folder, "AM", mod.HEADER_FIRST_ROW, [], mod.OVERFLOW_PARTS, False,
        lambda *a: None, lambda *a: None)
    rows = read_xlsx(paths[0])
    records = as_records(rows)
    sheets = {(r["Source File"], r["Source Sheet"]) for r in records}

    check_true("the decoy sheet IS combined in first-row mode",
               any(s == "Notes" for _, s in sheets))
    check_true("the decoy file IS combined in first-row mode",
               any("echo" in f for f, _ in sheets))
    check_true("pivot sheet is still excluded",
               not any(s == "PivotTable1" for _, s in sheets))
    check_true("hidden sheet is still excluded",
               not any(s == "HiddenStuff" for _, s in sheets))
    check_true("first-row mode takes in more than by-name mode", written > 11)
    check("alpha's row 1 became its header - so its own header row is data",
          rows[0][2], "Fabricated payroll extract - test data only")


def scenario_overflow_parts(mod, folder):
    print("")
    print("-" * 72)
    print("TOO BIG FOR ONE WORKSHEET - SPLIT INTO PARTS")
    print("-" * 72)
    keep_max, keep_part = mod.EXCEL_MAX_ROWS, mod.PART_ROWS
    mod.EXCEL_MAX_ROWS, mod.PART_ROWS = 8, 4
    try:
        _, paths, _, written = mod.combine(
            folder, "AM", mod.HEADER_BY_NAME, ["First Name", "SSN"],
            mod.OVERFLOW_PARTS, False, lambda *a: None, lambda *a: None)
    finally:
        mod.EXCEL_MAX_ROWS, mod.PART_ROWS = keep_max, keep_part

    check_true("more than one part written", len(paths) > 1)
    check_true("every part is named 'part NN'",
               all("data part" in os.path.basename(p) for p in paths))

    total = 0
    where = {}
    for number, path in enumerate(paths, 1):
        rows = read_xlsx(path)
        check("part %d repeats the header row" % number, rows[0][2], "First Name")
        for record in as_records(rows):
            total += 1
            where.setdefault(record["Source File"], set()).add(number)
        check_true("part %d holds no more than %d rows" % (number, 4),
                   len(rows) - 1 <= 4)
    check("no row lost across the parts", total, written)

    split = sorted(name for name, parts in where.items() if len(parts) > 1)
    check("no source file was split across two parts", split, [])


def scenario_overflow_csv(mod, folder):
    print("")
    print("-" * 72)
    print("TOO BIG FOR ONE WORKSHEET - ONE CSV INSTEAD")
    print("-" * 72)
    keep_max, keep_part = mod.EXCEL_MAX_ROWS, mod.PART_ROWS
    mod.EXCEL_MAX_ROWS, mod.PART_ROWS = 8, 4
    try:
        _, paths, _, written = mod.combine(
            folder, "AM", mod.HEADER_BY_NAME, ["First Name", "SSN"],
            mod.OVERFLOW_CSV, False, lambda *a: None, lambda *a: None)
    finally:
        mod.EXCEL_MAX_ROWS, mod.PART_ROWS = keep_max, keep_part

    check("exactly one file written", len(paths), 1)
    check("it is a CSV", paths[0].lower().endswith(".csv"), True)
    rows = read_csv_file(paths[0])
    check("the CSV holds every row past the Excel limit", len(rows) - 1, written)
    records = as_records(rows)
    ada = find(records, "Ada")
    check("the CSV keeps leading zeros as text", ada.get("Client ID"), "007123")
    check("the CSV keeps the SSN mask", ada.get("SSN"), "123-45-6789")


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

def scenario_date_formats(mod, parent):
    """Dates are copied exactly as the cell shows them - no year is rewritten.

    The exception that has to be handled is Excel's LOCALE-LINKED format, the
    one Format Cells marks "*3/14/2012". The file stores it as the code
    mm-dd-yy, but Excel draws the cell with the machine's short date, so it
    reads 7/7/2001 on screen. Copying the stored code would shorten the year to
    07-07-01, which is the one case where trusting the code is wrong."""
    print("")
    print("-" * 72)
    print("DATE FORMATS ARE COPIED AS SHOWN")
    print("-" * 72)
    folder = os.path.join(parent, "dates")
    os.makedirs(folder, exist_ok=True)

    path = os.path.join(folder, "260902 AM test dates.xlsx")
    book = Workbook()
    sheet = book.active
    for column, name in enumerate(
            ["First Name", "SSN", "Two Digit", "Four Digit", "Month Name",
             "Locale Linked"], 1):
        put(sheet, 1, column, name)
    put(sheet, 2, 1, "Ada")
    put(sheet, 2, 2, 111223333, "000-00-0000")
    put(sheet, 2, 3, datetime(1999, 1, 1), "dd/mm/yy")
    put(sheet, 2, 4, datetime(1999, 1, 1), "dd/mm/yyyy")
    put(sheet, 2, 5, datetime(2001, 12, 31), "d-mmm-yy")
    put(sheet, 2, 6, datetime(2001, 7, 7), "mm-dd-yy")     # Excel builtin 14
    book.save(path)
    book.close()

    csv_path = os.path.join(folder, "260902 AM test dates.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["First Name", "SSN", "Two Digit", "Four Digit"])
        writer.writerow(["Katherine", "333-44-5555", "7/4/76", "01/01/1999"])

    _, paths, _, _ = mod.combine(
        folder, "AM", mod.HEADER_BY_NAME, ["First Name", "SSN"],
        mod.OVERFLOW_PARTS, False, lambda *a: None, lambda *a: None)
    records = as_records(read_xlsx(paths[0]))

    ada = find(records, "Ada")
    check("a dd/mm/yy cell keeps its two-digit year, untouched",
          ada.get("Two Digit"), "01/01/99")
    check("a dd/mm/yyyy cell keeps its four-digit year, untouched",
          ada.get("Four Digit"), "01/01/1999")
    check("a d-mmm-yy cell is left exactly as it reads",
          ada.get("Month Name"), "31-Dec-01")

    # The locale-linked format is the one case the stored code lies about.
    shown = ada.get("Locale Linked")
    short_date = mod.LOCALE_LINKED_FORMATS.get("mm-dd-yy")
    check_true("locale-linked date is not the raw stored mm-dd-yy code",
               shown != "07-07-01")
    check_true("locale-linked date shows the year Excel shows (%r)" % short_date,
               "2001" in shown)
    if short_date == "m/d/yyyy":
        check("locale-linked date matches the US short date on screen",
              shown, "7/7/2001")

    katherine = find(records, "Katherine")
    check("a short text date in a CSV is NOT expanded",
          katherine.get("Two Digit"), "7/4/76")
    check("a full text date in a CSV is copied byte for byte",
          katherine.get("Four Digit"), "01/01/1999")


def main():
    print("=" * 72)
    print("SELF TEST - %s" % TOOL)
    print("All names, SSNs and client numbers below are invented test data.")
    print("=" * 72)

    mod = load_tool()
    folder = tempfile.mkdtemp(prefix="combiner selftest ")
    print("")
    print("Building test files in %s" % folder)
    try:
        for path in build_all(folder):
            print("  %s" % os.path.basename(path))
        scenario_by_name(mod, folder)
        scenario_first_row(mod, folder)
        scenario_overflow_parts(mod, folder)
        scenario_overflow_csv(mod, folder)
        scenario_date_formats(mod, folder)
    except Exception:
        print("")
        print(traceback.format_exc())
        FAILURES.append("the test itself blew up")
    finally:
        shutil.rmtree(folder, ignore_errors=True)

    print("")
    print("=" * 72)
    if FAILURES:
        print("%d of %d check(s) FAILED:" % (len(FAILURES), CHECKS[0]))
        for label in FAILURES:
            print("  - %s" % label)
    else:
        print("ALL %d CHECKS PASSED" % CHECKS[0])
    print("=" * 72)
    return 1 if FAILURES else 0


if __name__ == "__main__":
    code = main()
    if not sys.stdout.isatty():
        sys.exit(code)
