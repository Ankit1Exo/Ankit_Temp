"""
extract_wage_form.py
---------------------
Extracts the following fields from NY Quarterly Employee/Payee Wage
Reporting PDFs (Part C section):

    a  Social Security Number
    b  Last Name
    c  First Name
    d  MI
    g  Gross Federal Wages or Distribution

Supports searchable (text-layer) PDFs. Falls back to OCR automatically
if no text layer is found.

Output: CSV saved next to this script, plus a console summary.

Requirements:
    pip install pdfplumber pdf2image pytesseract pillow pandas openpyxl

Run:
    python extract_wage_form.py
"""

import os
import re
import sys
import csv

try:
    import pdfplumber
    import pandas as pd
except ImportError:
    sys.exit("Run: pip install pdfplumber pandas openpyxl")


# ── File / Folder Picker ──────────────────────────────────────────────────────

def pick_source() -> list:
    """
    GUI dialog: choose a folder (all PDFs inside) or individual PDF files.
    Falls back to console input in headless environments.
    """
    try:
        import tkinter as tk
        from tkinter import filedialog, simpledialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        mode = simpledialog.askstring(
            "Select Source",
            "Type  folder  → all PDFs in a folder\n"
            "Type  files   → pick individual PDFs",
            parent=root,
        )
        mode = (mode or "").strip().lower()

        if "folder" in mode:
            folder = filedialog.askdirectory(title="Select folder containing PDFs")
            root.destroy()
            if not folder:
                sys.exit("No folder selected.")
            paths = sorted([
                os.path.join(folder, f) for f in os.listdir(folder)
                if f.lower().endswith(".pdf") and not f.startswith("~")
            ])
            if not paths:
                sys.exit(f"No PDFs found in: {folder}")
        else:
            paths = list(filedialog.askopenfilenames(
                title="Select PDF files (Ctrl/Cmd+click for multiple)",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            ))
            root.destroy()
            if not paths:
                sys.exit("No files selected.")

        return paths

    except Exception:
        # Headless fallback
        print("\nHow to select files?")
        print("  1 - Folder path")
        print("  2 - Individual file paths")
        choice = input("Choice [1/2]: ").strip()
        if choice == "1":
            folder = input("Folder path: ").strip()
            paths = sorted([
                os.path.join(folder, f) for f in os.listdir(folder)
                if f.lower().endswith(".pdf")
            ])
            if not paths:
                sys.exit(f"No PDFs found in: {folder}")
            return paths
        else:
            paths = []
            print("Enter PDF paths (blank line to finish):")
            while True:
                p = input("  Path: ").strip()
                if not p:
                    break
                if os.path.exists(p):
                    paths.append(p)
                else:
                    print(f"  ⚠ Not found, skipping: {p}")
            if not paths:
                sys.exit("No valid files entered.")
            return paths


# ── Text Extraction Strategy ──────────────────────────────────────────────────

def get_page_words(page) -> list:
    """
    Return a list of word dicts with keys: text, x0, top, x1, bottom.
    Uses pdfplumber's word extraction which preserves spatial positions.
    """
    words = page.extract_words(
        x_tolerance=3,
        y_tolerance=3,
        keep_blank_chars=False,
        use_text_flow=False,
    )
    return words


def words_to_text(page) -> str:
    """Plain text dump for regex fallback."""
    return page.extract_text(x_tolerance=3, y_tolerance=3) or ""


# ── Field Detection — Layout-Aware ───────────────────────────────────────────
#
# The NY-45 Part C form has a repeating row structure per employee:
#
#   ROW 1 (label row):
#     "a Social Security number" | "b Last name" | "c First name" | "d MI" | "e Wage type R/O"
#
#   ROW 2 (data row):
#     <SSN>  |  <last name>  |  <first name>  |  <MI>  |  <wage type>
#
#   ROW 3 (label row):
#     "f Total UI remuneration..." | "g Gross federal wages..." | "h Total NYS..." | ...
#
#   ROW 4 (data row):
#     <UI amt>  |  <gross fed wages>  |  <NYS withheld>  |  ...
#
# Strategy:
#   1. Group words by their vertical (y) position into logical rows.
#   2. Detect "anchor" label rows that contain "Social Security".
#   3. The immediately following data row = employee name / SSN values.
#   4. Two rows later = the gross wages label row.
#   5. One more row = the gross wages data row.
# ─────────────────────────────────────────────────────────────────────────────

# Column x-ranges (approximate, as % of page width — scaled at runtime)
# These are calibrated from the NY-45 form layout visible in the screenshot.
# They work for standard letter-size NY-45 forms.
COL_RANGES = {
    "ssn":        (0.00, 0.18),   # a - Social Security Number
    "last_name":  (0.18, 0.38),   # b - Last Name
    "first_name": (0.38, 0.56),   # c - First Name
    "mi":         (0.56, 0.64),   # d - MI
    # "wage_type" skipped (e)
    "gross_wages": (0.32, 0.55),  # g - Gross Federal Wages (on the 3rd label row)
}


def words_in_band(words: list, y_min: float, y_max: float) -> list:
    """Return words whose vertical centre falls within [y_min, y_max]."""
    return [w for w in words if y_min <= (w["top"] + w["bottom"]) / 2 <= y_max]


def words_in_col(words: list, x_min: float, x_max: float) -> list:
    """Return words whose horizontal centre falls within [x_min, x_max]."""
    return [w for w in words if x_min <= (w["x0"] + w["x1"]) / 2 <= x_max]


def join_words(words: list) -> str:
    """Sort by x position and join into a string."""
    return " ".join(w["text"] for w in sorted(words, key=lambda w: w["x0"])).strip()


def group_into_rows(words: list, gap: float = 4.0) -> list:
    """
    Cluster words into horizontal rows by their top coordinate.
    Returns list of (avg_y, [words]) sorted top-to-bottom.
    """
    if not words:
        return []

    rows = []
    current_y   = words[0]["top"]
    current_row = [words[0]]

    for w in words[1:]:
        if abs(w["top"] - current_y) <= gap:
            current_row.append(w)
        else:
            rows.append((current_y, current_row))
            current_y   = w["top"]
            current_row = [w]
    rows.append((current_y, current_row))

    return sorted(rows, key=lambda r: r[0])


def is_label_row(row_words: list) -> bool:
    """True if this row contains the 'Social Security' label."""
    text = " ".join(w["text"].lower() for w in row_words)
    return "social" in text and "security" in text


def is_gross_label_row(row_words: list) -> bool:
    """True if this row contains the 'gross federal' wages label."""
    text = " ".join(w["text"].lower() for w in row_words)
    return "gross" in text and "federal" in text


def extract_records_from_page(page) -> list:
    """
    Extract employee records from one page of the form.
    Returns list of dicts with keys:
        ssn, last_name, first_name, mi, gross_wages
    """
    pw = float(page.width)
    words = get_page_words(page)

    # Sort all words by top then left
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))

    rows = group_into_rows(words, gap=5.0)
    records = []

    i = 0
    while i < len(rows):
        _, row_words = rows[i]

        if is_label_row(row_words):
            # Next row should be the SSN / Name data row
            if i + 1 < len(rows):
                _, data_row = rows[i + 1]

                def col(key):
                    lo, hi = COL_RANGES[key]
                    return join_words(words_in_col(data_row, pw * lo, pw * hi))

                ssn        = col("ssn")
                last_name  = col("last_name")
                first_name = col("first_name")
                mi         = col("mi")

                # Now find the gross wages: look ahead for the gross label row
                gross_wages = ""
                for j in range(i + 2, min(i + 6, len(rows))):
                    _, candidate = rows[j]
                    if is_gross_label_row(candidate):
                        # Data row immediately after the gross label
                        if j + 1 < len(rows):
                            _, wage_row = rows[j + 1]
                            lo, hi = COL_RANGES["gross_wages"]
                            gross_wages = join_words(
                                words_in_col(wage_row, pw * lo, pw * hi)
                            )
                        break

                # Only add if at least a name was found
                if last_name or first_name:
                    records.append({
                        "ssn":        clean_ssn(ssn),
                        "last_name":  last_name,
                        "first_name": first_name,
                        "mi":         mi,
                        "gross_wages": clean_number(gross_wages),
                    })

                i += 2   # skip the data row we just consumed
                continue

        i += 1

    return records


# ── OCR Fallback ──────────────────────────────────────────────────────────────

def ocr_fallback(pdf_path: str) -> list:
    """
    When no text layer is found, convert pages to images and run Tesseract.
    Returns the same list-of-dicts structure as extract_records_from_page.
    """
    try:
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import ImageEnhance, ImageFilter
    except ImportError:
        print("  ⚠ OCR requires: pip install pdf2image pytesseract pillow")
        return []

    print("  → No text layer. Running OCR fallback...")
    pages = convert_from_path(pdf_path, dpi=300)
    all_records = []

    for img in pages:
        img = img.resize((img.width * 2, img.height * 2))
        img = ImageEnhance.Contrast(img).enhance(2.0)
        img = img.filter(ImageFilter.SHARPEN)
        text = pytesseract.image_to_string(img, config="--psm 6")
        all_records.extend(parse_text_fallback(text))

    return all_records


def parse_text_fallback(text: str) -> list:
    """
    Regex-based parser used when only plain OCR text is available.
    Less precise than the coordinate approach but handles scanned PDFs.
    """
    # Pattern: lines near "Social Security" followed by name data
    SSN_RE     = re.compile(r"\b(\d{3}[-\s]?\d{2}[-\s]?\d{4})\b")
    NAME_RE    = re.compile(
        r"(?:Last\s+name|b\s+Last)[^\n]*\n([A-Z][A-Z\s,.\-]+?)"
        r"\s+([A-Z][A-Za-z\s.\-]+?)(?:\s+([A-Z]))?\s*$",
        re.MULTILINE | re.IGNORECASE,
    )
    GROSS_RE   = re.compile(
        r"(?:gross\s+federal[^\n]*)\n\s*([\d,.\s]+)",
        re.IGNORECASE,
    )

    records = []
    ssns    = SSN_RE.findall(text)
    grosses = [m.strip() for m in GROSS_RE.findall(text)]

    for nm in NAME_RE.finditer(text):
        idx = len(records)
        records.append({
            "ssn":         ssns[idx] if idx < len(ssns) else "",
            "last_name":   nm.group(1).strip(),
            "first_name":  nm.group(2).strip(),
            "mi":          nm.group(3).strip() if nm.group(3) else "",
            "gross_wages": clean_number(grosses[idx]) if idx < len(grosses) else "",
        })

    return records


# ── Value Cleaning ────────────────────────────────────────────────────────────

def clean_ssn(raw: str) -> str:
    """Normalise SSN to ###-##-#### format if digits present."""
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 9:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return raw   # return as-is if unexpected format


def clean_number(raw: str) -> str:
    """Strip stray characters from a wage number."""
    return re.sub(r"[^\d.,]", "", raw).strip()


# ── Per-file Driver ───────────────────────────────────────────────────────────

def process_pdf(pdf_path: str) -> list:
    """
    Open a PDF, extract records from every page, return list of dicts
    each tagged with the source filename.
    """
    filename = os.path.basename(pdf_path)
    records  = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            has_text = any(
                (p.extract_text() or "").strip() for p in pdf.pages
            )

            if not has_text:
                raw = ocr_fallback(pdf_path)
                for r in raw:
                    r["file_name"] = filename
                return raw

            for page_num, page in enumerate(pdf.pages, start=1):
                page_records = extract_records_from_page(page)
                for r in page_records:
                    r["file_name"] = filename
                    r["page"]      = page_num
                records.extend(page_records)

    except Exception as e:
        print(f"  ⚠ Error reading {filename}: {e}")

    return records


# ── Output ────────────────────────────────────────────────────────────────────

COLUMNS = [
    "file_name",
    "page",
    "ssn",
    "last_name",
    "first_name",
    "mi",
    "gross_wages",
]

COLUMN_LABELS = {
    "file_name":   "File Name",
    "page":        "Page",
    "ssn":         "a Social Security Number",
    "last_name":   "b Last Name",
    "first_name":  "c First Name",
    "mi":          "d MI",
    "gross_wages": "g Gross Federal Wages",
}


def save_csv(records: list, out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=COLUMNS,
            extrasaction="ignore",
        )
        # Write human-readable header
        writer.writerow(COLUMN_LABELS)
        writer.writerows(records)
    print(f"\nCSV saved: {out_path}")


def save_excel(records: list, out_path: str) -> None:
    df = pd.DataFrame(records, columns=COLUMNS)
    df.rename(columns=COLUMN_LABELS, inplace=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Wage Data")

        ws = writer.sheets["Wage Data"]

        # Header styling
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        HEADER_FILL = PatternFill("solid", start_color="1F4E79")
        HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        DATA_FONT   = Font(name="Arial", size=10)
        ALT_FILL    = PatternFill("solid", start_color="DEEAF1")

        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.font = DATA_FONT
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL

        for col in ws.columns:
            max_w = max(
                (len(str(c.value)) for c in col if c.value), default=8
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 4, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"Excel saved: {out_path}")


def print_summary(all_records: list) -> None:
    files = sorted({r["file_name"] for r in all_records})
    print(f"\n{'='*60}")
    print(f"  Extraction Summary")
    print(f"{'='*60}")
    print(f"  Files processed : {len(files)}")
    print(f"  Total records   : {len(all_records)}")
    print(f"\n  Records per file:")
    for f in files:
        n = sum(1 for r in all_records if r["file_name"] == f)
        print(f"    • {f}: {n} employee(s)")

    missing_ssn = sum(1 for r in all_records if not r.get("ssn"))
    if missing_ssn:
        print(f"\n  ⚠ {missing_ssn} record(s) with no SSN detected "
              f"(field may be redacted or blank in source PDF)")
    print(f"{'='*60}\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  NY Quarterly Wage Form — Field Extractor")
    print("=" * 60)

    pdf_paths = pick_source()
    print(f"\nFiles selected: {len(pdf_paths)}")
    for p in pdf_paths:
        print(f"  • {os.path.basename(p)}")

    all_records = []
    for path in pdf_paths:
        print(f"\nProcessing: {os.path.basename(path)}")
        records = process_pdf(path)
        print(f"  → {len(records)} record(s) found")
        all_records.extend(records)

    if not all_records:
        print("\nNo records extracted. Check PDF content or try OCR mode.")
        sys.exit(1)

    print_summary(all_records)

    # Save outputs in the same folder as the first selected PDF
    out_dir  = os.path.dirname(os.path.abspath(pdf_paths[0]))
    csv_path = os.path.join(out_dir, "wage_extraction.csv")
    xls_path = os.path.join(out_dir, "wage_extraction.xlsx")

    # Avoid overwriting existing files
    for base_path in [csv_path, xls_path]:
        if os.path.exists(base_path):
            root, ext = os.path.splitext(base_path)
            n = 1
            while os.path.exists(base_path):
                base_path = f"{root}_{n}{ext}"
                n += 1

    save_csv(all_records, csv_path)
    save_excel(all_records, xls_path)


if __name__ == "__main__":
    main()
