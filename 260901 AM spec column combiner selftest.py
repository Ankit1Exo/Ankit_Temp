"""
260901 AM spec column combiner selftest.py

Builds a folder of FABRICATED workbooks - no real data of any kind - runs
"260901 AM spec column excel combiner.py" over them headlessly, and checks
every rule the combiner claims to follow.

Decoys are planted on purpose: a sheet missing one spec column, a sheet whose
headings differ only in case and spacing, a file with two matching sheets, a
blank header row, duplicate headings, a heading gap, text that merely LOOKS
like a formula, a real formula with no cached value, a spec column that is
blank all the way down, values a naive tool would turn into numbers or dates,
and a completely blank row.

Run it from IDLE with F5. It prints PASS / FAIL per check and leaves the test
folder behind so the output can be opened in Excel.
"""

import csv
import importlib.util
import os
import shutil
import sys
import tempfile
from datetime import datetime

from openpyxl import Workbook, load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
TARGET = os.path.join(HERE, "260901 AM spec column excel combiner.py")

SPEC = ["Employee ID", "Full Name", "Amount"]

results = []


def check(label, got, want):
    ok = got == want
    results.append(ok)
    print("%-6s %-58s got %r" % ("PASS" if ok else "FAIL", label, got)
          + ("" if ok else "   want %r" % (want,)))


def load_tool():
    spec = importlib.util.spec_from_file_location("combiner", TARGET)
    module = importlib.util.module_from_spec(spec)
    sys.modules["combiner"] = module
    spec.loader.exec_module(module)
    return module


def sheet(workbook, title, rows, first=False):
    ws = workbook.active if first else workbook.create_sheet()
    ws.title = title
    for row in rows:
        ws.append(row)
    return ws


def force_text(ws, coordinate):
    """Store a value that starts with '=' as TEXT rather than as a formula,
    which is what a real export would contain."""
    ws[coordinate].data_type = "s"


def build(folder):
    """Every workbook below is invented for this test.

    Files are named so that they sort: blank header, clean, duplicate heading,
    missing amount, two matching sheets, zz blank amount column."""

    # 1. Clean file: exact spec + one extra column. Values chosen to break a
    #    tool that "helpfully" converts types.
    wb = Workbook()
    ws = sheet(wb, "Data", [
        ["Employee ID", "Full Name", "Amount", "Notes"],
        ["00123", "Ada", "0012.50", "leading zeros"],
        ["00007", " padded name ", 1234, "spaces kept"],
        ["8e9", "Bob", 12.5, "not a float"],
        [None, None, None, None],                       # blank row - dropped
        ["=1+1", "Cy", datetime(2026, 1, 5), "=UPPER(B6)"],
    ], first=True)
    force_text(ws, "A6")          # text that only LOOKS like a formula
    # D6 is left as a REAL formula with no cached value on purpose.
    wb.save(os.path.join(folder, "clean.xlsx"))

    # 2. Two matching sheets in one file, headings differing only by case and
    #    spacing, plus a third sheet missing a spec column.
    wb = Workbook()
    sheet(wb, "Jan", [
        ["EMPLOYEE  ID", "full name", " Amount "],
        ["001", "Dee", "10"],
    ], first=True)
    sheet(wb, "Feb", [
        ["Employee ID", "Full Name", "Amount", "Region"],
        ["002", "Eve", "20", "North"],
    ])
    sheet(wb, "Summary", [
        ["Employee ID", "Amount"],                      # Full Name missing
        ["003", "30"],
    ])
    wb.save(os.path.join(folder, "two matching sheets.xlsx"))

    # 3. Nothing matches: one spec column absent.
    wb = Workbook()
    sheet(wb, "Sheet1", [
        ["Employee ID", "Full Name"],
        ["004", "Fay"],
    ], first=True)
    wb.save(os.path.join(folder, "missing amount.xlsx"))

    # 4. Blank header row, and a header-only sheet with no data.
    wb = Workbook()
    sheet(wb, "Blank", [[None, None], ["x", "y"]], first=True)
    sheet(wb, "HeaderOnly", [["Employee ID", "Full Name", "Amount"]])
    wb.save(os.path.join(folder, "blank header.xlsx"))

    # 5. Duplicate heading and a gap in the heading row that still has data.
    wb = Workbook()
    sheet(wb, "Odd", [
        ["Employee ID", "Full Name", "Amount", "Notes", None, "Notes"],
        ["005", "Gus", "40", "first notes", "gap value", "second notes"],
    ], first=True)
    wb.save(os.path.join(folder, "duplicate heading.xlsx"))

    # 6. A spec column that is blank all the way down - what an uncached
    #    formula column looks like once it has been read.
    wb = Workbook()
    sheet(wb, "Data", [
        ["Employee ID", "Full Name", "Amount"],
        ["006", "Hal", None],
        ["007", "Ivy", None],
    ], first=True)
    wb.save(os.path.join(folder, "zz blank amount column.xlsx"))

    # 7. Files that must be ignored entirely.
    wb = Workbook()
    sheet(wb, "Data", [["Employee ID", "Full Name", "Amount"], ["999", "Ghost", "1"]],
          first=True)
    wb.save(os.path.join(folder, "~$locked.xlsx"))
    with open(os.path.join(folder, "notes.txt"), "w", encoding="utf-8") as handle:
        handle.write("ignore me")


def read_csv(path):
    with open(path, "r", newline="", encoding="utf-8-sig") as handle:
        return list(csv.reader(handle))


def summary_text(sheet_obj):
    return "\n".join(
        " | ".join("" if v is None else str(v) for v in row)
        for row in sheet_obj.iter_rows(values_only=True))


REAL_WORLD = [
    "Doc ID",
    "Student ID Number (CO, WA, DC)",
    "Employee ID and Password (ND,SD)",
    "Digital Cryptographic Signatures (AZ, NC, ND, WA)",
    "Parent's Birth Name (ND, NC)",
    "Username/Email Address and Password for Non-Fin Acc",
    "Driver's License Number",
    "Fin Acct Number ONLY",
]


def spec_checks(tool):
    """The column list must survive being typed out vertically. Six of the
    names below contain commas, which is what broke the first version."""
    print("-- reading the typed column list --")
    check("a vertical list with commas in the names is not split",
          tool.parse_spec("\n".join(REAL_WORLD)), REAL_WORLD)
    check("blank lines are ignored",
          tool.parse_spec("\n\nDoc ID\n\n  \nCity\n"), ["Doc ID", "City"])
    check("a paste carrying extra columns keeps only the first",
          tool.parse_spec("Doc ID\tTRUE\tDoc ID\nFirst Name\tTRUE\tFirst Name"),
          ["Doc ID", "First Name"])
    check("one single line is still split on commas",
          tool.parse_spec("Doc ID, First Name; Last Name"),
          ["Doc ID", "First Name", "Last Name"])
    check("a lone name with a comma in it survives on its own line",
          tool.parse_spec("Student ID Number (CO, WA, DC)\nDoc ID"),
          ["Student ID Number (CO, WA, DC)", "Doc ID"])
    check("case-insensitive duplicates are dropped",
          tool.parse_spec("Doc ID\ndoc id\nCity"), ["Doc ID", "City"])
    check("a curly apostrophe matches a straight one",
          tool.match_key("Driver’s License Number"),
          tool.match_key("Driver's License Number"))
    check("a wrapped heading with a line break matches",
          tool.match_key("Fin Acct\nNumber ONLY"),
          tool.match_key("Fin Acct Number ONLY"))
    check("a non-breaking space matches an ordinary one",
          tool.match_key("Doc\xa0ID"), tool.match_key("Doc ID"))
    print()


def punctuation_checks(tool, root):
    """A heading that differs only in punctuation must be explained, not just
    reported missing - and the tick box must then make it match."""
    print("\n-- punctuation and look-alike characters --")
    folder = os.path.join(root, "punctuation")
    os.makedirs(folder, exist_ok=True)
    spec = ["Doc ID", "Driver's License Number", "Fin Acct Number ONLY"]

    wb = Workbook()
    sheet(wb, "Data", [
        ["Doc ID", "Driver’s License Number", "Fin Acct Number ONLY"],
        ["D1", "L1", "F1"],
    ], first=True)
    wb.save(os.path.join(folder, "curly apostrophe.xlsx"))

    wb = Workbook()
    sheet(wb, "Data", [
        ["Doc ID", "Driver's License Number", "Fin-Acct Number ONLY"],
        ["D2", "L2", "F2"],
    ], first=True)
    wb.save(os.path.join(folder, "hyphenated.xlsx"))

    _, csv_path, _, summary_path, rows = tool.combine(
        folder, "ZZ", spec, lambda t="": None, lambda a, b: None)
    body = read_csv(csv_path)[1:]
    check("the curly-apostrophe file matched anyway", rows, 1)
    check("only the look-alike file came through",
          [r[0] for r in body], ["curly apostrophe.xlsx"])

    detail = {}
    for row in load_workbook(summary_path)["Sheet Details"].iter_rows(
            min_row=4, values_only=True):
        if row[0]:
            detail[row[0]] = row
    check("the hyphen difference is reported as missing",
          detail["hyphenated.xlsx"][6], "Fin Acct Number ONLY")
    check("and the near match names the real heading",
          detail["hyphenated.xlsx"][7],
          "Fin Acct Number ONLY ~ Fin-Acct Number ONLY")

    # Now the escape hatch.
    tool.IGNORE_PUNCTUATION = True
    try:
        _, csv_path, _, _, rows = tool.combine(
            folder, "ZZ", spec, lambda t="": None, lambda a, b: None)
        check("ignoring punctuation lets both files through", rows, 2)
        check("both files contributed",
              sorted(r[0] for r in read_csv(csv_path)[1:]),
              ["curly apostrophe.xlsx", "hyphenated.xlsx"])
    finally:
        tool.IGNORE_PUNCTUATION = False


def main():
    tool = load_tool()
    spec_checks(tool)

    root = os.path.join(tempfile.gettempdir(), "spec combiner selftest")
    shutil.rmtree(root, ignore_errors=True)
    os.makedirs(root)
    build(root)
    print("Test folder: %s\n" % root)

    log_lines = []
    out_folder, csv_path, parts, summary_path, rows = tool.combine(
        root, "ZZ", SPEC, log_lines.append, lambda a, b: None)

    table = read_csv(csv_path)
    headers = table[0]
    body = table[1:]
    index = {name: position for position, name in enumerate(headers)}

    # ---- shape ------------------------------------------------------------
    check("tracking columns lead the table",
          headers[:3], ["Source File", "Source Sheet", "Source Row"])
    check("spec columns keep the order given", headers[3:6], SPEC)
    check("extra columns land at the end, in discovery order",
          headers[6:], ["Notes", "Column E", "Notes (2)", "Region"])

    # ---- which sheets qualified ------------------------------------------
    pairs = sorted({(row[0], row[1]) for row in body})
    check("only fully matching sheets contributed", pairs, [
        ("clean.xlsx", "Data"),
        ("duplicate heading.xlsx", "Odd"),
        ("two matching sheets.xlsx", "Feb"),
        ("two matching sheets.xlsx", "Jan"),
        ("zz blank amount column.xlsx", "Data"),
    ])
    check("ignored the ~$ lock file and the .txt",
          any(r[0] in ("~$locked.xlsx", "notes.txt") for r in body), False)

    # ---- row handling -----------------------------------------------------
    check("blank row was dropped, everything else kept", len(body), 9)
    check("row count reported matches the file", rows, len(body))

    clean = [row for row in body if row[0] == "clean.xlsx"]
    check("source row numbers are the real Excel rows",
          [row[2] for row in clean], ["2", "3", "4", "6"])

    # ---- values are copied as they are ------------------------------------
    ids = [row[index["Employee ID"]] for row in clean]
    check("leading zeros survive", ids[0], "00123")
    check("00007 not turned into 7", ids[1], "00007")
    check("8e9 not turned into a number", ids[2], "8e9")
    check("text that looks like a formula stays text", ids[3], "=1+1")
    check("0012.50 keeps its trailing zero",
          clean[0][index["Amount"]], "0012.50")
    check("surrounding spaces are not trimmed",
          clean[1][index["Full Name"]], " padded name ")
    check("a whole number is not given a .0", clean[1][index["Amount"]], "1234")
    check("a real float keeps its point", clean[2][index["Amount"]], "12.5")
    check("a real date becomes YYYY-MM-DD", clean[3][index["Amount"]], "2026-01-05")
    # Documented limitation, asserted so it cannot change unnoticed.
    check("a real formula with no cached value comes through blank",
          clean[3][index["Notes"]], "")

    # ---- loose header matching, duplicates, gaps --------------------------
    jan = [row for row in body if row[1] == "Jan"][0]
    check("EMPLOYEE  ID matched Employee ID", jan[index["Employee ID"]], "001")
    check("' Amount ' matched Amount", jan[index["Amount"]], "10")
    check("a sheet without an extra leaves it blank", jan[index["Region"]], "")

    odd = [row for row in body if row[1] == "Odd"][0]
    check("duplicate heading kept apart", odd[index["Notes (2)"]], "second notes")
    check("gap column named after its letter", odd[index["Column E"]], "gap value")

    # ---- the xlsx copy ----------------------------------------------------
    check("one workbook was written", len(parts), 1)
    ws = load_workbook(parts[0], read_only=True)["Combined"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    check("xlsx has the same header", list(grid[0]), headers)
    check("xlsx has the same row count", len(grid) - 1, len(body))
    check("xlsx kept the leading zeros as text",
          grid[1][index["Employee ID"]], "00123")
    check("xlsx did not store =1+1 as a formula",
          grid[4][index["Employee ID"]], "=1+1")

    # ---- the summary workbook --------------------------------------------
    book = load_workbook(summary_path)
    check("summary has the three sheets",
          book.sheetnames, ["Summary", "Sheet Details", "Extra Columns"])

    detail = {}
    for row in book["Sheet Details"].iter_rows(min_row=4, values_only=True):
        if row[0]:
            detail[(row[0], row[1])] = row

    missing = detail[("two matching sheets.xlsx", "Summary")]
    check("skipped sheet has the missing column named",
          (missing[2], missing[6]), ("Skipped - missing columns", "Full Name"))
    check("file with no matching sheet is reported",
          detail[("missing amount.xlsx", "Sheet1")][6], "Amount")
    check("blank header row is reported",
          detail[("blank header.xlsx", "Blank")][2], "Skipped - no header row")
    check("header-only sheet appended zero rows",
          detail[("blank header.xlsx", "HeaderOnly")][4], 0)
    check("blank row is counted in the summary",
          detail[("clean.xlsx", "Data")][5], 1)
    check("a spec column blank all the way down is flagged",
          detail[("zz blank amount column.xlsx", "Data")][10], "Amount")
    # The flag is for a column blank ALL the way down; one uncached formula
    # inside an otherwise populated column is not - and must not be - flagged.
    check("a partly filled column is not flagged",
          detail[("clean.xlsx", "Data")][10], None)
    check("a healthy sheet flags nothing",
          detail[("two matching sheets.xlsx", "Jan")][10], None)

    text = summary_text(book["Summary"])
    check("multi-sheet file is flagged",
          "two matching sheets.xlsx | 2 | Jan, Feb" in text, True)
    check("split-into-parts line says No", "Split into parts | No" in text, True)

    # ---- splitting when the data will not fit -----------------------------
    print("\n-- forcing a split (PART_ROWS = 3) --")
    tool.PART_ROWS = 3
    split_folder = os.path.join(root, "split test")
    os.makedirs(split_folder, exist_ok=True)
    for name in ("clean.xlsx", "two matching sheets.xlsx", "duplicate heading.xlsx"):
        shutil.copy2(os.path.join(root, name), split_folder)
    _, split_csv, split_parts, split_summary, split_rows = tool.combine(
        split_folder, "ZZ", SPEC, log_lines.append, lambda a, b: None)
    check("run did not stop when the limit was hit", split_rows, 7)
    check("data was split into 3 parts", len(split_parts), 3)
    check("part files are numbered",
          [os.path.basename(p).split("combined ")[1] for p in split_parts],
          ["data part 01.xlsx", "data part 02.xlsx", "data part 03.xlsx"])
    counts = [len(list(load_workbook(p, read_only=True)["Combined"].iter_rows())) - 1
              for p in split_parts]
    check("every part carries its own header and <= 3 rows", counts, [3, 3, 1])
    check("the csv still holds every row in one piece",
          len(read_csv(split_csv)) - 1, 7)
    check("summary says it was split",
          "Yes - data exceeded Excel's row limit"
          in summary_text(load_workbook(split_summary)["Summary"]), True)

    # ---- a second run must not overwrite the first ------------------------
    print("\n-- running twice into the same folder --")
    tool.PART_ROWS = 1_000_000
    again_folder = os.path.join(root, "twice")
    os.makedirs(again_folder, exist_ok=True)
    shutil.copy2(os.path.join(root, "clean.xlsx"), again_folder)
    first = tool.combine(again_folder, "ZZ", SPEC, log_lines.append, lambda a, b: None)
    second = tool.combine(again_folder, "ZZ", SPEC, log_lines.append, lambda a, b: None)
    check("the second run wrote its own file", first[1] != second[1], True)
    check("the first run's file is still there", os.path.exists(first[1]), True)

    # ---- an empty folder must fail politely, not crash --------------------
    empty = os.path.join(root, "nothing here")
    os.makedirs(empty, exist_ok=True)
    try:
        tool.combine(empty, "ZZ", SPEC, log_lines.append, lambda a, b: None)
        check("empty folder raises a readable message", "no error", "an error")
    except RuntimeError as error:
        check("empty folder raises a readable message",
              str(error), "No .xlsx or .xlsm files in that folder.")

    print("\n%d of %d checks passed." % (sum(results), len(results)))
    print("Output left in: %s" % out_folder)
    return 0 if all(results) else 1


if __name__ == "__main__":
    sys.exit(main())
