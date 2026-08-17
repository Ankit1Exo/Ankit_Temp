"""
260817 AM combine excel selftest.py
-----------------------------------
Self-test for "260817 AM combine excel sheets.py".

Builds a set of FABRICATED workbooks (no real data of any kind) containing
deliberate traps, runs the combiner over them, and checks that the output is
exactly what it should be.

Traps planted:
    * a hidden sheet, with a hidden row and a hidden column   -> must be INCLUDED
    * a sheet with a bar chart on it                          -> must be SKIPPED
    * a chart-only sheet (chartsheet)                         -> must be SKIPPED
    * a sheet whose header sits in row 3 under a title        -> must be SKIPPED
    * a sheet whose first row is numbers, not headers         -> must be SKIPPED
    * a worksheet with a pivot table part attached            -> must be DETECTED
    * the same person recorded with different header spellings,
      different number formats and different date formats     -> must be DE-DUPED
Every trap row carries a "DECOY ..." payer name, so if any skipped sheet leaks
into the output the test fails loudly.

Run it from IDLE with F5.
"""

from __future__ import annotations

import importlib.util
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference

HERE = Path(__file__).resolve().parent
COMBINER = HERE / "260817 AM combine excel sheets.py"

FAILURES: list[str] = []


def check(condition: bool, message: str) -> None:
    if condition:
        print(f"  PASS  {message}")
    else:
        print(f"  FAIL  {message}")
        FAILURES.append(message)


def load_combiner():
    spec = importlib.util.spec_from_file_location("combiner", COMBINER)
    module = importlib.util.module_from_spec(spec)
    sys.modules["combiner"] = module
    spec.loader.exec_module(module)
    return module


# ---------------------------------------------------------------------------
# Fabricated test data
# ---------------------------------------------------------------------------

def build_file_alpha(path: Path) -> None:
    wb = openpyxl.Workbook()

    # -- sheet 1: the plain case -------------------------------------------
    ws = wb.active
    ws.title = "Data1"
    ws.append(["Per Num", "Pat Name", "Enc Num", "Enc Dt", "Payer Name"])
    ws.append([1001, "Ann Alpha", "E-1", datetime(2026, 1, 5), "Payer One"])
    ws.append([1002, "Bob Bravo", "E-2", datetime(2026, 1, 6), "Payer One"])
    ws.append([1003, "Cara Charlie", "E-3", datetime(2026, 1, 7), "Payer Two"])

    # -- sheet 2: different spellings + a same-file duplicate ---------------
    ws = wb.create_sheet("Data2")
    ws.append(["Person Number", "Patient Name", "Encounter", "Servc Dt", "Payer Name"])
    ws.append([" 1001.0 ", "ann alpha", "e-1", "2026-01-05", "Payer One"])   # duplicate
    ws.append([1004, "Dan Delta", "E-4", "01/08/2026", "Payer Two"])
    ws.append([1005, "Eve Echo", "E-5", "2026-01-09", "Payer Two"])

    # -- sheet 3: hidden sheet, hidden row, hidden column -------------------
    ws = wb.create_sheet("HiddenSheet")
    ws.append(["Per Number", "Pt Name", "Enc Nbr", "Encounter Date", "Payer Name"])
    ws.append([1006, "Fay Foxtrot", "E-6", datetime(2026, 1, 10), "Payer Hidden"])
    ws.sheet_state = "hidden"
    ws.row_dimensions[2].hidden = True
    ws.column_dimensions["C"].hidden = True

    # -- sheet 4: has a chart on it -> must be skipped ----------------------
    ws = wb.create_sheet("WithChart")
    ws.append(["Per Num", "Pat Name", "Enc Num", "Enc Dt", "Payer Name"])
    ws.append([9001, "Zoe Chart", "X-1", datetime(2026, 2, 1), "DECOY CHART"])
    ws.append([9002, "Yan Chart", "X-2", datetime(2026, 2, 2), "DECOY CHART"])
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=2, max_row=3))
    ws.add_chart(chart, "H2")

    # -- sheet 5: header is in row 3 -> must be skipped ---------------------
    ws = wb.create_sheet("TitleFirst")
    ws["A1"] = "Monthly Encounter Report - Q1"
    ws.append([])
    ws.append(["Per Num", "Pat Name", "Enc Num", "Enc Dt", "Payer Name"])
    ws.append([9003, "Xu Title", "X-3", datetime(2026, 2, 3), "DECOY TITLE"])

    # -- sheet 6: first row is data, not headers -> must be skipped ---------
    ws = wb.create_sheet("NumbersOnly")
    ws.append([9004, 1234, 5678, 91011])
    ws.append([9005, 2234, 6678, 91012])

    # -- sheet 6b: two blank rows above the header -> must be skipped -------
    ws = wb.create_sheet("BlankTop")
    ws.append([])
    ws.append([])
    ws.append(["Per Num", "Pat Name", "Enc Num", "Enc Dt", "Payer Name"])
    ws.append([9006, "Wes Blank", "X-6", datetime(2026, 2, 4), "DECOY BLANKTOP"])

    # -- sheet 7: chart-only sheet -> must be skipped -----------------------
    cs = wb.create_chartsheet("GraphOnly")
    only_chart = BarChart()
    only_chart.add_data(Reference(wb["Data1"], min_col=1, min_row=2, max_row=4))
    cs.add_chart(only_chart)

    wb.save(path)
    wb.close()


def build_file_bravo(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Per Nbr", "Pt.Name", "Enc Number", "Encounter Dt", "Payer Name", "Location"])
    ws.append([1003, "Cara Charlie", "E-3", datetime(2026, 1, 7), "Payer Two", "Site A"])  # cross-file dup
    ws.append([1007, "Gus Golf", "E-7", datetime(2026, 1, 11), "Payer Three", "Site A"])
    ws.append([1008, "Hal Hotel", "E-8", datetime(2026, 1, 12), "Payer Three", "Site B"])
    ws.append([1008, "Hal Hotel", "E-8", datetime(2026, 1, 12), "Payer Three", "Site B"])  # same-file dup
    wb.save(path)
    wb.close()


def build_file_charlie(path: Path) -> None:
    path.write_text(
        "PERSON NUM,PATIENT NAME,ENCOUNTER NUMBER,DATE OF SERVICE,Payer Name\n"
        "1007,Gus Golf,E-7,2026-01-11,Payer Three\n"      # cross-file duplicate of bravo
        "1009,Ivy India,E-9,2026-01-13,Payer Four\n",
        encoding="utf-8",
    )


def plant_pivot_decoy(source: Path, target: Path) -> str:
    """
    Copy a workbook and attach a pivot table part to its first worksheet, the
    same way Excel does. Returns the sheet name that should be detected.
    """
    with zipfile.ZipFile(source) as zin:
        items = [(i, zin.read(i.filename)) for i in zin.infolist()]

    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Target="../pivotTables/pivotTable1.xml" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable"/>'
        "</Relationships>"
    )
    skip = {"xl/worksheets/_rels/sheet1.xml.rels", "xl/pivotTables/pivotTable1.xml"}

    with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, data in items:
            if info.filename in skip:
                continue
            zout.writestr(info.filename, data)
        zout.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)
        zout.writestr("xl/pivotTables/pivotTable1.xml", '<pivotTableDefinition/>')

    wb = openpyxl.load_workbook(source)
    name = wb.sheetnames[0]
    wb.close()
    return name


# ---------------------------------------------------------------------------
# Test run
# ---------------------------------------------------------------------------

def main() -> int:
    if not COMBINER.exists():
        print(f"Cannot find the combiner script at {COMBINER}")
        return 1

    combiner = load_combiner()
    work = Path(tempfile.mkdtemp(prefix="combine_selftest_"))
    src = work / "input"
    out = work / "output"
    src.mkdir()

    print(f"Working folder: {work}\n")

    try:
        alpha = src / "260817 AM test alpha.xlsx"
        bravo = src / "260817 AM test bravo.xlsx"
        charlie = src / "260817 AM test charlie.csv"
        build_file_alpha(alpha)
        build_file_bravo(bravo)
        build_file_charlie(charlie)

        # ---- header mapping unit checks --------------------------------
        print("Header mapping")
        mapping_cases = {
            "Per Num": "Person Number", "person number": "Person Number",
            " Per  Number ": "Person Number", "Person Num": "Person Number",
            "Per Nbr": "Person Number",
            "Pat Name": "Patient Name", "Patient Name": "Patient Name",
            "Pt Name": "Patient Name", "Pt.Name": "Patient Name",
            "Encounter Number": "Encounter Number", "Enc Num": "Encounter Number",
            "Encounter": "Encounter Number", "Enc Nbr": "Encounter Number",
            "Date of Service": "Date of Service", "Enc Dt": "Date of Service",
            "Enc Date": "Date of Service", "Encounter Dt": "Date of Service",
            "Servc Dt": "Date of Service", "DOS": "Date of Service",
        }
        bad = {k: combiner.canonical_header(k) for k, v in mapping_cases.items()
               if combiner.canonical_header(k) != v}
        check(not bad, f"all {len(mapping_cases)} header spellings map correctly "
                       f"{'' if not bad else '- wrong: ' + str(bad)}")

        # decoys that must NOT be swallowed by the mapping rules
        decoys = {
            "Pat Age in Years (DOS)": None,
            "Rendering Provider Name": None,
            "Payer Name": None,
            "Location": None,
            "Fst Consult": None,
        }
        wrong = {k: combiner.canonical_header(k) for k, v in decoys.items()
                 if combiner.canonical_header(k) != v}
        check(not wrong, f"look-alike headers are left alone "
                         f"{'' if not wrong else '- wrong: ' + str(wrong)}")

        # ---- pivot decoy ------------------------------------------------
        print("\nPivot detection")
        decoy_path = work / "pivot decoy.xlsx"
        expected_sheet = plant_pivot_decoy(bravo, decoy_path)
        detected = combiner.sheets_with_pivots_from_package(decoy_path)
        check(expected_sheet in detected,
              f"planted pivot table detected on sheet '{expected_sheet}' (got {detected or 'nothing'})")
        check(not combiner.sheets_with_pivots_from_package(bravo),
              "a workbook with no pivot table reports none")

        # ---- full run ---------------------------------------------------
        print("\nCombine run")
        combiner.combine(src, out)

        parts = sorted(out.glob("* combined data*.xlsx"))
        check(len(parts) == 1, f"one output part written (got {len(parts)})")
        if not parts:
            return 1

        wb = openpyxl.load_workbook(parts[0])
        ws = wb["Combined"]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        header, data = rows[0], rows[1:]

        expected_cols = ["Person Number", "Patient Name", "Encounter Number", "Date of Service",
                         "Payer Name", "Location", "Source File", "Source Sheet"]
        check(header == expected_cols, f"columns are {expected_cols} (got {header})")

        check(len(data) == 9, f"9 unique rows after de-duplication (got {len(data)})")

        idx = {name: i for i, name in enumerate(header)}
        people = sorted(str(r[idx["Person Number"]]).strip() for r in data)
        check(people == [str(n) for n in range(1001, 1010)],
              f"person numbers 1001-1009 each appear exactly once (got {people})")

        payers = {str(r[idx["Payer Name"]]) for r in data}
        leaked = sorted(p for p in payers if p.startswith("DECOY"))
        check(not leaked, f"no rows leaked from skipped sheets {'' if not leaked else leaked}")

        names = {str(r[idx["Patient Name"]]) for r in data}
        check("Fay Foxtrot" in names, "hidden sheet / hidden row / hidden column was unhidden and included")

        sheets_used = {str(r[idx["Source Sheet"]]) for r in data}
        check(sheets_used == {"Data1", "Data2", "HiddenSheet", "Sheet1", "260817 AM test charlie"},
              f"only valid sheets contributed rows (got {sorted(sheets_used)})")

        loc = {str(r[idx["Person Number"]]).strip(): r[idx["Location"]] for r in data}
        check(loc.get("1008") == "Site B" and loc.get("1001") is None,
              "columns that exist in only one file are kept, others left blank")

        # ---- skip report ------------------------------------------------
        print("\nSkip report")
        reports = sorted(out.glob("* combine report.xlsx"))
        check(len(reports) == 1, "report written")
        wb = openpyxl.load_workbook(reports[0])
        report_rows = [list(r) for r in wb["Report"].iter_rows(values_only=True)][1:]
        wb.close()

        skipped = {str(r[1]): str(r[3]) for r in report_rows if r[2] == "SHEET SKIPPED"}
        check("WithChart" in skipped and "chart" in skipped["WithChart"].lower(),
              f"chart sheet skipped for the right reason ({skipped.get('WithChart')})")
        check("GraphOnly" in skipped, f"chart-only sheet skipped ({skipped.get('GraphOnly')})")
        check("TitleFirst" in skipped and "row 1" in skipped["TitleFirst"],
              f"header-not-in-row-1 sheet skipped ({skipped.get('TitleFirst')})")
        check("NumbersOnly" in skipped, f"headerless sheet skipped ({skipped.get('NumbersOnly')})")
        check("BlankTop" in skipped and "row 3" in skipped["BlankTop"],
              f"blank rows above the header are reported precisely ({skipped.get('BlankTop')})")

        print()
        if FAILURES:
            print(f"{len(FAILURES)} CHECK(S) FAILED:")
            for f in FAILURES:
                print(f"  - {f}")
            return 1
        print("All checks passed.")
        return 0

    finally:
        shutil.rmtree(work, ignore_errors=True)


if __name__ == "__main__":
    code = main()
    try:
        input("\nPress Enter to close...")
    except EOFError:
        pass
    sys.exit(code)
