"""
Business Tools -- simple tabbed GUI.

Seven tools, one window (tabs like a browser):

1. File Mover / Copier -- paste file names (no extension needed) and move or
   copy the matching files -- whatever their extension -- from one folder to
   another.

2. Split Files -- splits one Excel file into separate .xlsx files, one per
   unique value in a chosen column (e.g. one file per "Document" number),
   each written to an "Import Template" sheet.

3. Sheet Inspector -- scans every Excel file in a folder and reports, per
   sheet: visibility, sheet/workbook protection, filters, hidden rows/cols,
   and row/column counts. Can also unhide sheets/rows/columns, remove
   filters, and unprotect sheets/workbooks -- with a preview mode (default)
   so nothing is overwritten until "Save changes to files" is ticked. The
   report table can be exported to an .xlsx file.

4. Allocate & Move -- paste two columns copied from Excel (file name without
   extension, then an allocation/category value), pick the folder that has
   all the files, and this creates one subfolder per allocation value
   inside that same folder and moves each matching file into it.

5. PDF Searchable Check -- scans every PDF in a folder and reports, per
   page: whether it has a real text layer ("Yes" -- searchable), effectively
   none ("No" -- likely a scanned image that would need OCR to read), or
   very little ("Partial" -- worth a manual look), plus a word count. No OCR
   is performed -- this only detects and flags which pages need it. Export
   to Excel writes two sheets: "File Summary" (one row per file, with
   searchable/scanned/partial page counts) and "Page Detail" (one row per
   page).

6. Merge Files -- merges every .xlsx/.xlsm/.xls/.csv file in a folder into
   one combined file per extension (all .xlsx files merge together, all
   .csv files merge together, etc), with a "Source File" column added to
   every row so it can always be traced back to its original file. Two
   merge modes:
     - Append as-is: uses the first file's header row and pastes every
       other file's data underneath it unchanged, column position for
       column position (e.g. if file 1 is Name/ID and file 2 is ID/Name,
       file 2's data lands under Name/ID as-is, not realigned).
     - Match headers: trims/cleans each file's header text and matches
       columns by name (case-insensitive) across all files, so ID lines up
       with ID and Name with Name regardless of column order; any header
       only some files have is still included as its own column.

7. Text from PDF -- extracts text from a single PDF or every PDF in a
   folder, one row per line of text (file, page number, text, method).
   Uses each PDF's own text layer by default; can optionally fall back to
   (or force) Tesseract OCR for scanned pages with no text layer. Results
   show in a table and can be exported to Excel.

Usage:
    python business_tools.py

Requires: `pip install openpyxl` (used by all tools except PDF Searchable
Check and Text from PDF) and `pip install pdfplumber` (used by Text from
PDF). Optional: `pip install xlrd` to read legacy .xls files, `pip install
PyMuPDF` for PDF Searchable Check, and `pip install pdf2image pytesseract`
plus a separately installed Tesseract binary for Text from PDF's OCR modes
-- all degrade gracefully if not installed. .xlsb files are not supported by
any tool here -- save as .xlsx first.

Note: Sheet Inspector can overwrite Excel files in place, and Allocate &
Move rearranges files on disk. Review file contents before applying
changes, especially for anything containing student or personal data.
"""
import csv
import os
import threading
import shutil
import tkinter as tk
from concurrent.futures import ThreadPoolExecutor, as_completed
from tkinter import filedialog, messagebox, ttk

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection

try:
    import xlrd  # only needed to read legacy .xls files
except ImportError:
    xlrd = None

try:
    import fitz  # PyMuPDF -- only needed for the PDF Searchable Check tab.
    # Used instead of pypdf: PyMuPDF's text extraction is a compiled C
    # extension and is dramatically faster per page, which matters here
    # since a folder scan can mean extracting text from thousands of pages.
except ImportError:
    fitz = None

try:
    import pdfplumber  # only needed for the Text from PDF tab.
except ImportError:
    pdfplumber = None

WORKER_THREADS = 4
# Sheet Inspector's scan only opens files read-for-inspection (no writing), and
# most of the wait per file is I/O (opening/reading), not CPU -- especially
# when the folder is on a network share -- so it benefits from much higher
# concurrency than the write-heavy tools above.
SCAN_WORKER_THREADS = min(32, (os.cpu_count() or 4) * 4)
EXCEL_FILETYPES = [("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
INVALID_FILENAME_CHARS = '<>:"/\\|?*'


# ------------------------------------------------------------------
# Shared helpers
# ------------------------------------------------------------------

def sanitize_filename(text):
    text = (text or "").strip()
    if not text:
        return "(blank)"
    cleaned = "".join("_" if ch in INVALID_FILENAME_CHARS else ch for ch in text)
    cleaned = cleaned.rstrip(". ")
    return cleaned or "(blank)"


def unique_filename(used_counts, base):
    key = base.lower()
    if key not in used_counts:
        used_counts[key] = 1
        return base
    used_counts[key] += 1
    return f"{base} ({used_counts[key]})"


def cell_display(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def autofit_columns(ws, max_width=60):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 8), max_width)


class LogPanel(ttk.Frame):
    """A read-only scrolling log + one-line status, reused by every tab."""

    def __init__(self, parent, height=12):
        super().__init__(parent)
        text_frame = ttk.Frame(self)
        text_frame.pack(fill="both", expand=True)
        self.text = tk.Text(
            text_frame, height=height, state="disabled", wrap="word",
            font=("Consolas", 9), bg=PALETTE["card"], fg=PALETTE["text"],
            insertbackground=PALETTE["text"], relief="flat",
            borderwidth=1, highlightthickness=1,
            highlightbackground=PALETTE["border"], highlightcolor=PALETTE["accent"],
            padx=8, pady=6,
        )
        scroll = ttk.Scrollbar(text_frame, command=self.text.yview)
        self.text.configure(yscrollcommand=scroll.set)
        self.text.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        self.progress_var = tk.IntVar(value=0)
        self.progress_bar = ttk.Progressbar(
            self, orient="horizontal", mode="determinate", variable=self.progress_var,
        )
        self.progress_bar.pack(fill="x", pady=(6, 0))

        bottom_row = ttk.Frame(self)
        bottom_row.pack(fill="x", pady=(4, 0))
        self.status_var = tk.StringVar(value="Ready.")
        ttk.Label(
            bottom_row, textvariable=self.status_var, anchor="w",
            foreground=PALETTE["text_muted"], font=("Segoe UI", 9),
        ).pack(side="left", fill="x", expand=True)
        ttk.Button(bottom_row, text="Export Log...", command=self._export_log).pack(side="right")

    def log(self, message):
        self.text.configure(state="normal")
        self.text.insert("end", str(message) + "\n")
        self.text.see("end")
        self.text.configure(state="disabled")

    def clear(self):
        self.text.configure(state="normal")
        self.text.delete("1.0", "end")
        self.text.configure(state="disabled")
        self.set_progress(0, 0)

    def set_status(self, message):
        self.status_var.set(message)

    def set_progress(self, current, total):
        """Determinate progress bar -- total=0 shows an empty bar (idle). Always
        driven by a plain item count (files, in every caller), never by data
        volume (e.g. lines/rows extracted), so it stays a meaningful fraction."""
        self.progress_bar.configure(maximum=max(total, 1))
        self.progress_var.set(current)

    def _export_log(self):
        content = self.text.get("1.0", "end-1c")
        if not content.strip():
            messagebox.showinfo("Export Log", "Nothing to export yet -- the log is empty.")
            return
        path = filedialog.asksaveasfilename(
            title="Save log as",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile="log.txt",
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)
        except OSError as exc:
            messagebox.showerror("Export Log", f"Could not save log: {exc}")
            return
        messagebox.showinfo("Export Log", f"Log saved to:\n{path}")


def build_name_index(folder, recursive):
    """Scans folder ONCE and maps lowercase file stem -> full paths, so
    matching N requested names costs one scan instead of N."""
    index = {}

    def add(full_path, filename):
        stem = os.path.splitext(filename)[0].lower()
        index.setdefault(stem, []).append(full_path)

    if recursive:
        for root_dir, _dirs, files in os.walk(folder):
            for filename in files:
                add(os.path.join(root_dir, filename), filename)
    else:
        with os.scandir(folder) as entries:
            for entry in entries:
                if entry.is_file():
                    add(entry.path, entry.name)

    return index


def lookup_name(index, name):
    """Looks up `name` in a stem -> paths index built by build_name_index().
    Tries an exact match first (handles stems that legitimately contain a
    dot, e.g. "abc.001"); if that misses and `name` itself has a trailing
    extension (e.g. someone pasted "abc.xlsx" instead of just "abc"), also
    tries it with that last extension stripped off."""
    key = name.lower()
    hits = index.get(key)
    if hits:
        return hits
    stem = os.path.splitext(name)[0]
    if stem and stem.lower() != key:
        hits = index.get(stem.lower())
        if hits:
            return hits
    return []


# ------------------------------------------------------------------
# Tab 1: File Mover / Copier
# ------------------------------------------------------------------

class FileMoverTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.source_var = tk.StringVar()
        self.dest_var = tk.StringVar()
        self.action_var = tk.StringVar(value="copy")
        self.recursive_var = tk.BooleanVar(value=False)
        self.overwrite_var = tk.BooleanVar(value=False)
        self.use_subfolder_var = tk.BooleanVar(value=False)
        self._build_layout()

    def _build_layout(self):
        pad = {"padx": 10, "pady": 6}

        folder_frame = ttk.Frame(self)
        folder_frame.pack(fill="x", **pad)
        folder_frame.columnconfigure(1, weight=1)

        ttk.Label(folder_frame, text="Source folder:").grid(row=0, column=0, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.source_var).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(folder_frame, text="Browse...", command=self._pick_source).grid(row=0, column=2)

        ttk.Label(folder_frame, text="Destination folder:").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(folder_frame, textvariable=self.dest_var).grid(row=1, column=1, sticky="ew", padx=6, pady=(8, 0))
        ttk.Button(folder_frame, text="Browse...", command=self._pick_dest).grid(row=1, column=2, pady=(8, 0))

        self.names_label_var = tk.StringVar(value="File names (no extension needed), one per line:")
        ttk.Label(self, textvariable=self.names_label_var).pack(anchor="w", **pad)

        text_frame = ttk.Frame(self)
        text_frame.pack(fill="both", expand=True, padx=10)
        self.names_box = tk.Text(text_frame, height=10, wrap="none")
        names_scroll = ttk.Scrollbar(text_frame, command=self.names_box.yview)
        self.names_box.configure(yscrollcommand=names_scroll.set)
        self.names_box.pack(side="left", fill="both", expand=True)
        names_scroll.pack(side="right", fill="y")

        options_frame = ttk.Frame(self)
        options_frame.pack(fill="x", **pad)

        ttk.Radiobutton(options_frame, text="Copy", variable=self.action_var, value="copy").pack(side="left")
        ttk.Radiobutton(options_frame, text="Move", variable=self.action_var, value="move").pack(side="left", padx=(10, 0))
        ttk.Checkbutton(options_frame, text="Search subfolders too", variable=self.recursive_var).pack(side="left", padx=(20, 0))
        ttk.Checkbutton(options_frame, text="Overwrite existing files at destination", variable=self.overwrite_var).pack(side="left", padx=(20, 0))

        subfolder_frame = ttk.Frame(self)
        subfolder_frame.pack(fill="x", padx=10)
        ttk.Checkbutton(
            subfolder_frame,
            text="Paste 2 columns instead: file name, then the subfolder (under source) it's in",
            variable=self.use_subfolder_var,
            command=self._update_names_label,
        ).pack(side="left")

        button_row = ttk.Frame(self)
        button_row.pack(pady=(4, 8))
        self.run_button = ttk.Button(button_row, text="Run", command=self._run_clicked, style="Accent.TButton")
        self.run_button.pack(side="left", padx=(0, 8))
        ttk.Button(button_row, text="Reset", command=self._reset_clicked).pack(side="left")

        ttk.Label(self, text="Log:").pack(anchor="w", padx=10)
        self.log_panel = LogPanel(self, height=12)
        self.log_panel.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _update_names_label(self):
        if self.use_subfolder_var.get():
            self.names_label_var.set(
                "Paste 2 columns from Excel -- File Name (no extension), then Subfolder -- one pair per line:"
            )
        else:
            self.names_label_var.set("File names (no extension needed), one per line:")

    def _reset_clicked(self):
        self.source_var.set("")
        self.dest_var.set("")
        self.names_box.delete("1.0", "end")
        self.action_var.set("copy")
        self.recursive_var.set(False)
        self.overwrite_var.set(False)
        self.use_subfolder_var.set(False)
        self._update_names_label()
        self.log_panel.clear()
        self.log_panel.set_status("Ready.")
        self.run_button.configure(state="normal")

    def _pick_source(self):
        folder = filedialog.askdirectory(title="Select source folder")
        if folder:
            self.source_var.set(folder)

    def _pick_dest(self):
        folder = filedialog.askdirectory(title="Select destination folder")
        if folder:
            self.dest_var.set(folder)

    def _run_clicked(self):
        source = self.source_var.get().strip()
        dest = self.dest_var.get().strip()
        raw_names = self.names_box.get("1.0", "end")
        use_subfolder = self.use_subfolder_var.get()

        if not source or not os.path.isdir(source):
            messagebox.showerror("File Mover", "Please choose a valid source folder.")
            return
        if not dest:
            messagebox.showerror("File Mover", "Please choose a destination folder.")
            return

        if use_subfolder:
            pairs, bad_lines = parse_name_allocation_pairs(raw_names)
            if not pairs:
                messagebox.showerror(
                    "File Mover",
                    "Please paste at least one 'file name, subfolder' pair (one per line).",
                )
                return
            if bad_lines:
                proceed = messagebox.askyesno(
                    "File Mover",
                    f"{len(bad_lines)} line(s) don't look like 'file name + subfolder' and will be skipped. Continue?",
                )
                if not proceed:
                    return
            names = pairs
        else:
            names = []
            seen = set()
            for line in raw_names.splitlines():
                name = line.strip()
                if name and name.lower() not in seen:
                    seen.add(name.lower())
                    names.append(name)

            if not names:
                messagebox.showerror("File Mover", "Please paste at least one file name.")
                return

        os.makedirs(dest, exist_ok=True)

        self.log_panel.clear()
        self.run_button.configure(state="disabled")
        self.log_panel.set_status("Working...")

        thread = threading.Thread(
            target=self._process,
            args=(source, dest, names, self.action_var.get(), self.recursive_var.get(), self.overwrite_var.get(), use_subfolder),
            daemon=True,
        )
        thread.start()

    def _process(self, source, dest, names, action, recursive, overwrite, use_subfolder):
        # results_by_name["<label>"] collects one entry per matched file (or a
        # single NOT FOUND entry) so the closing summary can report success/
        # failure per requested name, with a reason for anything that didn't
        # go through -- not just a raw count.
        results_by_name = {}
        paths_to_transfer = []

        self.after(0, self.log_panel.set_status, "Scanning source folder...")

        if use_subfolder:
            # names is a list of (file_name, subfolder) pairs -- build one
            # index per unique subfolder mentioned, so a subfolder used by
            # many rows is only scanned once.
            subfolder_indexes = {}
            for name, subfolder in names:
                key = subfolder.lower()
                if key not in subfolder_indexes:
                    subfolder_path = os.path.join(source, subfolder)
                    subfolder_indexes[key] = (
                        build_name_index(subfolder_path, recursive)
                        if os.path.isdir(subfolder_path) else None
                    )

                label = f"{name} [{subfolder}]"
                index = subfolder_indexes[key]
                if index is None:
                    results_by_name[label] = [("not_found", None, f"subfolder '{subfolder}' does not exist under the source folder")]
                    self.after(0, self.log_panel.log, f"NOT FOUND: {label} -- subfolder does not exist")
                    continue

                hits = lookup_name(index, name)
                if not hits:
                    results_by_name[label] = [("not_found", None, f"no file with this name found in subfolder '{subfolder}'")]
                    self.after(0, self.log_panel.log, f"NOT FOUND: {label}")
                else:
                    results_by_name[label] = []
                    paths_to_transfer.extend((label, path) for path in hits)
        else:
            name_index = build_name_index(source, recursive)
            for name in names:
                hits = lookup_name(name_index, name)
                if not hits:
                    results_by_name[name] = [("not_found", None, "no file with this name found in source folder")]
                    self.after(0, self.log_panel.log, f"NOT FOUND: {name}")
                else:
                    results_by_name[name] = []
                    paths_to_transfer.extend((name, path) for path in hits)

        lock = threading.Lock()

        total = len(paths_to_transfer)
        done_count = 0
        self.after(0, self.log_panel.set_progress, 0, total)

        with ThreadPoolExecutor(max_workers=WORKER_THREADS) as pool:
            futures = [
                pool.submit(self._transfer_one, name, path, dest, action, overwrite)
                for name, path in paths_to_transfer
            ]
            for future in as_completed(futures):
                name, filename, status, detail = future.result()
                label = {"done": "MOVED" if action == "move" else "COPIED", "skipped": "SKIPPED", "error": "ERROR"}[status]
                suffix = f" -- {detail}" if detail else ""
                self.after(0, self.log_panel.log, f"{label}: {filename}{suffix}")
                with lock:
                    results_by_name[name].append((status, filename, detail))
                    done_count += 1
                    progress = done_count
                self.after(0, self.log_panel.set_progress, progress, total)
                self.after(0, self.log_panel.set_status, f"Processing file {progress} out of {total}...")

        self.after(0, self._log_final_status, results_by_name)

        succeeded = sum(1 for entries in results_by_name.values() for s, *_ in entries if s == "done")
        issues = sum(len(entries) for entries in results_by_name.values()) - succeeded
        summary = f"Done. {succeeded} file(s) transferred, {issues} with issues (see status above)."
        self.after(0, self._finish, summary)

    def _log_final_status(self, results_by_name):
        self.log_panel.log("")
        self.log_panel.log("===== STATUS =====")
        for name in sorted(results_by_name):
            entries = results_by_name[name]
            if all(status == "done" for status, _, _ in entries):
                files = ", ".join(filename for _, filename, _ in entries)
                self.log_panel.log(f"[OK]    {name}  -> {files}")
            else:
                self.log_panel.log(f"[ISSUE] {name}")
                for status, filename, detail in entries:
                    if status == "done":
                        self.log_panel.log(f"          - transferred OK: {filename}")
                    elif status == "not_found":
                        self.log_panel.log(f"          - not found: {detail}")
                    elif status == "skipped":
                        self.log_panel.log(f"          - skipped ({filename}): {detail}")
                    else:
                        self.log_panel.log(f"          - error ({filename}): {detail}")

    @staticmethod
    def _transfer_one(name, path, dest, action, overwrite):
        filename = os.path.basename(path)
        dest_path = os.path.join(dest, filename)

        if os.path.exists(dest_path) and not overwrite:
            return name, filename, "skipped", "a file with this name already exists at the destination"

        try:
            if action == "move":
                shutil.move(path, dest_path)
            else:
                shutil.copy2(path, dest_path)
            return name, filename, "done", None
        except OSError as exc:
            return name, filename, "error", str(exc)

    def _finish(self, summary):
        self.log_panel.log(summary)
        self.log_panel.set_status(summary)
        self.run_button.configure(state="normal")


# ------------------------------------------------------------------
# Tab 2: Split Files
# ------------------------------------------------------------------

def load_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [row for row in csv.reader(f)]

    if ext == ".xls":
        if xlrd is None:
            raise RuntimeError("Reading .xls files requires: pip install xlrd")
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        return [sheet.row_values(r) for r in range(sheet.nrows)]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def peek_header_row(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        if xlrd is None:
            raise RuntimeError("Reading .xls files requires: pip install xlrd")
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        return list(sheet.row_values(0)) if sheet.nrows else []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return list(row)
    finally:
        wb.close()


def group_rows_by_column(all_rows, column_index):
    header = all_rows[0] if all_rows else []
    order = []
    groups = {}
    for row in all_rows[1:]:
        key = cell_display(row[column_index]) if column_index < len(row) else ""
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)
    return header, order, groups


def write_group_file(header, rows, dest_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"
    if header:
        ws.append(header)
    for row in rows:
        ws.append(row)
    autofit_columns(ws)
    wb.save(dest_path)


class SplitFilesTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.input_file_var = tk.StringVar()
        self.output_dir_var = tk.StringVar()
        self.column_var = tk.StringVar(value="A")
        self.header_var = tk.StringVar(value="Document")
        self._build_layout()

    def _build_layout(self):
        pad = {"padx": 10, "pady": 6}

        file_frame = ttk.Frame(self)
        file_frame.pack(fill="x", **pad)
        file_frame.columnconfigure(1, weight=1)

        ttk.Label(file_frame, text="Input Excel file:").grid(row=0, column=0, sticky="w")
        ttk.Entry(file_frame, textvariable=self.input_file_var).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(file_frame, text="Browse...", command=self._pick_input).grid(row=0, column=2)

        ttk.Label(file_frame, text="Output folder:").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(file_frame, textvariable=self.output_dir_var).grid(row=1, column=1, sticky="ew", padx=6, pady=(8, 0))
        ttk.Button(file_frame, text="Browse...", command=self._pick_output).grid(row=1, column=2, pady=(8, 0))

        options_frame = ttk.Frame(self)
        options_frame.pack(fill="x", **pad)
        ttk.Label(options_frame, text="Split by column:").pack(side="left")
        ttk.Entry(options_frame, textvariable=self.column_var, width=4).pack(side="left", padx=(4, 16))
        ttk.Label(options_frame, text="Expected header text:").pack(side="left")
        ttk.Entry(options_frame, textvariable=self.header_var, width=18).pack(side="left", padx=(4, 0))

        button_row = ttk.Frame(self)
        button_row.pack(pady=(4, 8))
        self.run_button = ttk.Button(button_row, text="Split File", command=self._run_clicked, style="Accent.TButton")
        self.run_button.pack(side="left", padx=(0, 8))
        ttk.Button(button_row, text="Reset", command=self._reset_clicked).pack(side="left")

        ttk.Label(self, text="Log:").pack(anchor="w", padx=10)
        self.log_panel = LogPanel(self, height=10)
        self.log_panel.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _reset_clicked(self):
        self.input_file_var.set("")
        self.output_dir_var.set("")
        self.column_var.set("A")
        self.header_var.set("Document")
        self.log_panel.clear()
        self.log_panel.set_status("Ready.")
        self.run_button.configure(state="normal")

    def _pick_input(self):
        path = filedialog.askopenfilename(title="Select input Excel file", filetypes=EXCEL_FILETYPES)
        if path:
            self.input_file_var.set(path)

    def _pick_output(self):
        folder = filedialog.askdirectory(title="Select folder to save output files")
        if folder:
            self.output_dir_var.set(folder)

    def _run_clicked(self):
        path = self.input_file_var.get().strip()
        out_dir = self.output_dir_var.get().strip()
        column_letter = self.column_var.get().strip().upper() or "A"
        expected_header = self.header_var.get().strip()

        if not path or not os.path.isfile(path):
            messagebox.showerror("Split Files", "Please choose a valid input Excel file.")
            return
        if os.path.splitext(path)[1].lower() not in (".xlsx", ".xlsm", ".xls"):
            messagebox.showerror("Split Files", "Only .xlsx, .xlsm, or .xls files are supported (not .xlsb).")
            return
        if not out_dir:
            messagebox.showerror("Split Files", "Please choose an output folder.")
            return

        try:
            column_index = column_index_from_string(column_letter) - 1
        except ValueError:
            messagebox.showerror("Split Files", f"'{column_letter}' is not a valid column letter.")
            return

        try:
            header_row = peek_header_row(path)
        except Exception as exc:
            messagebox.showerror("Split Files", f"Could not read the file: {exc}")
            return

        if expected_header:
            actual = cell_display(header_row[column_index]) if column_index < len(header_row) else ""
            if actual != expected_header:
                proceed = messagebox.askyesno(
                    "Split Files",
                    f"The header in column {column_letter} is '{actual or '(empty)'}', "
                    f"not '{expected_header}'. Proceed anyway?",
                )
                if not proceed:
                    return

        os.makedirs(out_dir, exist_ok=True)
        self.log_panel.clear()
        self.run_button.configure(state="disabled")
        self.log_panel.set_status("Working...")

        thread = threading.Thread(
            target=self._process,
            args=(path, out_dir, column_index),
            daemon=True,
        )
        thread.start()

    def _process(self, path, out_dir, column_index):
        self.after(0, self.log_panel.set_status, "Reading source file...")
        try:
            all_rows = load_rows(path)
            header, order, groups = group_rows_by_column(all_rows, column_index)
        except Exception as exc:
            self.after(0, self.log_panel.log, f"ERROR reading source file: {exc}")
            self.after(0, self._finish, "Failed to read source file.")
            return

        used_names = {}
        jobs = []
        for key in order:
            filename = unique_filename(used_names, sanitize_filename(key)) + ".xlsx"
            jobs.append((key, filename, groups[key]))

        total = len(jobs)
        lock = threading.Lock()
        results = {}
        done_count = 0
        self.after(0, self.log_panel.set_progress, 0, total)

        with ThreadPoolExecutor(max_workers=WORKER_THREADS) as pool:
            futures = {
                pool.submit(write_group_file, header, rows, os.path.join(out_dir, filename)): (key, filename, len(rows))
                for key, filename, rows in jobs
            }
            for future in as_completed(futures):
                key, filename, row_count = futures[future]
                try:
                    future.result()
                    self.after(0, self.log_panel.log, f"CREATED: {filename}  ({row_count} row(s))")
                    with lock:
                        results[key] = ("done", filename, row_count)
                except Exception as exc:
                    self.after(0, self.log_panel.log, f"ERROR: {filename} -- {exc}")
                    with lock:
                        results[key] = ("error", filename, str(exc))
                with lock:
                    done_count += 1
                    progress = done_count
                self.after(0, self.log_panel.set_progress, progress, total)
                self.after(0, self.log_panel.set_status, f"Processing file {progress} out of {total}...")

        self.after(0, self._log_final_status, results)
        succeeded = sum(1 for status, *_ in results.values() if status == "done")
        summary = f"Done. {succeeded}/{len(results)} file(s) created in {out_dir}."
        self.after(0, self._finish, summary)

    def _log_final_status(self, results):
        self.log_panel.log("")
        self.log_panel.log("===== STATUS =====")
        for key in sorted(results):
            status, filename, detail = results[key]
            label = key if key else "(blank)"
            if status == "done":
                self.log_panel.log(f"[OK]    {label}  -> {filename}  ({detail} row(s))")
            else:
                self.log_panel.log(f"[ISSUE] {label}  -> {filename}: {detail}")

    def _finish(self, summary):
        self.log_panel.log(summary)
        self.log_panel.set_status(summary)
        self.run_button.configure(state="normal")


# ------------------------------------------------------------------
# Tab 3: Sheet Inspector (identify + unhide/unprotect)
# ------------------------------------------------------------------

def _blank_scan_row(filename, note):
    return {
        "file": filename, "sheet": "", "visibility": "", "sheet_protected": "",
        "wb_protected": "", "filter": "", "hidden": "", "rows": "", "cols": "", "note": note,
    }


def scan_workbook(path):
    filename = os.path.basename(path)
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        return _scan_xls(path, filename)

    try:
        wb = openpyxl.load_workbook(path, keep_vba=(ext == ".xlsm"))
    except Exception as exc:
        return [_blank_scan_row(filename, f"could not open (possibly password-protected): {exc}")]

    wb_protected = "Yes" if wb.security and wb.security.lockStructure else "No"
    rows = []
    for ws in wb.worksheets:
        visibility = {"visible": "Visible", "hidden": "Hidden", "veryHidden": "Very Hidden"}.get(ws.sheet_state, ws.sheet_state)
        hidden_rows = sum(1 for d in ws.row_dimensions.values() if d.hidden)
        hidden_cols = sum(1 for d in ws.column_dimensions.values() if d.hidden)
        hidden = f"{hidden_rows} row(s), {hidden_cols} col(s)" if (hidden_rows or hidden_cols) else "None"
        rows.append({
            "file": filename, "sheet": ws.title, "visibility": visibility,
            "sheet_protected": "Yes" if ws.protection.sheet else "No",
            "wb_protected": wb_protected,
            "filter": "Yes" if ws.auto_filter.ref else "No",
            "hidden": hidden, "rows": ws.max_row, "cols": ws.max_column, "note": "",
        })
    wb.close()
    return rows


def _scan_xls(path, filename):
    if xlrd is None:
        return [_blank_scan_row(filename, "cannot read .xls -- install xlrd (pip install xlrd)")]
    try:
        book = xlrd.open_workbook(path)
    except Exception as exc:
        return [_blank_scan_row(filename, f"could not open: {exc}")]
    rows = []
    for sheet in book.sheets():
        rows.append({
            "file": filename, "sheet": sheet.name, "visibility": "n/a",
            "sheet_protected": "n/a", "wb_protected": "n/a", "filter": "n/a",
            "hidden": "n/a", "rows": sheet.nrows, "cols": sheet.ncols,
            "note": "legacy .xls -- protection/filter info not available",
        })
    return rows


def apply_workbook_actions(path, opts):
    ext = os.path.splitext(path)[1].lower()
    try:
        wb = openpyxl.load_workbook(path, keep_vba=(ext == ".xlsm"))
    except Exception as exc:
        return "error", f"could not open: {exc}"

    changed = []
    for ws in wb.worksheets:
        if opts["unhide_sheets"] and ws.sheet_state != "visible":
            ws.sheet_state = "visible"
            changed.append(f"unhid sheet '{ws.title}'")

        if opts["unhide_rowcols"]:
            hidden_rows = sum(1 for d in ws.row_dimensions.values() if d.hidden)
            hidden_cols = sum(1 for d in ws.column_dimensions.values() if d.hidden)
            for dim in ws.row_dimensions.values():
                dim.hidden = False
            for dim in ws.column_dimensions.values():
                dim.hidden = False
            if hidden_rows or hidden_cols:
                changed.append(f"unhid {hidden_rows} row(s)/{hidden_cols} col(s) on '{ws.title}'")

        if opts["remove_filters"] and ws.auto_filter.ref:
            ws.auto_filter.ref = None
            changed.append(f"removed filter on '{ws.title}'")

        if opts["unprotect_sheets"] and ws.protection.sheet:
            ws.protection = SheetProtection()
            changed.append(f"unprotected sheet '{ws.title}'")

    if opts["unprotect_workbook"] and wb.security and wb.security.lockStructure:
        wb.security = WorkbookProtection()
        changed.append("unprotected workbook structure")

    if not changed:
        return "skipped", "nothing to change"

    if opts["save"]:
        try:
            wb.save(path)
        except Exception as exc:
            return "error", f"could not save: {exc}"
        return "done", "; ".join(changed)

    return "preview", "; ".join(changed) + " (not saved -- tick 'Save changes' to apply)"


class SheetInspectorTab(ttk.Frame):
    COLUMNS = ("file", "sheet", "visibility", "sheet_protected", "wb_protected", "filter", "hidden", "rows", "cols", "note")
    HEADINGS = {
        "file": "File", "sheet": "Sheet", "visibility": "Visibility",
        "sheet_protected": "Sheet Protected", "wb_protected": "Workbook Protected",
        "filter": "Filter Applied", "hidden": "Hidden Rows/Cols", "rows": "Rows", "cols": "Columns", "note": "Note",
    }

    def __init__(self, parent):
        super().__init__(parent)
        self.folder_var = tk.StringVar()
        self.recursive_var = tk.BooleanVar(value=False)
        self.unhide_sheets_var = tk.BooleanVar(value=True)
        self.unhide_rowcols_var = tk.BooleanVar(value=True)
        self.remove_filters_var = tk.BooleanVar(value=True)
        self.unprotect_sheets_var = tk.BooleanVar(value=True)
        self.unprotect_workbook_var = tk.BooleanVar(value=True)
        self.save_changes_var = tk.BooleanVar(value=False)
        self.records = []
        self._build_layout()

    def _build_layout(self):
        pad = {"padx": 10, "pady": 6}

        folder_frame = ttk.Frame(self)
        folder_frame.pack(fill="x", **pad)
        folder_frame.columnconfigure(1, weight=1)
        ttk.Label(folder_frame, text="Folder with Excel files:").grid(row=0, column=0, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.folder_var).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(folder_frame, text="Browse...", command=self._pick_folder).grid(row=0, column=2)
        ttk.Checkbutton(folder_frame, text="Include subfolders", variable=self.recursive_var).grid(row=1, column=1, sticky="w", pady=(6, 0))

        scan_buttons = ttk.Frame(self)
        scan_buttons.pack(pady=(2, 8))
        ttk.Button(scan_buttons, text="Scan Files", command=self._scan_clicked, style="Accent.TButton").pack(side="left", padx=(0, 8))
        ttk.Button(scan_buttons, text="Export Report to Excel...", command=self._export_clicked).pack(side="left")

        self.preview_label = ttk.Label(self, text="Preview (first 50 rows):")
        self.preview_label.pack(anchor="w", padx=10, pady=(0, 2))

        table_frame = ttk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10)
        self.tree = ttk.Treeview(table_frame, columns=self.COLUMNS, show="headings", height=8)
        style_tree_rows(self.tree)
        for col in self.COLUMNS:
            self.tree.heading(col, text=self.HEADINGS[col])
            self.tree.column(col, width=100, anchor="w", stretch=True)
        self.tree.column("file", width=170)
        self.tree.column("note", width=200)
        tree_scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        tree_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        actions_frame = ttk.LabelFrame(self, text="Unhide / Unprotect")
        actions_frame.pack(fill="x", padx=10, pady=(10, 6))
        ttk.Checkbutton(actions_frame, text="Unhide hidden sheets", variable=self.unhide_sheets_var).grid(row=0, column=0, sticky="w", padx=8, pady=4)
        ttk.Checkbutton(actions_frame, text="Unhide hidden rows/columns", variable=self.unhide_rowcols_var).grid(row=0, column=1, sticky="w", padx=8, pady=4)
        ttk.Checkbutton(actions_frame, text="Remove filters", variable=self.remove_filters_var).grid(row=0, column=2, sticky="w", padx=8, pady=4)
        ttk.Checkbutton(actions_frame, text="Unprotect sheets", variable=self.unprotect_sheets_var).grid(row=1, column=0, sticky="w", padx=8, pady=4)
        ttk.Checkbutton(actions_frame, text="Unprotect workbook structure", variable=self.unprotect_workbook_var).grid(row=1, column=1, sticky="w", padx=8, pady=4)
        ttk.Checkbutton(actions_frame, text="Save changes to files (modifies originals!)", variable=self.save_changes_var).grid(row=1, column=2, sticky="w", padx=8, pady=4)

        button_row = ttk.Frame(self)
        button_row.pack(pady=(2, 8))
        self.apply_button = ttk.Button(button_row, text="Apply", command=self._apply_clicked, style="Accent.TButton")
        self.apply_button.pack(side="left", padx=(0, 8))
        ttk.Button(button_row, text="Reset", command=self._reset_clicked).pack(side="left")

        ttk.Label(self, text="Log:").pack(anchor="w", padx=10)
        self.log_panel = LogPanel(self, height=8)
        self.log_panel.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _reset_clicked(self):
        self.folder_var.set("")
        self.recursive_var.set(False)
        self.unhide_sheets_var.set(True)
        self.unhide_rowcols_var.set(True)
        self.remove_filters_var.set(True)
        self.unprotect_sheets_var.set(True)
        self.unprotect_workbook_var.set(True)
        self.save_changes_var.set(False)
        for row_id in self.tree.get_children():
            self.tree.delete(row_id)
        self.records = []
        self.preview_label.configure(text="Preview (first 50 rows):")
        self.log_panel.clear()
        self.log_panel.set_status("Ready.")
        self.apply_button.configure(state="normal")

    def _pick_folder(self):
        folder = filedialog.askdirectory(title="Select folder with Excel files")
        if folder:
            self.folder_var.set(folder)

    @staticmethod
    def _collect_files(folder, recursive):
        files = []
        if recursive:
            for root_dir, _dirs, names in os.walk(folder):
                for name in names:
                    if name.lower().endswith((".xlsx", ".xlsm", ".xls")) and not name.startswith("~$"):
                        files.append(os.path.join(root_dir, name))
        else:
            with os.scandir(folder) as entries:
                for entry in entries:
                    if entry.is_file() and entry.name.lower().endswith((".xlsx", ".xlsm", ".xls")) and not entry.name.startswith("~$"):
                        files.append(entry.path)
        return files

    def _scan_clicked(self):
        folder = self.folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Sheet Inspector", "Please choose a valid folder.")
            return

        files = self._collect_files(folder, self.recursive_var.get())
        if not files:
            messagebox.showinfo("Sheet Inspector", "No Excel files found in that folder.")
            return

        for row_id in self.tree.get_children():
            self.tree.delete(row_id)
        self.records = []
        self.preview_label.configure(text="Preview (first 50 rows):")
        self.log_panel.clear()
        self.log_panel.set_status(f"Scanning {len(files)} file(s)...")

        thread = threading.Thread(target=self._scan_worker, args=(files,), daemon=True)
        thread.start()

    def _export_clicked(self):
        if not self.records:
            messagebox.showinfo("Sheet Inspector", "Nothing to export yet -- run Scan Files first.")
            return

        path = filedialog.asksaveasfilename(
            title="Save report as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="sheet_inspector_report.xlsx",
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet Inspector Report"
            ws.append([self.HEADINGS[col] for col in self.COLUMNS])
            for row in self.records:
                ws.append([row.get(col, "") for col in self.COLUMNS])
            autofit_columns(ws)
            wb.save(path)
        except Exception as exc:
            messagebox.showerror("Sheet Inspector", f"Could not save report: {exc}")
            return

        messagebox.showinfo("Sheet Inspector", f"Report saved to:\n{path}")

    def _scan_worker(self, files):
        total = len(files)
        done_count = 0
        displayed_count = 0
        total_rows = 0
        self.after(0, self.log_panel.log, f"Found {total} file(s). Scanning with {SCAN_WORKER_THREADS} threads...")
        self.after(0, self.log_panel.set_progress, 0, total)
        with ThreadPoolExecutor(max_workers=SCAN_WORKER_THREADS) as pool:
            futures = [pool.submit(scan_workbook, path) for path in files]
            for future in as_completed(futures):
                rows = future.result()
                self.records.extend(rows)
                total_rows += len(rows)
                for row in rows:
                    if displayed_count < 50:
                        self.after(0, self._add_row, row)
                        displayed_count += 1
                    if row.get("note"):
                        self.after(0, self.log_panel.log, f"{row['file']}: {row['note']}")
                done_count += 1
                self.after(0, self.log_panel.set_progress, done_count, total)
                self.after(0, self.log_panel.set_status, f"Scanning file {done_count} out of {total}...")
                if done_count % 100 == 0:
                    self.after(0, self.log_panel.log, f"...scanned {done_count} of {total} file(s)")
        self.after(0, self.log_panel.set_status, f"Scan complete -- {len(files)} file(s).")
        self.after(0, self.log_panel.log, f"Scan complete. {len(files)} file(s) scanned.")
        if total_rows > 50:
            self.after(0, self._set_preview_label, total_rows)

    def _set_preview_label(self, total_rows):
        self.preview_label.configure(text=f"Preview (first 50 of {total_rows} rows -- use Export for the rest):")

    def _add_row(self, row):
        self.tree.insert("", "end", values=[row.get(col, "") for col in self.COLUMNS], tags=(next_row_tag(self.tree),))

    def _apply_clicked(self):
        folder = self.folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Sheet Inspector", "Please choose a valid folder.")
            return

        opts = {
            "unhide_sheets": self.unhide_sheets_var.get(),
            "unhide_rowcols": self.unhide_rowcols_var.get(),
            "remove_filters": self.remove_filters_var.get(),
            "unprotect_sheets": self.unprotect_sheets_var.get(),
            "unprotect_workbook": self.unprotect_workbook_var.get(),
            "save": self.save_changes_var.get(),
        }
        if not any(opts[k] for k in ("unhide_sheets", "unhide_rowcols", "remove_filters", "unprotect_sheets", "unprotect_workbook")):
            messagebox.showerror("Sheet Inspector", "Please select at least one action.")
            return

        files = [f for f in self._collect_files(folder, self.recursive_var.get()) if not f.lower().endswith(".xls")]
        if not files:
            messagebox.showinfo("Sheet Inspector", "No .xlsx/.xlsm files found in that folder (legacy .xls can't be modified).")
            return

        if opts["save"]:
            proceed = messagebox.askyesno(
                "Sheet Inspector",
                f"This will open and OVERWRITE up to {len(files)} file(s) in:\n{folder}\n\nThis cannot be undone. Continue?",
            )
            if not proceed:
                return
        else:
            messagebox.showinfo(
                "Sheet Inspector",
                "Preview mode: changes will be logged but NOT saved.\nTick 'Save changes to files' to actually apply them.",
            )

        self.log_panel.clear()
        self.apply_button.configure(state="disabled")
        self.log_panel.set_status("Working...")

        thread = threading.Thread(target=self._apply_worker, args=(files, opts), daemon=True)
        thread.start()

    def _apply_worker(self, files, opts):
        results = {}
        total = len(files)
        done_count = 0
        self.after(0, self.log_panel.set_progress, 0, total)
        with ThreadPoolExecutor(max_workers=WORKER_THREADS) as pool:
            futures = {pool.submit(apply_workbook_actions, path, opts): path for path in files}
            for future in as_completed(futures):
                filename = os.path.basename(futures[future])
                status, detail = future.result()
                results[filename] = (status, detail)
                label = {"done": "DONE", "preview": "PREVIEW", "skipped": "SKIPPED", "error": "ERROR"}[status]
                self.after(0, self.log_panel.log, f"{label}: {filename} -- {detail}")
                done_count += 1
                self.after(0, self.log_panel.set_progress, done_count, total)
                self.after(0, self.log_panel.set_status, f"Processing file {done_count} out of {total}...")

        self.after(0, self._log_final_status, results)
        done = sum(1 for status, _ in results.values() if status in ("done", "preview"))
        summary = f"Done. {done}/{len(results)} file(s) updated."
        self.after(0, self._finish, summary)

    def _log_final_status(self, results):
        self.log_panel.log("")
        self.log_panel.log("===== STATUS =====")
        for filename in sorted(results):
            status, detail = results[filename]
            if status in ("done", "preview"):
                self.log_panel.log(f"[OK]    {filename}  -> {detail}")
            else:
                self.log_panel.log(f"[ISSUE] {filename}  -> {detail}")

    def _finish(self, summary):
        self.log_panel.log(summary)
        self.log_panel.set_status(summary)
        self.apply_button.configure(state="normal")


# ------------------------------------------------------------------
# Tab 4: Allocate & Move
# ------------------------------------------------------------------

def parse_name_allocation_pairs(raw_text):
    """Parses lines pasted from a 2-column Excel selection (tab-separated)
    -- also accepts comma-separated as a manual-entry fallback. Returns
    (pairs, bad_lines) so the caller can warn about anything unreadable."""
    pairs = []
    bad_lines = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "\t" in line:
            parts = line.split("\t")
        elif "," in line:
            parts = line.split(",")
        else:
            parts = line.split(None, 1)

        parts = [p.strip() for p in parts]
        if len(parts) < 2 or not parts[0] or not parts[1]:
            bad_lines.append(line)
            continue
        pairs.append((parts[0], parts[1]))
    return pairs, bad_lines


def allocate_one(base_folder, allocation, path, overwrite):
    filename = os.path.basename(path)
    folder_name = sanitize_filename(allocation)
    dest_dir = os.path.join(base_folder, folder_name)
    os.makedirs(dest_dir, exist_ok=True)
    dest_path = os.path.join(dest_dir, filename)

    if os.path.exists(dest_path) and not overwrite:
        return filename, "skipped", f"a file with this name already exists in '{folder_name}'"

    try:
        shutil.move(path, dest_path)
        return filename, "done", folder_name
    except OSError as exc:
        return filename, "error", str(exc)


class AllocateFilesTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.folder_var = tk.StringVar()
        self.recursive_var = tk.BooleanVar(value=False)
        self.overwrite_var = tk.BooleanVar(value=False)
        self._build_layout()

    def _build_layout(self):
        pad = {"padx": 10, "pady": 6}

        folder_frame = ttk.Frame(self)
        folder_frame.pack(fill="x", **pad)
        folder_frame.columnconfigure(1, weight=1)
        ttk.Label(folder_frame, text="Folder with the files:").grid(row=0, column=0, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.folder_var).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(folder_frame, text="Browse...", command=self._pick_folder).grid(row=0, column=2)

        ttk.Label(
            self,
            text="Paste 2 columns from Excel -- File Name (no extension), then Allocation -- one pair per line:",
        ).pack(anchor="w", padx=10, pady=(6, 0))

        text_frame = ttk.Frame(self)
        text_frame.pack(fill="both", expand=True, padx=10, pady=(4, 0))
        self.pairs_box = tk.Text(text_frame, height=8, wrap="none")
        pairs_scroll = ttk.Scrollbar(text_frame, command=self.pairs_box.yview)
        self.pairs_box.configure(yscrollcommand=pairs_scroll.set)
        self.pairs_box.pack(side="left", fill="both", expand=True)
        pairs_scroll.pack(side="right", fill="y")

        options_frame = ttk.Frame(self)
        options_frame.pack(fill="x", **pad)
        ttk.Checkbutton(options_frame, text="Also search subfolders", variable=self.recursive_var).pack(side="left")
        ttk.Checkbutton(options_frame, text="Overwrite existing files at destination", variable=self.overwrite_var).pack(side="left", padx=(20, 0))

        button_row = ttk.Frame(self)
        button_row.pack(pady=(4, 8))
        self.run_button = ttk.Button(button_row, text="Allocate & Move", command=self._run_clicked, style="Accent.TButton")
        self.run_button.pack(side="left", padx=(0, 8))
        ttk.Button(button_row, text="Reset", command=self._reset_clicked).pack(side="left")

        ttk.Label(self, text="Log:").pack(anchor="w", padx=10)
        self.log_panel = LogPanel(self, height=10)
        self.log_panel.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _reset_clicked(self):
        self.folder_var.set("")
        self.pairs_box.delete("1.0", "end")
        self.recursive_var.set(False)
        self.overwrite_var.set(False)
        self.log_panel.clear()
        self.log_panel.set_status("Ready.")
        self.run_button.configure(state="normal")

    def _pick_folder(self):
        folder = filedialog.askdirectory(title="Select folder with the files")
        if folder:
            self.folder_var.set(folder)

    def _run_clicked(self):
        folder = self.folder_var.get().strip()
        raw = self.pairs_box.get("1.0", "end")

        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Allocate & Move", "Please choose a valid folder.")
            return

        pairs, bad_lines = parse_name_allocation_pairs(raw)
        if not pairs:
            messagebox.showerror(
                "Allocate & Move",
                "Please paste at least one 'file name, allocation' pair (one per line).",
            )
            return
        if bad_lines:
            proceed = messagebox.askyesno(
                "Allocate & Move",
                f"{len(bad_lines)} line(s) don't look like 'file name + allocation' and will be skipped. Continue?",
            )
            if not proceed:
                return

        self.log_panel.clear()
        self.run_button.configure(state="disabled")
        self.log_panel.set_status("Working...")

        thread = threading.Thread(
            target=self._process,
            args=(folder, pairs, self.recursive_var.get(), self.overwrite_var.get()),
            daemon=True,
        )
        thread.start()

    def _process(self, folder, pairs, recursive, overwrite):
        self.after(0, self.log_panel.set_status, "Scanning folder...")
        name_index = build_name_index(folder, recursive)

        results = {}
        jobs = []
        for name, allocation in pairs:
            hits = lookup_name(name_index, name)
            if not hits:
                results[name] = [("not_found", None, None, "no file with this name found in folder")]
                self.after(0, self.log_panel.log, f"NOT FOUND: {name}")
            else:
                results[name] = []
                jobs.extend((name, allocation, path) for path in hits)

        total = len(jobs)
        done_count = 0
        lock = threading.Lock()

        self.after(0, self.log_panel.set_status, f"Allocating {total} file(s) using {WORKER_THREADS} threads...")

        with ThreadPoolExecutor(max_workers=WORKER_THREADS) as pool:
            futures = {
                pool.submit(allocate_one, folder, allocation, path, overwrite): (name, allocation)
                for name, allocation, path in jobs
            }
            for future in as_completed(futures):
                name, allocation = futures[future]
                filename, status, detail = future.result()
                if status == "done":
                    self.after(0, self.log_panel.log, f"MOVED: {filename}  -> {detail}\\")
                else:
                    label = "SKIPPED" if status == "skipped" else "ERROR"
                    self.after(0, self.log_panel.log, f"{label}: {filename} -- {detail}")
                with lock:
                    results[name].append((status, filename, allocation, detail))
                    done_count += 1
                    progress = done_count
                self.after(0, self.log_panel.set_status, f"Processing file {progress} out of {total}...")

        self.after(0, self._log_final_status, results)
        succeeded = sum(1 for entries in results.values() for s, *_ in entries if s == "done")
        issues = sum(len(entries) for entries in results.values()) - succeeded
        summary = f"Done. {succeeded} file(s) moved, {issues} with issues (see status above)."
        self.after(0, self._finish, summary)

    def _log_final_status(self, results):
        self.log_panel.log("")
        self.log_panel.log("===== STATUS =====")
        for name in sorted(results):
            entries = results[name]
            if all(status == "done" for status, _, _, _ in entries):
                dests = ", ".join(f"{filename} -> {allocation}\\" for _, filename, allocation, _ in entries)
                self.log_panel.log(f"[OK]    {name}  -> {dests}")
            else:
                self.log_panel.log(f"[ISSUE] {name}")
                for status, filename, allocation, detail in entries:
                    if status == "done":
                        self.log_panel.log(f"          - moved OK: {filename} -> {allocation}\\")
                    elif status == "not_found":
                        self.log_panel.log(f"          - not found: {detail}")
                    elif status == "skipped":
                        self.log_panel.log(f"          - skipped ({filename}): {detail}")
                    else:
                        self.log_panel.log(f"          - error ({filename}): {detail}")

    def _finish(self, summary):
        self.log_panel.log(summary)
        self.log_panel.set_status(summary)
        self.run_button.configure(state="normal")


# ------------------------------------------------------------------
# Tab 5: PDF Searchable Check
# ------------------------------------------------------------------

def analyze_pdf_searchability(path):
    """Per-page searchable-vs-scanned check for one PDF: tries to extract
    each page's own text layer (via PyMuPDF -- a compiled C extension,
    much faster per page than pure-Python pypdf, which matters when
    scanning a whole folder). Real word content found -- the page is
    text-searchable. Nothing (or only a couple of stray characters) found --
    the page is very likely a scanned image with no text layer at all, and
    would need OCR to be read (no OCR is performed here -- this only
    detects and flags which pages need it). Returns a list of one dict per
    page: file, page, searchable ("Yes"/"Partial"/"No"), word_count, note."""
    file_name = os.path.basename(path)
    try:
        doc = fitz.open(path)
    except Exception as exc:
        return [{"file": file_name, "page": "", "searchable": "", "word_count": "",
                  "note": f"could not open PDF: {exc}"}]

    rows = []
    with doc:
        for i, page in enumerate(doc, start=1):
            try:
                text = page.get_text()
            except Exception as exc:
                rows.append({"file": file_name, "page": i, "searchable": "",
                             "word_count": "", "note": f"error reading page: {exc}"})
                continue
            word_count = len(text.split())
            if word_count == 0:
                searchable, note = "No", "no text layer found -- likely a scanned image, needs OCR"
            elif word_count < 5:
                searchable, note = "Partial", "very little text found -- verify manually"
            else:
                searchable, note = "Yes", ""
            rows.append({"file": file_name, "page": i, "searchable": searchable,
                         "word_count": word_count, "note": note})
    return rows


def summarize_pdf_rows_by_file(rows):
    """Collapse the per-page rows into one row per file: total pages and
    the searchable/partial/scanned breakdown -- the file-wise view asked
    for alongside the page-wise detail."""
    order = []
    counts = {}
    for row in rows:
        file_name = row.get("file", "")
        if file_name not in counts:
            counts[file_name] = {"total": 0, "yes": 0, "partial": 0, "no": 0, "errors": 0}
            order.append(file_name)
        c = counts[file_name]
        searchable = row.get("searchable")
        if searchable == "":
            c["errors"] += 1
            continue
        c["total"] += 1
        if searchable == "Yes":
            c["yes"] += 1
        elif searchable == "Partial":
            c["partial"] += 1
        elif searchable == "No":
            c["no"] += 1

    summary = []
    for file_name in order:
        c = counts[file_name]
        summary.append({
            "file": file_name, "total_pages": c["total"], "searchable_pages": c["yes"],
            "partial_pages": c["partial"], "scanned_pages": c["no"],
            "note": "could not open/read file" if c["errors"] and c["total"] == 0 else "",
        })
    return summary


class PdfSearchableTab(ttk.Frame):
    COLUMNS = ("file", "page", "searchable", "word_count", "note")
    HEADINGS = {
        "file": "File", "page": "Page No", "searchable": "Searchable",
        "word_count": "Word Count", "note": "Note",
    }

    def __init__(self, parent):
        super().__init__(parent)
        self.folder_var = tk.StringVar()
        self.recursive_var = tk.BooleanVar(value=False)
        self._build_layout()

    def _build_layout(self):
        pad = {"padx": 10, "pady": 6}

        folder_frame = ttk.Frame(self)
        folder_frame.pack(fill="x", **pad)
        folder_frame.columnconfigure(1, weight=1)
        ttk.Label(folder_frame, text="Folder with PDF files:").grid(row=0, column=0, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.folder_var).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(folder_frame, text="Browse...", command=self._pick_folder).grid(row=0, column=2)
        ttk.Checkbutton(folder_frame, text="Include subfolders", variable=self.recursive_var).grid(row=1, column=1, sticky="w", pady=(6, 0))

        scan_buttons = ttk.Frame(self)
        scan_buttons.pack(pady=(2, 8))
        ttk.Button(scan_buttons, text="Scan Files", command=self._scan_clicked, style="Accent.TButton").pack(side="left", padx=(0, 8))
        ttk.Button(scan_buttons, text="Export Report to Excel...", command=self._export_clicked).pack(side="left", padx=(0, 8))
        ttk.Button(scan_buttons, text="Reset", command=self._reset_clicked).pack(side="left")

        table_frame = ttk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10)
        self.tree = ttk.Treeview(table_frame, columns=self.COLUMNS, show="headings", height=10)
        style_tree_rows(self.tree)
        for col in self.COLUMNS:
            self.tree.heading(col, text=self.HEADINGS[col])
            self.tree.column(col, width=100, anchor="w", stretch=True)
        self.tree.column("file", width=220)
        self.tree.column("note", width=260)
        tree_scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        tree_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        ttk.Label(self, text="Log:").pack(anchor="w", padx=10, pady=(10, 0))
        self.log_panel = LogPanel(self, height=8)
        self.log_panel.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _pick_folder(self):
        folder = filedialog.askdirectory(title="Select folder with PDF files")
        if folder:
            self.folder_var.set(folder)

    def _reset_clicked(self):
        self.folder_var.set("")
        self.recursive_var.set(False)
        for row_id in self.tree.get_children():
            self.tree.delete(row_id)
        self.log_panel.clear()
        self.log_panel.set_status("Ready.")

    @staticmethod
    def _collect_files(folder, recursive):
        files = []
        if recursive:
            for root_dir, _dirs, names in os.walk(folder):
                for name in names:
                    if name.lower().endswith(".pdf") and not name.startswith("~$"):
                        files.append(os.path.join(root_dir, name))
        else:
            with os.scandir(folder) as entries:
                for entry in entries:
                    if entry.is_file() and entry.name.lower().endswith(".pdf") and not entry.name.startswith("~$"):
                        files.append(entry.path)
        return files

    def _scan_clicked(self):
        if fitz is None:
            messagebox.showerror(
                "PDF Searchable Check",
                "This tool requires 'PyMuPDF'. Install it with: pip install PyMuPDF",
            )
            return

        folder = self.folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("PDF Searchable Check", "Please choose a valid folder.")
            return

        files = self._collect_files(folder, self.recursive_var.get())
        if not files:
            messagebox.showinfo("PDF Searchable Check", "No PDF files found in that folder.")
            return

        for row_id in self.tree.get_children():
            self.tree.delete(row_id)
        self.log_panel.clear()
        self.log_panel.set_status(f"Scanning {len(files)} file(s)...")

        thread = threading.Thread(target=self._scan_worker, args=(files,), daemon=True)
        thread.start()

    def _export_clicked(self):
        row_ids = self.tree.get_children()
        if not row_ids:
            messagebox.showinfo("PDF Searchable Check", "Nothing to export yet -- run Scan Files first.")
            return

        path = filedialog.asksaveasfilename(
            title="Save report as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="pdf_searchable_check_report.xlsx",
        )
        if not path:
            return

        page_rows = [dict(zip(self.COLUMNS, self.tree.item(row_id, "values"))) for row_id in row_ids]
        file_summary_cols = ["file", "total_pages", "searchable_pages", "partial_pages", "scanned_pages", "note"]
        file_summary_headings = ["File", "Total Pages", "Searchable Pages", "Partial Pages", "Scanned Pages", "Note"]

        try:
            wb = openpyxl.Workbook()
            ws_summary = wb.active
            ws_summary.title = "File Summary"
            ws_summary.append(file_summary_headings)
            for file_row in summarize_pdf_rows_by_file(page_rows):
                ws_summary.append([file_row.get(c, "") for c in file_summary_cols])
            autofit_columns(ws_summary)

            ws_detail = wb.create_sheet("Page Detail")
            ws_detail.append([self.HEADINGS[col] for col in self.COLUMNS])
            for row_id in row_ids:
                ws_detail.append(self.tree.item(row_id, "values"))
            autofit_columns(ws_detail)

            wb.save(path)
        except Exception as exc:
            messagebox.showerror("PDF Searchable Check", f"Could not save report: {exc}")
            return

        messagebox.showinfo("PDF Searchable Check", f"Report saved to:\n{path}")

    def _scan_worker(self, files):
        total = len(files)
        done_count = 0
        with ThreadPoolExecutor(max_workers=WORKER_THREADS) as pool:
            futures = [pool.submit(analyze_pdf_searchability, path) for path in files]
            for future in as_completed(futures):
                for row in future.result():
                    self.after(0, self._add_row, row)
                done_count += 1
                self.after(0, self.log_panel.set_status, f"Scanning file {done_count} out of {total}...")
        self.after(0, self.log_panel.set_status, f"Scan complete -- {len(files)} file(s).")
        self.after(0, self.log_panel.log, f"Scan complete. {len(files)} file(s) scanned.")

    def _add_row(self, row):
        self.tree.insert("", "end", values=[row.get(col, "") for col in self.COLUMNS], tags=(next_row_tag(self.tree),))
        if row.get("note"):
            self.log_panel.log(f"{row['file']}: page {row.get('page', '')} -- {row['note']}")


# ------------------------------------------------------------------
# Tab 6: Merge Files
# ------------------------------------------------------------------

MERGE_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".csv")


def collect_files_by_extensions(folder, recursive, extensions):
    files = []
    if recursive:
        for root_dir, _dirs, names in os.walk(folder):
            for name in names:
                if name.lower().endswith(extensions) and not name.startswith("~$"):
                    files.append(os.path.join(root_dir, name))
    else:
        with os.scandir(folder) as entries:
            for entry in entries:
                if entry.is_file() and entry.name.lower().endswith(extensions) and not entry.name.startswith("~$"):
                    files.append(entry.path)
    return files


def merge_rows_positional(paths, read_results, results, log_fn):
    """Append mode: the first readable file's header row becomes the output
    header, and every file's data rows are pasted underneath unchanged --
    column position for column position -- with no attempt to line up
    columns by name across files."""
    master_header = None
    output_rows = []
    for path in paths:
        filename = os.path.basename(path)
        rows = read_results.get(path)
        if rows is None:
            continue  # read failed -- already logged as an error

        if not rows:
            results[filename] = ("done", "0 row(s) (file was empty)")
            continue

        header, data_rows = rows[0], rows[1:]
        if master_header is None:
            master_header = list(header)
        elif [cell_display(c) for c in header] != [cell_display(c) for c in master_header]:
            log_fn(f"NOTE: {filename} has different column headers -- rows still appended under the first file's headers.")

        for row in data_rows:
            output_rows.append(list(row) + [filename])

        results[filename] = ("done", f"{len(data_rows)} row(s)")
        log_fn(f"MERGED: {filename}  ({len(data_rows)} row(s))")

    return master_header, output_rows


def build_header_union(headers_by_file):
    """Returns an ordered list of (key, display) pairs covering every header
    cell seen across all files -- key is the trimmed, lower-cased text used
    for matching; display is the trimmed text from the file it first
    appeared in, in first-seen order (so files with extra columns just add
    theirs at the end)."""
    order = []
    display_by_key = {}
    for header in headers_by_file:
        for cell in header:
            display = cell_display(cell).strip()
            if not display:
                continue
            key = display.lower()
            if key not in display_by_key:
                display_by_key[key] = display
                order.append(key)
    return [(key, display_by_key[key]) for key in order]


def merge_rows_by_header(paths, read_results, results, log_fn):
    """Match-headers mode: trims/cleans every file's header text and matches
    columns by name (case-insensitive) so the same field lines up in the
    same output column regardless of the order it appears in each source
    file. Any header only some files have still gets its own output column
    -- blank for files that don't have it."""
    headers = [read_results[path][0] for path in paths if read_results.get(path)]
    union = build_header_union(headers)
    key_order = [key for key, _display in union]
    master_header = [display for _key, display in union]

    output_rows = []
    for path in paths:
        filename = os.path.basename(path)
        rows = read_results.get(path)
        if rows is None:
            continue  # read failed -- already logged as an error

        if not rows:
            results[filename] = ("done", "0 row(s) (file was empty)")
            continue

        header, data_rows = rows[0], rows[1:]
        column_by_key = {}
        for idx, cell in enumerate(header):
            display = cell_display(cell).strip()
            if display:
                column_by_key.setdefault(display.lower(), idx)

        for row in data_rows:
            new_row = []
            for key in key_order:
                idx = column_by_key.get(key)
                new_row.append(row[idx] if idx is not None and idx < len(row) else "")
            new_row.append(filename)
            output_rows.append(new_row)

        results[filename] = ("done", f"{len(data_rows)} row(s)")
        log_fn(f"MERGED: {filename}  ({len(data_rows)} row(s))")

    return master_header, output_rows


class MergeFilesTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.folder_var = tk.StringVar()
        self.output_dir_var = tk.StringVar()
        self.recursive_var = tk.BooleanVar(value=False)
        self.merge_mode_var = tk.StringVar(value="append")
        self._build_layout()

    def _build_layout(self):
        pad = {"padx": 10, "pady": 6}

        folder_frame = ttk.Frame(self)
        folder_frame.pack(fill="x", **pad)
        folder_frame.columnconfigure(1, weight=1)

        ttk.Label(folder_frame, text="Folder with files to merge:").grid(row=0, column=0, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.folder_var).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(folder_frame, text="Browse...", command=self._pick_folder).grid(row=0, column=2)

        ttk.Label(folder_frame, text="Output folder:").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(folder_frame, textvariable=self.output_dir_var).grid(row=1, column=1, sticky="ew", padx=6, pady=(8, 0))
        ttk.Button(folder_frame, text="Browse...", command=self._pick_output).grid(row=1, column=2, pady=(8, 0))

        ttk.Checkbutton(folder_frame, text="Include subfolders", variable=self.recursive_var).grid(row=2, column=1, sticky="w", pady=(6, 0))

        ttk.Label(
            self,
            text=(
                "Merges every .xlsx / .xlsm / .xls / .csv file in the folder into one combined file per "
                "extension (e.g. all .xlsx files merge together, all .csv files merge together), with a "
                "\"Source File\" column added so every row can be traced back to its original file."
            ),
            wraplength=680, justify="left",
        ).pack(anchor="w", padx=10, pady=(4, 0))

        mode_frame = ttk.LabelFrame(self, text="Merge mode")
        mode_frame.pack(fill="x", padx=10, pady=(8, 0))
        ttk.Radiobutton(
            mode_frame, text="Append as-is", variable=self.merge_mode_var, value="append",
        ).grid(row=0, column=0, sticky="w", padx=8, pady=(6, 0))
        ttk.Label(
            mode_frame,
            text=(
                "Uses the first file's header row; every other file's data is pasted underneath unchanged, "
                "column position for column position (e.g. file 1 = Name/ID, file 2 = ID/Name -- file 2's "
                "data still lands under Name/ID, not realigned)."
            ),
            wraplength=660, justify="left", foreground="#555555",
        ).grid(row=1, column=0, sticky="w", padx=28, pady=(0, 6))

        ttk.Radiobutton(
            mode_frame, text="Match headers", variable=self.merge_mode_var, value="match",
        ).grid(row=2, column=0, sticky="w", padx=8)
        ttk.Label(
            mode_frame,
            text=(
                "Trims/cleans each file's header text and matches columns by name (case-insensitive) across "
                "all files, so ID lines up with ID and Name with Name no matter the column order. Any header "
                "only some files have still gets added as its own column."
            ),
            wraplength=660, justify="left", foreground="#555555",
        ).grid(row=3, column=0, sticky="w", padx=28, pady=(0, 6))

        button_row = ttk.Frame(self)
        button_row.pack(pady=(6, 8))
        self.run_button = ttk.Button(button_row, text="Merge Files", command=self._run_clicked, style="Accent.TButton")
        self.run_button.pack(side="left", padx=(0, 8))
        ttk.Button(button_row, text="Reset", command=self._reset_clicked).pack(side="left")

        ttk.Label(self, text="Log:").pack(anchor="w", padx=10)
        self.log_panel = LogPanel(self, height=10)
        self.log_panel.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _reset_clicked(self):
        self.folder_var.set("")
        self.output_dir_var.set("")
        self.recursive_var.set(False)
        self.merge_mode_var.set("append")
        self.log_panel.clear()
        self.log_panel.set_status("Ready.")
        self.run_button.configure(state="normal")

    def _pick_folder(self):
        folder = filedialog.askdirectory(title="Select folder with files to merge")
        if folder:
            self.folder_var.set(folder)

    def _pick_output(self):
        folder = filedialog.askdirectory(title="Select output folder")
        if folder:
            self.output_dir_var.set(folder)

    def _run_clicked(self):
        folder = self.folder_var.get().strip()
        out_dir = self.output_dir_var.get().strip()

        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Merge Files", "Please choose a valid folder.")
            return
        if not out_dir:
            messagebox.showerror("Merge Files", "Please choose an output folder.")
            return

        files = collect_files_by_extensions(folder, self.recursive_var.get(), MERGE_EXTENSIONS)
        if not files:
            messagebox.showinfo("Merge Files", "No .xlsx/.xlsm/.xls/.csv files found in that folder.")
            return

        groups = {}
        for path in files:
            ext = os.path.splitext(path)[1].lower()
            groups.setdefault(ext, []).append(path)

        os.makedirs(out_dir, exist_ok=True)
        self.log_panel.clear()
        self.run_button.configure(state="disabled")
        self.log_panel.set_status("Working...")

        thread = threading.Thread(target=self._process, args=(groups, out_dir, self.merge_mode_var.get()), daemon=True)
        thread.start()

    def _process(self, groups, out_dir, mode):
        total_files = sum(len(paths) for paths in groups.values())
        done_count = 0
        lock = threading.Lock()
        results = {}
        output_files = []

        for ext in sorted(groups):
            paths = sorted(groups[ext])
            self.after(0, self.log_panel.set_status, f"Reading {len(paths)} {ext} file(s)...")

            read_results = {}
            with ThreadPoolExecutor(max_workers=WORKER_THREADS) as pool:
                futures = {pool.submit(load_rows, path): path for path in paths}
                for future in as_completed(futures):
                    path = futures[future]
                    filename = os.path.basename(path)
                    try:
                        read_results[path] = future.result()
                    except Exception as exc:
                        results[filename] = ("error", f"could not read: {exc}")
                        read_results[path] = None
                    with lock:
                        done_count += 1
                        progress = done_count
                    self.after(0, self.log_panel.set_status, f"Processing file {progress} out of {total_files}...")

            out_name = f"merged_{ext.lstrip('.')}.xlsx"
            out_path = os.path.join(out_dir, out_name)
            log_fn = lambda msg: self.after(0, self.log_panel.log, msg)

            if mode == "match":
                master_header, output_rows = merge_rows_by_header(paths, read_results, results, log_fn)
            else:
                master_header, output_rows = merge_rows_positional(paths, read_results, results, log_fn)

            if master_header:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Merged"
                ws.append(list(master_header) + ["Source File"])
                for row in output_rows:
                    ws.append(row)
                autofit_columns(ws)
                try:
                    wb.save(out_path)
                    output_files.append(out_path)
                    self.after(0, self.log_panel.log, f"CREATED: {out_name}")
                except Exception as exc:
                    self.after(0, self.log_panel.log, f"ERROR saving {out_name}: {exc}")
            else:
                self.after(0, self.log_panel.log, f"SKIPPED: no readable {ext} files to merge.")

        self.after(0, self._log_final_status, results)
        succeeded = sum(1 for status, _ in results.values() if status == "done")
        summary = f"Done. {succeeded}/{len(results)} file(s) merged into {len(output_files)} output file(s) in {out_dir}."
        self.after(0, self._finish, summary)

    def _log_final_status(self, results):
        self.log_panel.log("")
        self.log_panel.log("===== STATUS =====")
        for filename in sorted(results):
            status, detail = results[filename]
            if status == "done":
                self.log_panel.log(f"[OK]    {filename}  -> {detail}")
            else:
                self.log_panel.log(f"[ISSUE] {filename}  -> {detail}")

    def _finish(self, summary):
        self.log_panel.log(summary)
        self.log_panel.set_status(summary)
        self.run_button.configure(state="normal")


# ------------------------------------------------------------------
# Tab 7: Text from PDF
# ------------------------------------------------------------------

def resolve_pdf_files(path):
    """Return list of absolute PDF paths from a file or folder, or raise ValueError."""
    if os.path.isfile(path):
        if not path.lower().endswith(".pdf"):
            raise ValueError(f"'{path}' is not a PDF file.")
        return [os.path.abspath(path)]
    elif os.path.isdir(path):
        files = [
            os.path.abspath(os.path.join(path, f))
            for f in os.listdir(path)
            if f.lower().endswith(".pdf")
        ]
        if not files:
            raise ValueError(f"No PDF files found in '{path}'.")
        return sorted(files)
    else:
        raise ValueError(f"Path does not exist: '{path}'")


def check_ocr_deps():
    """Return an error message if pdf2image/pytesseract aren't importable, else None."""
    missing = []
    try:
        import pdf2image  # noqa: F401
    except ImportError:
        missing.append("pdf2image")
    try:
        import pytesseract  # noqa: F401
    except ImportError:
        missing.append("pytesseract")

    if missing:
        return (
            f"OCR requires these packages: {', '.join(missing)}. Install with: "
            "pip install pdf2image pytesseract -- and make sure the Tesseract OCR "
            "binary is installed separately (https://github.com/UB-Mannheim/tesseract/wiki)."
        )
    return None


def ocr_page_image(pdf_path, page_num, lang, dpi):
    """Render a single PDF page to an image and return OCR text."""
    from pdf2image import convert_from_path
    import pytesseract

    images = convert_from_path(pdf_path, dpi=dpi, first_page=page_num, last_page=page_num)
    if not images:
        return ""
    return pytesseract.image_to_string(images[0], lang=lang)


MIN_PAGES_PER_CHUNK = 25


def plan_pdf_chunks(pdf_files, num_workers):
    """
    Open each PDF just far enough to get its page count, then split its
    pages into up to num_workers contiguous chunks so one large file can be
    spread across the whole thread pool instead of tying up a single
    worker for its entire length. Small files (fewer than
    MIN_PAGES_PER_CHUNK pages) stay as one chunk.

    Returns a list of dicts: file_index, pdf_path, total_pages, page_start,
    page_end, n_chunks, chunk_no.
    """
    chunks = []
    for file_index, pdf_path in enumerate(pdf_files, start=1):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
        except Exception:
            total_pages = 0

        n_splits = min(num_workers, max(1, total_pages // MIN_PAGES_PER_CHUNK)) if total_pages else 1
        base, rem = divmod(total_pages, n_splits)
        start = 0
        for i in range(n_splits):
            size = base + (1 if i < rem else 0)
            end = start + size
            chunks.append({
                "file_index": file_index, "pdf_path": pdf_path, "total_pages": total_pages,
                "page_start": start, "page_end": end, "n_chunks": n_splits, "chunk_no": i + 1,
            })
            start = end
    return chunks


def extract_pdf_text(pdf_path, ocr_mode, lang, dpi, preserve_layout=False,
                      page_start=0, page_end=None, on_page=None):
    """
    Extract text lines from a page range [page_start, page_end) of one PDF
    (page_end=None means "to the end"). A whole file is just the range
    covering all of its pages.

    ocr_mode:
      'never'  -- pdfplumber only
      'auto'   -- pdfplumber first; OCR fallback for pages with no text
      'always' -- OCR every page (skips pdfplumber)

    preserve_layout: when True, keeps the PDF's original horizontal spacing
    (e.g. the gap between side-by-side columns) by using pdfplumber's
    layout mode and only trimming trailing whitespace; when False, each
    line is fully trimmed and words are joined with normal single spaces.

    on_page, when given, is called after every page (including skipped/
    errored ones) so callers can track live per-document progress.

    Returns (records, stats) where stats counts native/ocr/empty/error pages.
    """
    records = []
    stats = {"native": 0, "ocr": 0, "empty": 0, "error": 0}
    pdf_file = os.path.basename(pdf_path)

    with pdfplumber.open(pdf_path) as pdf:
        pages = pdf.pages[page_start:page_end]
        for offset, page in enumerate(pages):
            page_num = page_start + offset + 1
            try:
                text = ""
                method = "native"

                if ocr_mode != "always":
                    if preserve_layout:
                        text = page.extract_text(layout=True) or ""
                    else:
                        text = page.extract_text() or ""

                if not text.strip() and ocr_mode != "never":
                    text = ocr_page_image(pdf_path, page_num, lang, dpi)
                    method = "ocr"

                lines_added = 0
                for line in text.split("\n"):
                    line_text = line.rstrip() if preserve_layout and method == "native" else line.strip()
                    if line_text.strip():
                        records.append({
                            "file": pdf_file, "page": page_num,
                            "text": line_text, "method": method,
                        })
                        lines_added += 1

                if lines_added == 0:
                    stats["empty"] += 1
                elif method == "ocr":
                    stats["ocr"] += 1
                else:
                    stats["native"] += 1

            except Exception:
                stats["error"] += 1
            finally:
                if on_page is not None:
                    on_page()

    return records, stats


class TextFromPdfTab(ttk.Frame):
    COLUMNS = ("file", "page", "text", "method")
    HEADINGS = {
        "file": "File Name", "page": "Page Number",
        "text": "Extracted Text", "method": "Method",
    }

    def __init__(self, parent):
        super().__init__(parent)
        self.input_var = tk.StringVar()
        self.ocr_var = tk.StringVar(value="never")
        self.lang_var = tk.StringVar(value="eng")
        self.dpi_var = tk.StringVar(value="300")
        self.workers_var = tk.StringVar(value="4")
        self.layout_var = tk.BooleanVar(value=False)
        self.records = []
        self.progress_rows = {}
        self.file_progress = {}
        self._build_layout()

    def _build_layout(self):
        pad = {"padx": 10, "pady": 6}

        input_frame = ttk.Frame(self)
        input_frame.pack(fill="x", **pad)
        input_frame.columnconfigure(1, weight=1)
        ttk.Label(input_frame, text="PDF file or folder:").grid(row=0, column=0, sticky="w")
        ttk.Entry(input_frame, textvariable=self.input_var).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(input_frame, text="Browse File...", command=self._pick_file).grid(row=0, column=2)
        ttk.Button(input_frame, text="Browse Folder...", command=self._pick_folder).grid(row=0, column=3, padx=(4, 0))

        options_frame = ttk.Frame(self)
        options_frame.pack(fill="x", **pad)
        ttk.Label(options_frame, text="OCR:").pack(side="left")
        ttk.Combobox(
            options_frame, textvariable=self.ocr_var, state="readonly", width=8,
            values=("never", "auto", "always"),
        ).pack(side="left", padx=(4, 16))
        ttk.Label(options_frame, text="Language(s):").pack(side="left")
        ttk.Entry(options_frame, textvariable=self.lang_var, width=10).pack(side="left", padx=(4, 16))
        ttk.Label(options_frame, text="OCR DPI:").pack(side="left")
        ttk.Entry(options_frame, textvariable=self.dpi_var, width=6).pack(side="left", padx=(4, 16))
        ttk.Label(options_frame, text="Workers:").pack(side="left")
        ttk.Entry(options_frame, textvariable=self.workers_var, width=4).pack(side="left", padx=(4, 0))
        ttk.Checkbutton(
            options_frame, text="Preserve PDF spacing (layout)", variable=self.layout_var,
        ).pack(side="left", padx=(16, 0))

        ttk.Label(
            self,
            text=(
                "Extracts each page's text into one row per line of text. 'never' uses only the PDF's own "
                "text layer (fast); 'auto' falls back to Tesseract OCR for pages with no text layer "
                "(scanned pages); 'always' forces OCR on every page. OCR requires: pip install pdf2image "
                "pytesseract, plus the Tesseract binary installed separately. 'Preserve PDF spacing' keeps "
                "the original gaps between side-by-side columns (e.g. \"Student   Borrower\") instead of "
                "collapsing them to a single space; it only applies to native (non-OCR) text."
            ),
            wraplength=680, justify="left",
        ).pack(anchor="w", padx=10, pady=(0, 4))

        button_row = ttk.Frame(self)
        button_row.pack(pady=(2, 8))
        self.run_button = ttk.Button(button_row, text="Extract Text", command=self._run_clicked, style="Accent.TButton")
        self.run_button.pack(side="left", padx=(0, 8))
        ttk.Button(button_row, text="Reset", command=self._reset_clicked).pack(side="left")

        ttk.Label(self, text="Document progress:").pack(anchor="w", padx=10, pady=(4, 0))
        progress_frame = ttk.Frame(self)
        progress_frame.pack(fill="x", padx=10)
        self.progress_tree = ttk.Treeview(
            progress_frame, columns=("doc", "pages", "status"), show="headings", height=5,
        )
        style_tree_rows(self.progress_tree)
        self.progress_tree.heading("doc", text="File Name")
        self.progress_tree.heading("pages", text="Pages Extracted / Total")
        self.progress_tree.heading("status", text="Status")
        self.progress_tree.column("doc", width=260, anchor="w", stretch=True)
        self.progress_tree.column("pages", width=180, anchor="w")
        self.progress_tree.column("status", width=100, anchor="w")
        progress_scroll_y = ttk.Scrollbar(progress_frame, orient="vertical", command=self.progress_tree.yview)
        self.progress_tree.configure(yscrollcommand=progress_scroll_y.set)
        self.progress_tree.pack(side="left", fill="x", expand=True)
        progress_scroll_y.pack(side="left", fill="y")

        self.preview_label = ttk.Label(self, text="Preview (first 50 rows):")
        self.preview_label.pack(anchor="w", padx=10, pady=(4, 0))

        table_frame = ttk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10)
        self.tree = ttk.Treeview(table_frame, columns=self.COLUMNS, show="headings", height=10)
        style_tree_rows(self.tree)
        for col in self.COLUMNS:
            self.tree.heading(col, text=self.HEADINGS[col])
            self.tree.column(col, width=100, anchor="w", stretch=True)
        self.tree.column("file", width=160)
        self.tree.column("text", width=340)
        tree_scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        tree_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        ttk.Button(self, text="Export Results to Excel...", command=self._export_clicked).pack(pady=(6, 0))

        ttk.Label(self, text="Log:").pack(anchor="w", padx=10, pady=(10, 0))
        self.log_panel = LogPanel(self, height=8)
        self.log_panel.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _pick_file(self):
        path = filedialog.askopenfilename(
            title="Select a PDF file", filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if path:
            self.input_var.set(path)

    def _pick_folder(self):
        folder = filedialog.askdirectory(title="Select a folder of PDF files")
        if folder:
            self.input_var.set(folder)

    def _reset_clicked(self):
        self.input_var.set("")
        self.ocr_var.set("never")
        self.lang_var.set("eng")
        self.dpi_var.set("300")
        self.workers_var.set("4")
        self.layout_var.set(False)
        for row_id in self.tree.get_children():
            self.tree.delete(row_id)
        for row_id in self.progress_tree.get_children():
            self.progress_tree.delete(row_id)
        self.records = []
        self.progress_rows = {}
        self.file_progress = {}
        self.preview_label.configure(text="Preview (first 50 rows):")
        self.log_panel.clear()
        self.log_panel.set_status("Ready.")
        self.run_button.configure(state="normal")

    def _run_clicked(self):
        if pdfplumber is None:
            messagebox.showerror(
                "Text from PDF", "This tool requires 'pdfplumber'. Install it with: pip install pdfplumber",
            )
            return

        input_path = self.input_var.get().strip()
        if not input_path or not os.path.exists(input_path):
            messagebox.showerror("Text from PDF", "Please choose a valid PDF file or folder.")
            return

        ocr_mode = self.ocr_var.get()
        preserve_layout = self.layout_var.get()
        lang = self.lang_var.get().strip() or "eng"
        try:
            dpi = int(self.dpi_var.get().strip())
        except ValueError:
            messagebox.showerror("Text from PDF", "OCR DPI must be a whole number.")
            return
        try:
            workers = int(self.workers_var.get().strip())
            if workers < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("Text from PDF", "Workers must be a whole number of 1 or more.")
            return

        if ocr_mode != "never":
            error = check_ocr_deps()
            if error:
                messagebox.showerror("Text from PDF", error)
                return

        try:
            pdf_files = resolve_pdf_files(input_path)
        except ValueError as exc:
            messagebox.showerror("Text from PDF", str(exc))
            return

        for row_id in self.tree.get_children():
            self.tree.delete(row_id)
        for row_id in self.progress_tree.get_children():
            self.progress_tree.delete(row_id)
        self.records = []
        self.progress_rows = {}
        self.file_progress = {}
        self.preview_label.configure(text="Preview (first 50 rows):")
        self.log_panel.clear()
        self.run_button.configure(state="disabled")
        self.log_panel.set_status(f"Extracting text from {len(pdf_files)} file(s)...")

        thread = threading.Thread(
            target=self._process, args=(pdf_files, ocr_mode, lang, dpi, workers, preserve_layout), daemon=True,
        )
        thread.start()

    def _process(self, pdf_files, ocr_mode, lang, dpi, workers, preserve_layout):
        total = len(pdf_files)
        done_count = 0
        total_lines = 0
        displayed_count = 0

        self.after(0, self.log_panel.set_status, "Scanning page counts...")
        chunks = plan_pdf_chunks(pdf_files, workers)
        num_workers = min(workers, len(chunks)) if chunks else 1

        for file_index, path in enumerate(pdf_files, start=1):
            total_pages = next((c["total_pages"] for c in chunks if c["file_index"] == file_index), 0)
            self.after(0, self._init_progress_row, file_index, os.path.basename(path), total_pages)

        pending = {
            c["file_index"]: {"total": c["n_chunks"], "parts": {}} for c in chunks
        }

        self.after(0, self.log_panel.set_status, f"Extracting with {num_workers} thread(s) across {len(chunks)} chunk(s)...")

        with ThreadPoolExecutor(max_workers=num_workers) as pool:
            futures = {}
            for chunk in chunks:
                fi = chunk["file_index"]
                fut = pool.submit(
                    extract_pdf_text, chunk["pdf_path"], ocr_mode, lang, dpi, preserve_layout,
                    chunk["page_start"], chunk["page_end"],
                    lambda fi=fi: self.after(0, self._tick_progress, fi),
                )
                futures[fut] = chunk

            for future in as_completed(futures):
                chunk = futures[future]
                file_index = chunk["file_index"]
                path = chunk["pdf_path"]
                filename = os.path.basename(path)
                try:
                    records, stats = future.result()
                except Exception as exc:
                    self.after(0, self.log_panel.log, f"ERROR: {filename} -- {exc}")
                    records, stats = [], {"native": 0, "ocr": 0, "empty": 0, "error": 0}

                pending[file_index]["parts"][chunk["chunk_no"]] = (records, stats)
                if len(pending[file_index]["parts"]) != pending[file_index]["total"]:
                    continue

                merged_records = []
                merged_stats = {"native": 0, "ocr": 0, "empty": 0, "error": 0}
                for chunk_no in sorted(pending[file_index]["parts"]):
                    r, s = pending[file_index]["parts"][chunk_no]
                    merged_records.extend(r)
                    for k in merged_stats:
                        merged_stats[k] += s[k]

                self.records.extend(merged_records)
                for row in merged_records:
                    if displayed_count < 50:
                        self.after(0, self._add_row, row)
                        displayed_count += 1
                total_lines += len(merged_records)

                parts = [f"{len(merged_records)} line(s)"]
                if merged_stats["ocr"]:
                    parts.append(f"{merged_stats['ocr']} pg OCR")
                if merged_stats["empty"]:
                    parts.append(f"{merged_stats['empty']} pg empty")
                if merged_stats["error"]:
                    parts.append(f"{merged_stats['error']} pg error")
                self.after(0, self.log_panel.log, f"{filename}: {', '.join(parts)}")
                self.after(0, self._mark_file_done, file_index)

                done_count += 1
                self.after(0, self.log_panel.set_status, f"Processing file {done_count} out of {total}...")

        summary = f"Done. {total_lines} line(s) extracted from {total} file(s)."
        self.after(0, self.log_panel.log, summary)
        self.after(0, self._finish, summary, total_lines)

    def _init_progress_row(self, file_index, filename, total_pages):
        pages_label = f"0 / {total_pages}" if total_pages else "0 / ?"
        row_id = self.progress_tree.insert(
            "", "end", values=(filename, pages_label, "Queued"), tags=(next_row_tag(self.progress_tree),),
        )
        self.progress_rows[file_index] = row_id
        self.file_progress[file_index] = {"done": 0, "total": total_pages}

    def _tick_progress(self, file_index):
        row_id = self.progress_rows.get(file_index)
        if row_id is None:
            return
        prog = self.file_progress[file_index]
        prog["done"] += 1
        total_label = prog["total"] if prog["total"] else "?"
        self.progress_tree.set(row_id, "pages", f"{prog['done']} / {total_label}")
        self.progress_tree.set(row_id, "status", "Processing")

    def _mark_file_done(self, file_index):
        row_id = self.progress_rows.get(file_index)
        if row_id is None:
            return
        self.progress_tree.set(row_id, "status", "Done")

    def _add_row(self, row):
        self.tree.insert("", "end", values=[row.get(col, "") for col in self.COLUMNS], tags=(next_row_tag(self.tree),))

    def _export_clicked(self):
        if not self.records:
            messagebox.showinfo("Text from PDF", "Nothing to export yet -- run Extract Text first.")
            return

        path = filedialog.asksaveasfilename(
            title="Save extracted text as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="extracted_text.xlsx",
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Text"
            ws.append([self.HEADINGS[col] for col in self.COLUMNS])
            for row in self.records:
                ws.append([row.get(col, "") for col in self.COLUMNS])
            autofit_columns(ws)
            wb.save(path)
        except Exception as exc:
            messagebox.showerror("Text from PDF", f"Could not save file: {exc}")
            return

        messagebox.showinfo("Text from PDF", f"Extracted text saved to:\n{path}")

    def _finish(self, summary, total_lines):
        self.log_panel.set_status(summary)
        if total_lines > 50:
            self.preview_label.configure(text=f"Preview (first 50 of {total_lines} rows -- use Export for the rest):")
        self.run_button.configure(state="normal")


# ------------------------------------------------------------------
# App shell -- Chrome-style tab bar
# ------------------------------------------------------------------

PALETTE = {
    "shell": "#eef1f6",
    "card": "#ffffff",
    "accent": "#2f6fed",
    "accent_hover": "#1f5be0",
    "accent_pressed": "#1a4bbd",
    "accent_disabled": "#a9c0f5",
    "text": "#1f2430",
    "text_muted": "#5f6b7a",
    "border": "#d9dee6",
    "header_bg": "#f5f7fb",
    "row_alt": "#f5f8fd",
    "selection": "#dbe6fd",
}


def apply_browser_style(root):
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    p = PALETTE
    strip_bg = p["shell"]
    active_bg = p["card"]
    text_color = p["text"]
    text_active = p["text"]

    root.configure(bg=strip_bg)
    style.configure(".", font=("Segoe UI", 10), background=active_bg, foreground=text_color)
    style.configure("TFrame", background=active_bg)
    style.configure("TLabelframe", background=active_bg, bordercolor=p["border"], relief="solid")
    style.configure("TLabelframe.Label", background=active_bg, foreground=text_color, font=("Segoe UI", 10, "bold"))
    style.configure("TLabel", background=active_bg, foreground=text_color)
    style.configure("TCheckbutton", background=active_bg, foreground=text_color)
    style.configure("TRadiobutton", background=active_bg, foreground=text_color)

    style.configure(
        "TEntry",
        padding=6,
        relief="flat",
        fieldbackground=active_bg,
        bordercolor=p["border"],
        lightcolor=p["border"],
        darkcolor=p["border"],
    )
    style.map(
        "TEntry",
        bordercolor=[("focus", p["accent"])],
        lightcolor=[("focus", p["accent"])],
        darkcolor=[("focus", p["accent"])],
    )

    style.configure(
        "TCombobox",
        padding=5,
        relief="flat",
        fieldbackground=active_bg,
        background=active_bg,
        arrowcolor=p["text_muted"],
        bordercolor=p["border"],
    )
    style.map("TCombobox", bordercolor=[("focus", p["accent"])])

    # Secondary (default) buttons -- neutral, for Browse/Export actions.
    style.configure(
        "TButton",
        padding=(12, 7),
        relief="flat",
        borderwidth=0,
        background=p["header_bg"],
        foreground=text_color,
        font=("Segoe UI", 10),
    )
    style.map(
        "TButton",
        background=[("pressed", p["border"]), ("active", "#e9edf5")],
    )

    # Primary buttons -- one clear call-to-action per tab (Run/Apply/Scan/...).
    style.configure(
        "Accent.TButton",
        padding=(14, 8),
        relief="flat",
        borderwidth=0,
        background=p["accent"],
        foreground="#ffffff",
        font=("Segoe UI", 10, "bold"),
    )
    style.map(
        "Accent.TButton",
        background=[
            ("disabled", p["accent_disabled"]),
            ("pressed", p["accent_pressed"]),
            ("active", p["accent_hover"]),
        ],
        foreground=[("disabled", "#f0f4fe")],
    )

    style.configure("TNotebook", background=strip_bg, borderwidth=0, tabmargins=(8, 8, 8, 0))
    style.configure("TNotebook.Tab", background=strip_bg, foreground=text_color, padding=(16, 9), font=("Segoe UI", 10))
    style.map(
        "TNotebook.Tab",
        background=[("selected", active_bg)],
        foreground=[("selected", text_active)],
        expand=[("selected", (1, 1, 1, 0))],
    )

    style.configure(
        "Treeview",
        background=active_bg,
        fieldbackground=active_bg,
        foreground=text_color,
        bordercolor=p["border"],
        borderwidth=1,
        relief="solid",
        rowheight=24,
    )
    style.configure(
        "Treeview.Heading",
        background=p["header_bg"],
        foreground=p["text_muted"],
        font=("Segoe UI", 9, "bold"),
        relief="flat",
        borderwidth=1,
    )
    style.map(
        "Treeview.Heading",
        background=[("active", "#e9edf5")],
    )
    style.map(
        "Treeview",
        background=[("selected", p["selection"])],
        foreground=[("selected", text_color)],
    )

    for orient in ("Vertical", "Horizontal"):
        style.configure(
            f"{orient}.TScrollbar",
            background=p["header_bg"],
            troughcolor=active_bg,
            bordercolor=active_bg,
            arrowcolor=p["text_muted"],
            relief="flat",
        )
        style.map(f"{orient}.TScrollbar", background=[("active", p["border"])])


def style_tree_rows(tree):
    """Configures the alternating-row tags shared by every results table."""
    tree.tag_configure("evenrow", background=PALETTE["card"])
    tree.tag_configure("oddrow", background=PALETTE["row_alt"])


def next_row_tag(tree):
    return "oddrow" if len(tree.get_children()) % 2 else "evenrow"


class BusinessToolsApp:
    def __init__(self, root):
        self.root = root
        root.title("Business Tools")
        root.geometry("860x700")
        root.minsize(680, 540)
        apply_browser_style(root)

        header = tk.Frame(root, bg=PALETTE["accent"])
        header.pack(fill="x", side="top")
        tk.Label(
            header, text="Business Tools", bg=PALETTE["accent"], fg="#ffffff",
            font=("Segoe UI", 16, "bold"), anchor="w",
        ).pack(side="left", padx=(16, 6), pady=(12, 12))
        tk.Label(
            header, text="File, spreadsheet & PDF utilities",
            bg=PALETTE["accent"], fg="#dbe6fd",
            font=("Segoe UI", 10), anchor="w",
        ).pack(side="left", padx=(0, 16), pady=(12, 12))

        notebook = ttk.Notebook(root)
        notebook.pack(fill="both", expand=True, padx=8, pady=8)

        notebook.add(FileMoverTab(notebook), text="  \U0001F4C1  File Mover / Copier  ")
        notebook.add(SplitFilesTab(notebook), text="  ✂  Split Files  ")
        notebook.add(SheetInspectorTab(notebook), text="  \U0001F50E  Sheet Inspector  ")
        notebook.add(AllocateFilesTab(notebook), text="  \U0001F5C2  Allocate & Move  ")
        notebook.add(PdfSearchableTab(notebook), text="  \U0001F50D  PDF Searchable Check  ")
        notebook.add(MergeFilesTab(notebook), text="  \U0001F517  Merge Files  ")
        notebook.add(TextFromPdfTab(notebook), text="  \U0001F4DD  Text from PDF  ")


if __name__ == "__main__":
    root = tk.Tk()
    BusinessToolsApp(root)
    root.mainloop()
