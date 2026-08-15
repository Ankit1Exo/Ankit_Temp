"""
260814 AM combine contribution files.py

Combines contribution data from a folder of mixed-format spreadsheets.

WHAT IT DOES
    1. Walks an input folder (recursively by default) and finds every spreadsheet
       (.xlsx .xlsm .xls .xlsb .ods .csv .tsv and template variants).
    2. Opens EVERY sheet in each workbook, including hidden and very-hidden
       sheets. Hidden rows and hidden columns are read as normal data, so
       nothing needs to be unhidden -- the source files are never modified.
    3. Locates the header row (searched in the first 25 rows, so title/logo
       rows above the header are tolerated) and matches these columns:
           EmployeeIdentifier, ContributionDate, ContributionDescription,
           ContributionAmount, PlanName
       Matching ignores case, spaces, underscores and column order
       ("employee identifier", "EMPLOYEE_IDENTIFIER", "Employee-Identifier"
       all match).
    4. A sheet is combined when at least MIN_MATCHES (default 2) of those
       columns are present. When a sheet qualifies, EVERY column on it is
       carried through: the five target columns first, then all additional
       columns appended at the end, pooled across all files.
    5. Any file where no sheet reaches the threshold is excluded. Excluded file
       names and the reason are written to a second sheet, "Excluded Files",
       inside the workbook -- and also to a standalone CSV and a text summary.
    6. Writes one combined CSV, then XLSX output: a single workbook if the data
       fits Excel's 1,048,576-row limit, otherwise split into numbered parts.

DATA HANDLING NOTE
    Output contains employee identifiers (SSNs) and contribution amounts.
    Write the output to an access-controlled location, not a local desktop or
    Downloads folder. The audit log records file names, column names and row
    counts only -- never row-level data.

USAGE
    python "260814 AM combine contribution files.py" -i "<input folder>"
    python "260814 AM combine contribution files.py" -i "<in>" -o "<out>"

    -i / --input            Folder containing the source files (required)
    -o / --output           Output folder (default: <input>\\_combined)
    --no-recursive          Only the top-level input folder, no subfolders
    --min-matches N         Target columns required to accept a sheet (default 2)
    --targets-only          Output only the 5 target columns, dropping extras
    --keep-empty-columns    Keep additional columns that turned out fully blank
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
from decimal import Decimal
from itertools import chain
from pathlib import Path

# --------------------------------------------------------------------------
# CONFIGURATION
# --------------------------------------------------------------------------

# Canonical output columns, keyed by their normalised form.
TARGETS: dict[str, str] = {
    "employeeidentifier": "EmployeeIdentifier",
    "contributiondate": "ContributionDate",
    "contributiondescription": "ContributionDescription",
    "contributionamount": "ContributionAmount",
    "planname": "PlanName",
}
CANONICAL_ORDER = [
    "EmployeeIdentifier",
    "ContributionDate",
    "ContributionDescription",
    "ContributionAmount",
    "PlanName",
]

# Extra header spellings that should map onto a canonical column. Keys must be
# already-normalised (lowercase, alphanumeric only). Empty by default so the
# script never guesses; add entries here if your files use other names, e.g.
#   "ssn": "EmployeeIdentifier",
#   "socialsecuritynumber": "EmployeeIdentifier",
#   "paydate": "ContributionDate",
ALIASES: dict[str, str] = {}

MIN_MATCHES = 2             # how many target columns a sheet must have
HEADER_SCAN_ROWS = 25       # how deep to search for the header row
DATE_FORMAT = "%m/%d/%Y"    # matches the source files' displayed format
DATETIME_FORMAT = "%m/%d/%Y %H:%M:%S"

EXCEL_MAX_DATA_ROWS = 1_048_575         # 1,048,576 minus the header row
EXCEL_MAX_COLUMNS = 16_384              # Excel's hard column limit
SERIAL_DATE_MIN, SERIAL_DATE_MAX = 20000, 80000   # ~1954 to ~2119

PROVENANCE = ["SourceFile", "SourceSheet", "SheetVisibility", "HeaderRow", "SourceRow"]
EXCLUDED_SHEET_TITLE = "Excluded Files"
EXCLUDED_SHEET_HEADER = ["#", "ExcludedFile", "Reason"]

SPREADSHEET_EXTS = {
    ".xlsx", ".xlsm", ".xltx", ".xltm",
    ".xls", ".xlt",
    ".xlsb",
    ".ods",
    ".csv", ".tsv",
    # ".txt" is deliberately NOT included: this script's own summary is a .txt
    # and its "Match rule: at least 2 of EmployeeIdentifier, ContributionDate,
    # ..." line parses as a valid header row. Add it back only if you really
    # have .txt data files, and keep the output folder outside the input tree.
}

# Output naming. Files whose stem starts with OUTPUT_STEM, and anything inside a
# folder named DEFAULT_OUTPUT_DIRNAME, are skipped on input so repeat runs never
# re-ingest their own results.
OUTPUT_STEM = "260814 AM combined contributions"
DEFAULT_OUTPUT_DIRNAME = "_combined"

SSML_NS = "urn:schemas-microsoft-com:office:spreadsheet"

# Keys used in the staging file for provenance, kept short to bound its size.
STAGE_KEYS = ["_f", "_s", "_v", "_h", "_r"]


# --------------------------------------------------------------------------
# VALUE / HEADER NORMALISATION
# --------------------------------------------------------------------------

_non_alnum = re.compile(r"[^0-9a-z]+")


def normalise(value) -> str:
    """Reduce a header cell to a comparable key: lowercase, alphanumeric only."""
    if value is None:
        return ""
    return _non_alnum.sub("", str(value).strip().lower())


def to_text(value) -> str:
    """Convert any cell value to a string without corrupting it.

    Floats that are whole numbers become plain integers (so an ID read as
    123456789.0 is written as 123456789, not 123456789.0). Datetimes at
    midnight lose the meaningless 00:00:00.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, dt.datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.strftime(DATE_FORMAT)
        return value.strftime(DATETIME_FORMAT)
    if isinstance(value, dt.date):
        return value.strftime(DATE_FORMAT)
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):   # NaN / inf
            return ""
        if value.is_integer():
            return str(int(value))
        return repr(value)
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (bytes, bytearray)):
        return value.decode("utf-8", errors="replace").strip()
    text = str(value).strip()
    return "" if text.lower() in ("nan", "nat", "none") else text


def coerce_serial_date(text: str) -> tuple[str, bool]:
    """Turn a bare Excel date serial in the date column into a real date.

    Some engines (and some source files) hand back the raw serial number for a
    date cell. Only applied to the ContributionDate column, and only inside a
    plausible year range, so real numbers are left alone.
    """
    if not text:
        return text, False
    try:
        number = float(text)
    except ValueError:
        return text, False
    if not (SERIAL_DATE_MIN <= number <= SERIAL_DATE_MAX):
        return text, False
    try:
        base = dt.datetime(1899, 12, 30)      # Excel's 1900 system, leap-bug aware
        return (base + dt.timedelta(days=number)).strftime(DATE_FORMAT), True
    except (OverflowError, ValueError):
        return text, False


# --------------------------------------------------------------------------
# FORMAT DETECTION
# --------------------------------------------------------------------------

def sniff(path: Path) -> str:
    """Identify a file by its real content, not its extension.

    Returns one of: zip, ole, xml, html, text, empty, unknown.
    """
    try:
        with open(path, "rb") as handle:
            head = handle.read(8192)
    except OSError:
        return "unknown"
    if not head:
        return "empty"
    if head[:2] == b"PK":
        return "zip"
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "ole"
    if head[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "text"                                   # UTF-16 text, e.g. a UTF-16 CSV
    probe = head.decode("utf-8", errors="replace").lstrip().lower()
    if probe.startswith("<?xml"):
        return "html" if "<html" in probe[:4096] else "xml"
    if probe.startswith(("<html", "<!doctype html", "<table", "<meta", "<body")):
        return "html"
    if "<table" in probe[:4096] or "<html" in probe[:4096]:
        return "html"
    if b"\x00" in head[:512]:
        return "unknown"
    return "text"


# --------------------------------------------------------------------------
# SHEET READERS -- each yields (sheet_name, visibility, row_iterator)
# --------------------------------------------------------------------------

def read_openpyxl(path: Path):
    """xlsx / xlsm / xltx / xltm. Includes hidden and very-hidden sheets."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        for sheet in workbook.worksheets:
            visibility = getattr(sheet, "sheet_state", "visible") or "visible"

            def rows(sheet=sheet):
                for row in sheet.iter_rows(values_only=True):
                    yield [to_text(cell) for cell in row]

            yield sheet.title, visibility, rows()
    finally:
        workbook.close()


def read_xls(path: Path):
    """Genuine BIFF .xls via xlrd. Includes hidden and very-hidden sheets."""
    import xlrd

    book = xlrd.open_workbook(path, formatting_info=False, on_demand=False)
    states = {0: "visible", 1: "hidden", 2: "veryHidden"}
    for sheet in book.sheets():
        visibility = states.get(getattr(sheet, "visibility", 0), "visible")

        def rows(sheet=sheet, datemode=book.datemode):
            for index in range(sheet.nrows):
                out = []
                for cell in sheet.row(index):
                    kind, value = cell.ctype, cell.value
                    if kind in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        out.append("")
                    elif kind == xlrd.XL_CELL_DATE:
                        try:
                            out.append(to_text(xlrd.xldate.xldate_as_datetime(value, datemode)))
                        except Exception:
                            out.append(to_text(value))
                    elif kind == xlrd.XL_CELL_BOOLEAN:
                        out.append("TRUE" if value else "FALSE")
                    elif kind == xlrd.XL_CELL_ERROR:
                        out.append("")
                    else:
                        out.append(to_text(value))
                yield out

        yield sheet.name, visibility, rows()


def read_xlsb(path: Path):
    """Binary .xlsb via pyxlsb."""
    from pyxlsb import open_workbook as open_xlsb

    with open_xlsb(str(path)) as book:
        for name in book.sheets:
            with book.get_sheet(name) as sheet:
                collected = []
                for row in sheet.rows():
                    collected.append([to_text(cell.v) for cell in row])
                yield name, "visible", iter(collected)


def read_ods(path: Path):
    """OpenDocument .ods via pandas + odfpy."""
    import pandas as pd

    frames = pd.read_excel(path, sheet_name=None, header=None, dtype=object, engine="odf")
    for name, frame in frames.items():
        rows = [[to_text(cell) for cell in record] for record in frame.itertuples(index=False)]
        yield str(name), "visible", iter(rows)


def read_spreadsheetml(path: Path):
    """Excel 2003 XML Spreadsheet saved with an .xls/.xml extension."""
    from lxml import etree

    tree = etree.parse(str(path))
    ns = {"ss": SSML_NS}
    for worksheet in tree.iter("{%s}Worksheet" % SSML_NS):
        name = worksheet.get("{%s}Name" % SSML_NS) or "Sheet"
        collected: list[list[str]] = []
        for table in worksheet.findall("ss:Table", ns):
            row_cursor = 0
            for row in table.findall("ss:Row", ns):
                row_index = row.get("{%s}Index" % SSML_NS)
                row_cursor = int(row_index) - 1 if row_index else row_cursor
                cells: list[str] = []
                col_cursor = 0
                for cell in row.findall("ss:Cell", ns):
                    cell_index = cell.get("{%s}Index" % SSML_NS)
                    if cell_index:
                        col_cursor = int(cell_index) - 1
                    while len(cells) < col_cursor:
                        cells.append("")
                    data = cell.find("ss:Data", ns)
                    raw = "" if data is None else "".join(data.itertext())
                    kind = data.get("{%s}Type" % SSML_NS) if data is not None else None
                    if kind == "DateTime" and raw:
                        try:
                            raw = to_text(dt.datetime.fromisoformat(raw))
                        except ValueError:
                            pass
                    elif kind == "Number" and raw:
                        try:
                            raw = to_text(float(raw))   # 1000.0 -> 1000
                        except ValueError:
                            pass
                    elif kind == "Boolean":
                        raw = "TRUE" if raw.strip() in ("1", "true", "TRUE") else "FALSE"
                    cells.append(raw.strip())
                    col_cursor += 1
                while len(collected) < row_cursor:
                    collected.append([])
                collected.append(cells)
                row_cursor += 1
        yield name, "visible", iter(collected)


def read_html_tables(path: Path):
    """HTML table exports mislabelled as .xls -- a very common portal export.

    Parsed with lxml rather than pandas.read_html on purpose: read_html
    promotes a <th> row to DataFrame column labels (so the header row never
    reaches the row stream and the file looks like it has no target columns),
    and it type-infers every cell, which silently strips the leading zeros off
    identifiers such as 012345678.
    """
    from lxml import html as lxml_html

    with open(path, "rb") as handle:
        tree = lxml_html.fromstring(handle.read())

    for position, table in enumerate(tree.xpath("//table"), start=1):
        collected: list[list[str]] = []
        for row in table.xpath(".//tr"):
            owner = row.xpath("ancestor::table[1]")
            if owner and owner[0] is not table:
                continue                                # belongs to a nested table
            cells: list[str] = []
            for cell in row.xpath("./td | ./th"):
                cells.append(" ".join(cell.itertext()).replace("\xa0", " ").strip())
                span = cell.get("colspan")
                if span and span.strip().isdigit():
                    cells.extend([""] * (int(span) - 1))
            collected.append(cells)
        if collected:
            yield f"Table{position}", "visible", iter(collected)


def read_delimited(path: Path):
    """CSV / TXT / TSV with encoding and delimiter detection."""
    with open(path, "rb") as handle:
        bom = handle.read(4)
    candidates = ["utf-8-sig", "cp1252", "latin-1"]
    if bom[:2] in (b"\xff\xfe", b"\xfe\xff"):
        candidates.insert(0, "utf-16")

    sample, encoding = None, None
    for candidate in candidates:
        try:
            with open(path, "r", encoding=candidate, newline="") as handle:
                sample = handle.read(65536)
            encoding = candidate
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if encoding is None:
        encoding, sample = "latin-1", ""
    if sample and "\x00" in sample:
        raise ValueError("binary content - not a delimited text file")

    delimiter = ","
    if sample:
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except csv.Error:
            if path.suffix.lower() == ".tsv" or sample.count("\t") > sample.count(","):
                delimiter = "\t"

    def rows():
        with open(path, "r", encoding=encoding, errors="replace", newline="") as handle:
            for record in csv.reader(handle, delimiter=delimiter):
                yield [to_text(cell) for cell in record]

    yield path.stem, "visible", rows()


def open_sheets(path: Path):
    """Route a file to the right reader, falling back on disguised formats."""
    kind = sniff(path)
    ext = path.suffix.lower()

    if kind == "empty":
        raise ValueError("file is empty (0 bytes)")

    if kind == "zip":
        if ext == ".ods":
            attempts = [("ods", read_ods), ("xlsx", read_openpyxl)]
        elif ext == ".xlsb":
            attempts = [("xlsb", read_xlsb), ("xlsx", read_openpyxl)]
        else:
            attempts = [("xlsx", read_openpyxl), ("xlsb", read_xlsb), ("ods", read_ods)]
    elif kind == "ole":
        attempts = [("xls", read_xls)]
    elif kind == "xml":
        attempts = [("xml2003", read_spreadsheetml), ("html", read_html_tables)]
    elif kind == "html":
        attempts = [("html", read_html_tables), ("xml2003", read_spreadsheetml)]
    elif kind == "text":
        attempts = [("csv", read_delimited), ("html", read_html_tables)]
    else:
        # Binary content we could not identify. Do NOT try the CSV reader here --
        # it would "succeed" on garbage and the file would be reported as
        # "columns not found" instead of the truthful "unreadable".
        attempts = [("xls", read_xls), ("xlsx", read_openpyxl), ("xlsb", read_xlsb)]

    errors = []
    for label, reader in attempts:
        try:
            generator = reader(path)
            first = next(generator, None)               # force the open to happen now
            if first is None:
                errors.append(f"{label}: no sheets")
                continue
            return chain([first], generator), f"{kind}/{label}"
        except Exception as exc:                        # noqa: BLE001 - report, don't crash
            errors.append(f"{label}: {type(exc).__name__}: {exc}")

    hint = ""
    if kind == "ole" and any("encrypt" in e.lower() or "password" in e.lower() for e in errors):
        hint = " (looks password-protected)"
    raise ValueError(f"could not read as {kind}{hint}; tried -> " + " | ".join(errors))


# --------------------------------------------------------------------------
# HEADER DETECTION
# --------------------------------------------------------------------------

def build_columns(row: list[str]) -> tuple[dict[int, str], list[tuple[int, str, str]]]:
    """Map every cell of a candidate header row to an output column.

    Returns (targets, columns) where targets maps column index -> canonical
    name (used only to score the row), and columns is the full list of
    (column_index, output_key, display_name) for EVERY column on the sheet --
    target columns and additional columns alike.

    Duplicate header names get a " (2)", " (3)" suffix so no column is lost.
    Blank headers become "Unnamed <position>".
    """
    targets: dict[int, str] = {}
    columns: list[tuple[int, str, str]] = []
    used: dict[str, int] = {}

    for position, cell in enumerate(row):
        text = "" if cell is None else str(cell).strip()
        key = normalise(text)
        canonical = TARGETS.get(key) or ALIASES.get(key) if key else None

        if canonical:
            base_key, display = canonical, canonical
        elif key:
            base_key, display = key, text
        else:
            base_key, display = f"unnamed{position + 1}", f"Unnamed {position + 1}"

        seen = used.get(base_key, 0) + 1
        used[base_key] = seen
        if seen == 1:
            out_key, out_display = base_key, display
            if canonical:
                targets[position] = canonical
        else:
            out_key, out_display = f"{base_key}#{seen}", f"{display} ({seen})"

        columns.append((position, out_key, out_display))

    return targets, columns


def find_header(buffer: list[list[str]], min_matches: int):
    """Find the header row within the first rows of a sheet.

    Picks the row carrying the most target columns; ties keep the earliest row.
    Returns (row_index, targets, columns) or None when no row reaches
    min_matches.
    """
    best = None
    for index, row in enumerate(buffer):
        targets, columns = build_columns(row)
        if len(targets) >= min_matches:
            if best is None or len(targets) > len(best[1]):
                best = (index, targets, columns)
            if len(targets) == len(CANONICAL_ORDER):
                break
    return best


# --------------------------------------------------------------------------
# FILE DISCOVERY
# --------------------------------------------------------------------------

def discover_files(root: Path, recursive: bool, output_dir: Path) -> list[Path]:
    """Find candidate spreadsheets, never picking up this script's own output."""
    pattern = "**/*" if recursive else "*"
    output_resolved = output_dir.resolve()
    stem_prefix = OUTPUT_STEM.lower()
    found = []
    for path in sorted(root.glob(pattern)):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):                  # Excel lock file
            continue
        if path.suffix.lower() not in SPREADSHEET_EXTS:
            continue
        if path.stem.lower().startswith(stem_prefix):
            continue                                    # a previous run's output file
        try:
            resolved = path.resolve()
        except OSError:
            found.append(path)
            continue
        if output_resolved == resolved.parent or output_resolved in resolved.parents:
            continue                                    # inside this run's output folder
        if any(part.lower() == DEFAULT_OUTPUT_DIRNAME for part in resolved.parts[:-1]):
            continue                                    # inside any run's output folder
        found.append(path)
    return found


# --------------------------------------------------------------------------
# PASS 1 -- read every sheet, stage rows, learn the full column set
# --------------------------------------------------------------------------

def stage_rows(files: list[Path], input_dir: Path, staging_path: Path, args):
    """Read every qualifying sheet into a staging file, one JSON object per row.

    Two passes are unavoidable because the complete set of additional columns is
    only known once every file has been read, and those columns must appear in
    the output. Staging to disk keeps memory flat regardless of input size.
    """
    extra_display: dict[str, str] = {}      # output_key -> display name, first-seen order
    non_empty: set[str] = set()             # output keys that carried at least one value
    log_rows: list[dict] = []
    included_files: set[str] = set()
    excluded_files: list[tuple[str, str]] = []
    total_rows = 0
    coerced_dates = 0

    with open(staging_path, "w", encoding="utf-8", newline="") as staging:
        for position, path in enumerate(files, start=1):
            relative = str(path.relative_to(input_dir))
            print(f"[{position}/{len(files)}] {relative}", flush=True)

            try:
                sheets, route = open_sheets(path)
            except Exception as exc:                    # noqa: BLE001
                reason = f"unreadable: {exc}"
                excluded_files.append((relative, reason))
                log_rows.append({
                    "File": relative, "Sheet": "", "Status": "ERROR", "Reason": str(exc),
                    "Route": "", "HeaderRow": "", "RowsAdded": 0,
                    "MatchedColumns": "", "MissingColumns": "", "AdditionalColumns": "",
                })
                print(f"        !! {reason}")
                continue

            file_had_match = False
            sheet_seen = False

            while True:
                try:
                    item = next(sheets, None)
                except Exception as exc:                # noqa: BLE001
                    log_rows.append({
                        "File": relative, "Sheet": "", "Status": "ERROR",
                        "Reason": f"sheet iteration failed: {type(exc).__name__}: {exc}",
                        "Route": route, "HeaderRow": "", "RowsAdded": 0,
                        "MatchedColumns": "", "MissingColumns": "", "AdditionalColumns": "",
                    })
                    break
                if item is None:
                    break
                sheet_name, visibility, rows = item
                sheet_seen = True

                buffer: list[list[str]] = []
                try:
                    for row in rows:
                        buffer.append(row)
                        if len(buffer) >= HEADER_SCAN_ROWS:
                            break
                except Exception as exc:                # noqa: BLE001
                    log_rows.append({
                        "File": relative, "Sheet": sheet_name, "Status": "ERROR",
                        "Reason": f"read failed: {type(exc).__name__}: {exc}",
                        "Route": route, "HeaderRow": "", "RowsAdded": 0,
                        "MatchedColumns": "", "MissingColumns": "", "AdditionalColumns": "",
                    })
                    continue

                found = find_header(buffer, args.min_matches)
                if found is None:
                    present = sorted({
                        TARGETS[normalise(c)] for row in buffer for c in row
                        if normalise(c) in TARGETS
                    })
                    log_rows.append({
                        "File": relative, "Sheet": sheet_name, "Status": "SKIPPED",
                        "Reason": (
                            f"fewer than {args.min_matches} target columns found in the "
                            f"first {HEADER_SCAN_ROWS} rows"
                            + (f"; only found: {', '.join(present)}" if present else "")
                        ),
                        "Route": route, "HeaderRow": "", "RowsAdded": 0,
                        "MatchedColumns": ", ".join(present), "MissingColumns": "",
                        "AdditionalColumns": "",
                    })
                    continue

                header_index, targets, columns = found
                matched = [targets[k] for k in sorted(targets)]
                missing = [c for c in CANONICAL_ORDER if c not in matched]
                has_date_column = "ContributionDate" in matched

                # Additional (non-target) columns on this sheet.
                extras = [(i, k, d) for i, k, d in columns if d not in CANONICAL_ORDER]
                if not args.targets_only:
                    for _, key, display in extras:
                        extra_display.setdefault(key, display)
                carried = columns if not args.targets_only else [
                    (i, k, d) for i, k, d in columns if d in CANONICAL_ORDER
                ]

                sheet_rows = 0
                source_row = header_index + 1           # 1-based row after the header
                stream = chain(buffer[header_index + 1:], rows)
                try:
                    for row in stream:
                        source_row += 1
                        record: dict[str, str] = {}
                        for column_index, key, _ in carried:
                            if column_index < len(row):
                                value = row[column_index]
                                if value != "":
                                    record[key] = value
                        if not record:
                            continue                    # entirely blank row
                        if has_date_column and record.get("ContributionDate"):
                            fixed, changed = coerce_serial_date(record["ContributionDate"])
                            if changed:
                                record["ContributionDate"] = fixed
                                coerced_dates += 1
                        non_empty.update(record)
                        record["_f"] = relative
                        record["_s"] = sheet_name
                        record["_v"] = visibility
                        record["_h"] = str(header_index + 1)
                        record["_r"] = str(source_row)
                        staging.write(json.dumps(record, ensure_ascii=False) + "\n")
                        sheet_rows += 1
                except Exception as exc:                # noqa: BLE001
                    log_rows.append({
                        "File": relative, "Sheet": sheet_name, "Status": "PARTIAL",
                        "Reason": f"stopped mid-sheet: {type(exc).__name__}: {exc}",
                        "Route": route, "HeaderRow": header_index + 1,
                        "RowsAdded": sheet_rows, "MatchedColumns": ", ".join(matched),
                        "MissingColumns": ", ".join(missing),
                        "AdditionalColumns": "; ".join(d for _, _, d in extras),
                    })
                    total_rows += sheet_rows
                    file_had_match = file_had_match or sheet_rows > 0
                    print(f"        ~  {sheet_name}: {sheet_rows} rows (stopped early)")
                    continue

                total_rows += sheet_rows
                file_had_match = True
                log_rows.append({
                    "File": relative, "Sheet": sheet_name,
                    "Status": "INCLUDED" if sheet_rows else "INCLUDED (no data rows)",
                    "Reason": "", "Route": route, "HeaderRow": header_index + 1,
                    "RowsAdded": sheet_rows, "MatchedColumns": ", ".join(matched),
                    "MissingColumns": ", ".join(missing),
                    "AdditionalColumns": "; ".join(d for _, _, d in extras),
                })
                print(
                    f"        OK {sheet_name} [{visibility}]: {sheet_rows} rows, "
                    f"matched {len(matched)}/5, +{len(extras)} extra column(s)"
                )

            if file_had_match:
                included_files.add(relative)
            else:
                reason = (
                    f"no sheet contained at least {args.min_matches} of the target columns"
                    if sheet_seen else "no readable sheets"
                )
                excluded_files.append((relative, reason))

    return {
        "extra_display": extra_display,
        "non_empty": non_empty,
        "log_rows": log_rows,
        "included_files": included_files,
        "excluded_files": excluded_files,
        "total_rows": total_rows,
        "coerced_dates": coerced_dates,
    }


# --------------------------------------------------------------------------
# PASS 2 -- write the CSV using the full column set
# --------------------------------------------------------------------------

def write_csv(staging_path: Path, csv_path: Path, header: list[str],
              keys: list[str]) -> int:
    """Stream the staging file into the combined CSV, aligned to the full header."""
    written = 0
    with open(staging_path, "r", encoding="utf-8") as staging, \
            open(csv_path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        for line in staging:
            if not line.strip():
                continue
            record = json.loads(line)
            writer.writerow([record.get(key, "") for key in keys])
            written += 1
    return written


# --------------------------------------------------------------------------
# PASS 3 -- write XLSX, splitting into parts and adding the Excluded sheet
# --------------------------------------------------------------------------

def write_xlsx(csv_path: Path, output_dir: Path, stem: str, header: list[str],
               total_rows: int, excluded_files: list[tuple[str, str]]) -> list[Path]:
    """Write the combined CSV out as XLSX, splitting into parts if needed.

    Every workbook produced gets a second sheet, "Excluded Files", listing the
    files that were not combined and why, so each part is self-contained.
    """
    from openpyxl import Workbook

    truncated: list[str] = []
    if len(header) > EXCEL_MAX_COLUMNS:
        truncated = header[EXCEL_MAX_COLUMNS:]
        header = header[:EXCEL_MAX_COLUMNS]
        print(f"\nWARNING: {len(truncated)} column(s) exceed Excel's "
              f"{EXCEL_MAX_COLUMNS:,}-column limit and are omitted from the XLSX "
              f"(they remain in the CSV).")

    if total_rows == 0:
        # Still produce a workbook so the excluded list has somewhere to live.
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Combined"
        sheet.append(header)
        add_excluded_sheet(workbook, excluded_files)
        path = output_dir / f"{stem}.xlsx"
        workbook.save(path)
        print(f"\nNo data rows - wrote {path.name} with headers and the "
              f"{EXCLUDED_SHEET_TITLE} sheet only.")
        return [path]

    parts_needed = (total_rows + EXCEL_MAX_DATA_ROWS - 1) // EXCEL_MAX_DATA_ROWS
    if parts_needed <= 1:
        print(f"\n{total_rows:,} rows fit Excel's limit - writing a single workbook.")
    else:
        print(f"\n{total_rows:,} rows exceed Excel's {EXCEL_MAX_DATA_ROWS + 1:,}-row "
              f"limit - splitting into {parts_needed} part(s).")

    written: list[Path] = []
    workbook = None
    sheet = None
    part = 0
    in_part = 0
    width = len(header)

    def start_part():
        nonlocal workbook, sheet, part, in_part
        part += 1
        in_part = 0
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Combined")
        sheet.append(header)

    def finish_part():
        nonlocal workbook
        if workbook is None:
            return
        add_excluded_sheet(workbook, excluded_files)
        name = f"{stem}.xlsx" if parts_needed <= 1 else f"{stem} part {part:02d}.xlsx"
        path = output_dir / name
        workbook.save(path)
        workbook.close()
        written.append(path)
        print(f"  wrote {name} ({in_part:,} rows)")

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        next(reader, None)                              # discard the CSV header
        start_part()
        for row in reader:
            if in_part >= EXCEL_MAX_DATA_ROWS:
                finish_part()
                start_part()
            sheet.append(row[:width])
            in_part += 1
        finish_part()

    return written


def add_excluded_sheet(workbook, excluded_files: list[tuple[str, str]]) -> None:
    """Append the 'Excluded Files' sheet listing every file that was not combined."""
    sheet = workbook.create_sheet(title=EXCLUDED_SHEET_TITLE)
    sheet.append(EXCLUDED_SHEET_HEADER)
    if not excluded_files:
        sheet.append(["", "None - every file contained the required columns.", ""])
        return
    for index, (name, reason) in enumerate(excluded_files, start=1):
        sheet.append([index, name, reason])


# --------------------------------------------------------------------------
# ORCHESTRATION
# --------------------------------------------------------------------------

def combine(args) -> int:
    input_dir = Path(args.input).expanduser().resolve()
    if not input_dir.is_dir():
        print(f"ERROR: input folder not found: {input_dir}", file=sys.stderr)
        return 2

    output_dir = (
        Path(args.output).expanduser().resolve() if args.output
        else input_dir / DEFAULT_OUTPUT_DIRNAME
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    stem = OUTPUT_STEM
    csv_path = output_dir / f"{stem}.csv"
    excluded_csv_path = output_dir / f"{stem} excluded files.csv"
    log_path = output_dir / f"{stem} log.csv"
    summary_path = output_dir / f"{stem} summary.txt"
    staging_path = output_dir / f"{stem}.staging.jsonl"

    files = discover_files(input_dir, not args.no_recursive, output_dir)
    if not files:
        print(f"ERROR: no spreadsheet files found under {input_dir}", file=sys.stderr)
        return 2

    print(f"Scanning {len(files)} file(s) under {input_dir}\n")

    try:
        result = stage_rows(files, input_dir, staging_path, args)

        extra_display = result["extra_display"]
        non_empty = result["non_empty"]
        excluded_files = result["excluded_files"]
        total_rows = result["total_rows"]

        # Full column set: provenance, the five targets, then additional columns
        # in first-seen order. Additional columns that never carried a value are
        # dropped unless --keep-empty-columns was given.
        extra_keys = [
            key for key in extra_display
            if args.keep_empty_columns or key in non_empty
        ]
        dropped_empty = [
            extra_display[key] for key in extra_display if key not in extra_keys
        ]
        keys = STAGE_KEYS + CANONICAL_ORDER + extra_keys
        header = PROVENANCE + CANONICAL_ORDER + [extra_display[k] for k in extra_keys]

        written_rows = write_csv(staging_path, csv_path, header, keys)
        if written_rows != total_rows:
            print(f"WARNING: staged {total_rows} rows but wrote {written_rows}",
                  file=sys.stderr)

        xlsx_paths = write_xlsx(csv_path, output_dir, stem, header,
                                written_rows, excluded_files)
    finally:
        staging_path.unlink(missing_ok=True)

    # ---------------- excluded-files CSV ----------------
    with open(excluded_csv_path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(EXCLUDED_SHEET_HEADER)
        for index, (name, reason) in enumerate(excluded_files, start=1):
            writer.writerow([index, name, reason])

    # ---------------- per-sheet audit log ----------------
    log_fields = [
        "File", "Sheet", "Status", "Reason", "Route", "HeaderRow",
        "RowsAdded", "MatchedColumns", "MissingColumns", "AdditionalColumns",
    ]
    with open(log_path, "w", encoding="utf-8-sig", newline="") as handle:
        log_writer = csv.DictWriter(handle, fieldnames=log_fields)
        log_writer.writeheader()
        log_writer.writerows(result["log_rows"])

    # ---------------- summary ----------------
    sheets_combined = sum(
        1 for r in result["log_rows"] if r["Status"].startswith("INCLUDED")
    )
    with open(summary_path, "w", encoding="utf-8") as handle:
        handle.write("COMBINE CONTRIBUTION FILES - RUN SUMMARY\n")
        handle.write("=" * 72 + "\n\n")
        handle.write(f"Input folder     : {input_dir}\n")
        handle.write(f"Output folder    : {output_dir}\n")
        handle.write(f"Recursive        : {not args.no_recursive}\n")
        handle.write(f"Match rule       : at least {args.min_matches} of the "
                     f"{len(CANONICAL_ORDER)} target columns\n")
        handle.write(f"Column mode      : "
                     f"{'target columns only' if args.targets_only else 'all columns carried through'}\n\n")
        handle.write(f"Files scanned    : {len(files)}\n")
        handle.write(f"Files combined   : {len(result['included_files'])}\n")
        handle.write(f"Files excluded   : {len(excluded_files)}\n")
        handle.write(f"Sheets combined  : {sheets_combined}\n")
        handle.write(f"Data rows written: {written_rows:,}\n")
        handle.write(f"Columns written  : {len(header)} "
                     f"({len(PROVENANCE)} provenance + {len(CANONICAL_ORDER)} target "
                     f"+ {len(extra_keys)} additional)\n")
        if result["coerced_dates"]:
            handle.write(f"Date serials converted: {result['coerced_dates']:,}\n")
        if dropped_empty:
            handle.write(f"\nAdditional columns dropped as entirely empty "
                         f"({len(dropped_empty)}): {', '.join(dropped_empty)}\n")
            handle.write("Re-run with --keep-empty-columns to retain them.\n")

        handle.write(f"\nCSV output       : {csv_path.name}\n")
        for xlsx in xlsx_paths:
            handle.write(f"XLSX output      : {xlsx.name}\n")
        handle.write(f"Excluded list    : {excluded_csv_path.name} "
                     f"(also the '{EXCLUDED_SHEET_TITLE}' sheet in each workbook)\n")
        handle.write(f"Per-sheet log    : {log_path.name}\n")

        handle.write("\n\nCOLUMN LAYOUT\n")
        handle.write("-" * 72 + "\n")
        for index, name in enumerate(header, start=1):
            tag = ("provenance" if name in PROVENANCE
                   else "target" if name in CANONICAL_ORDER else "additional")
            handle.write(f"{index:>5}. [{tag}] {name}\n")

        handle.write("\n\nEXCLUDED FILES\n")
        handle.write("-" * 72 + "\n")
        if not excluded_files:
            handle.write("None - every file contained the required columns.\n")
        else:
            handle.write(
                f"These {len(excluded_files)} file(s) were NOT combined because none of "
                f"their sheets contained at least {args.min_matches} of the target\n"
                "columns (or the file could not be read):\n\n"
            )
            for index, (name, reason) in enumerate(excluded_files, start=1):
                handle.write(f"{index:>4}. {name}\n         reason: {reason}\n")

        handle.write("\n\nHANDLING NOTE\n")
        handle.write("-" * 72 + "\n")
        handle.write(
            "The combined output contains employee identifiers and contribution\n"
            "amounts. Store it in an access-controlled location only, and retain\n"
            "this summary and the per-sheet log as the processing audit trail.\n"
        )

    # ---------------- console report ----------------
    print("\n" + "=" * 72)
    print(f"Files scanned    : {len(files)}")
    print(f"Files combined   : {len(result['included_files'])}")
    print(f"Files excluded   : {len(excluded_files)}")
    print(f"Sheets combined  : {sheets_combined}")
    print(f"Data rows written: {written_rows:,}")
    print(f"Columns written  : {len(header)} "
          f"({len(CANONICAL_ORDER)} target + {len(extra_keys)} additional)")
    if dropped_empty:
        print(f"Empty extra columns dropped: {len(dropped_empty)} "
              f"(use --keep-empty-columns to retain)")
    print(f"\nCSV      : {csv_path}")
    for xlsx in xlsx_paths:
        print(f"XLSX     : {xlsx}")
    print(f"Excluded : {excluded_csv_path}")
    print(f"Log      : {log_path}")
    print(f"Summary  : {summary_path}")

    if excluded_files:
        print(f"\nEXCLUDED - target columns not found ({len(excluded_files)}):")
        for name, reason in excluded_files:
            print(f"  - {name}   [{reason}]")
    print("=" * 72)
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Combine contribution spreadsheets of mixed formats.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("-i", "--input", required=True, help="folder containing source files")
    parser.add_argument("-o", "--output", default=None,
                        help="output folder (default: <input>\\_combined)")
    parser.add_argument("--no-recursive", action="store_true",
                        help="do not descend into subfolders")
    parser.add_argument("--min-matches", type=int, default=MIN_MATCHES,
                        help=f"target columns required to accept a sheet (default {MIN_MATCHES})")
    parser.add_argument("--targets-only", action="store_true",
                        help="output only the 5 target columns, dropping additional columns")
    parser.add_argument("--keep-empty-columns", action="store_true",
                        help="keep additional columns that turned out to be entirely blank")
    args = parser.parse_args()

    if args.min_matches < 1:
        print("ERROR: --min-matches must be at least 1", file=sys.stderr)
        return 2

    try:
        return combine(args)
    except KeyboardInterrupt:
        print("\nInterrupted.", file=sys.stderr)
        return 130


if __name__ == "__main__":
    sys.exit(main())
