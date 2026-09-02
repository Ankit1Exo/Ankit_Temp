# ============================================================
#  Excel File Compiler — Dynamic Headers
#  - Compiles all Excel files (.xlsx, .xls, .xlsm) into one
#  - Supports multiple sheets per file
#  - Adds "Source File" and "Sheet Name" columns
#  - Auto-discovers all headers across all files/sheets
#  - New headers added as new columns (blank if not in sheet)
#  - Skips pivot table sheets automatically
# ============================================================
#
#  Requirements:
#    pip install openpyxl pandas xlrd
# ============================================================

import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

import openpyxl
import pandas as pd

# ── Hide root Tkinter window ─────────────────────────────────
root = tk.Tk()
root.withdraw()
root.attributes("-topmost", True)

print("=" * 60)
print("        Excel File Compiler — Dynamic Headers")
print("=" * 60)

# ── Step 1: Select input folder ──────────────────────────────
print("\n📂 Select INPUT folder...")
input_folder = filedialog.askdirectory(title="Select Folder Containing Excel Files")
if not input_folder:
    messagebox.showerror("Cancelled", "No input folder selected. Exiting.")
    exit()

SUPPORTED_EXTS = (".xlsx", ".xlsm", ".xls")
excel_files = [
    f for f in os.listdir(input_folder)
    if f.lower().endswith(SUPPORTED_EXTS) and not f.startswith("~$")
]

if not excel_files:
    messagebox.showerror("No Excel Files Found", f"No Excel files found in:\n{input_folder}")
    exit()

print(f"✅ Found {len(excel_files)} Excel file(s)")
for f in excel_files:
    print(f"   • {f}")

# ── Step 2: Select output folder ─────────────────────────────
print("\n📁 Select OUTPUT folder...")
output_folder = filedialog.askdirectory(title="Select Output Folder")
if not output_folder:
    messagebox.showerror("Cancelled", "No output folder selected. Exiting.")
    exit()

# ── Step 3: Output file name ─────────────────────────────────
while True:
    output_name = simpledialog.askstring(
        "Output File Name",
        "Enter output file name (without extension):",
        parent=root
    )
    if not output_name:
        messagebox.showerror("Error", "File name cannot be empty.")
        continue
    output_name  = output_name.strip()
    output_path  = os.path.join(output_folder, output_name + ".xlsx")
    summary_path = os.path.join(output_folder, output_name + "_summary.xlsx")
    if os.path.exists(output_path):
        if messagebox.askyesno("File Exists", f"'{output_name}.xlsx' already exists. Overwrite?"):
            break
    else:
        break

# ── Step 4: Progress Window ───────────────────────────────────
progress_win = tk.Toplevel(root)
progress_win.title("Compiling Excel Files...")
progress_win.geometry("660x380")
progress_win.resizable(False, False)
progress_win.attributes("-topmost", True)

tk.Label(progress_win, text="Excel File Compiler", font=("Segoe UI", 13, "bold")).pack(pady=(14, 4))
tk.Label(progress_win, text="Overall Progress:", font=("Segoe UI", 9)).pack(anchor="w", padx=20)

overall_bar = ttk.Progressbar(progress_win, length=615, mode="determinate", maximum=len(excel_files))
overall_bar.pack(padx=20, pady=(2, 4))

overall_label = tk.Label(
    progress_win,
    text=f"0 of {len(excel_files)} files  |  {len(excel_files)} pending",
    font=("Segoe UI", 9), fg="#555"
)
overall_label.pack()

current_label = tk.Label(progress_win, text="Starting...", font=("Segoe UI", 9, "italic"), fg="#333", wraplength=620)
current_label.pack(pady=(6, 2))

stats_frame = tk.Frame(progress_win, bd=1, relief="groove", padx=10, pady=8)
stats_frame.pack(padx=20, pady=10, fill="x")

def stat_col(parent, label, var, col):
    tk.Label(parent, text=label, font=("Segoe UI", 8), fg="#777").grid(row=0, column=col, padx=12)
    tk.Label(parent, textvariable=var, font=("Segoe UI", 12, "bold")).grid(row=1, column=col, padx=12)

v_total       = tk.StringVar(value=str(len(excel_files)))
v_done        = tk.StringVar(value="0")
v_pending     = tk.StringVar(value=str(len(excel_files)))
v_sheets      = tk.StringVar(value="0")
v_rows        = tk.StringVar(value="0")
v_headers     = tk.StringVar(value="0")
v_pivot_skip  = tk.StringVar(value="0")
v_errors      = tk.StringVar(value="0")

stat_col(stats_frame, "Total Files",    v_total,      0)
stat_col(stats_frame, "Processed",      v_done,       1)
stat_col(stats_frame, "Pending",        v_pending,    2)
stat_col(stats_frame, "Sheets Read",    v_sheets,     3)
stat_col(stats_frame, "Rows Written",   v_rows,       4)
stat_col(stats_frame, "Unique Headers", v_headers,    5)
stat_col(stats_frame, "Pivot Skipped",  v_pivot_skip, 6)
stat_col(stats_frame, "Errors",         v_errors,     7)

progress_win.update()

# ── Helper: detect pivot sheets ───────────────────────────────
def is_pivot_sheet(ws):
    try:
        return len(ws._pivots) > 0
    except Exception:
        return False

# ── Step 5: Pass 1 — Collect ALL headers across all files ────
print("\n🔍 Pass 1: Scanning all files for headers...")
current_label.config(text="Pass 1: Scanning headers across all files & sheets...")
progress_win.update()

master_headers = []   # ordered, preserving discovery order
header_set     = set()
file_meta      = {}   # file -> list of {sheet_name, headers, pivot}
total_pivot_skipped = 0

for excel_file in excel_files:
    file_path = os.path.join(input_folder, excel_file)
    file_meta[excel_file] = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if is_pivot_sheet(ws):
                print(f"   ⏭️  {excel_file} / '{sheet_name}' — pivot table, skipping")
                file_meta[excel_file].append({"sheet": sheet_name, "headers": [], "pivot": True})
                total_pivot_skipped += 1
                continue

            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c).strip() if c is not None else "" for c in row]
                break
            # Remove trailing empty headers
            while headers and headers[-1] == "":
                headers.pop()

            for h in headers:
                if h and h not in header_set:
                    master_headers.append(h)
                    header_set.add(h)

            file_meta[excel_file].append({"sheet": sheet_name, "headers": headers, "pivot": False})
            print(f"   • {excel_file} / '{sheet_name}' — {len(headers)} headers")
        wb.close()
    except Exception as e:
        print(f"   ⚠️  Could not scan {excel_file}: {e}")

print(f"\n✅ Total unique headers discovered: {len(master_headers)}")
v_headers.set(str(len(master_headers)))
v_pivot_skip.set(str(total_pivot_skipped))
progress_win.update()

# ── Step 6: Pass 2 — Compile all rows ────────────────────────
print("\n📝 Pass 2: Compiling all rows...")
current_label.config(text="Pass 2: Compiling rows into output Excel...")
progress_win.update()

all_columns   = ["Source File", "Sheet Name"] + master_headers
compiled_rows = []
total_rows    = 0
total_errors  = 0
total_sheets  = 0
file_summary  = []

for idx, excel_file in enumerate(excel_files, start=1):
    file_path = os.path.join(input_folder, excel_file)
    current_label.config(text=f"Compiling ({idx}/{len(excel_files)}): {excel_file}")
    overall_bar["value"] = idx - 1
    overall_label.config(text=f"{idx-1} of {len(excel_files)} files  |  {len(excel_files)-idx+1} pending")
    progress_win.update()

    print(f"\n📄 {excel_file}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)

        for sheet_info in file_meta[excel_file]:
            sheet_name = sheet_info["sheet"]
            status     = "Skipped (Pivot)" if sheet_info["pivot"] else "Success"
            row_count  = 0

            if sheet_info["pivot"]:
                file_summary.append({
                    "File Name"   : excel_file,
                    "Sheet Name"  : sheet_name,
                    "Headers"     : 0,
                    "Rows"        : 0,
                    "Status"      : status,
                    "Error"       : ""
                })
                continue

            try:
                ws      = wb[sheet_name]
                headers = sheet_info["headers"]

                rows_iter = ws.iter_rows(min_row=2, values_only=True)
                for row in rows_iter:
                    row_dict = {"Source File": excel_file, "Sheet Name": sheet_name}
                    for col_idx, h in enumerate(headers):
                        row_dict[h] = row[col_idx] if col_idx < len(row) else None
                    compiled_rows.append(row_dict)
                    row_count  += 1
                    total_rows += 1

                total_sheets += 1
                print(f"   ✅ '{sheet_name}' — {len(headers)} headers | {row_count} rows")

            except Exception as e:
                print(f"   ❌ '{sheet_name}': {e}")
                status      = "Error"
                total_errors += 1
                file_summary.append({
                    "File Name"  : excel_file,
                    "Sheet Name" : sheet_name,
                    "Headers"    : len(sheet_info["headers"]),
                    "Rows"       : row_count,
                    "Status"     : status,
                    "Error"      : str(e)
                })
                continue

            file_summary.append({
                "File Name"  : excel_file,
                "Sheet Name" : sheet_name,
                "Headers"    : len(headers),
                "Rows"       : row_count,
                "Status"     : status,
                "Error"      : ""
            })

        wb.close()

    except Exception as e:
        print(f"   ❌ Could not open {excel_file}: {e}")
        total_errors += 1
        file_summary.append({
            "File Name"  : excel_file,
            "Sheet Name" : "N/A",
            "Headers"    : 0,
            "Rows"       : 0,
            "Status"     : "Error",
            "Error"      : str(e)
        })

    v_done.set(str(idx))
    v_pending.set(str(len(excel_files) - idx))
    v_sheets.set(str(total_sheets))
    v_rows.set(str(total_rows))
    v_errors.set(str(total_errors))
    overall_bar["value"] = idx
    overall_label.config(text=f"{idx} of {len(excel_files)} files  |  {len(excel_files)-idx} pending")
    progress_win.update()

# ── Step 7: Write compiled output Excel ──────────────────────
current_label.config(text="Writing output Excel file...")
progress_win.update()

df_out = pd.DataFrame(compiled_rows, columns=all_columns)
df_out.to_excel(output_path, index=False, engine="openpyxl")

# ── Step 8: Write summary Excel ──────────────────────────────
current_label.config(text="Writing summary Excel file...")
progress_win.update()

with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:

    # Overall summary
    overall_data = {
        "Metric": [
            "Total Excel Files", "Files Processed OK", "Files Errored",
            "Total Sheets Read", "Pivot Sheets Skipped",
            "Total Unique Headers", "Total Rows Compiled"
        ],
        "Value": [
            len(excel_files), len(excel_files) - total_errors, total_errors,
            total_sheets, total_pivot_skipped,
            len(master_headers), total_rows
        ]
    }
    pd.DataFrame(overall_data).to_excel(writer, sheet_name="Overall Summary", index=False)

    # Headers discovered
    header_rows = []
    for i, h in enumerate(master_headers, start=1):
        first_file  = next((f for f, sheets in file_meta.items()
                            for s in sheets if h in s["headers"]), "")
        first_sheet = next((s["sheet"] for f, sheets in file_meta.items()
                            for s in sheets if h in s["headers"] and f == first_file), "")
        header_rows.append({"#": i, "Header Name": h,
                             "First Seen In File": first_file,
                             "First Seen In Sheet": first_sheet})
    pd.DataFrame(header_rows).to_excel(writer, sheet_name="All Headers", index=False)

    # File/sheet level summary
    pd.DataFrame(file_summary).to_excel(writer, sheet_name="File Summary", index=False)

# ── Done ─────────────────────────────────────────────────────
progress_win.destroy()

summary_msg = (
    f"✅ Compilation Complete!\n\n"
    f"📁 Files Found           : {len(excel_files)}\n"
    f"✅ Files Processed       : {len(excel_files) - total_errors}\n"
    f"❌ Errors                : {total_errors}\n"
    f"📋 Sheets Read           : {total_sheets}\n"
    f"⏭️  Pivot Sheets Skipped  : {total_pivot_skipped}\n"
    f"🏷️  Unique Headers        : {len(master_headers)}\n"
    f"📝 Total Rows Written    : {total_rows}\n\n"
    f"💾 Compiled Excel : {output_path}\n"
    f"📊 Summary Excel  : {summary_path}"
)

print("\n" + "=" * 60)
print(summary_msg)
print("=" * 60)

messagebox.showinfo("Compilation Complete ✅", summary_msg)
root.destroy()
