#!/usr/bin/env python3
"""
Append every sheet of every .xlsx/.xlsm workbook in a chosen folder into one table.

Behaviour
---------
* Folder picker (tkinter); falls back to a typed path if no GUI is available.
* Unhides everything -- very-hidden/hidden sheets, hidden rows, hidden columns,
  outline groups and autofilter criteria -- and saves UNHIDDEN COPIES.
  Original files are never modified.
* Skips a sheet only if that specific sheet has a pivot table on it. Every other
  sheet, hidden or not, is appended.
* Column A = source file name, Column B = sheet name.
* Headers that match are appended into the same column. Headers that do not
  match are added as new columns at the end, in first-seen order.
* Nothing is de-duplicated. Only completely blank rows are dropped.
* Falls back to CSV automatically if the result exceeds Excel's row limit.

Usage
-----
    python "<this file>.py"              # opens the folder picker
    python "<this file>.py" "C:\\path"    # skips the picker

Requires: openpyxl  (pip install openpyxl)
Note: .xls and .xlsb are NOT supported by openpyxl. Re-save them as .xlsx first.
"""

from __future__ import annotations

import csv
import os
import re
import sys
import traceback
import zipfile
from datetime import date, datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font

# ----------------------------------------------------------------------------
# CONFIGURATION -- adjust these if needed
# ----------------------------------------------------------------------------

DOC_TYPE = "re"          # 2-letter document type code for the output file name
AUTHOR_CODE = ""         # your 2-letter initials; leave "" to be prompted
DESCRIPTION = "consolidated workbook data"

FILE_EXTENSIONS = (".xlsx", ".xlsm")
RECURSE_SUBFOLDERS = False   # True to also pick up workbooks in subfolders

MAKE_UNHIDDEN_COPIES = True  # write unhidden copies of each source workbook
SKIP_BLANK_ROWS = True       # drop rows where every cell is empty
NORMALIZE_HEADERS = True     # "Total Amount" == "total  amount " when matching
ADD_SOURCE_ROW_COLUMN = False  # adds a 3rd column with the original row number

# Header row detection: "first_non_empty" or an integer (1-based) to force a row
HEADER_ROW = "first_non_empty"

EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLS = 16_384

FIXED_COLUMNS = ["source file name", "sheet name"]

# ----------------------------------------------------------------------------
# XML namespaces used for pivot-table detection
# ----------------------------------------------------------------------------

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PIVOT_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable"
)


# ----------------------------------------------------------------------------
# Small helpers
# ----------------------------------------------------------------------------

def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def clean_value(value):
    """Strip characters Excel refuses to store."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def header_key(header: str) -> str:
    """The key used to decide whether two headers are 'the same' column."""
    text = str(header)
    if not NORMALIZE_HEADERS:
        return text
    return re.sub(r"\s+", " ", text).strip().casefold()


def stringify_header(value, position: int) -> str:
    """Turn a raw header cell into a usable column name."""
    if is_blank(value):
        return f"column {position}"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


# ----------------------------------------------------------------------------
# Pivot table detection -- reads the package relationships directly, so it maps
# each pivot table back to the exact sheet it lives on.
# ----------------------------------------------------------------------------

def sheets_with_pivot_tables(path: Path) -> set[str]:
    pivot_sheets: set[str] = set()
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            if "xl/workbook.xml" not in names:
                return pivot_sheets
            if not any(n.startswith("xl/pivotTables/") for n in names):
                return pivot_sheets  # no pivot tables anywhere in this workbook

            workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
            rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            rid_to_target = {rel.get("Id"): rel.get("Target") for rel in rels_xml}

            for sheet_el in workbook_xml.iter(f"{{{NS_MAIN}}}sheet"):
                sheet_name = sheet_el.get("name")
                target = rid_to_target.get(sheet_el.get(f"{{{NS_R}}}id"))
                if not sheet_name or not target:
                    continue

                target = target.replace("\\", "/").lstrip("/")
                part = target if target.startswith("xl/") else f"xl/{target}"
                part = os.path.normpath(part).replace("\\", "/")

                rels_part = (
                    f"{os.path.dirname(part)}/_rels/{os.path.basename(part)}.rels"
                )
                if rels_part not in names:
                    continue

                sheet_rels = ET.fromstring(archive.read(rels_part))
                for rel in sheet_rels:
                    if rel.get("Type") == PIVOT_REL_TYPE:
                        pivot_sheets.add(sheet_name)
                        break
    except (zipfile.BadZipFile, ET.ParseError, KeyError):
        pass  # fall back to openpyxl's own detection during the scan
    return pivot_sheets


# ----------------------------------------------------------------------------
# Unhide everything and save a copy (originals are left untouched)
# ----------------------------------------------------------------------------

def write_unhidden_copy(source: Path, destination_dir: Path, log: list[str]) -> None:
    try:
        workbook = openpyxl.load_workbook(
            source, data_only=False, keep_vba=source.suffix.lower() == ".xlsm"
        )
    except Exception as exc:
        log.append(f"  ! could not open for unhiding: {exc}")
        return

    changes = 0
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            sheet.sheet_state = "visible"
            changes += 1

        for dimension in sheet.row_dimensions.values():
            if dimension.hidden or dimension.outlineLevel:
                dimension.hidden = False
                dimension.outlineLevel = 0
                dimension.collapsed = False
                changes += 1

        for dimension in sheet.column_dimensions.values():
            if dimension.hidden or dimension.outlineLevel:
                dimension.hidden = False
                dimension.outlineLevel = 0
                dimension.collapsed = False
                changes += 1

        # Clear filter criteria so Excel does not re-hide the rows on open
        try:
            if sheet.auto_filter is not None and sheet.auto_filter.filterColumn:
                sheet.auto_filter.filterColumn = []
                changes += 1
        except Exception:
            pass

    # Chart sheets and dialog sheets can be hidden too
    for sheet in workbook._sheets:
        if getattr(sheet, "sheet_state", "visible") != "visible":
            sheet.sheet_state = "visible"
            changes += 1

    destination_dir.mkdir(parents=True, exist_ok=True)
    target = destination_dir / f"{source.stem} unhidden{source.suffix}"
    try:
        workbook.save(target)
        log.append(f"  unhidden copy saved ({changes} item(s) unhidden): {target.name}")
    except Exception as exc:
        log.append(f"  ! could not save unhidden copy: {exc}")
    finally:
        workbook.close()


# ----------------------------------------------------------------------------
# Pass 1 -- scan every sheet, collect headers and row counts
# ----------------------------------------------------------------------------

class SheetPlan:
    __slots__ = ("file_path", "sheet_name", "header_row", "headers", "row_count")

    def __init__(self, file_path, sheet_name, header_row, headers, row_count):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.headers = headers          # list of display names, by position
        self.row_count = row_count


def scan_sheet(worksheet) -> tuple[int, list[str], int]:
    """Return (header_row_index, headers, data_row_count) for one worksheet."""
    header_row_index = 0
    headers: list[str] = []
    data_rows = 0

    forced = HEADER_ROW if isinstance(HEADER_ROW, int) else None

    for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not header_row_index:
            if forced is not None:
                if index < forced:
                    continue
                header_row_index = index
            elif all(is_blank(cell) for cell in row):
                continue
            else:
                header_row_index = index

            raw = list(row)
            while raw and is_blank(raw[-1]):
                raw.pop()  # trim trailing empty header cells

            seen: dict[str, int] = {}
            for position, cell in enumerate(raw, start=1):
                name = stringify_header(cell, position)
                key = header_key(name)
                if key in seen:
                    seen[key] += 1
                    name = f"{name}.{seen[key]}"
                else:
                    seen[key] = 0
                headers.append(name)
            continue

        if SKIP_BLANK_ROWS and all(is_blank(cell) for cell in row):
            continue
        data_rows += 1

    return header_row_index, headers, data_rows


def scan_workbooks(files: list[Path], log: list[str]) -> tuple[list[SheetPlan], list[str], dict[str, int]]:
    plans: list[SheetPlan] = []
    column_order: list[str] = []          # display names, first-seen order
    column_index: dict[str, int] = {}     # header_key -> position in column_order

    for path in files:
        log.append(f"\n{path.name}")
        pivot_sheets = sheets_with_pivot_tables(path)

        try:
            workbook = openpyxl.load_workbook(
                path, data_only=True, read_only=True, keep_links=False
            )
        except Exception as exc:
            log.append(f"  ! SKIPPED FILE -- could not open: {exc}")
            continue

        try:
            for worksheet in workbook.worksheets:
                name = worksheet.title
                if not hasattr(worksheet, "iter_rows"):
                    log.append(f"  - {name}: skipped (chart/dialog sheet, no cells)")
                    continue
                if name in pivot_sheets:
                    log.append(f"  - {name}: SKIPPED (pivot table on this sheet)")
                    continue

                state = getattr(worksheet, "sheet_state", "visible")
                header_row_index, headers, row_count = scan_sheet(worksheet)

                if not headers:
                    log.append(f"  - {name}: skipped (sheet is empty)")
                    continue

                for header in headers:
                    key = header_key(header)
                    if key not in column_index:
                        column_index[key] = len(column_order)
                        column_order.append(header)

                plans.append(SheetPlan(path, name, header_row_index, headers, row_count))
                flag = "" if state == "visible" else f" [was {state}]"
                log.append(
                    f"  + {name}{flag}: header on row {header_row_index}, "
                    f"{row_count} data row(s), {len(headers)} column(s)"
                )
        finally:
            workbook.close()

    return plans, column_order, column_index


# ----------------------------------------------------------------------------
# Pass 2 -- write the combined table
# ----------------------------------------------------------------------------

def iter_output_rows(plans, column_order, column_index, log):
    width = len(column_order)
    for plan in plans:
        try:
            workbook = openpyxl.load_workbook(
                plan.file_path, data_only=True, read_only=True, keep_links=False
            )
        except Exception as exc:
            log.append(f"  ! could not re-open {plan.file_path.name}: {exc}")
            continue

        try:
            worksheet = workbook[plan.sheet_name]
            # Precompute where each of this sheet's columns lands in the output
            mapping = [column_index[header_key(h)] for h in plan.headers]

            for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if index <= plan.header_row:
                    continue
                if SKIP_BLANK_ROWS and all(is_blank(cell) for cell in row):
                    continue

                buffer = [None] * width
                for position, value in enumerate(row):
                    if position >= len(mapping):
                        break  # value sits beyond this sheet's header width
                    buffer[mapping[position]] = clean_value(value)

                prefix = [plan.file_path.name, plan.sheet_name]
                if ADD_SOURCE_ROW_COLUMN:
                    prefix.append(index)
                yield prefix + buffer
        finally:
            workbook.close()


def write_xlsx(target: Path, header_row: list[str], rows) -> int:
    workbook = openpyxl.Workbook(write_only=True)
    try:
        workbook._named_styles["Normal"].font = Font(name="Calibri", size=11)
    except Exception:
        pass

    sheet = workbook.create_sheet("appended data")
    sheet.freeze_panes = "A2"

    bold = Font(name="Calibri", size=11, bold=True)
    header_cells = []
    for text in header_row:
        cell = openpyxl.cell.WriteOnlyCell(sheet, value=text)
        cell.font = bold
        header_cells.append(cell)
    sheet.append(header_cells)

    written = 0
    for row in rows:
        sheet.append(row)
        written += 1

    workbook.save(target)
    return written


def write_csv(target: Path, header_row: list[str], rows) -> int:
    written = 0
    with open(target, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(header_row)
        for row in rows:
            writer.writerow(row)
            written += 1
    return written


# ----------------------------------------------------------------------------
# Folder picker and author code
# ----------------------------------------------------------------------------

def pick_folder() -> Path | None:
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).expanduser()
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        chosen = filedialog.askdirectory(title="Select the folder containing the workbooks")
        root.destroy()
        return Path(chosen) if chosen else None
    except Exception:
        typed = input("Folder containing the workbooks: ").strip().strip('"')
        return Path(typed).expanduser() if typed else None


def get_author_code() -> str:
    code = AUTHOR_CODE.strip()
    while not re.fullmatch(r"[A-Za-z]{2}", code):
        try:
            import tkinter as tk
            from tkinter import simpledialog

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            answer = simpledialog.askstring(
                "Author code",
                "Please provide your 2-letter author code (your initials)",
            )
            root.destroy()
        except Exception:
            answer = input(
                "Please provide your 2-letter author code (your initials): "
            )
        if answer is None:
            print("An author code is required for the file name. Exiting.")
            sys.exit(1)
        code = answer.strip()
    return code.lower()


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main() -> None:
    folder = pick_folder()
    if not folder or not folder.is_dir():
        print("No valid folder selected. Exiting.")
        return

    pattern = "**/*" if RECURSE_SUBFOLDERS else "*"
    files = sorted(
        p for p in folder.glob(pattern)
        if p.is_file()
        and p.suffix.lower() in FILE_EXTENSIONS
        and not p.name.startswith("~$")          # ignore Excel lock files
        and "unhidden" not in p.stem.lower()     # ignore our own copies
    )
    if not files:
        print(f"No {' / '.join(FILE_EXTENSIONS)} files found in {folder}")
        return

    author = get_author_code()
    stamp = date.today().strftime("%y%m%d")
    base_name = f"{stamp} {DOC_TYPE} {author} {DESCRIPTION}".lower()

    output_dir = folder / "appended output"
    output_dir.mkdir(parents=True, exist_ok=True)

    log: list[str] = [
        f"Workbook append run  {datetime.now():%Y-%m-%d %H:%M:%S}",
        f"Source folder: {folder}",
        f"Files found: {len(files)}",
    ]

    print(f"Found {len(files)} workbook(s) in {folder}\n")

    if MAKE_UNHIDDEN_COPIES:
        print("Step 1/3  writing unhidden copies ...")
        log.append("\n--- UNHIDDEN COPIES ---")
        copies_dir = output_dir / "unhidden copies"
        for path in files:
            log.append(f"\n{path.name}")
            write_unhidden_copy(path, copies_dir, log)

    print("Step 2/3  scanning sheets and collecting headers ...")
    log.append("\n--- SHEET SCAN ---")
    plans, column_order, column_index = scan_workbooks(files, log)

    if not plans:
        print("Nothing to append -- every sheet was empty, a pivot, or unreadable.")
        (output_dir / f"{base_name} log.txt").write_text("\n".join(log), encoding="utf-8")
        return

    header_row = list(FIXED_COLUMNS)
    if ADD_SOURCE_ROW_COLUMN:
        header_row.append("source row number")
    header_row += column_order

    if len(header_row) > EXCEL_MAX_COLS:
        print(f"! {len(header_row)} columns exceeds Excel's limit -- writing CSV.")

    total_rows = sum(plan.row_count for plan in plans)
    use_csv = (total_rows + 1 > EXCEL_MAX_ROWS) or (len(header_row) > EXCEL_MAX_COLS)

    print(
        f"          {len(plans)} sheet(s) to append, "
        f"~{total_rows:,} data row(s), {len(header_row)} column(s)"
    )
    print("Step 3/3  writing output ...")

    rows = iter_output_rows(plans, column_order, column_index, log)
    if use_csv:
        target = output_dir / f"{base_name}.csv"
        written = write_csv(target, header_row, rows)
    else:
        target = output_dir / f"{base_name}.xlsx"
        written = write_xlsx(target, header_row, rows)

    log.append("\n--- RESULT ---")
    log.append(f"Sheets appended: {len(plans)}")
    log.append(f"Rows written:    {written}")
    log.append(f"Columns:         {len(header_row)}")
    log.append(f"Output file:     {target}")
    log.append("\nColumn order:")
    for position, name in enumerate(header_row, start=1):
        log.append(f"  {position:>4}  {name}")

    log_path = output_dir / f"{base_name} log.txt"
    log_path.write_text("\n".join(log), encoding="utf-8")

    print(f"\nDone. {written:,} row(s) written to:\n  {target}\nLog:\n  {log_path}")
    print("\nSave the output in the appropriate Global Insider folder.")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        input("\nAn error occurred. Press Enter to close.")
