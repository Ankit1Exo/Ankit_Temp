"""
count_excel_rows.py

Counts the number of rows containing data in every sheet of every Excel file
(.xlsx, .xlsm, .xls, .xlsb) inside a folder (including subfolders), and writes
a summary CSV:

    File Name | Sheet Name | Entity Count per sheet | Entity count per file

Designed for large batches (thousands of files) using multiprocessing.

USAGE:
    python count_excel_rows.py "C:\path\to\folder" -o output.csv

REQUIREMENTS:
    pip install openpyxl xlrd pyxlsb
    (xlrd only needed for legacy .xls, pyxlsb only needed for .xlsb)
"""

import argparse
import csv
import os
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

SUPPORTED_EXT = {".xlsx", ".xlsm", ".xls", ".xlsb"}


def count_rows_xlsx_or_xlsm(path):
    """Fast row-with-data count for .xlsx/.xlsm using openpyxl read-only mode."""
    import openpyxl

    results = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = 0
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    count += 1
            results.append((sheet_name, count))
    finally:
        wb.close()
    return results


def count_rows_xls(path):
    """Row-with-data count for legacy .xls using xlrd."""
    import xlrd

    results = []
    wb = xlrd.open_workbook(path, on_demand=True)
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        count = 0
        for r in range(ws.nrows):
            row_vals = ws.row_values(r)
            if any(str(v).strip() != "" for v in row_vals):
                count += 1
        results.append((sheet_name, count))
        wb.unload_sheet(sheet_name)
    return results


def count_rows_xlsb(path):
    """Row-with-data count for .xlsb using pyxlsb."""
    from pyxlsb import open_workbook

    results = []
    with open_workbook(path) as wb:
        for sheet_name in wb.sheets:
            with wb.get_sheet(sheet_name) as ws:
                count = 0
                for row in ws.rows():
                    if any(
                        c.v is not None and str(c.v).strip() != "" for c in row
                    ):
                        count += 1
                results.append((sheet_name, count))
    return results


def process_file(path_str):
    path = Path(path_str)
    ext = path.suffix.lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            sheet_counts = count_rows_xlsx_or_xlsm(path)
        elif ext == ".xls":
            sheet_counts = count_rows_xls(path)
        elif ext == ".xlsb":
            sheet_counts = count_rows_xlsb(path)
        else:
            return path.name, None, f"Unsupported extension: {ext}"

        total = sum(c for _, c in sheet_counts)
        rows = [
            (path.name, sheet_name, count, total)
            for sheet_name, count in sheet_counts
        ]
        return path.name, rows, None
    except Exception as e:
        return path.name, None, str(e)


def find_excel_files(folder):
    for root, _dirs, files in os.walk(folder):
        for f in files:
            # skip temp/lock files like ~$file.xlsx
            if f.startswith("~$"):
                continue
            ext = os.path.splitext(f)[1].lower()
            if ext in SUPPORTED_EXT:
                yield os.path.join(root, f)


def main():
    parser = argparse.ArgumentParser(description="Count data rows per sheet across Excel files.")
    parser.add_argument("folder", help="Folder containing the Excel files (searched recursively)")
    parser.add_argument("-o", "--output", default="excel_row_counts.csv", help="Output CSV path")
    parser.add_argument("-w", "--workers", type=int, default=os.cpu_count(), help="Parallel worker processes")
    args = parser.parse_args()

    files = list(find_excel_files(args.folder))
    print(f"Found {len(files)} Excel files. Starting scan with {args.workers} workers...")

    all_rows = []
    errors = []
    done = 0

    with ProcessPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(process_file, f): f for f in files}
        for future in as_completed(futures):
            fname, rows, err = future.result()
            done += 1
            if err:
                errors.append((fname, err))
            else:
                all_rows.extend(rows)

            if done % 100 == 0 or done == len(files):
                print(f"  processed {done}/{len(files)}", flush=True)

    # write main output
    with open(args.output, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["File Name", "Sheet Name", "Entity Count per sheet", "Entity count per file"])
        writer.writerows(all_rows)

    print(f"\nDone. Wrote {len(all_rows)} sheet-rows to {args.output}")

    # write error log if any
    if errors:
        err_path = os.path.splitext(args.output)[0] + "_errors.csv"
        with open(err_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["File Name", "Error"])
            writer.writerows(errors)
        print(f"{len(errors)} files failed to process. See {err_path}")


if __name__ == "__main__":
    main()
