"""
CSV / Excel  ->  SQL Server uploader  (Windows Authentication)

What it does
------------
* Pick a .csv, .xlsx, .xlsm or .xls file with a file dialog.
* Type the server, database, schema and table name.
* If the table does not exist it is CREATED with every column as NVARCHAR(MAX).
* The table is TRUNCATED, then every row of the file is inserted as text.
  Whatever the source data type is (number, date, boolean, text) it is
  converted to a string, so mismatched types can never fail the load.

Nothing is written to disk - the file is read straight into memory and
pushed to SQL Server over the connection. No staging CSV, no BULK INSERT,
no file share on the server.

Requirements
------------
    pip install pyodbc openpyxl
    (xlrd is only needed for legacy .xls files:  pip install xlrd)
    Plus a Microsoft ODBC Driver for SQL Server (17 or 18).

Run it: open in IDLE and press F5, or double-click. No arguments needed.

Security note: this uses your own Windows login. It never asks for,
stores or writes out a password.
"""

import csv
import os
import queue
import threading
import traceback
from datetime import date, datetime, time as dtime
from decimal import Decimal

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

APP_TITLE = "CSV / Excel  ->  SQL Server uploader (Windows auth)"
BATCH_SIZE = 1000
CSV_EXT = {".csv", ".txt", ".tsv"}
XLSX_EXT = {".xlsx", ".xlsm"}
XLS_EXT = {".xls"}

DELIMITERS = {
    "Auto detect": None,
    "Comma  ,": ",",
    "Semicolon  ;": ";",
    "Tab": "\t",
    "Pipe  |": "|",
}
ENCODINGS = ["Auto detect", "utf-8-sig", "utf-8", "cp1252", "latin-1", "utf-16"]


# --------------------------------------------------------------------------
# value / identifier helpers
# --------------------------------------------------------------------------
def bracket(name):
    """Wrap an identifier in [] and escape any closing bracket."""
    return "[" + str(name).replace("]", "]]") + "]"


def cell_to_text(value, trim=True, blank_is_null=True):
    """Turn any cell value into a string suitable for an NVARCHAR(MAX) column."""
    if value is None:
        return None

    if isinstance(value, str):
        text = value.strip() if trim else value
    elif isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if value != value:                       # NaN
            return None
        if value in (float("inf"), float("-inf")):
            return None
        text = str(int(value)) if value.is_integer() and abs(value) < 1e15 else repr(value)
    elif isinstance(value, Decimal):
        text = format(value, "f")
    elif isinstance(value, datetime):
        # Excel stores plain dates as midnight datetimes - keep those date-only.
        text = value.strftime("%Y-%m-%d") if value.time() == dtime(0, 0) \
            else value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, date):
        text = value.isoformat()
    elif isinstance(value, dtime):
        text = value.strftime("%H:%M:%S")
    else:
        text = str(value)

    if blank_is_null and not text.strip():
        return None
    return text


def clean_headers(raw_headers):
    """Make a usable, unique column name list out of the header row."""
    names, seen = [], {}
    for index, raw in enumerate(raw_headers, start=1):
        name = "" if raw is None else str(raw)
        name = " ".join(name.split())            # collapse newlines / runs of spaces
        if not name:
            name = "Column_%d" % index
        name = name[:120]
        key = name.lower()
        if key in seen:
            seen[key] += 1
            name = "%s_%d" % (name, seen[key])
            seen.setdefault(name.lower(), 1)
        else:
            seen[key] = 1
        names.append(name)
    return names


# --------------------------------------------------------------------------
# file readers
# --------------------------------------------------------------------------
def sniff_csv(path, delimiter_choice, encoding_choice):
    """Work out the encoding + delimiter and return them with the raw header row."""
    candidates = [encoding_choice] if encoding_choice != "Auto detect" \
        else ["utf-8-sig", "cp1252", "latin-1"]

    last_error = None
    for encoding in candidates:
        try:
            with open(path, "r", newline="", encoding=encoding) as handle:
                sample = handle.read(64 * 1024)
                handle.seek(0)
                delimiter = DELIMITERS.get(delimiter_choice)
                if delimiter is None:
                    try:
                        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
                    except csv.Error:
                        delimiter = "\t" if path.lower().endswith(".tsv") else ","
                reader = csv.reader(handle, delimiter=delimiter)
                for row in reader:
                    if any(cell.strip() for cell in row):
                        return encoding, delimiter, row
                raise ValueError("The file has no header row - it looks empty.")
        except UnicodeDecodeError as error:
            last_error = error
    raise ValueError(
        "Could not decode the file. Tried: %s.\nPick a specific encoding and retry.\n(%s)"
        % (", ".join(candidates), last_error)
    )


def read_csv_rows(path, encoding, delimiter):
    """Return every row of the CSV, header included."""
    with open(path, "r", newline="", encoding=encoding) as handle:
        return [row for row in csv.reader(handle, delimiter=delimiter)]


def excel_sheet_names(path):
    extension = os.path.splitext(path)[1].lower()
    if extension in XLSX_EXT:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    import xlrd
    return list(xlrd.open_workbook(path, on_demand=True).sheet_names())


def read_excel_rows(path, sheet_name):
    """Return every row of the sheet, header included."""
    extension = os.path.splitext(path)[1].lower()
    if extension in XLSX_EXT:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
            return [list(row) for row in worksheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    import xlrd
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)
    rows = []
    for row_index in range(sheet.nrows):
        row = []
        for column_index in range(sheet.ncols):
            cell = sheet.cell(row_index, column_index)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                row.append(bool(cell.value))
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


def read_all_rows(path, sheet_name, encoding, delimiter):
    if os.path.splitext(path)[1].lower() in CSV_EXT:
        return read_csv_rows(path, encoding, delimiter)
    return read_excel_rows(path, sheet_name)


def peek_headers(path, sheet_name, delimiter_choice, encoding_choice):
    """Read only the header row so the GUI stays responsive.

    Returns (clean column names, encoding, delimiter). The last two are None
    for Excel sources.
    """
    if os.path.splitext(path)[1].lower() in CSV_EXT:
        encoding, delimiter, raw = sniff_csv(path, delimiter_choice, encoding_choice)
        return clean_headers(raw), encoding, delimiter

    extension = os.path.splitext(path)[1].lower()
    if extension in XLSX_EXT:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
            for row in worksheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    return clean_headers(row), None, None
        finally:
            workbook.close()
        raise ValueError("That sheet has no header row - it looks empty.")

    for row in read_excel_rows(path, sheet_name):
        if any(cell is not None and str(cell).strip() for cell in row):
            return clean_headers(row), None, None
    raise ValueError("That sheet has no header row - it looks empty.")


# --------------------------------------------------------------------------
# SQL Server helpers
# --------------------------------------------------------------------------
def pick_driver(pyodbc):
    installed = pyodbc.drivers()
    for preferred in ("ODBC Driver 18 for SQL Server",
                      "ODBC Driver 17 for SQL Server",
                      "ODBC Driver 13.1 for SQL Server",
                      "ODBC Driver 13 for SQL Server",
                      "ODBC Driver 11 for SQL Server",
                      "SQL Server Native Client 11.0",
                      "SQL Server"):
        if preferred in installed:
            return preferred
    if installed:
        return installed[0]
    raise RuntimeError(
        "No ODBC driver found. Install 'Microsoft ODBC Driver 18 for SQL Server'."
    )


def build_connection_string(server, database, driver):
    parts = [
        "DRIVER={%s}" % driver,
        "SERVER=%s" % server,
        "DATABASE=%s" % database,
        "Trusted_Connection=yes",
        "APP=CSV Excel uploader",
    ]
    if "18" in driver or "17" in driver:
        # Driver 18 encrypts by default; without this a self-signed cert fails.
        parts.append("TrustServerCertificate=yes")
    return ";".join(parts) + ";"


def connect(server, database, driver):
    import pyodbc
    return pyodbc.connect(build_connection_string(server, database, driver),
                          autocommit=False, timeout=30)


def fetch_table_columns(cursor, schema, table):
    """Return [(name, type_name, max_length)] or None when the table does not exist."""
    full_name = "%s.%s" % (bracket(schema), bracket(table))
    cursor.execute("SELECT OBJECT_ID(?, 'U')", full_name)
    if cursor.fetchone()[0] is None:
        return None
    cursor.execute(
        """
        SELECT c.name, t.name, c.max_length
        FROM sys.columns AS c
        JOIN sys.types   AS t ON t.user_type_id = c.user_type_id
        WHERE c.object_id = OBJECT_ID(?, 'U')
        ORDER BY c.column_id
        """,
        full_name,
    )
    return [(row[0], row[1], row[2]) for row in cursor.fetchall()]


# --------------------------------------------------------------------------
# the upload worker (runs off the GUI thread)
# --------------------------------------------------------------------------
class UploadWorker(threading.Thread):
    def __init__(self, plan, messages, cancel_event):
        threading.Thread.__init__(self, daemon=True)
        self.plan = plan
        self.messages = messages
        self.cancel_event = cancel_event

    def say(self, text):
        self.messages.put(("log", text))

    def run(self):
        plan = self.plan
        connection = None
        try:
            self.say("Reading %s ..." % os.path.basename(plan["path"]))
            rows = read_all_rows(plan["path"], plan["sheet"], plan["encoding"], plan["delimiter"])
            if not rows:
                raise ValueError("The file is empty.")

            header_index = 0
            for index, row in enumerate(rows):
                if any(cell is not None and str(cell).strip() for cell in row):
                    header_index = index
                    break
            data_rows = rows[header_index + 1:]
            self.say("%d data rows, %d columns." % (len(data_rows), len(plan["file_columns"])))

            # file column position -> target column name (unmapped columns are skipped)
            positions = [position for position, _ in plan["mapping"]]
            target_columns = [name for _, name in plan["mapping"]]

            self.say("Converting values to text ...")
            payload, blank_rows = [], 0
            trim, blank_is_null = plan["trim"], plan["blank_is_null"]
            for row in data_rows:
                if not any(cell is not None and str(cell).strip() for cell in row):
                    blank_rows += 1
                    continue
                payload.append([
                    cell_to_text(row[position] if position < len(row) else None,
                                 trim, blank_is_null)
                    for position in positions
                ])
            if blank_rows:
                self.say("Skipped %d completely blank row(s)." % blank_rows)
            if not payload:
                raise ValueError("There are no data rows to upload.")

            self.messages.put(("max", len(payload)))

            self.say("Connecting to %s / %s ..." % (plan["server"], plan["database"]))
            connection = connect(plan["server"], plan["database"], plan["driver"])
            cursor = connection.cursor()

            full_name = "%s.%s" % (bracket(plan["schema"]), bracket(plan["table"]))

            if plan["create_table"]:
                columns_sql = ",\n    ".join("%s NVARCHAR(MAX) NULL" % bracket(name)
                                             for name in plan["file_columns"])
                cursor.execute("CREATE TABLE %s (\n    %s\n)" % (full_name, columns_sql))
                self.say("Created table %s with %d NVARCHAR(MAX) columns."
                         % (full_name, len(plan["file_columns"])))
            else:
                self.say("Using existing table %s." % full_name)

            self.say("Emptying the table ...")
            try:
                cursor.execute("TRUNCATE TABLE %s" % full_name)
            except Exception as error:
                self.say("TRUNCATE refused (%s) - falling back to DELETE."
                         % str(error).split("\n")[0][:120])
                cursor.execute("DELETE FROM %s" % full_name)

            insert_sql = "INSERT INTO %s (%s) VALUES (%s)" % (
                full_name,
                ", ".join(bracket(name) for name in target_columns),
                ", ".join("?" for _ in target_columns),
            )

            self.say("Inserting %d rows in batches of %d ..." % (len(payload), BATCH_SIZE))
            inserted = self._insert(cursor, insert_sql, payload, len(target_columns))

            if self.cancel_event.is_set():
                connection.rollback()
                self.messages.put(("done", False,
                                   "Cancelled - rolled back, the table is unchanged."))
                return

            connection.commit()
            self.say("Committed.")
            self.messages.put(("done", True,
                               "Uploaded %d rows into %s." % (inserted, full_name)))

        except Exception as error:
            if connection is not None:
                try:
                    connection.rollback()
                    self.say("Rolled back - the table is unchanged.")
                except Exception:
                    pass
            detail = str(error).strip() or error.__class__.__name__
            self.messages.put(("log", "ERROR: " + detail))
            self.messages.put(("trace", traceback.format_exc()))
            self.messages.put(("done", False, detail))
        finally:
            if connection is not None:
                try:
                    connection.close()
                except Exception:
                    pass

    def _insert(self, cursor, insert_sql, payload, column_count):
        """executemany in batches, with a safe fallback for very long values."""
        import pyodbc

        inserted = 0
        for start in range(0, len(payload), BATCH_SIZE):
            if self.cancel_event.is_set():
                return inserted
            batch = payload[start:start + BATCH_SIZE]

            widths = [0] * column_count
            for record in batch:
                for index, value in enumerate(record):
                    if value is not None and len(value) > widths[index]:
                        widths[index] = len(value)

            # fast_executemany cannot bind values wider than an nvarchar(4000)
            # parameter, so batches containing long text go the ordinary route.
            fast = max(widths) <= 4000 if widths else True
            try:
                cursor.fast_executemany = fast
                if fast:
                    cursor.setinputsizes([(pyodbc.SQL_WVARCHAR, max(1, width), 0)
                                          for width in widths])
                else:
                    cursor.setinputsizes(None)
                cursor.executemany(insert_sql, batch)
            except Exception:
                # Retry the batch one row at a time so the failing row is named.
                cursor.fast_executemany = False
                try:
                    cursor.setinputsizes(None)
                except Exception:
                    pass
                for offset, record in enumerate(batch):
                    try:
                        cursor.execute(insert_sql, record)
                    except Exception as error:
                        raise ValueError("Row %d failed: %s"
                                         % (start + offset + 2, str(error).strip()))
            inserted += len(batch)
            self.messages.put(("progress", inserted))
        return inserted


# --------------------------------------------------------------------------
# GUI
# --------------------------------------------------------------------------
class UploaderApp:
    def __init__(self, root):
        self.root = root
        root.title(APP_TITLE)
        root.geometry("880x660")
        root.minsize(780, 580)

        self.path_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.delimiter_var = tk.StringVar(value="Auto detect")
        self.encoding_var = tk.StringVar(value="Auto detect")
        self.server_var = tk.StringVar()
        self.database_var = tk.StringVar()
        self.schema_var = tk.StringVar(value="dbo")
        self.table_var = tk.StringVar()
        self.trim_var = tk.BooleanVar(value=True)
        self.blank_null_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="Ready.")

        self.messages = queue.Queue()
        self.cancel_event = threading.Event()
        self.worker = None

        self._build_widgets()
        self._toggle_source_fields()
        self.root.after(120, self._drain_messages)

    # ---------------- layout ----------------
    def _build_widgets(self):
        padding = {"padx": 8, "pady": 4}

        source = ttk.LabelFrame(self.root, text="1.  Source file")
        source.pack(fill="x", **padding)
        source.columnconfigure(1, weight=1)

        ttk.Label(source, text="File").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(source, textvariable=self.path_var).grid(row=0, column=1, sticky="ew", pady=6)
        ttk.Button(source, text="Browse...", command=self.browse) \
            .grid(row=0, column=2, padx=8, pady=6)

        ttk.Label(source, text="Sheet").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        self.sheet_box = ttk.Combobox(source, textvariable=self.sheet_var,
                                      state="disabled", values=[])
        self.sheet_box.grid(row=1, column=1, sticky="w", pady=4)

        csv_row = ttk.Frame(source)
        csv_row.grid(row=2, column=0, columnspan=3, sticky="w", padx=8, pady=4)
        ttk.Label(csv_row, text="CSV delimiter").pack(side="left")
        self.delimiter_box = ttk.Combobox(csv_row, textvariable=self.delimiter_var, width=14,
                                          state="readonly", values=list(DELIMITERS))
        self.delimiter_box.pack(side="left", padx=(6, 18))
        ttk.Label(csv_row, text="Encoding").pack(side="left")
        self.encoding_box = ttk.Combobox(csv_row, textvariable=self.encoding_var, width=14,
                                         state="readonly", values=ENCODINGS)
        self.encoding_box.pack(side="left", padx=6)

        target = ttk.LabelFrame(
            self.root, text="2.  SQL Server destination  (Windows authentication)")
        target.pack(fill="x", **padding)
        target.columnconfigure(1, weight=1)
        target.columnconfigure(3, weight=1)

        ttk.Label(target, text="Server").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(target, textvariable=self.server_var).grid(row=0, column=1, sticky="ew", pady=6)
        ttk.Label(target, text="Database").grid(row=0, column=2, sticky="w", padx=8, pady=6)
        ttk.Entry(target, textvariable=self.database_var).grid(row=0, column=3, sticky="ew",
                                                               padx=(0, 8), pady=6)

        ttk.Label(target, text="Schema").grid(row=1, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(target, textvariable=self.schema_var).grid(row=1, column=1, sticky="ew", pady=6)
        ttk.Label(target, text="Table").grid(row=1, column=2, sticky="w", padx=8, pady=6)
        ttk.Entry(target, textvariable=self.table_var).grid(row=1, column=3, sticky="ew",
                                                            padx=(0, 8), pady=6)

        ttk.Button(target, text="Test connection", command=self.test_connection) \
            .grid(row=2, column=3, sticky="e", padx=(0, 8), pady=(0, 8))

        options = ttk.LabelFrame(self.root, text="3.  Options")
        options.pack(fill="x", **padding)
        ttk.Checkbutton(options, text="Trim leading / trailing spaces",
                        variable=self.trim_var).pack(side="left", padx=10, pady=6)
        ttk.Checkbutton(options, text="Store blank cells as NULL",
                        variable=self.blank_null_var).pack(side="left", padx=10, pady=6)
        ttk.Label(options,
                  text="Missing table is created as NVARCHAR(MAX), then emptied before loading.",
                  foreground="#555").pack(side="left", padx=14)

        actions = ttk.Frame(self.root)
        actions.pack(fill="x", **padding)
        self.upload_button = ttk.Button(actions, text="Upload to SQL Server",
                                        command=self.start_upload)
        self.upload_button.pack(side="left")
        self.cancel_button = ttk.Button(actions, text="Cancel", command=self.cancel,
                                        state="disabled")
        self.cancel_button.pack(side="left", padx=8)
        self.progress = ttk.Progressbar(actions, mode="determinate", length=380)
        self.progress.pack(side="right", padx=4)

        ttk.Label(self.root, textvariable=self.status_var, anchor="w").pack(fill="x", padx=10)

        log_frame = ttk.LabelFrame(self.root, text="Log")
        log_frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        self.log = tk.Text(log_frame, height=12, wrap="word", state="disabled")
        scrollbar = ttk.Scrollbar(log_frame, command=self.log.yview)
        self.log.configure(yscrollcommand=scrollbar.set)
        self.log.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    # ---------------- small helpers ----------------
    def write_log(self, text):
        self.log.configure(state="normal")
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _toggle_source_fields(self):
        extension = os.path.splitext(self.path_var.get())[1].lower()
        is_csv = extension in CSV_EXT or not self.path_var.get()
        state = "readonly" if is_csv else "disabled"
        self.delimiter_box.configure(state=state)
        self.encoding_box.configure(state=state)

    # ---------------- events ----------------
    def browse(self):
        path = filedialog.askopenfilename(
            title="Select the CSV or Excel file to upload",
            filetypes=[("Spreadsheets and CSV", "*.csv *.txt *.tsv *.xlsx *.xlsm *.xls"),
                       ("CSV files", "*.csv *.txt *.tsv"),
                       ("Excel files", "*.xlsx *.xlsm *.xls"),
                       ("All files", "*.*")],
        )
        if not path:
            return
        self.path_var.set(path)
        self._toggle_source_fields()

        extension = os.path.splitext(path)[1].lower()
        if extension in XLSX_EXT or extension in XLS_EXT:
            try:
                sheets = excel_sheet_names(path)
            except ImportError:
                self.sheet_box.configure(values=[], state="disabled")
                messagebox.showerror(
                    "Missing module",
                    "Legacy .xls files need the xlrd module:\n\n    pip install xlrd\n\n"
                    "Or save the workbook as .xlsx and try again.")
                return
            except Exception as error:
                self.sheet_box.configure(values=[], state="disabled")
                messagebox.showerror("Cannot read workbook", str(error))
                return
            self.sheet_box.configure(values=sheets, state="readonly")
            self.sheet_var.set(sheets[0] if sheets else "")
        else:
            self.sheet_box.configure(values=[], state="disabled")
            self.sheet_var.set("")

        if not self.table_var.get():
            suggestion = os.path.splitext(os.path.basename(path))[0]
            suggestion = "".join(character if character.isalnum() else "_"
                                 for character in suggestion).strip("_")
            self.table_var.set(suggestion or "UploadedData")

        self.write_log("Selected %s" % path)

    def _require_pyodbc(self):
        try:
            import pyodbc                                   # noqa: F401
            return True
        except ImportError:
            messagebox.showerror(
                "Missing module",
                "pyodbc is not installed.\n\nOpen a command prompt and run:\n\n"
                "    pip install pyodbc\n\nThen restart this script.")
            return False

    def _validate(self, need_file=True):
        if need_file:
            if not self.path_var.get().strip():
                messagebox.showwarning("Missing file", "Choose a CSV or Excel file first.")
                return False
            if not os.path.isfile(self.path_var.get()):
                messagebox.showwarning("Missing file", "That file does not exist any more.")
                return False
        for label, variable in (("server", self.server_var),
                                ("database", self.database_var),
                                ("schema", self.schema_var)):
            if not variable.get().strip():
                messagebox.showwarning("Missing detail", "Enter the %s name." % label)
                return False
        if need_file and not self.table_var.get().strip():
            messagebox.showwarning("Missing detail", "Enter the table name.")
            return False
        return True

    def test_connection(self):
        if not self._require_pyodbc() or not self._validate(need_file=False):
            return
        import pyodbc
        try:
            driver = pick_driver(pyodbc)
            connection = connect(self.server_var.get().strip(),
                                 self.database_var.get().strip(), driver)
            cursor = connection.cursor()
            cursor.execute("SELECT SUSER_SNAME(), DB_NAME(), @@VERSION")
            login, database, version = cursor.fetchone()
            connection.close()
            self.write_log("Connected via %s as %s to %s" % (driver, login, database))
            messagebox.showinfo(
                "Connection OK",
                "Driver:   %s\nLogin:    %s\nDatabase: %s\n\n%s"
                % (driver, login, database, version.split("\n")[0]))
        except Exception as error:
            self.write_log("Connection failed: %s" % error)
            messagebox.showerror("Connection failed", str(error))

    def start_upload(self):
        if self.worker is not None and self.worker.is_alive():
            return
        if not self._require_pyodbc() or not self._validate():
            return

        import pyodbc
        path = self.path_var.get().strip()
        sheet = self.sheet_var.get() or None
        server = self.server_var.get().strip()
        database = self.database_var.get().strip()
        schema = self.schema_var.get().strip()
        table = self.table_var.get().strip()

        # 1. header row -------------------------------------------------
        try:
            file_columns, encoding, delimiter = peek_headers(
                path, sheet, self.delimiter_var.get(), self.encoding_var.get())
        except ImportError:
            messagebox.showerror("Missing module",
                                 "Legacy .xls files need xlrd:\n\n    pip install xlrd")
            return
        except Exception as error:
            messagebox.showerror("Cannot read the file", str(error))
            return

        if encoding:
            self.write_log("CSV encoding %s, delimiter %r" % (encoding, delimiter))
        self.write_log("File columns (%d): %s" % (len(file_columns), ", ".join(file_columns)))

        # 2. inspect the destination -----------------------------------
        try:
            driver = pick_driver(pyodbc)
            connection = connect(server, database, driver)
            cursor = connection.cursor()
            existing = fetch_table_columns(cursor, schema, table)
            connection.close()
        except Exception as error:
            self.write_log("Connection failed: %s" % error)
            messagebox.showerror("Connection failed", str(error))
            return

        # 3. decide the column mapping ---------------------------------
        if existing is None:
            create_table = True
            mapping = list(enumerate(file_columns))
            summary = ("Table %s.%s does not exist.\n\nIt will be created with %d columns, "
                       "all NVARCHAR(MAX), then loaded with the file's rows.\n\nContinue?"
                       % (schema, table, len(file_columns)))
        else:
            create_table = False
            by_lower = {name.lower(): (name, type_name, max_length)
                        for name, type_name, max_length in existing}
            mapping, unmatched, narrow = [], [], []
            for position, column in enumerate(file_columns):
                found = by_lower.get(column.lower())
                if found is None:
                    unmatched.append(column)
                    continue
                mapping.append((position, found[0]))
                is_text = found[1] in ("nvarchar", "varchar", "ntext", "text", "xml")
                if not is_text or (found[1] in ("nvarchar", "varchar") and found[2] != -1):
                    narrow.append("%s (%s)" % (found[0], found[1]))
            if not mapping:
                messagebox.showerror(
                    "No matching columns",
                    "None of the file's column names match the columns in %s.%s.\n\n"
                    "Table columns: %s"
                    % (schema, table, ", ".join(name for name, _, _ in existing)))
                return
            file_lower = {column.lower() for column in file_columns}
            missing_in_file = [name for name, _, _ in existing if name.lower() not in file_lower]
            lines = ["Table %s.%s already exists." % (schema, table),
                     "",
                     "It will be EMPTIED, then %d of the file's %d columns will be loaded."
                     % (len(mapping), len(file_columns))]
            if unmatched:
                lines += ["", "File columns with no matching table column "
                              "(their data will NOT be uploaded):",
                          "  " + ", ".join(unmatched)]
            if missing_in_file:
                lines += ["", "Table columns not present in the file (left NULL):",
                          "  " + ", ".join(missing_in_file)]
            if narrow:
                lines += ["", "Columns that are not NVARCHAR(MAX) - long values may be rejected:",
                          "  " + ", ".join(narrow)]
            lines += ["", "Continue?"]
            summary = "\n".join(lines)

        if not messagebox.askokcancel("Confirm upload", summary):
            self.write_log("Cancelled at the confirmation step.")
            return

        # 4. hand over to the worker -----------------------------------
        plan = {
            "path": path, "sheet": sheet, "encoding": encoding, "delimiter": delimiter,
            "server": server, "database": database, "schema": schema, "table": table,
            "driver": driver, "file_columns": file_columns, "mapping": mapping,
            "create_table": create_table,
            "trim": self.trim_var.get(), "blank_is_null": self.blank_null_var.get(),
        }

        self.cancel_event.clear()
        self.progress.configure(value=0, maximum=100)
        self.upload_button.configure(state="disabled")
        self.cancel_button.configure(state="normal")
        self.status_var.set("Working...")
        self.write_log("-" * 70)
        self.worker = UploadWorker(plan, self.messages, self.cancel_event)
        self.worker.start()

    def cancel(self):
        if self.worker is not None and self.worker.is_alive():
            self.cancel_event.set()
            self.status_var.set("Cancelling - the load will be rolled back...")
            self.write_log("Cancel requested.")

    def _drain_messages(self):
        try:
            while True:
                message = self.messages.get_nowait()
                kind = message[0]
                if kind == "log":
                    self.write_log(message[1])
                    self.status_var.set(message[1])
                elif kind == "trace":
                    self.write_log(message[1])
                elif kind == "max":
                    self.progress.configure(maximum=max(1, message[1]), value=0)
                elif kind == "progress":
                    self.progress.configure(value=message[1])
                elif kind == "done":
                    ok, text = message[1], message[2]
                    self.upload_button.configure(state="normal")
                    self.cancel_button.configure(state="disabled")
                    self.status_var.set(text)
                    self.write_log(text)
                    if ok:
                        messagebox.showinfo("Upload finished", text)
                    else:
                        messagebox.showerror("Upload failed", text)
        except queue.Empty:
            pass
        self.root.after(120, self._drain_messages)


def main():
    root = tk.Tk()
    try:
        ttk.Style().theme_use("vista")
    except tk.TclError:
        pass
    UploaderApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
