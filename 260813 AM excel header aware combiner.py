"""
260813 AM excel header aware combiner.py

Combines every sheet of every Excel/CSV file in a folder into ONE table, even
when the header row is not row 1.

HOW THE HEADER ROW IS FOUND
    Each sheet is scanned from row 1 down to row 30 (HEADER_SCAN_ROWS). Every
    row holding a cell that reads "EmployeeName" is a candidate; the candidate
    with the MOST filled cells wins, and a tie goes to the LOWER row.
    Everything above the winner (PayDate, PeriodStart, PeriodEnd, report
    titles, blank rows, ...) is discarded.

    Picking by width matters because some layouts print the report parameters
    HORIZONTALLY - one row reading "EmployeeName | PeriodStart | PeriodEnd |
    PayDate" - which also contains the word EmployeeName. That parameter row
    is only a handful of cells wide, so the real header row beats it.

    Matching is LOOSE: case, spaces, underscores and punctuation are ignored,
    so EmployeeName / "Employee Name" / EMPLOYEE_NAME / "employee-name" all
    match. A header that merely STARTS with the keyword also matches, e.g.
    "EmployeeName (Last, First)".

    If no such row is found in the first 30 rows, the sheet is SKIPPED and a
    line is written into the NOTES block at the bottom of the output table.

    Use the "Diagnose headers" button to print the top rows of every sheet and
    see exactly which row was chosen and which rivals it beat.

HIDDEN DATA
    Nothing is filtered out on visibility. Hidden and very-hidden SHEETS,
    hidden ROWS and hidden COLUMNS are all read and combined - the effect is
    the same as unhiding everything first, but the source files are never
    modified.

COLUMN MATCHING
    Header text is matched TRIMMED and CASE-INSENSITIVELY. Columns that match
    line up under the same output column; columns never seen before are
    appended at the RIGHT-HAND END of the table. Sheets without a given column
    are left blank there. The first spelling seen becomes the output heading.

    By default spacing and punctuation are ignored too, because these reports
    use wrapped headers whose spacing drifts between exports - "AmountCurrent"
    in one file and "Amount Current" in the next are the same column, as are
    "SSN_SIN" / "SSN SIN" and "EmployeeNumber" / "Employee Number". Without
    this you get near-duplicate columns side by side. Set
    COLUMN_MATCH_IGNORES_SPACING = False for strict trim + case matching only.

    Duplicate headers inside the SAME sheet are kept apart as "Name",
    "Name (2)". A data column whose header cell is blank is named after its
    Excel column letter, e.g. "Column T". Columns that are blank in both the
    header and every data row are dropped.

    The employee-name column is the one exception to the strict matching: any
    spelling that identified the header row (EmployeeName, "Employee Name",
    EMPLOYEE_NAME, ...) is always folded into a single output column named
    "EmployeeName", so names never end up split across look-alike columns.

OUTPUT
    Column 1 = Source File, Column 2 = Source Sheet, then the data columns.
    Written as .xlsx when it fits inside Excel's limits (1,048,576 rows /
    16,384 columns). If it does not fit, the output is written instead as
    numbered CSV parts of up to 1,000,000 data rows each:
        <name> part 1 of 3.csv, <name> part 2 of 3.csv, ...
    The NOTES block is appended at the end of the table (and, for .xlsx, also
    repeated on a separate "Notes" sheet).

SUPPORTED INPUT
    .xlsx .xlsm .xltx .xltm   - built in (openpyxl)
    .xls                      - needs:  pip install xlrd
    .xlsb                     - needs:  pip install pyxlsb
    .csv                      - built in
    Missing optional libraries are reported in the NOTES block, not crashed on.

USAGE
    python "260813 AM excel header aware combiner.py"
    (Tkinter GUI: pick the source folder and the output file, then Run.)

NOTE ON FORMULAS
    Cells are read as cached VALUES. A workbook that has never been opened and
    saved by Excel may have no cached values, in which case formula cells come
    through blank.

NOTE ON SENSITIVE DATA
    These registers typically contain employee names, SSN/SIN and health-plan
    deduction data. Run this locally and save the output only to the approved
    Global Insider folder. Do not e-mail or upload the combined file.
"""

import csv
import os
import re
import threading
import traceback
from datetime import date, datetime, time

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

# How far down each sheet to look for the header row.
HEADER_SCAN_ROWS = 30

# A candidate header row must hold at least this many filled cells. Stops a
# one-off label such as a lone "EmployeeName:" caption from being mistaken for
# the header row of the table.
MIN_HEADER_CELLS = 2

# A row is the header row when one of its cells matches (or starts with) one of
# these keywords, after lowercasing and stripping every non-alphanumeric
# character. Add alternates here if some files use a different label.
HEADER_KEYWORDS = ("employeename",)

# Column matching is always trimmed and case-insensitive. Leaving this True
# also ignores spaces, line breaks and punctuation, so a header that wraps as
# "Amount Current" in one export and "AmountCurrent" in the next still lands in
# one column. Set it to False for strict trim + case matching only.
COLUMN_MATCH_IGNORES_SPACING = True

# Every spelling of the employee-name header is written into this one column.
EMPLOYEE_COLUMN_NAME = "EmployeeName"
EMPLOYEE_COLUMN_KEY = "\x00employee-name-column"

SOURCE_FILE_COLUMN = "Source File"
SOURCE_SHEET_COLUMN = "Source Sheet"

EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLS = 16_384
CSV_ROWS_PER_PART = 1_000_000

OPENPYXL_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm")
SUPPORTED_EXTENSIONS = OPENPYXL_EXTENSIONS + (".xls", ".xlsb", ".csv")

DEFAULT_OUTPUT_NAME = "260813 AM combined output.xlsx"


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------

def is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def keyword_key(value):
    """Loose key used to spot the header row: lowercase, every non-alphanumeric
    character removed. 'Employee Name' and 'EMPLOYEE_NAME' both -> employeename."""
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def match_key(header_text):
    """Key used to line columns up across sheets: trimmed and lowercased, with
    spacing and punctuation dropped as well unless strict matching is asked
    for. Header cells in these reports wrap, so "Amount Current" and
    "AmountCurrent" have to reach the same key."""
    text = "" if header_text is None else str(header_text)
    if COLUMN_MATCH_IGNORES_SPACING:
        return re.sub(r"[^a-z0-9]+", "", text.lower())
    return " ".join(text.split()).lower()


def is_employee_header(value):
    """True when this single cell is the employee-name header, ignoring case,
    spaces and punctuation - 'EmployeeName (Last, First)' counts too."""
    key = keyword_key(value)
    if not key:
        return False
    return any(key == keyword or key.startswith(keyword) for keyword in HEADER_KEYWORDS)


def row_is_header(row_values):
    """True when any cell in the row names the employee-name column."""
    return any(is_employee_header(v) for v in row_values)


def row_is_blank(row_values):
    return all(is_blank(v) for v in row_values)


def csv_text(value):
    """Render a cell for CSV output without losing leading zeros or date parts."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value


# ---------------------------------------------------------------------------
# Sheet readers - each yields (sheet_name, rows, hidden_flag)
# rows is a list of lists of raw cell values, every row padded to equal width.
# Hidden sheets / rows / columns are NOT filtered out anywhere below.
# ---------------------------------------------------------------------------

def read_openpyxl(path, notes, log):
    """.xlsx / .xlsm / .xltx / .xltm - read-only streaming, all sheets."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            state = getattr(worksheet, "sheet_state", "visible")
            rows = []
            width = 0
            for row in worksheet.iter_rows(values_only=True):
                values = list(row)
                rows.append(values)
                width = max(width, len(values))
            for values in rows:
                if len(values) < width:
                    values.extend([None] * (width - len(values)))
            yield worksheet.title, rows, state != "visible"
    finally:
        workbook.close()


def read_xls(path, notes, log):
    """Legacy .xls via xlrd. Reads the full used range of every sheet."""
    try:
        import xlrd
    except ImportError:
        notes.append(
            f"{path.name if hasattr(path, 'name') else os.path.basename(path)}: "
            "legacy .xls file skipped - the xlrd library is not installed "
            "(run: pip install xlrd)"
        )
        return

    book = xlrd.open_workbook(str(path))
    for sheet in book.sheets():
        if sheet.nrows == 0 or sheet.ncols == 0:
            continue

        def cell_value(r, c):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
            if cell.ctype == xlrd.XL_CELL_BOOLEAN:
                return bool(cell.value)
            if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                return None
            return cell.value

        rows = [[cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        # xlrd exposes sheet visibility as 0 = visible, 1 = hidden, 2 = very hidden.
        yield sheet.name, rows, getattr(sheet, "visibility", 0) != 0


def read_xlsb(path, notes, log):
    """.xlsb via pyxlsb."""
    try:
        from pyxlsb import open_workbook as open_xlsb
    except ImportError:
        notes.append(
            f"{os.path.basename(str(path))}: .xlsb file skipped - the pyxlsb "
            "library is not installed (run: pip install pyxlsb)"
        )
        return

    with open_xlsb(str(path)) as book:
        for sheet_name in book.sheets:
            with book.get_sheet(sheet_name) as sheet:
                rows = []
                width = 0
                for row in sheet.rows():
                    values = [c.v for c in row]
                    rows.append(values)
                    width = max(width, len(values))
                for values in rows:
                    if len(values) < width:
                        values.extend([None] * (width - len(values)))
            yield sheet_name, rows, False


def read_csv_file(path, notes, log):
    """A CSV is treated as a single sheet named 'CSV'."""
    rows = []
    width = 0
    with open(str(path), "r", newline="", encoding="utf-8-sig", errors="replace") as handle:
        for values in csv.reader(handle):
            rows.append(list(values))
            width = max(width, len(values))
    for values in rows:
        if len(values) < width:
            values.extend([None] * (width - len(values)))
    yield "CSV", rows, False


def read_any(path, notes, log):
    ext = os.path.splitext(str(path))[1].lower()
    if ext in OPENPYXL_EXTENSIONS:
        yield from read_openpyxl(path, notes, log)
    elif ext == ".xls":
        yield from read_xls(path, notes, log)
    elif ext == ".xlsb":
        yield from read_xlsb(path, notes, log)
    elif ext == ".csv":
        yield from read_csv_file(path, notes, log)
    else:
        raise ValueError(f"Unsupported extension: {ext}")


# ---------------------------------------------------------------------------
# Turning one sheet into a header + data block
# ---------------------------------------------------------------------------

def header_candidates(rows):
    """Every row in the scan window that could be the header row, as a list of
    (row_number, filled_cell_count), row_number being 1-based."""
    found = []
    for index in range(min(len(rows), HEADER_SCAN_ROWS)):
        row = rows[index]
        if not row_is_header(row):
            continue
        filled = sum(1 for value in row if not is_blank(value))
        if filled < MIN_HEADER_CELLS:
            continue
        found.append((index + 1, filled))
    return found


def choose_header_row(rows):
    """Pick the best candidate: widest row wins, ties go to the lower row.

    A horizontally printed parameter block ("EmployeeName | PeriodStart |
    PeriodEnd | PayDate") also contains the keyword but is only a few cells
    wide, so the real header row - which spans every column of the table -
    outscores it. Returns the 1-based row number, or None."""
    found = header_candidates(rows)
    if not found:
        return None
    best = max(found, key=lambda item: (item[1], item[0]))
    return best[0]


def split_header_and_data(rows):
    """Return (header_row, data_rows, chosen_row_number, candidates).
    header_row is None when no usable EmployeeName row exists in the first
    HEADER_SCAN_ROWS rows."""
    found = header_candidates(rows)
    chosen = choose_header_row(rows)
    if chosen is None:
        return None, None, None, found
    return rows[chosen - 1], rows[chosen:], chosen, found


def build_column_names(header_row, data_rows):
    """Name every column, drop the ones that are blank top to bottom, and make
    duplicate names inside this sheet unique.

    Returns (names, keep_indexes, is_employee_flags). The flag marks the FIRST
    employee-name column in this sheet, which is folded into the single
    EMPLOYEE_COLUMN_NAME output column; a second one is left as its own column
    so nothing is silently overwritten."""
    width = len(header_row)
    for row in data_rows:
        width = max(width, len(row))

    keep = []
    for index in range(width):
        header_blank = index >= len(header_row) or is_blank(header_row[index])
        if not header_blank:
            keep.append(index)
            continue
        if any(index < len(row) and not is_blank(row[index]) for row in data_rows):
            keep.append(index)

    names = []
    flags = []
    seen = {}
    employee_taken = False
    for index in keep:
        raw = header_row[index] if index < len(header_row) else None
        text = f"Column {get_column_letter(index + 1)}" if is_blank(raw) else str(raw).strip()

        is_employee = is_employee_header(raw) and not employee_taken
        if is_employee:
            employee_taken = True

        if text in seen:
            seen[text] += 1
            text = f"{text} ({seen[text]})"
        else:
            seen[text] = 1

        names.append(text)
        flags.append(is_employee)
    return names, keep, flags


# ---------------------------------------------------------------------------
# Combine
# ---------------------------------------------------------------------------

def collect_files(source_folder, include_subfolders, exclude_paths):
    """Every supported file in the folder, skipping Excel lock files (~$...)
    and anything we are about to write as output."""
    found = []
    if include_subfolders:
        for root, _dirs, names in os.walk(source_folder):
            for name in names:
                found.append(os.path.join(root, name))
    else:
        for name in sorted(os.listdir(source_folder)):
            full = os.path.join(source_folder, name)
            if os.path.isfile(full):
                found.append(full)

    keep = []
    for full in found:
        name = os.path.basename(full)
        if name.startswith("~$"):
            continue
        if os.path.splitext(name)[1].lower() not in SUPPORTED_EXTENSIONS:
            continue
        if os.path.abspath(full) in exclude_paths:
            continue
        keep.append(full)
    return sorted(keep)


def combine(source_folder, dest_path, include_subfolders, log, progress, status=None):
    def say(message):
        if status:
            status(message)

    if not os.path.isdir(source_folder):
        raise ValueError(f"Source folder not found: {source_folder}")

    files = collect_files(source_folder, include_subfolders, {os.path.abspath(dest_path)})
    if not files:
        raise ValueError(
            "No .xlsx / .xlsm / .xltx / .xltm / .xls / .xlsb / .csv files found "
            "in the source folder"
        )

    log(f"Found {len(files)} file(s).")
    progress(0, len(files))

    notes = []
    blocks = []            # (source file label, sheet name, column names, rows)
    column_order = []      # output column names, in first-seen order
    key_to_column = {}     # match key -> output column name
    used_names = {SOURCE_FILE_COLUMN: 1, SOURCE_SHEET_COLUMN: 1}
    total_rows = 0
    sheets_combined = 0
    hidden_sheets = 0

    for file_index, full_path in enumerate(files, start=1):
        label = (
            os.path.relpath(full_path, source_folder) if include_subfolders
            else os.path.basename(full_path)
        )
        say(f"Opening {label} ...")
        log(f"{label}")

        try:
            sheets = list(read_any(full_path, notes, log))
        except Exception as exc:
            message = f"{label}: could not be opened - {exc}"
            log(f"  ! {message}")
            notes.append(message)
            progress(file_index, len(files))
            continue

        if not sheets:
            log("  - no sheets read")

        for sheet_name, rows, hidden in sheets:
            if hidden:
                hidden_sheets += 1

            hidden_tag = " (hidden sheet, included)" if hidden else ""

            if not rows or all(row_is_blank(r) for r in rows):
                log(f"  - [{sheet_name}]{hidden_tag}: blank sheet, skipped")
                continue

            header_row, data_rows, chosen, candidates = split_header_and_data(rows)
            if header_row is None:
                message = (
                    f"{label} [{sheet_name}]: EmployeeName column not found in "
                    f"rows 1-{HEADER_SCAN_ROWS} - sheet skipped, no rows added"
                )
                log(f"  ! {message}")
                notes.append(message)
                continue

            if len(candidates) > 1:
                rivals = ", ".join(f"row {r} ({n} cells)" for r, n in candidates if r != chosen)
                log(f"    (header row {chosen} chosen over {rivals})")

            data_rows = [r for r in data_rows if not row_is_blank(r)]
            if not data_rows:
                log(f"  - [{sheet_name}]{hidden_tag}: header found but no data rows, skipped")
                continue

            names, keep, flags = build_column_names(header_row, data_rows)
            trimmed = [[row[i] if i < len(row) else None for i in keep] for row in data_rows]

            output_names = []
            new_columns = []
            for name, is_employee in zip(names, flags):
                if is_employee:
                    key, display = EMPLOYEE_COLUMN_KEY, EMPLOYEE_COLUMN_NAME
                else:
                    key, display = match_key(name), name
                if key not in key_to_column:
                    # Two different match keys must never share one output name,
                    # or rows would be written into the wrong column.
                    if display in used_names:
                        used_names[display] += 1
                        display = f"{display} ({used_names[display]})"
                    used_names[display] = 1
                    key_to_column[key] = display
                    column_order.append(display)
                    new_columns.append(display)
                output_names.append(key_to_column[key])

            blocks.append((label, sheet_name, output_names, trimmed))
            total_rows += len(trimmed)
            sheets_combined += 1

            note = f" | {len(new_columns)} new column(s): {', '.join(new_columns)}" if new_columns else ""
            log(f"  - [{sheet_name}]{hidden_tag}: header on row {chosen}, "
                f"{len(trimmed)} row(s), {len(names)} column(s){note}")

        progress(file_index, len(files))

    if not blocks:
        notes.insert(0, "No sheet in any file produced usable data.")
        raise ValueError(
            "No usable sheets found - see the log. "
            "The most likely cause is that no sheet has an EmployeeName header "
            f"within its first {HEADER_SCAN_ROWS} rows."
        )

    headers = [SOURCE_FILE_COLUMN, SOURCE_SHEET_COLUMN] + column_order
    summary = [
        f"Files scanned: {len(files)}",
        f"Sheets combined: {sheets_combined}",
        f"Hidden / very-hidden sheets included: {hidden_sheets}",
        f"Data rows: {total_rows}",
        f"Data columns: {len(column_order)} (plus Source File and Source Sheet)",
    ]

    say("Writing output ...")
    written = write_output(headers, blocks, notes, summary, dest_path, total_rows, log, say)

    progress(len(files), len(files))
    say("Done.")
    return written, total_rows, len(column_order), notes


# ---------------------------------------------------------------------------
# Diagnose - why a given sheet picked the header row it picked
# ---------------------------------------------------------------------------

DIAGNOSE_CELL_WIDTH = 22
DIAGNOSE_MAX_CELLS = 14


def diagnose(source_folder, include_subfolders, log, progress, status=None):
    """Print the top rows of every sheet, flagging the candidate header rows
    and the winner, without writing any output file. Everything stays in this
    window on this machine - nothing is saved or sent anywhere."""
    def say(message):
        if status:
            status(message)

    if not os.path.isdir(source_folder):
        raise ValueError(f"Source folder not found: {source_folder}")

    files = collect_files(source_folder, include_subfolders, set())
    if not files:
        raise ValueError("No supported files found in the source folder")

    progress(0, len(files))
    log(f"Header diagnostic - first {HEADER_SCAN_ROWS} rows of each sheet.")
    log("  >> = chosen header row   ?  = candidate that lost")
    log("")

    notes = []
    for file_index, full_path in enumerate(files, start=1):
        label = (
            os.path.relpath(full_path, source_folder) if include_subfolders
            else os.path.basename(full_path)
        )
        say(f"Scanning {label} ...")
        log(f"=== {label}")

        try:
            sheets = list(read_any(full_path, notes, log))
        except Exception as exc:
            log(f"  ! could not be opened - {exc}")
            progress(file_index, len(files))
            continue

        for sheet_name, rows, hidden in sheets:
            hidden_tag = " (hidden sheet)" if hidden else ""
            chosen = choose_header_row(rows)
            candidates = {row for row, _ in header_candidates(rows)}

            if chosen is None:
                log(f"  --- [{sheet_name}]{hidden_tag}: NO EmployeeName row in the "
                    f"first {HEADER_SCAN_ROWS} rows -> sheet would be skipped")
            else:
                log(f"  --- [{sheet_name}]{hidden_tag}: header row {chosen}")

            for index in range(min(len(rows), HEADER_SCAN_ROWS)):
                row = rows[index]
                number = index + 1
                filled = sum(1 for value in row if not is_blank(value))
                cells = [
                    str(value).strip()[:DIAGNOSE_CELL_WIDTH]
                    for value in row if not is_blank(value)
                ][:DIAGNOSE_MAX_CELLS]
                more = "" if filled <= DIAGNOSE_MAX_CELLS else f" ... (+{filled - DIAGNOSE_MAX_CELLS})"
                mark = ">>" if number == chosen else ("? " if number in candidates else "  ")
                log(f"    {mark} row {number:>2} [{filled:>3} cells] {' | '.join(cells)}{more}")
            log("")

        progress(file_index, len(files))

    say("Diagnostic complete.")
    log("Diagnostic complete - no files were written.")


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def note_lines(notes, summary):
    lines = ["NOTES"]
    lines.extend(summary)
    if notes:
        lines.append(f"Issues: {len(notes)}")
        lines.extend(notes)
    else:
        lines.append("Issues: none - every sheet had an EmployeeName header row.")
    return lines


def write_output(headers, blocks, notes, summary, dest_path, total_rows, log, say):
    """Write .xlsx when everything fits, otherwise numbered CSV parts."""
    lines = note_lines(notes, summary)
    needed_rows = 1 + total_rows + 1 + len(lines)   # header + data + spacer + notes
    force_csv = dest_path.lower().endswith(".csv")

    if not force_csv and needed_rows <= EXCEL_MAX_ROWS and len(headers) <= EXCEL_MAX_COLS:
        return [write_xlsx(headers, blocks, lines, dest_path, say)]

    if len(headers) > EXCEL_MAX_COLS:
        log(f"  ! {len(headers)} columns exceeds Excel's {EXCEL_MAX_COLS} column limit "
            "- writing CSV instead.")
    elif not force_csv:
        log(f"  ! {needed_rows:,} rows exceeds Excel's {EXCEL_MAX_ROWS:,} row limit "
            "- writing CSV parts instead.")
    return write_csv_parts(headers, blocks, lines, dest_path, total_rows, log, say)


def write_xlsx(headers, blocks, lines, dest_path, say):
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Combined")
    sheet.freeze_panes = "A2"

    header_cells = []
    for text in headers:
        cell = WriteOnlyCell(sheet, value=text)
        cell.font = Font(bold=True)
        header_cells.append(cell)
    sheet.append(header_cells)

    position = {name: index for index, name in enumerate(headers)}
    for label, sheet_name, names, rows in blocks:
        say(f"Writing {label} [{sheet_name}] ...")
        targets = [position[name] for name in names]
        for row in rows:
            out = [None] * len(headers)
            out[0] = label
            out[1] = sheet_name
            for target, value in zip(targets, row):
                out[target] = value
            sheet.append(out)

    sheet.append([])
    for index, text in enumerate(lines):
        cell = WriteOnlyCell(sheet, value=text)
        if index == 0:
            cell.font = Font(bold=True)
        sheet.append([cell])

    notes_sheet = workbook.create_sheet("Notes")
    for index, text in enumerate(lines):
        cell = WriteOnlyCell(notes_sheet, value=text)
        if index == 0:
            cell.font = Font(bold=True)
        notes_sheet.append([cell])

    say("Saving workbook ...")
    workbook.save(dest_path)
    return dest_path


def write_csv_parts(headers, blocks, lines, dest_path, total_rows, log, say):
    folder = os.path.dirname(os.path.abspath(dest_path))
    stem = os.path.splitext(os.path.basename(dest_path))[0]
    parts = max(1, -(-total_rows // CSV_ROWS_PER_PART))   # ceiling division

    position = {name: index for index, name in enumerate(headers)}
    written = []
    handle = None
    writer = None
    part = 0
    rows_in_part = CSV_ROWS_PER_PART   # forces the first part to open immediately

    def close_part():
        if handle:
            handle.close()

    try:
        for label, sheet_name, names, rows in blocks:
            say(f"Writing {label} [{sheet_name}] ...")
            targets = [position[name] for name in names]
            for row in rows:
                if rows_in_part >= CSV_ROWS_PER_PART:
                    close_part()
                    part += 1
                    path = os.path.join(folder, f"{stem} part {part} of {parts}.csv")
                    handle = open(path, "w", newline="", encoding="utf-8-sig")
                    writer = csv.writer(handle)
                    writer.writerow(headers)
                    written.append(path)
                    log(f"  - writing {os.path.basename(path)}")
                    rows_in_part = 0

                out = [None] * len(headers)
                out[0] = label
                out[1] = sheet_name
                for target, value in zip(targets, row):
                    out[target] = value
                writer.writerow([csv_text(v) for v in out])
                rows_in_part += 1

        if writer is None:   # no data rows at all
            path = os.path.join(folder, f"{stem} part 1 of 1.csv")
            handle = open(path, "w", newline="", encoding="utf-8-sig")
            writer = csv.writer(handle)
            writer.writerow(headers)
            written.append(path)

        writer.writerow([])
        for text in lines:
            writer.writerow([text])
    finally:
        close_part()

    return written


# ---------------------------------------------------------------------------
# Tkinter GUI
# ---------------------------------------------------------------------------

class CombinerApp:
    def __init__(self, root):
        self.root = root
        root.title("Excel Header-Aware Combiner")
        root.geometry("760x540")

        self.source_var = tk.StringVar()
        self.dest_var = tk.StringVar()
        self.subfolders_var = tk.BooleanVar(value=False)

        frame = tk.Frame(root)
        frame.pack(fill="x", padx=8, pady=8)

        tk.Label(frame, text="Source folder:").grid(row=0, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.source_var, width=68).grid(row=0, column=1, padx=6)
        tk.Button(frame, text="Browse...", command=self.browse_source).grid(row=0, column=2)

        tk.Label(frame, text="Output file:").grid(row=1, column=0, sticky="w", pady=(8, 0))
        tk.Entry(frame, textvariable=self.dest_var, width=68).grid(row=1, column=1, padx=6, pady=(8, 0))
        tk.Button(frame, text="Save As...", command=self.browse_dest).grid(row=1, column=2, pady=(8, 0))

        tk.Checkbutton(
            frame, text="Include subfolders", variable=self.subfolders_var
        ).grid(row=2, column=1, sticky="w", pady=(8, 0))

        tk.Label(
            root,
            text=("Header row = the first row within the first 30 rows containing "
                  "\"EmployeeName\". Hidden sheets, rows and columns are all included."),
            fg="#444", wraplength=730, justify="left",
        ).pack(fill="x", padx=8)

        buttons = tk.Frame(root)
        buttons.pack(pady=10)
        self.run_button = tk.Button(buttons, text="Run", width=20, command=self.on_run)
        self.run_button.pack(side="left", padx=4)
        self.diagnose_button = tk.Button(
            buttons, text="Diagnose headers", width=20, command=self.on_diagnose
        )
        self.diagnose_button.pack(side="left", padx=4)

        bar = tk.Frame(root)
        bar.pack(fill="x", padx=8)
        self.progress_bar = ttk.Progressbar(bar, mode="determinate")
        self.progress_bar.pack(fill="x", side="left", expand=True)
        self.progress_label = tk.Label(bar, text="0 / 0", width=10)
        self.progress_label.pack(side="left", padx=(8, 0))

        self.status_var = tk.StringVar(value="Idle")
        tk.Label(root, textvariable=self.status_var, anchor="w", fg="#444").pack(
            fill="x", padx=8, pady=(4, 0)
        )

        tk.Label(root, text="Log:").pack(anchor="w", padx=8, pady=(8, 0))
        self.log_box = scrolledtext.ScrolledText(root, height=18, state="disabled")
        self.log_box.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    def browse_source(self):
        path = filedialog.askdirectory(title="Select the folder holding the Excel files")
        if path:
            self.source_var.set(path)
            if not self.dest_var.get().strip():
                self.dest_var.set(os.path.join(path, DEFAULT_OUTPUT_NAME))

    def browse_dest(self):
        path = filedialog.asksaveasfilename(
            title="Save combined output as",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("CSV", "*.csv")],
            initialfile=DEFAULT_OUTPUT_NAME,
        )
        if path:
            self.dest_var.set(path)

    def log(self, message):
        def append():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", message + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.root.after(0, append)

    def progress(self, done, total):
        def update():
            self.progress_bar.configure(maximum=max(total, 1), value=done)
            self.progress_label.configure(text=f"{done} / {total}")
        self.root.after(0, update)

    def status(self, message):
        self.root.after(0, lambda: self.status_var.set(message))

    def _start(self, target, args):
        self.run_button.configure(state="disabled")
        self.diagnose_button.configure(state="disabled")
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self.progress(0, 0)
        self.status("Starting ...")
        threading.Thread(target=target, args=args, daemon=True).start()

    def _finish(self):
        self.run_button.configure(state="normal")
        self.diagnose_button.configure(state="normal")

    def on_run(self):
        source = self.source_var.get().strip()
        dest = self.dest_var.get().strip()
        if not source or not dest:
            messagebox.showerror("Missing path", "Pick both a source folder and an output file.")
            return
        self._start(self._worker, (source, dest, self.subfolders_var.get()))

    def on_diagnose(self):
        source = self.source_var.get().strip()
        if not source:
            messagebox.showerror("Missing path", "Pick a source folder first.")
            return
        self._start(self._diagnose_worker, (source, self.subfolders_var.get()))

    def _diagnose_worker(self, source, include_subfolders):
        try:
            diagnose(source, include_subfolders, self.log, self.progress, status=self.status)
        except Exception as exc:
            message = str(exc)
            self.status(f"Failed: {message}")
            self.log("ERROR: " + message)
            self.log(traceback.format_exc())
            self.root.after(0, lambda: messagebox.showerror("Error", message))
        finally:
            self.root.after(0, self._finish)

    def _worker(self, source, dest, include_subfolders):
        try:
            written, rows, columns, notes = combine(
                source, dest, include_subfolders,
                self.log, self.progress, status=self.status,
            )
            self.log("")
            self.log(f"Done. {rows:,} row(s), {columns} data column(s).")
            for path in written:
                self.log(f"  -> {path}")
            if notes:
                self.log(f"{len(notes)} note(s) written at the end of the table:")
                for note in notes:
                    self.log(f"  * {note}")

            files = "\n".join(written)
            extra = f"\n\n{len(notes)} note(s) - see the end of the table." if notes else ""
            self.root.after(0, lambda: messagebox.showinfo(
                "Done", f"{rows:,} row(s) combined.\n\nSaved to:\n{files}{extra}"
            ))
        except Exception as exc:
            message = str(exc)
            self.status(f"Failed: {message}")
            self.log("ERROR: " + message)
            self.log(traceback.format_exc())
            self.root.after(0, lambda: messagebox.showerror("Error", message))
        finally:
            self.root.after(0, self._finish)


def main():
    root = tk.Tk()
    CombinerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
